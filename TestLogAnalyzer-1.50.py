import sys
import os
import re
import fnmatch
import tempfile
import math
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import matplotlib.pyplot as plt
# 配置matplotlib中文字体支持
import matplotlib
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']  # 设置支持中文的字体
plt.rcParams['axes.unicode_minus'] = False  # 解决保存图像时负号'-'显示为方块的问题
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import time
import logging
import datetime

# 配置全局日志记录
class LogConfig:
    @staticmethod
    def _create_log_directory():
        logs_dir = os.path.join(os.getcwd(), 'logs')
        os.makedirs(logs_dir, exist_ok=True)
        return logs_dir
    
    @staticmethod
    def _get_log_filepath():
        logs_dir = LogConfig._create_log_directory()
        log_filename = datetime.datetime.now().strftime('TestLogAnalyzer_%Y%m%d_%H%M%S.log')
        return os.path.abspath(os.path.join(logs_dir, os.path.basename(log_filename)))
    
    @staticmethod
    def _create_handlers(log_filepath):
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        
        file_handler = logging.FileHandler(log_filepath, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        
        return file_handler, console_handler
    
    @staticmethod
    def setup_logger():
        log_filepath = LogConfig._get_log_filepath()
        
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)
        logger.handlers.clear()
        
        for handler in LogConfig._create_handlers(log_filepath):
            logger.addHandler(handler)
        
        logger.info(f"Log file created: {log_filepath}")
        return logger

# 初始化日志记录器
try:
    logger = LogConfig.setup_logger()
except (OSError, PermissionError, IOError) as e:
    print(f"Log file initialization failed: {str(e)}")
    logger = None
except Exception as e:
    print(f"Unexpected error during log initialization: {type(e).__name__}: {str(e)}")
    logger = None

# 配置matplotlib日志级别，抑制字体相关日志
logging.getLogger('matplotlib.font_manager').setLevel(logging.ERROR)
logging.getLogger('matplotlib').setLevel(logging.WARNING)

# 设置matplotlib中文字体
plt.rcParams["font.family"] = ["SimHei", "WenQuanYi Micro Hei", "Heiti TC", "Arial Unicode MS", "Microsoft YaHei"]
plt.rcParams["axes.unicode_minus"] = False  # 解决负号显示问题

class TestLogAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("pDOT Test Log Analyzer v1.50 -  Edwarlyu@20251017 ")
        # 初始化logger
        self.logger = logging.getLogger("TestLogAnalyzer")
        
        # 设置窗口图标
        try:
            self.root.iconbitmap("NEW7.ico")
        except (tk.TclError, FileNotFoundError):
            pass  # 如果没有图标文件，继续执行
        
        # 启动最大化
        try:
            self.root.state("zoomed")
        except:
            self.root.attributes("-zoomed", True)
        
        # 初始化选项卡样式和状态
        self.setup_tab_styles()
        self.tab_last_update = {}  # 存储每个选项卡最后更新时间
        self.tab_status = {}  # 存储每个选项卡的状态
        
        # 数据存储（保留空的变量定义）
        self.data_frames = {}
        self.selected_files = []
        self.processed_data = None
        self.spec_data = None  # 规格数据
        
        # 创建菜单栏
        self.menu_bar = tk.Menu(root)
        
        # 文件菜单
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Load Files", command=self.add_files)
        self.file_menu.add_command(label="Find File", command=self.find_file)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Clear All", command=self.clear_all)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.on_closing)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        
        # 数据处理菜单
        self.process_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.process_menu.add_command(label="Data Processing", command=self.data_processing_function)
        self.process_menu.add_command(label="Data Re-Processing", command=self.data_reprocessing_function)
        self.menu_bar.add_cascade(label="Data Processing", menu=self.process_menu)
        
        # 数据分析菜单
        self.analysis_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.analysis_menu.add_command(label="Yield Analysis", command=self.calculate_yield)
        self.analysis_menu.add_command(label="Top Defects", command=self.show_help)
        self.analysis_menu.add_command(label="Cpk", command=self.show_help)
        self.analysis_menu.add_command(label="Criteria", command=self.show_help)
        self.analysis_menu.add_command(label="Color Point Chart", command=self.show_help)
        self.menu_bar.add_cascade(label="Data Analysis", menu=self.analysis_menu)
        
        # 保存菜单
        self.save_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.save_menu.add_command(label="Save as CSV", command=self.save_processed_data)
        self.menu_bar.add_cascade(label="Save", menu=self.save_menu)
        
        # 帮助菜单
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.show_about)
        self.help_menu.add_command(label="Help", command=self.show_help)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)
        
        # 设置菜单栏到主窗口
        root.config(menu=self.menu_bar)
        
        # 初始化菜单项状态（禁用需要文件的菜单项）
        # 将在所有UI元素创建完成后调用update_menu_status方法
        
        # 创建主框架
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(fill="both", expand=True)
        
        # 文件列表区域
        self.file_frame = tk.LabelFrame(self.main_frame, text="File List", padx=5, pady=5)
        self.file_frame.pack(fill="x", padx=5, pady=5)
        
        # 文件操作按钮区域
        self.file_buttons_frame = tk.Frame(self.file_frame)
        self.file_buttons_frame.pack(side="top", fill="x", pady=(0, 5))
        
        # 添加Load Files按钮
        self.load_files_button = tk.Button(self.file_buttons_frame, text="Load Files", command=self.add_files)
        self.load_files_button.pack(side="left", padx=(0, 5))
        self.add_hover_effect(self.load_files_button)
        
        # 添加Find File按钮
        self.find_file_button = tk.Button(self.file_buttons_frame, text="Find File", command=self.find_file)
        self.find_file_button.pack(side="left", padx=(0, 5))
        self.add_hover_effect(self.find_file_button)
        
        # 添加Unload File按钮
        self.unload_file_button = tk.Button(self.file_buttons_frame, text="Unload File", command=self.unload_selected_files, state=tk.DISABLED)
        self.unload_file_button.pack(side="left", padx=(0, 5))
        self.add_hover_effect(self.unload_file_button)
        
        # 添加Select All按钮
        self.select_all_button = tk.Button(self.file_buttons_frame, text="Select All", command=self.select_all_files, state=tk.DISABLED)
        self.select_all_button.pack(side="left")
        self.add_hover_effect(self.select_all_button)
        
        # 文件列表显示区域
        self.file_list_frame = tk.Frame(self.file_frame)
        self.file_list_frame.pack(fill="both", expand=True)
        
        # 创建Canvas和Scrollbar用于文件列表
        self.file_list_canvas = tk.Canvas(self.file_list_frame, height=100)
        self.file_list_canvas.pack(side="left", fill="both", expand=True)
        
        self.file_list_scrollbar = tk.Scrollbar(self.file_list_frame, orient="vertical", command=self.file_list_canvas.yview)
        self.file_list_scrollbar.pack(side="right", fill="y")
        
        self.file_list_canvas.configure(yscrollcommand=self.file_list_scrollbar.set)
        
        # 创建文件列表容器框架
        self.file_list_container = tk.Frame(self.file_list_canvas)
        self.file_list_window = self.file_list_canvas.create_window((0, 0), window=self.file_list_container, anchor="nw", width=self.file_list_canvas.winfo_width())
        
        # 绑定事件，确保Canvas窗口正确调整宽度
        def on_configure(event):
            self.file_list_canvas.itemconfig(self.file_list_window, width=event.width)
        self.file_list_canvas.bind("<Configure>", on_configure)
        
        # 绑定事件，当内部容器大小改变时更新滚动区域
        def on_container_configure(event):
            self.file_list_canvas.configure(scrollregion=self.file_list_canvas.bbox("all"))
        self.file_list_container.bind("<Configure>", on_container_configure)
        
        # 文件选择变量
        self.file_vars = {}
        self.file_checks = {}
        self.file_labels = {}
        self.file_count = 0  # 用于文件编号计数
        
        # 数据处理区域
        self.process_frame = tk.LabelFrame(self.main_frame, text="Data Processing", padx=5, pady=5)
        self.process_frame.pack(fill="x", padx=5, pady=5)
        
        # 数据处理窗口保持存在，但内容已清空
        # 可以在这里添加一个占位标签或保留空白状态
        pass
        
        # 结果显示区域
        self.result_frame = tk.LabelFrame(self.main_frame, text="Analysis Results", padx=5, pady=5)
        self.result_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建按钮容器
        self.button_frame = tk.Frame(self.result_frame)
        self.button_frame.pack(side="top", anchor="nw", padx=5, pady=5, fill="x")
        
        # 统一按钮宽度
        button_width = 12
        
        # 在分析结果左侧添加文件预览按钮
        self.refresh_button = tk.Button(self.button_frame, text="File PreView", command=self.refresh_data, width=button_width)
        self.refresh_button.pack(side="left", padx=5)
        # 初始状态下禁用预览按钮
        self.refresh_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.refresh_button)
        
        # 在File PreView按钮后添加Review Criteria按钮
        self.review_criteria_button = tk.Button(self.button_frame, text="Read Criteria", command=self.review_criteria, width=button_width)
        self.review_criteria_button.pack(side="left", padx=5)
        # 初始状态下禁用按钮
        self.review_criteria_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.review_criteria_button)
        
        # 在Read Criteria按钮后添加Read ColorPoint Spec按钮
        self.read_colorpoint_spec_button = tk.Button(self.button_frame, text="ColorPointSpec", command=self.read_colorpoint_spec, width=button_width)
        self.read_colorpoint_spec_button.pack(side="left", padx=5)
        # 初始状态下禁用按钮
        self.read_colorpoint_spec_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.read_colorpoint_spec_button)
        
        # 添加数据处理按钮
        self.data_processing_button = tk.Button(self.button_frame, text="Data Processing", command=self.data_processing_function, width=button_width)
        self.data_processing_button.pack(side="left", padx=5)
        # 初始状态下禁用数据处理按钮
        self.data_processing_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.data_processing_button)
        
        # 添加不良率分析按钮
        self.yield_analysis_button = tk.Button(self.button_frame, text="Yield Analysis", command=self.yield_analysis, width=button_width)
        self.yield_analysis_button.pack(side="left", padx=5)
        # 初始状态下禁用不良率分析按钮
        self.yield_analysis_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.yield_analysis_button)
        
        # 添加Top 10按钮
        self.top10_button = tk.Button(self.button_frame, text="Top Defects", command=self.show_top10_tab, width=button_width)
        self.top10_button.pack(side="left", padx=5)
        # 初始状态下禁用Top 10按钮
        self.top10_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.top10_button)
        
        # 添加Cpk按钮
        self.cpk_button = tk.Button(self.button_frame, text="Cpk", command=self.show_cpk_tab, width=button_width)
        self.cpk_button.pack(side="left", padx=5)
        # 初始状态下禁用Cpk按钮
        self.cpk_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.cpk_button)
        

        
        # 添加Color Point Chart按钮
        self.color_point_chart_button = tk.Button(self.button_frame, text="Color Point Chart", command=self.show_color_point_chart, width=button_width)
        self.color_point_chart_button.pack(side="left", padx=5)
        # 初始状态下禁用按钮
        self.color_point_chart_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.color_point_chart_button)
        
        # 添加Clear Data按钮到右侧，设置淡黄色背景
        self.clear_button = tk.Button(self.button_frame, text="Clear Data", command=self.clear_all, width=button_width, bg="#FFFFCC")
        self.clear_button.pack(side="right", padx=5)
        # 初始状态下禁用清除按钮
        self.clear_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.clear_button)
        
        # Save as Excel按钮已移除
        
        # 添加数据保存按钮 - CSV格式
        self.temp_excel_button = tk.Button(self.button_frame, text="Save_as_Excel", command=self.save_processed_data_to_excel, width=button_width)
        self.temp_excel_button.pack(side="right", padx=5)
        # 初始状态下禁用临时存储excel按钮
        self.temp_excel_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.temp_excel_button)
        
        self.data_save_csv_button = tk.Button(self.button_frame, text="Save_as_CSV", command=self.save_processed_data, width=button_width)
        self.data_save_csv_button.pack(side="right", padx=5)
        # 初始状态下禁用保存按钮
        self.data_save_csv_button.config(state=tk.DISABLED)
        self.add_hover_effect(self.data_save_csv_button)
        
        # 初始状态下禁用按钮
        # 重置按钮状态
        self.color_point_chart_button.config(state=tk.DISABLED)
        
        # 创建选项卡控件
        self.tab_control = ttk.Notebook(self.result_frame)
        self.tab_control.pack(fill="both", expand=True)
        
        # 数据预览选项卡
        self.data_preview_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.data_preview_tab, text="Data Preview")
        
        # 在Data Preview选项卡后添加Review Criteria选项卡
        self.review_criteria_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.review_criteria_tab, text="Read Criteria")
        
        # 在Read Criteria选项卡后添加Read ColorPoint Spec选项卡
        self.colorpoint_spec_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.colorpoint_spec_tab, text="ColorPointSpec")
        
        # 创建数据预览表格
        self.create_data_preview_table()
        
        # 数据处理选项卡
        self.data_processing_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.data_processing_tab, text="Data Processing")
        
        # 数据重新处理选项卡
        self.data_reprocessing_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.data_reprocessing_tab, text="Data Re-Processing")
        
        # 不良率分析选项卡
        self.yield_analysis_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.yield_analysis_tab, text="Yield Analysis")
        
        # 创建不良率分析图表
        self.create_yield_analysis_chart()
        
        # Top 10选项卡
        self.top10_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.top10_tab, text="Top Defects")
        
        # 创建Top 10内容
        self.create_top10_content()
        
        # Cpk选项卡
        self.cpk_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.cpk_tab, text="Cpk")
        
        # 创建Cpk内容
        self.create_cpk_content()
        
        # 合并的Criteria选项卡

        
        # Color Point Chart选项卡
        self.color_point_chart_tab = ttk.Frame(self.tab_control)
        self.tab_control.add(self.color_point_chart_tab, text="Color Point Chart")
        
        # 创建Color Point Chart内容
        self.create_color_point_chart_content()
        
        # 绑定选项卡选择事件
        self.tab_control.bind("<<NotebookTabChanged>>", self.on_tab_selected)
        
        # 状态信息
        self.status_var = tk.StringVar(value="Ready")
        self.status_bar = tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor=tk.W)
        
        # 记录程序启动日志
        try:
            if logger:
                logger.info("Program started: pDOT Test Log Analyzer v1.50")
                logger.info(f"Working directory: {os.getcwd()}")
                logger.info(f"Python version: {sys.version}")
        except (OSError, AttributeError):
            pass  # 如果日志记录失败，不影响程序运行
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # 在所有UI元素创建完成后，统一初始化菜单项和按钮状态
        self.update_menu_status(has_files=False, has_selected_files=False, has_processed_data=False)
        
    def _setup_data_processing_tab(self):
        # 设置数据处理选项卡的布局"""
        # 清除数据处理选项卡中的所有控件
        for widget in self.data_processing_tab.winfo_children():
            widget.destroy()
        
        # 创建一个容器来放置Canvas和滚动条
        self.processing_preview_container = tk.Frame(self.data_processing_tab)
        self.processing_preview_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建Canvas和滚动条
        self.processing_preview_canvas = tk.Canvas(self.processing_preview_container)
        
        # 创建垂直滚动条
        self.processing_preview_vscroll = tk.Scrollbar(self.processing_preview_container, 
                                                      orient="vertical", 
                                                      command=self.processing_preview_canvas.yview)
        self.processing_preview_canvas.configure(yscrollcommand=self.processing_preview_vscroll.set)
        
        # 创建水平滚动条
        self.processing_preview_hscroll = tk.Scrollbar(self.processing_preview_container, 
                                                      orient="horizontal", 
                                                      command=self.processing_preview_canvas.xview)
        self.processing_preview_canvas.configure(xscrollcommand=self.processing_preview_hscroll.set)
        
        # 使用grid布局
        self.processing_preview_vscroll.grid(row=0, column=1, sticky="ns")
        self.processing_preview_hscroll.grid(row=1, column=0, sticky="ew")
        self.processing_preview_canvas.grid(row=0, column=0, sticky="nsew")
        
        # 设置容器的列和行权重
        self.processing_preview_container.grid_columnconfigure(0, weight=1)
        self.processing_preview_container.grid_rowconfigure(0, weight=1)
        
        # 创建表格容器
        self.processing_table_frame = tk.Frame(self.processing_preview_canvas)
        self.processing_preview_canvas_window = self.processing_preview_canvas.create_window(
            (0, 0), window=self.processing_table_frame, anchor="nw")
        
        # 绑定事件，确保Canvas窗口正确调整宽度和高度
        def on_configure(event):
            try:
                bbox = self.processing_preview_canvas.bbox("all")
                if bbox:
                    self.processing_preview_canvas.configure(scrollregion=bbox)
                else:
                    table_width = self.processing_table_frame.winfo_width()
                    table_height = self.processing_table_frame.winfo_height()
                    self.processing_preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
            except (tk.TclError, AttributeError):
                self.processing_preview_canvas.configure(scrollregion=(0, 0, 1000, 500))
        self.processing_table_frame.bind("<Configure>", on_configure)
        
        # 添加鼠标滚轮支持
        def on_mousewheel(event):
            if event.state & 0x0008:  # Shift键
                self.processing_preview_canvas.xview_scroll(-1 * (event.delta // 120), "units")
            else:
                self.processing_preview_canvas.yview_scroll(-1 * (event.delta // 120), "units")
        
        # 绑定鼠标滚轮事件
        self.processing_preview_canvas.bind("<MouseWheel>", on_mousewheel)
        self.processing_preview_canvas.bind("<Button-4>", lambda e: self.processing_preview_canvas.yview_scroll(-1, "units"))
        self.processing_preview_canvas.bind("<Button-5>", lambda e: self.processing_preview_canvas.yview_scroll(1, "units"))
        
        # 添加状态栏标签
        self.processing_status_var = tk.StringVar(value="Ready")
        self.processing_status_label = tk.Label(self.data_processing_tab, 
                                              textvariable=self.processing_status_var, 
                                              bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.processing_status_label.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)
    
    def _setup_data_reprocessing_tab(self):
        """设置数据重新处理选项卡的布局"""
        # 清除数据重新处理选项卡中的所有控件
        for widget in self.data_reprocessing_tab.winfo_children():
            widget.destroy()
        
        # 创建一个容器来放置Canvas和滚动条
        self.reprocessing_preview_container = tk.Frame(self.data_reprocessing_tab)
        self.reprocessing_preview_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建Canvas和滚动条
        self.reprocessing_preview_canvas = tk.Canvas(self.reprocessing_preview_container)
        
        # 创建垂直滚动条
        self.reprocessing_preview_vscroll = tk.Scrollbar(self.reprocessing_preview_container, 
                                                      orient="vertical", 
                                                      command=self.reprocessing_preview_canvas.yview)
        self.reprocessing_preview_canvas.configure(yscrollcommand=self.reprocessing_preview_vscroll.set)
        
        # 创建水平滚动条
        self.reprocessing_preview_hscroll = tk.Scrollbar(self.reprocessing_preview_container, 
                                                      orient="horizontal", 
                                                      command=self.reprocessing_preview_canvas.xview)
        self.reprocessing_preview_canvas.configure(xscrollcommand=self.reprocessing_preview_hscroll.set)
        
        # 使用grid布局
        self.reprocessing_preview_vscroll.grid(row=0, column=1, sticky="ns")
        self.reprocessing_preview_hscroll.grid(row=1, column=0, sticky="ew")
        self.reprocessing_preview_canvas.grid(row=0, column=0, sticky="nsew")
        
        # 设置容器的列和行权重
        self.reprocessing_preview_container.grid_columnconfigure(0, weight=1)
        self.reprocessing_preview_container.grid_rowconfigure(0, weight=1)
        
        # 创建表格容器
        self.reprocessing_table_frame = tk.Frame(self.reprocessing_preview_canvas)
        self.reprocessing_preview_canvas_window = self.reprocessing_preview_canvas.create_window(
            (0, 0), window=self.reprocessing_table_frame, anchor="nw")
        
        # 绑定事件，确保Canvas窗口正确调整宽度和高度
        def on_configure(event):
            try:
                bbox = self.reprocessing_preview_canvas.bbox("all")
                if bbox:
                    self.reprocessing_preview_canvas.configure(scrollregion=bbox)
                else:
                    table_width = self.reprocessing_table_frame.winfo_width()
                    table_height = self.reprocessing_table_frame.winfo_height()
                    self.reprocessing_preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
            except (tk.TclError, AttributeError):
                self.reprocessing_preview_canvas.configure(scrollregion=(0, 0, 1000, 500))
        self.reprocessing_table_frame.bind("<Configure>", on_configure)
        
        # 添加鼠标滚轮支持
        def on_mousewheel(event):
            if event.state & 0x0008:  # Shift键
                self.reprocessing_preview_canvas.xview_scroll(-1 * (event.delta // 120), "units")
            else:
                self.reprocessing_preview_canvas.yview_scroll(-1 * (event.delta // 120), "units")
        
        # 绑定鼠标滚轮事件
        self.reprocessing_preview_canvas.bind("<MouseWheel>", on_mousewheel)
        self.reprocessing_preview_canvas.bind("<Button-4>", lambda e: self.reprocessing_preview_canvas.yview_scroll(-1, "units"))
        self.reprocessing_preview_canvas.bind("<Button-5>", lambda e: self.reprocessing_preview_canvas.yview_scroll(1, "units"))
        
        # 添加状态栏标签
        self.reprocessing_status_var = tk.StringVar(value="Ready")
        self.reprocessing_status_label = tk.Label(self.data_reprocessing_tab, 
                                              textvariable=self.reprocessing_status_var, 
                                              anchor="w", font=("SimHei", 9))
        self.reprocessing_status_label.pack(side="bottom", fill="x", padx=10, pady=5)
        
        # 添加数据重新处理按钮
        button_frame = tk.Frame(self.data_reprocessing_tab)
        button_frame.pack(fill="x", padx=10, pady=5)
        
        self.reprocess_data_button = tk.Button(button_frame, text="Data_ReProssing", 
                                          command=self.data_reprocessing_function,
                                          bg="#4CAF50", fg="white", 
                                          activebackground="#45a049", font=("SimHei", 10, "bold"))
        self.reprocess_data_button.pack(side="left", padx=5, pady=5)
        
        # 添加清空按钮
        self.clear_reprocessing_button = tk.Button(button_frame, text="Clear_ReProcessing", 
                                             command=lambda: self._clear_reprocessing_tab(),
                                             bg="#f44336", fg="white", 
                                             activebackground="#d32f2f", font=("SimHei", 10, "bold"))
        self.clear_reprocessing_button.pack(side="right", padx=5, pady=5)
    
    def _clear_reprocessing_tab(self):
        """清空数据重新处理选项卡的内容"""
        try:
            # 清除表格容器中的所有控件
            for widget in self.reprocessing_table_frame.winfo_children():
                widget.destroy()
            
            # 更新状态栏
            self.reprocessing_status_var.set("Ready")
        except Exception as e:
            self.reprocessing_status_var.set(f"清空出错: {str(e)}")
            logging.error(f"清空数据重新处理选项卡时出错: {str(e)}", exc_info=True)
    
    def _calculate_name_similarity(self, name1, name2):
        """计算两个名称之间的相似度分数（0-1之间）
        
        使用基于关键词的相似度算法，提高列名匹配的准确性
        """
        try:
            # 转换为小写字符串
            name1_str = str(name1).lower()
            name2_str = str(name2).lower()
            
            # 预定义的精确匹配映射表
            predefined_mapping = {
                "white l": ["white l (cd/m^2)"],
                "white u": ["white u (%)"],
                "white dy": ["white dy (%/cm)"],
                "white ru": ["white ru"],
                "white rv": ["white rv"],
                "white du": ["white du"],
                "white dv": ["white dv"],
                "white dl*min": ["white dl*min (%/cm)"],
                "white dl*max": ["white dl*max (%/cm)"],
                "white demax": ["white demax (%/cm)"],
                "white metric2": ["w_m2_gradient"],
                "white metric3": ["w_m3_mediumrangeuniformity"],
                "white metric7": ["w_m7_yellowpatches"],
                "white metric13": ["w_m13_globaldelta"],
                "white metric14": ["w_m14_shortedgebightline"],
                "white metric15": ["w_m15_longedgehotspot"],
                "white metric16": ["w_m16_longedgebightline"],
                "mixed l": ["mixed l (cd/m^2)"],
                "mixed u": ["mixed u (%)"],
                "mixed dy": ["mixed dy (%/cm)"],
                "mixed ru": ["mixed ru"],
                "mixed rv": ["mixed rv"],
                "mixed du": ["mixed du"],
                "mixed dv": ["mixed dv"],
                "mixed dl*min": ["mixed dl*min (%/cm)"],
                "mixed dl*max": ["mixed dl*max (%/cm)"],
                "mixed demax": ["mixed demax (%/cm)"],
                "mixed metric2": ["m_m2_gradient"],
                "mixed metric3": ["m_m3_mediumrangeuniformity"],
                "mixed metric7": ["m_m7_yellowpatches"],
                "mixed metric13": ["m_m13_globaldelta"],
                "mixed metric14": ["m_m14_shortedgebightline"],
                "mixed metric15": ["m_m15_longedgehotspot"],
                "mixed metric16": ["m_m16_longedgebightline"]
            }
            
            # 检查是否是预定义映射中的匹配
            if name1_str in predefined_mapping:
                if name2_str in predefined_mapping[name1_str]:
                    similarity_score = 1.0
                    # 记录匹配结果到日志
                    try:
                        if logger:
                            logger.debug(f"预定义精确匹配: '{name1}' 与 '{name2}' 匹配成功")
                    except (OSError, AttributeError):
                        pass
                    return similarity_score
            
            # 反向检查映射
            for key, values in predefined_mapping.items():
                if name1_str in values and name2_str == key:
                    similarity_score = 1.0
                    # 记录匹配结果到日志
                    try:
                        if logger:
                            logger.debug(f"预定义精确匹配: '{name1}' 与 '{name2}' 匹配成功")
                    except (OSError, AttributeError):
                        pass
                    return similarity_score
            
            # 转换为小写并分词
            words1 = set(name1_str.split())
            words2 = set(name2_str.split())
            
            # 如果有共同单词，至少有一定相似度
            if words1 & words2:
                # 使用Jaccard相似度
                intersection = len(words1 & words2)
                union = len(words1 | words2)
                similarity_score = intersection / union if union > 0 else 0
            else:
                # 检查子字符串匹配
                if name1_str in name2_str or name2_str in name1_str:
                    similarity_score = 0.8  # 较高的相似度分数
                else:
                    similarity_score = 0
            
            # 记录匹配结果到日志
            try:
                if logger:
                    logger.debug(f"名称匹配: '{name1}' 与 '{name2}' 相似度分数 = {similarity_score:.4f}")
            except (OSError, AttributeError):
                pass
                
            return similarity_score
        except Exception as e:
            try:
                if logger:
                    logger.error(f"名称相似度计算出错: name1='{name1}', name2='{name2}', 错误: {str(e)}")
            except (OSError, AttributeError):
                pass
            return 0
    
    def show_reprocessing_result(self, df, message=None, criteria_dict=None):
        """在数据重新处理选项卡中显示处理结果，风格与Data Processing选项卡保持一致
        
        参数:
            df: 要显示的数据
            message: 可选的状态消息
            criteria_dict: 规格标准字典，用于实际数值超限检查
        """
        try:
            # 将重新处理的数据存储在类变量中，供Top Defect功能使用
            self.reprocessed_data = df.copy()
            
            # 清除表格内容
            for widget in self.reprocessing_table_frame.winfo_children():
                widget.destroy()
            
            # 更新状态栏信息
            if message:
                self.reprocessing_status_var.set(message)
            
            # 如果没有数据，显示空信息
            if df is None or df.empty:
                empty_label = tk.Label(self.reprocessing_table_frame, 
                                     text=message or "No data available to display", 
                                     font=('Arial', 12), fg="gray")
                empty_label.grid(row=0, column=0, padx=50, pady=50)
                return
            
            # 更新选项卡状态
            self.update_tab_status("Data Re-Processing")
            
            # 限制显示的行数，最多显示17行
            max_display_rows = 17
            display_df = df.head(max_display_rows)
            
            # 创建行号表头
            row_num_header = tk.Label(self.reprocessing_table_frame, 
                                     text="Row#.", 
                                     font=('Courier New', 10, 'bold'), 
                                     relief=tk.RAISED, 
                                     padx=5, 
                                     pady=2, 
                                     anchor="w", 
                                     width=6)
            row_num_header.grid(row=0, column=0, sticky="nsew", ipady=0)
            self.reprocessing_table_frame.columnconfigure(0, minsize=60, weight=0)
            
            # 创建数据列表头
            for j, column in enumerate(display_df.columns):
                header = tk.Label(self.reprocessing_table_frame, 
                                text=f"{j+1}. {column}", 
                                font=('Courier New', 10, 'bold'), 
                                relief=tk.RAISED, 
                                padx=5, 
                                pady=2, 
                                anchor="w", 
                                wraplength=0, 
                                justify='left')
                header.grid(row=0, column=j+1, sticky="nsew", ipady=0)
                self.reprocessing_table_frame.columnconfigure(j+1, minsize=140, weight=1)
            
            # 不再重置format_cells字典，它在data_reprocessing_function中已处理好所有行的数据
            
            # 创建行数据（包含行号）
            for i, (index, row) in enumerate(display_df.iterrows(), 1):
                # 首先检查该行是否包含Fail值
                is_fail_row = False
                for j, value in enumerate(row):
                    if 'Pass/Fail' in str(display_df.columns[j]) and str(value).upper() == 'FAIL':
                        is_fail_row = True
                        break
                        
                # 创建行号单元格
                # 如果是Fail行，则行号单元格也设置为淡蓝色背景
                row_num_bg = "#e6f2ff" if is_fail_row else "white"
                row_num_cell = tk.Label(self.reprocessing_table_frame, 
                                      text=str(i), 
                                      font=('Courier New', 10), 
                                      borderwidth=1, 
                                      relief=tk.SUNKEN, 
                                      padx=5, 
                                      pady=2, 
                                      anchor="w", 
                                      width=6,
                                      background=row_num_bg)
                row_num_cell.grid(row=i, column=0, sticky="nsew", ipady=0)
                
                # 创建数据单元格
                for j, value in enumerate(row):
                    # 检查单元格是否需要特殊背景色
                    is_pass_fail_cell = 'Pass/Fail' in str(display_df.columns[j])
                    is_fail_cell = is_pass_fail_cell and str(value).upper() == 'FAIL'
                    
                    # 如果是Fail行，默认使用淡蓝色背景，否则使用白色背景
                    bg_color = "#e6f2ff" if is_fail_row else "white"
                    
                    # 使用data_reprocessing_function中预先处理好的format_cells字典判断单元格是否需要淡黄色填充
                    # 注意：index是原始数据中的索引，而i是显示的行号（从1开始）
                    if index in self.format_cells and j in self.format_cells[index]:
                        bg_color = "#ffffcc"  # 设置淡黄色背景
                    # 对于Pass/Fail列中的FAIL单元格，也设置淡黄色背景
                    elif is_fail_cell:
                        bg_color = "#ffffcc"  # 设置淡黄色背景
                    
                    # 格式化单元格值
                    cell_value = str(value)
                    if len(cell_value) > 100:
                        cell_value = cell_value[:97] + "..."
                    
                    # 创建单元格标签
                    cell = tk.Label(self.reprocessing_table_frame, 
                                  text=cell_value, 
                                  font=('Courier New', 10), 
                                  borderwidth=1, 
                                  relief=tk.SUNKEN, 
                                  padx=5, 
                                  pady=2, 
                                  anchor="w", 
                                  wraplength=0, 
                                  justify='left',
                                  background=bg_color)
                    cell.grid(row=i, column=j+1, sticky="nsew", ipady=0)
            
            # 如果有更多行，显示提示
            if len(df) > max_display_rows:
                more_label = tk.Label(
                    self.reprocessing_table_frame, 
                    text=f"... and {len(df) - max_display_rows} more rows not displayed ...", 
                    font=('Arial', 10, 'italic'), 
                    fg="blue",
                    anchor="w"
                )
                more_label.grid(row=len(display_df) + 1, column=0, columnspan=len(display_df.columns) + 1, pady=5, sticky="w")
                
                # 更新状态栏，显示总行数和显示行数
                self.reprocessing_status_var.set(f"显示前{max_display_rows}行数据（共{len(df)}行）")
            
            
        except Exception as e:
            self.reprocessing_status_var.set(f"显示数据出错: {str(e)}")
            error_label = tk.Label(self.reprocessing_table_frame, text=f"显示数据时出错: {str(e)}", font=("SimHei", 10), fg="red")
            error_label.pack(padx=20, pady=20)
            logging.error(f"显示数据重新处理结果时出错: {str(e)}", exc_info=True)
    
    def is_point_in_polygon(self, point, polygon):
        """
        优化的Ray Casting算法判断点是否在多边形内部
        使用预计算和快速路径优化提高性能
        
        参数:
            point: 待判断的点(x, y)
            polygon: 多边形顶点列表 [(x1, y1), (x2, y2), ...]
            
        返回:
            bool: 如果点在多边形内部，返回True，否则返回False
        """
        x, y = point
        n = len(polygon)
        
        # 快速边界检查 - 如果点在多边形边界框外，直接返回False
        if n < 3:
            return False
            
        # 预计算边界框
        min_x = min(p[0] for p in polygon)
        max_x = max(p[0] for p in polygon)
        min_y = min(p[1] for p in polygon)
        max_y = max(p[1] for p in polygon)
        
        # 边界框检查
        if x < min_x or x > max_x or y < min_y or y > max_y:
            return False
        
        inside = False
        j = n - 1  # 上一个顶点索引
        
        # 优化的射线投射算法
        for i in range(n):
            xi, yi = polygon[i]
            xj, yj = polygon[j]
            
            # 快速跳过不相关的边
            if (yi > y) == (yj > y):
                j = i
                continue
                
            # 计算交点 - 避免除零
            dy = yj - yi
            if abs(dy) < 1e-10:  # 水平边
                j = i
                continue
                
            # 计算射线与边的交点x坐标
            x_intersect = xi + (y - yi) * (xj - xi) / dy
            
            # 检查交点是否在射线上
            if x < x_intersect:
                inside = not inside
                
            j = i
        
        return inside
    
    def data_reprocessing_function(self):
        """
        数据重新处理函数
        
        功能说明:
        - 按照Review Criteria中的标准对每个数据点进行评估
        - 实现精确的数据质量分析和可视化标记
        - 使用从ColorPointSpec选项卡提取的多边形顶点坐标进行区域判定
        
        处理流程:
        1. 读取数据
        2. 读取规格
        3. 匹配规格和数据
        4. 根据超限情况设置单元格填充色和Pass/Fail值
        
        多边形坐标来源:
        - CAFL0 (White) 多边形: 使用从ColorPointSpec提取的精确6顶点坐标
        - CAFL24 (Mixed) 多边形: 使用从ColorPointSpec提取的精确4顶点坐标
        
        更新说明:
        坐标数据于2025-10-28从ColorPointSpec选项卡提取，确保了多边形区域判定的准确性和完整性。
        """
        try:
            import pandas as pd  # 确保导入pandas
            import tkinter as tk  # 确保导入tkinter
            import time
            import logging
            import math  # 确保导入math模块
            global logger
            
            # 添加详细日志记录
            if logger:
                logger.info("====== 开始数据重新处理流程 ======")
            
            logging.info("开始数据重新处理流程")
            # 更新状态栏
            if hasattr(self, 'reprocessing_status_var'):
                self.reprocessing_status_var.set("准备重新处理数据...")
            
            # 1. 读取data processing数据
            if not hasattr(self, 'processed_data') or self.processed_data is None or self.processed_data.empty:
                if hasattr(self, 'reprocessing_status_var'):
                    self.reprocessing_status_var.set("没有已处理的数据可供重新处理")
                messagebox.showinfo("信息", "没有已处理的数据可供重新处理，请先处理数据。")
                return
            
            # 自动保存Review Criteria设置
            if hasattr(self, '_save_criteria_to_temp_file'):
                if hasattr(self, 'reprocessing_status_var'):
                    self.reprocessing_status_var.set("自动保存Review Criteria设置...")
                self._save_criteria_to_temp_file()
                if logger:
                    logger.info("Review Criteria设置已自动保存")
                if hasattr(self, 'reprocessing_status_var'):
                    self.reprocessing_status_var.set("Review Criteria设置保存完成")
                time.sleep(0.3)  # 短暂延迟
            
            # 创建进度窗口
            progress_window = tk.Toplevel(self.root)  # 创建独立的顶级窗口
            progress_window.title("数据重新处理进度")
            progress_window.geometry("400x100")
            progress_window.resizable(False, False)
            # 计算位置，使其在屏幕中央
            x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 200
            y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
            progress_window.geometry(f"400x100+{x}+{y}")
            progress_window.grab_set()  # 模态窗口
            
            # 创建进度条
            progress_var = tk.DoubleVar()
            progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
            progress_bar.pack(pady=20, padx=20, fill="x")
            
            # 创建状态标签
            status_var = tk.StringVar(value="开始重新处理数据...")
            status_label = tk.Label(progress_window, textvariable=status_var)
            status_label.pack(padx=20)
            
            progress_window.update_idletasks()
            
            # 更新进度条
            progress_var.set(10)
            status_var.set("复制数据进行重新处理...")
            progress_window.update_idletasks()
            
            # 创建数据副本进行重新处理
            reprocess_df = self.processed_data.copy()
            
            # 更新进度条
            progress_var.set(30)
            status_var.set("获取Review Criteria标准...")
            progress_window.update_idletasks()
            
            # 2. 从临时规格文件读取规格数据
            criteria_dict = {}
            try:
                # 使用_get_criteria_dict方法从临时文件读取规格数据
                if hasattr(self, '_get_criteria_dict'):
                    if logger:
                        logger.info("尝试使用_get_criteria_dict从临时规格文件读取规格数据")
                    criteria_dict = self._get_criteria_dict()
                    if logger:
                        logger.info(f"成功从临时规格文件获取 {len(criteria_dict) if criteria_dict else 0} 条规格数据")
                
                # 记录规格数据
                if logger:
                    logger.info(f"规格数据获取完成，共 {len(criteria_dict) if criteria_dict else 0} 条标准")
                    if criteria_dict:
                        logger.debug("获取到的规格数据:")
                        for k, v in list(criteria_dict.items())[:10]:  # 只记录前10条避免日志过长
                            logger.debug(f"  {k}: {v}")
                        if len(criteria_dict) > 10:
                            logger.debug(f"  ... 等 {len(criteria_dict) - 10} 条标准")
            except Exception as criteria_error:
                if logger:
                    logger.error(f"读取临时规格文件规格数据时出错: {str(criteria_error)}", exc_info=True)
                criteria_dict = {}
            
            # 如果没有标准数据，显示提示
            if not criteria_dict:
                progress_window.destroy()
                if hasattr(self, 'reprocessing_status_var'):
                    self.reprocessing_status_var.set("没有找到Review Criteria标准数据")
                messagebox.showinfo("提示", "请先设置Review Criteria")
                return
            
            # 更新进度条
            progress_var.set(50)
            status_var.set("基于Review Criteria重新判断数据...")
            progress_window.update_idletasks()
            
            # 优化：只加载一次ColorPointSpec数据，避免每行重复加载文件
            self.logger.info("尝试加载ColorPointSpec_Current.json文件以获取最新的多边形数据")
            cafl0_polygon = []
            cafl24_polygon = []
            
            try:
                # 只加载一次ColorPointSpec配置
                self._load_colorpoint_spec_from_temp_file()
                
                # 检查colorpoint_spec_data是否存在且格式正确
                if hasattr(self, 'colorpoint_spec_data') and isinstance(self.colorpoint_spec_data, dict):
                    # 获取White多边形数据
                    white_data = None
                    # 首先检查直接的White字段
                    if 'White' in self.colorpoint_spec_data:
                        if isinstance(self.colorpoint_spec_data['White'], dict) and 'coordinates' in self.colorpoint_spec_data['White']:
                            white_data = self.colorpoint_spec_data['White']['coordinates']
                        elif isinstance(self.colorpoint_spec_data['White'], list):
                            white_data = self.colorpoint_spec_data['White']
                    # 然后检查可能的嵌套data结构
                    elif 'data' in self.colorpoint_spec_data and isinstance(self.colorpoint_spec_data['data'], dict):
                        if 'White' in self.colorpoint_spec_data['data'] and isinstance(self.colorpoint_spec_data['data']['White'], dict):
                            if 'coordinates' in self.colorpoint_spec_data['data']['White']:
                                white_data = self.colorpoint_spec_data['data']['White']['coordinates']
                        elif 'White' in self.colorpoint_spec_data['data'] and isinstance(self.colorpoint_spec_data['data']['White'], list):
                            white_data = self.colorpoint_spec_data['data']['White']
                    
                    # 处理获取到的白色区域数据
                    if white_data:
                        cafl0_polygon = [p for p in white_data if isinstance(p, (list, tuple)) and len(p) >= 2 and p[0] is not None and p[1] is not None]
                        if cafl0_polygon:
                            self.logger.info(f"成功从ColorPointSpec_Current.json文件加载了{len(cafl0_polygon)}个有效的White多边形顶点")
                        else:
                            self.logger.warning("从ColorPointSpec_Current.json加载的White多边形数据中没有有效的坐标点")
                    else:
                        self.logger.warning("ColorPointSpec_Current.json数据中未找到有效的'White'多边形配置")
                    
                    # 获取Mixed多边形数据
                    mixed_data = None
                    # 首先检查直接的Mixed字段
                    if 'Mixed' in self.colorpoint_spec_data:
                        if isinstance(self.colorpoint_spec_data['Mixed'], dict) and 'coordinates' in self.colorpoint_spec_data['Mixed']:
                            mixed_data = self.colorpoint_spec_data['Mixed']['coordinates']
                        elif isinstance(self.colorpoint_spec_data['Mixed'], list):
                            mixed_data = self.colorpoint_spec_data['Mixed']
                    # 然后检查可能的嵌套data结构
                    elif 'data' in self.colorpoint_spec_data and isinstance(self.colorpoint_spec_data['data'], dict):
                        if 'Mixed' in self.colorpoint_spec_data['data'] and isinstance(self.colorpoint_spec_data['data']['Mixed'], dict):
                            if 'coordinates' in self.colorpoint_spec_data['data']['Mixed']:
                                mixed_data = self.colorpoint_spec_data['data']['Mixed']['coordinates']
                        elif 'Mixed' in self.colorpoint_spec_data['data'] and isinstance(self.colorpoint_spec_data['data']['Mixed'], list):
                            mixed_data = self.colorpoint_spec_data['data']['Mixed']
                    
                    # 处理获取到的混合色区域数据
                    if mixed_data:
                        cafl24_polygon = [p for p in mixed_data if isinstance(p, (list, tuple)) and len(p) >= 2 and p[0] is not None and p[1] is not None]
                        if cafl24_polygon:
                            self.logger.info(f"成功从ColorPointSpec_Current.json文件加载了{len(cafl24_polygon)}个有效的Mixed多边形顶点")
                        else:
                            self.logger.warning("从ColorPointSpec_Current.json加载的Mixed多边形数据中没有有效的坐标点")
                    else:
                        self.logger.warning("ColorPointSpec_Current.json数据中未找到有效的'Mixed'多边形配置")
                else:
                    self.logger.warning("colorpoint_spec_data不存在或格式不正确，可能是文件加载失败")
            except Exception as load_error:
                self.logger.error(f"加载ColorPointSpec_Current.json文件时发生错误: {str(load_error)}")
            
            # 如果没有有效的CAFL0多边形数据，使用默认的白色区域
            if not cafl0_polygon or len(cafl0_polygon) < 3:
                cafl0_polygon = [
                    (0.183000, 0.461000),  # 点1
                    (0.193500, 0.455000),  # 点2
                    (0.196780, 0.461590),  # 点3
                    (0.203090, 0.458130),  # 点4
                    (0.209100, 0.469180),  # 点5
                    (0.189200, 0.478400)  # 点6
                ]
                self.logger.info(f"使用默认的White多边形({len(cafl0_polygon)}个点)进行CAFL0检查")
            else:
                self.logger.info(f"使用从ColorPointSpec_Current.json文件加载的White多边形({len(cafl0_polygon)}个点)进行CAFL0检查")
            
            # 如果没有有效的CAFL24多边形数据，使用默认的混合色区域
            if not cafl24_polygon or len(cafl24_polygon) < 3:
                cafl24_polygon = [
                    (0.242700, 0.513900),  # 点1
                    (0.222600, 0.522200),  # 点2
                    (0.234200, 0.539600),  # 点3
                    (0.253900, 0.532500)  # 点4
                ]
                self.logger.info(f"使用默认的Mixed多边形({len(cafl24_polygon)}个点)进行CAFL24检查")
            else:
                self.logger.info(f"使用从ColorPointSpec_Current.json文件加载的Mixed多边形({len(cafl24_polygon)}个点)进行CAFL24检查")
            
            # 初始化Pass/Fail和Fail_Reason列
            if 'Pass/Fail' not in reprocess_df.columns:
                reprocess_df['Pass/Fail'] = 'Pass'  # 默认设为Pass
            else:
                reprocess_df['Pass/Fail'] = 'Pass'  # 默认设为Pass
                
            if 'Fail_Reason' not in reprocess_df.columns:
                reprocess_df['Fail_Reason'] = ''
            else:
                reprocess_df['Fail_Reason'] = ''
            
            # 保存原始format_cells字典
            # 注意：不再为'Avg'列继承数据处理阶段的填色逻辑，仅当坐标点不在多边形区域内时才添加淡黄色填充
            original_format_cells = getattr(self, 'format_cells', {}).copy()
            
            # 重置format_cells字典，准备新的格式设置
            self.format_cells = {}
            
            # 记录处理结果统计
            fail_count = 0
            total_count = len(reprocess_df)
            
            # 3. 匹配规格和数据，遍历所有单元格
            for idx, record in reprocess_df.iterrows():
                # 更新进度条
                progress_var.set(50 + (idx + 1) / total_count * 40)
                progress_window.update_idletasks()
                
                row_has_fail = False
                row_failed_criteria = []
                matched_columns_set = set()  # 用于跟踪当前行已匹配的列
                
                try:
                    # 遍历所有标准
                    for std_type, limits in criteria_dict.items():
                        # 确保limits是包含两个元素的元组或列表
                        try:
                            # 如果是列表或元组，取前两个元素作为上下限
                            if isinstance(limits, (list, tuple)) and len(limits) >= 2:
                                lower_str, upper_str = limits[0], limits[1]
                            else:
                                continue
                        except:
                            continue
                            
                        # 获取标准类型信息
                        std_info = self._standardize_criteria_type(std_type)
                        
                        # 查找匹配的列
                        column_name = None
                        if hasattr(self, '_get_best_matching_column'):
                            column_name = self._get_best_matching_column(std_info, record.index)
                            if column_name is None:
                                column_name = self._get_best_matching_column(std_type, record.index)
                        
                        # 如果找到匹配列
                        if column_name is not None and column_name in reprocess_df.columns:
                            matched_columns_set.add(column_name)
                            
                            # 获取记录值
                            value = record[column_name]
                            
                            # 跳过空值
                            if pd.isna(value):
                                continue
                            
                            # 尝试转换为数值
                            try:
                                value_float = float(value)
                            except:
                                continue
                            
                            # 判断是否超限
                            is_out_of_range = False
                            fail_description = None
                            
                            # math模块已在文件顶部导入
                            
                            # 检查下限
                            if lower_str:
                                try:
                                    lower_limit = float(lower_str)
                                    # 使用近似比较，但保持严格小于的判断逻辑
                                    # 如果value_float明显小于lower_limit，则认为超出范围
                                    if not math.isclose(value_float, lower_limit) and value_float < lower_limit:
                                        is_out_of_range = True
                                        fail_description = column_name  # 仅保留列名（不良项目名称）
                                except:
                                    pass
                            
                            # 检查上限
                            if upper_str and not is_out_of_range:
                                try:
                                    upper_limit = float(upper_str)
                                    # 使用近似比较，但保持严格大于的判断逻辑
                                    # 如果value_float明显大于upper_limit，则认为超出范围
                                    if not math.isclose(value_float, upper_limit) and value_float > upper_limit:
                                        is_out_of_range = True
                                        fail_description = column_name  # 仅保留列名（不良项目名称）
                                except:
                                    pass
                            
                            # 3.1 如果超限，设置单元格填充色为淡黄色，并标记行为Fail
                            if is_out_of_range:
                                row_has_fail = True
                                col_index = reprocess_df.columns.get_loc(column_name)
                                # 初始化该行的格式列表
                                if idx not in self.format_cells:
                                    self.format_cells[idx] = []
                                # 记录需要淡黄色填充的单元格位置
                                if col_index not in self.format_cells[idx]:
                                    self.format_cells[idx].append(col_index)
                                
                                # 记录失败原因（保持现有逻辑）
                                if fail_description:
                                    row_failed_criteria.append(fail_description)
                    
                    # 标准逻辑检查完成，继续进行多边形区域检测
                    
                    # 3.2 添加CAFL0和CAFL24多边形区域判定逻辑
                    # 检查"White u Avg"和"White v Avg"列是否存在
                    if 'White u Avg' in reprocess_df.columns and 'White v Avg' in reprocess_df.columns:
                        try:
                            # 获取当前行的White u/v值
                            white_u = record['White u Avg']
                            white_v = record['White v Avg']
                            
                            # 检查是否为有效数值
                            if pd.notna(white_u) and pd.notna(white_v):
                                white_u = float(white_u)
                                white_v = float(white_v)
                                
                                # 优化版：直接从ColorPointSpec_Current.json文件获取CAFL0多边形顶点坐标
                                # 优化：直接使用循环外已加载的cafl0_polygon数据
                                self.logger.info("使用循环外预加载的White多边形数据进行CAFL0检查")
                                # 记录使用的多边形数据来源，明确标识是从ColorPointSpec_Current.json加载还是使用默认值
                                if cafl0_polygon and len(cafl0_polygon) >= 3 and len(cafl0_polygon) != 6:
                                    # 如果有3个以上的点且不是默认的6个点，则认为是从文件加载的
                                    self.logger.info(f"使用从ColorPointSpec_Current.json文件加载的White多边形({len(cafl0_polygon)}个点)进行CAFL0检查")
                                    # 更新状态栏，提供实时反馈
                                    if hasattr(self, 'status_bar'):
                                        self.status_bar.config(text=f"使用最新的White多边形配置进行CAFL0检查")
                                else:
                                    # 使用默认多边形时的日志
                                    self.logger.info(f"使用默认的White多边形({len(cafl0_polygon)}个点)进行CAFL0检查")
                                    # 更新状态栏，提示用户使用了默认配置
                                    if hasattr(self, 'status_bar'):
                                        self.status_bar.config(text=f"未找到有效的White多边形配置，使用默认值进行CAFL0检查")
                            
                            # 判断点是否在CAFL0多边形内
                            try:
                                polygon_valid = cafl0_polygon and len(cafl0_polygon) >= 3
                                
                                if polygon_valid and not self.is_point_in_polygon((white_u, white_v), cafl0_polygon):
                                    # 点不在多边形内，需要标记单元格
                                    row_has_fail = True
                                    
                                    # 标记White u Avg列
                                    white_u_col_idx = reprocess_df.columns.get_loc('White u Avg')
                                    if idx not in self.format_cells:
                                        self.format_cells[idx] = []
                                    if white_u_col_idx not in self.format_cells[idx]:
                                        self.format_cells[idx].append(white_u_col_idx)
                                    
                                    # 标记White v Avg列
                                    white_v_col_idx = reprocess_df.columns.get_loc('White v Avg')
                                    if white_v_col_idx not in self.format_cells[idx]:
                                        self.format_cells[idx].append(white_v_col_idx)
                                    
                                    # 添加失败原因 - 仅保留不良项目名称
                                    fail_reason = "White u Avg; White v Avg"  # White点坐标不良
                                    if fail_reason not in row_failed_criteria:
                                        row_failed_criteria.append(fail_reason)
                                elif not polygon_valid:
                                    self.logger.warning(f"无效的CAFL0多边形数据，跳过对记录{idx}的White区域检查")
                            except Exception as polygon_error:
                                self.logger.error(f"执行CAFL0多边形区域检查时出错: {str(polygon_error)}")
                                # 出错时不影响整体处理流程，继续下一步
                        except Exception as e:
                            # 处理异常，记录日志
                            if logger:
                                logger.error(f"处理White u/v Avg数据时出错: {str(e)}")
                    
                    # 检查"Mixed u Avg"和"Mixed v Avg"列是否存在
                    if 'Mixed u Avg' in reprocess_df.columns and 'Mixed v Avg' in reprocess_df.columns:
                        try:
                            # 获取当前行的Mixed u/v值
                            mixed_u = record['Mixed u Avg']
                            mixed_v = record['Mixed v Avg']
                            
                            # 检查是否为有效数值
                            if pd.notna(mixed_u) and pd.notna(mixed_v):
                                mixed_u = float(mixed_u)
                                mixed_v = float(mixed_v)
                                
                                # 优化：直接使用循环外已加载的cafl24_polygon数据
                                self.logger.info("使用循环外预加载的Mixed多边形数据进行CAFL24检查")
                                # 记录使用的多边形数据来源，明确标识是从ColorPointSpec_Current.json加载还是使用默认值
                                if cafl24_polygon and len(cafl24_polygon) >= 3:
                                    self.logger.info(f"使用从ColorPointSpec_Current.json文件加载的Mixed多边形({len(cafl24_polygon)}个点)进行CAFL24检查")
                                    # 更新状态栏，提供实时反馈
                                    if hasattr(self, 'status_bar'):
                                        self.status_bar.config(text=f"使用最新的Mixed多边形配置进行CAFL24检查")
                                else:
                                    self.logger.info(f"使用默认的Mixed多边形({len(cafl24_polygon)}个点)进行CAFL24检查")
                                    # 更新状态栏，提示用户使用了默认配置
                                    if hasattr(self, 'status_bar'):
                                        self.status_bar.config(text=f"未找到有效的Mixed多边形配置，使用默认值进行CAFL24检查")
                                
                                # 判断点是否在CAFL24多边形内
                                try:
                                    polygon_valid = cafl24_polygon and len(cafl24_polygon) >= 3
                                    
                                    if polygon_valid and not self.is_point_in_polygon((mixed_u, mixed_v), cafl24_polygon):
                                        # 点不在多边形内，需要标记单元格
                                        row_has_fail = True
                                        
                                        # 标记Mixed u Avg列
                                        mixed_u_col_idx = reprocess_df.columns.get_loc('Mixed u Avg')
                                        if idx not in self.format_cells:
                                            self.format_cells[idx] = []
                                        if mixed_u_col_idx not in self.format_cells[idx]:
                                            self.format_cells[idx].append(mixed_u_col_idx)
                                        
                                        # 标记Mixed v Avg列
                                        mixed_v_col_idx = reprocess_df.columns.get_loc('Mixed v Avg')
                                        if mixed_v_col_idx not in self.format_cells[idx]:
                                            self.format_cells[idx].append(mixed_v_col_idx)
                                        
                                        # 添加失败原因 - 仅保留不良项目名称
                                        fail_reason = "Mixed u Avg; Mixed v Avg"  # Mixed点坐标不良
                                        if fail_reason not in row_failed_criteria:
                                            row_failed_criteria.append(fail_reason)
                                    elif not polygon_valid:
                                        self.logger.warning(f"无效的CAFL24多边形数据，跳过对记录{idx}的Mixed区域检查")
                                except Exception as polygon_error:
                                    self.logger.error(f"执行CAFL24多边形区域检查时出错: {str(polygon_error)}")
                                    # 出错时不影响整体处理流程，继续下一步
                        except Exception as e:
                            # 处理异常，记录日志
                            if logger:
                                logger.error(f"处理Mixed u/v Avg数据时出错: {str(e)}")
                    
                    # 3.3 其它未匹配的单元格，保持白色填充（不做任何处理，默认就是白色）
                    
                    # 在所有检测完成后，根据row_has_fail状态更新Pass/Fail列
                    reprocess_df.at[idx, 'Pass/Fail'] = 'Fail' if row_has_fail else 'Pass'
                    
                    # 更新失败计数
                    if row_has_fail:
                        if reprocess_df.at[idx, 'Pass/Fail'] == 'Fail' and idx not in [i for i, r in reprocess_df.iterrows() if r['Pass/Fail'] == 'Fail' and i < idx]:
                            fail_count += 1
                    
                    # 更新失败原因列 - 仅当有失败原因时才写入，否则保持为空
                    if row_failed_criteria:
                        reprocess_df.at[idx, 'Fail_Reason'] = '; '.join(row_failed_criteria)
                    # 无不良情况时，Fail_Reason列保持为空（不写入任何内容）
                    
                    # 记录Pass/Fail列的位置，用于show_reprocessing_result函数中特殊处理
                    for col_idx, col_name in enumerate(reprocess_df.columns):
                        if 'Pass/Fail' in str(col_name):
                            # 只有当值为'Fail'时才记录单元格位置，确保仅Fail值单元格被着色
                            if str(reprocess_df.at[idx, col_name]).strip().lower() == 'fail':
                                if idx not in self.format_cells:
                                    self.format_cells[idx] = []
                                if col_idx not in self.format_cells[idx]:
                                    self.format_cells[idx].append(col_idx)
                                
                except Exception as eval_error:
                    # 处理单个记录评估失败的情况
                    logging.error(f"评估记录 {idx} 时出错: {str(eval_error)}")
                    reprocess_df.at[idx, 'Pass/Fail'] = 'Fail'
                    # 评估失败时仅标记项目名称，不写入详细错误信息
                    reprocess_df.at[idx, 'Fail_Reason'] = '数据评估'  # 仅保留不良项目名称
                    fail_count += 1
            
            # 更新进度条
            progress_var.set(95)
            status_var.set("处理完成，准备显示结果...")
            progress_window.update_idletasks()
            
            # 延迟以确保用户看到完成状态
            time.sleep(0.3)
            
            # 更新进度条
            progress_var.set(100)
            status_var.set("处理完成，准备显示结果...")
            progress_window.update_idletasks()
            
            # 关闭进度窗口
            progress_window.destroy()
            
            # 显示处理结果统计
            pass_count = total_count - fail_count
            if hasattr(self, 'reprocessing_status_var'):
                self.reprocessing_status_var.set(f"Review Criteria判断完成 - 通过: {pass_count}, 失败: {fail_count}")
            
            # 保存重新处理后的数据
            self.reprocessed_data = reprocess_df
            
            # 显示重新处理后的结果
            self.show_reprocessing_result(reprocess_df)
            
            # 切换到数据重新处理选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'data_reprocessing_tab'):
                self.tab_control.select(self.data_reprocessing_tab)
                
        except Exception as e:
            # 显示错误信息
            messagebox.showerror("错误", f"数据重新处理失败: {str(e)}")
            if logger:
                logger.error(f"数据重新处理失败: {str(e)}", exc_info=True)
            
        finally:
            # 更新状态栏
            if hasattr(self, 'reprocessing_status_var'):
                self.reprocessing_status_var.set("数据处理完成")
            
            # 详细的日志记录
            if logger:
                logger.info(f"====== 数据重新处理完成 ======")
                try:
                    if 'reprocess_df' in locals():
                        total_count = len(reprocess_df)
                        pass_count = len(reprocess_df[reprocess_df['Pass/Fail'] == 'Pass']) if 'Pass/Fail' in reprocess_df.columns else total_count
                        fail_count = total_count - pass_count
                        logger.info(f"总记录数: {total_count}")
                        logger.info(f"通过记录数: {pass_count}")
                        logger.info(f"未通过记录数: {fail_count}")
                        
                        # 记录使用的规格数据来源
                        if hasattr(self, '_white_criteria_file') and hasattr(self, '_mixed_criteria_file'):
                            logger.info(f"使用临时文件处理规格数据:")
                            logger.info(f"  White窗口规格文件: {self._white_criteria_file}")
                            logger.info(f"  Mixed窗口规格文件: {self._mixed_criteria_file}")
                except Exception:
                    pass
    
    def show_processing_result(self, df, message=None, is_reprocessing=False):
        """在数据处理选项卡中显示处理结果
        
        参数:
            df: 要显示的数据
            message: 可选的状态消息
            is_reprocessing: 是否为重新处理数据模式
        """
        # 清除表格内容
        for widget in self.processing_table_frame.winfo_children():
            widget.destroy()
        
        # 更新状态栏信息
        if message:
            self.processing_status_var.set(message)
        
        # 如果没有数据，显示空信息
        if df is None or df.empty:
            empty_label = tk.Label(self.processing_table_frame, 
                                 text=message or "No data available to display", 
                                 font=('Arial', 12), fg="gray")
            empty_label.grid(row=0, column=0, padx=50, pady=50)
            return
        
        # 更新选项卡状态
        self.update_tab_status("Data Processing")
        
        # 限制显示的行数，最多显示50行
        max_display_rows = 15
        display_df = df.head(max_display_rows)
        
        # 创建行号表头
        row_num_header = tk.Label(self.processing_table_frame, 
                                 text="Row#.", 
                                 font=('Courier New', 10, 'bold'), 
                                 relief=tk.RAISED, 
                                 padx=5, 
                                 pady=2, 
                                 anchor="w", 
                                 width=6)
        row_num_header.grid(row=0, column=0, sticky="nsew", ipady=0)
        self.processing_table_frame.columnconfigure(0, minsize=60, weight=0)
        
        # 创建数据列表头
        for j, column in enumerate(display_df.columns):
            header = tk.Label(self.processing_table_frame, 
                            text=f"{j+1}. {column}", 
                            font=('Courier New', 10, 'bold'), 
                            relief=tk.RAISED, 
                            padx=5, 
                            pady=2, 
                            anchor="w", 
                            wraplength=0, 
                            justify='left')
            header.grid(row=0, column=j+1, sticky="nsew", ipady=0)
            self.processing_table_frame.columnconfigure(j+1, minsize=140, weight=1)
        
        # 创建行数据（包含行号）
        for i, (index, row) in enumerate(display_df.iterrows(), 1):
            # 创建行号单元格
            row_num_cell = tk.Label(self.processing_table_frame, 
                                  text=str(i), 
                                  font=('Courier New', 10), 
                                  borderwidth=1, 
                                  relief=tk.SUNKEN, 
                                  padx=5, 
                                  pady=2, 
                                  anchor="w", 
                                  width=6)
            row_num_cell.grid(row=i, column=0, sticky="nsew", ipady=0)
            
            # 检查是否是Fail行
            is_fail_row = False
            fail_column_index = -1
            for j, (column, value) in enumerate(zip(display_df.columns, row)):
                if 'Pass/Fail' in str(column) and str(value).upper() == 'FAIL':
                    is_fail_row = True
                    fail_column_index = j
                    break
            
            # 创建数据单元格
            for j, value in enumerate(row):
                # 格式化单元格值
                cell_value = str(value)
                if len(cell_value) > 100:
                    cell_value = cell_value[:97] + "..."
                
                # 设置单元格背景色
                bg_color = "white"  # 默认使用白色而不是空字符串
                
                # 优先检查是否是Fail单元格
                if j == fail_column_index and is_fail_row:
                    # Fail单元格使用淡黄色背景
                    bg_color = "#ffffcc"
                # 其次检查是否是hash标记的单元格
                elif not is_reprocessing and hasattr(self, 'hash_cells') and i-1 in self.hash_cells and j in self.hash_cells[i-1]:
                    # 如果是Data Processing模式且单元格曾包含#字符，则添加淡黄色背景
                    bg_color = "#ffffcc"
                # 最后检查是否是Fail行的其他单元格
                elif is_fail_row and j != fail_column_index:
                    # Fail行的其他单元格使用淡蓝色背景
                    bg_color = "#e6f2ff"
                
                # 创建单元格标签
                cell = tk.Label(self.processing_table_frame, 
                              text=cell_value, 
                              font=('Courier New', 10), 
                              borderwidth=1, 
                              relief=tk.SUNKEN, 
                              padx=5, 
                              pady=2, 
                              anchor="w", 
                              wraplength=0, 
                              justify='left',
                              background=bg_color)
                cell.grid(row=i, column=j+1, sticky="nsew", ipady=0)
        
        # 如果有更多行，显示提示
        if len(df) > max_display_rows:
            more_rows_label = tk.Label(self.processing_table_frame, 
                                     text=f"... and {len(df) - max_display_rows} more rows not displayed ...", 
                                     font=('Arial', 10, 'italic'), 
                                     fg="blue",
                                     anchor="w")
            more_rows_label.grid(row=len(display_df) + 1, column=0, columnspan=len(display_df.columns) + 1, pady=5, sticky="w")
    
    def setup_tab_styles(self):
        """设置选项卡样式和按钮悬浮效果"""
        style = ttk.Style()
        # 创建淡蓝色样式（数据更新时）
        style.configure("Tab.NewData.TNotebook.Tab", background="#e6f2ff")
        # 创建淡绿色样式（选中时）
        style.configure("Tab.Selected.TNotebook.Tab", background="#e6ffe6")
        # 默认样式
        style.configure("Tab.Default.TNotebook.Tab", background="")
        
        # 创建自定义按钮样式，实现鼠标悬浮效果
        # 定义普通状态下的按钮样式
        style.configure("HoverButton.TButton",
                       background="#d9d9d9",
                       foreground="#000000",
                       relief="raised",
                       padding=5)
        # 定义悬浮状态下的按钮样式（淡蓝色）
        style.map("HoverButton.TButton",
                  background=[("active", "#cce5ff")],
                  foreground=[("active", "#000000")])
    
    def add_hover_effect(self, button):
        """为单个tk.Button添加鼠标悬浮效果"""
        def on_enter(event):
            # 保存原始背景色
            if not hasattr(event.widget, 'original_bg'):
                event.widget.original_bg = event.widget['bg'] if event.widget['bg'] != 'SystemButtonFace' else '#d9d9d9'
            # 悬浮时改变背景色为淡蓝色
            event.widget.config(bg='#cce5ff')
            
        def on_leave(event):
            # 离开时恢复原始背景色
            if hasattr(event.widget, 'original_bg'):
                event.widget.config(bg=event.widget.original_bg)
        
        # 绑定鼠标事件
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)

    def on_tab_selected(self, event):
        """选项卡选中事件处理"""
        notebook = event.widget
        tab_idx = notebook.index(notebook.select())
        
        # 移除选项卡样式设置，避免不支持style属性的错误
        # 只记录选中状态，不修改样式
        pass
    
    def update_tab_status(self, tab_name, has_data=True):
        """更新选项卡状态"""
        # 记录更新时间
        if has_data:
            self.tab_last_update[tab_name] = time.time()
            
            # 移除选项卡样式设置，避免不支持style属性的错误
            # 只记录更新状态，不修改样式
            pass
    
    def update_status(self, message):
        """更新状态栏信息并记录日志"""
        timestamped_message = f"[{time.strftime('%H:%M:%S')}] {message}"
        self.status_var.set(timestamped_message)
        
        # 记录到日志文件
        try:
            if logger:
                # 根据消息内容判断日志级别
                if any(keyword in message.lower() for keyword in ['error', '失败', 'exception']):
                    logger.error(message)
                elif any(keyword in message.lower() for keyword in ['warning', '警告']):
                    logger.warning(message)
                else:
                    logger.info(message)
        except (OSError, AttributeError):
            pass  # 如果日志记录失败，不影响程序运行
            
        self.root.update_idletasks()
        
    def add_files(self):
        """添加CSV文件"""
        try:
            if logger:
                logger.info("Start adding CSV files")
        except (OSError, AttributeError):
            pass
            
        # 打开文件选择对话框，允许选择多个CSV文件
        file_paths = tk.filedialog.askopenfilenames(
            title="Add CSV Files",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        
        try:
            if logger:
                logger.info(f"Choose {len(file_paths)} CSV files")
        except (OSError, AttributeError):
            pass
            
        if not file_paths:
            return
        
        # 为新添加的文件创建复选框和标签
        first_file_path = None
        # 计算当前已有文件数量
        current_count = len(self.file_vars)
        
        for file_path in file_paths:
            # 只获取文件名（不含路径）
            file_name = os.path.basename(file_path)
            
            # 如果文件已存在，跳过
            if file_path in self.file_vars:
                continue
            
            # 记录第一个文件路径用于预览和保存目录
            if first_file_path is None:
                first_file_path = file_path
                # 保存为实例变量，供保存功能使用
                self.first_file_path = file_path
            
            # 增加文件计数
            self.file_count += 1
            
            # 创建变量存储复选框状态
            var = tk.BooleanVar(value=True)
            self.file_vars[file_path] = var
            
            # 计算网格位置（每行4个文件）
            total_files = len(self.file_vars)
            row = (total_files - 1) // 4
            col = (total_files - 1) % 4
            
            # 创建文件框架
            file_frame = tk.Frame(self.file_list_container)
            file_frame.grid(row=row, column=col, padx=5, pady=5, sticky="w")
            
            # 创建复选框
            check = tk.Checkbutton(file_frame, variable=var, 
                                  command=lambda p=file_path: self.on_checkbox_change(p))
            check.pack(side="left", padx=2)
            self.file_checks[file_path] = check
            
            # 创建编号标签
            number_label = tk.Label(file_frame, text=f"{self.file_count}.", width=4, anchor="w")
            number_label.pack(side="left")
            
            # 创建文件名标签
            label = tk.Label(file_frame, text=file_name, anchor="w")
            label.pack(side="left")
            self.file_labels[file_path] = label
            
            # 更新状态栏信息
            self.update_status(f"Successfully added file: {file_name}")
        
        # 重新配置滚动区域
        self.file_list_container.update_idletasks()
        self.file_list_canvas.configure(scrollregion=self.file_list_canvas.bbox("all"))
        
        # 保存选中文件列表，供Save as CSV功能使用
        self.selected_files = file_paths
        
        # 使用统一的方法更新所有菜单和按钮的状态
        self.update_menu_status(has_files=True, has_selected_files=True)
        
        # 不再自动触发任何其他按钮功能，等待用户手动操作
        
    def _post_file_loading_sequence(self):
        """文件加载后的自动化处理序列"""
        try:
            # 1. 切换到Data Process选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'data_processing_tab'):
                self.tab_control.select(self.data_processing_tab)
                self.update_status("切换到Data Processing选项卡")
                
            # 2. 执行Data Processing功能
            self.update_status("开始执行Data Processing功能")
            self.data_processing_function()
            
            # 3. 添加延迟，确保数据处理完成
            self.root.after(1000, self._switch_to_reprocessing)
            
        except Exception as e:
            self.update_status(f"自动化处理序列出错: {str(e)}")
            
    def _switch_to_reprocessing(self):
        """切换到重新处理数据选项卡并执行重新处理功能，然后继续执行后续分析功能"""
        try:
            # 1. 切换到Data Re-Processing选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'data_reprocessing_tab'):
                self.tab_control.select(self.data_reprocessing_tab)
                self.update_status("切换到Data Re-Processing选项卡")
                
            # 2. 执行重新处理数据功能
            self.update_status("开始执行重新处理数据功能")
            self.data_reprocessing_function()
            
            # 3. 添加延迟，确保重新处理完成后继续执行
            self.root.after(1000, self._execute_yield_analysis)
            
        except Exception as e:
            self.update_status(f"重新处理数据序列出错: {str(e)}")
    
    def _execute_yield_analysis(self):
        """执行Yield Analysis功能"""
        try:
            # 切换到Yield Analysis选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'yield_analysis_tab'):
                self.tab_control.select(self.yield_analysis_tab)
                self.update_status("切换到Yield Analysis选项卡")
            
            # 执行Yield Analysis功能
            self.update_status("开始执行Yield Analysis功能")
            self.yield_analysis()
            
            # 添加延迟后继续执行Top Defects功能
            self.root.after(1000, self._execute_top_defects)
            
        except Exception as e:
            self.update_status(f"执行Yield Analysis出错: {str(e)}")
            # 即使出错也要继续执行下一个功能
            self.root.after(1000, self._execute_top_defects)
    
    def _execute_top_defects(self):
        """执行Top Defects功能"""
        try:
            # 切换到Top Defects选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'top10_tab'):
                self.tab_control.select(self.top10_tab)
                self.update_status("切换到Top Defects选项卡")
            
            # 执行Top Defects功能
            self.update_status("开始执行Top Defects功能")
            self.show_top10_tab()
            
            # 添加延迟后继续执行Cpk功能
            self.root.after(1000, self._execute_cpk)
            
        except Exception as e:
            self.update_status(f"执行Top Defects出错: {str(e)}")
            # 即使出错也要继续执行下一个功能
            self.root.after(1000, self._execute_cpk)
    
    def _execute_cpk(self):
        """执行Cpk功能"""
        try:
            # 切换到Cpk选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'cpk_tab'):
                self.tab_control.select(self.cpk_tab)
                self.update_status("切换到Cpk选项卡")
            
            # 执行Cpk功能
            self.update_status("开始执行Cpk功能")
            self.show_cpk_tab()
            
            # 添加延迟后继续执行Color Point Chart功能
            self.root.after(1000, self._execute_color_point_chart)
            
        except Exception as e:
            self.update_status(f"执行Cpk出错: {str(e)}")
            # 即使出错也要继续执行下一个功能
            self.root.after(1000, self._execute_color_point_chart)
    
    def _execute_color_point_chart(self):
        """执行Color Point Chart功能"""
        try:
            # 切换到Color Point Chart选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'color_point_chart_tab'):
                self.tab_control.select(self.color_point_chart_tab)
                self.update_status("切换到Color Point Chart选项卡")
            
            # 执行Color Point Chart功能
            self.update_status("开始执行Color Point Chart功能")
            self.show_color_point_chart()
            
            # 所有功能执行完成
            self.update_status("自动化分析流程执行完成")
            
        except Exception as e:
            self.update_status(f"执行Color Point Chart出错: {str(e)}")
    
    def update_menu_status(self, has_files=False, has_selected_files=False, has_processed_data=False):
        """更新菜单和按钮的状态"""
        # 更新File菜单
        if hasattr(self, 'file_menu'):
            # Clear All菜单项的状态与Clear Data按钮保持一致
            if has_files:
                # 启用File Preview按钮
                if hasattr(self, 'refresh_button'):
                    self.refresh_button.config(state=tk.NORMAL)
                # 启用Clear Data按钮
                if hasattr(self, 'clear_button'):
                    self.clear_button.config(state=tk.NORMAL)
                # 启用Unload File按钮
                if hasattr(self, 'unload_file_button'):
                    self.unload_file_button.config(state=tk.NORMAL)
                # 启用Select All按钮
                if hasattr(self, 'select_all_button'):
                    self.select_all_button.config(state=tk.NORMAL)
                # 启用Review Criteria按钮
                if hasattr(self, 'review_criteria_button') and has_selected_files:
                    self.review_criteria_button.config(state=tk.NORMAL)
                # 启用ColorPointSpec按钮（只要有文件加载就启用）
                if hasattr(self, 'read_colorpoint_spec_button'):
                    self.read_colorpoint_spec_button.config(state=tk.NORMAL)
                # 找到Clear All菜单项并启用
                for i in range(self.file_menu.index('end') + 1):
                    try:
                        label = self.file_menu.entrycget(i, 'label')
                        if label == 'Clear All':
                            self.file_menu.entryconfig(i, state=tk.NORMAL)
                            break
                    except tk.TclError:
                        pass
            else:
                # 禁用File Preview按钮
                if hasattr(self, 'refresh_button'):
                    self.refresh_button.config(state=tk.DISABLED)
                # 禁用Clear Data按钮
                if hasattr(self, 'clear_button'):
                    self.clear_button.config(state=tk.DISABLED)
                # 禁用Unload File按钮
                if hasattr(self, 'unload_file_button'):
                    self.unload_file_button.config(state=tk.DISABLED)
                # 禁用Select All按钮
                if hasattr(self, 'select_all_button'):
                    self.select_all_button.config(state=tk.DISABLED)
                # 禁用Review Criteria按钮
                if hasattr(self, 'review_criteria_button'):
                    self.review_criteria_button.config(state=tk.DISABLED)
                # 禁用ColorPointSpec按钮
                if hasattr(self, 'read_colorpoint_spec_button'):
                    self.read_colorpoint_spec_button.config(state=tk.DISABLED)
                # 找到Clear All菜单项并禁用
                for i in range(self.file_menu.index('end') + 1):
                    try:
                        label = self.file_menu.entrycget(i, 'label')
                        if label == 'Clear All':
                            self.file_menu.entryconfig(i, state=tk.DISABLED)
                            break
                    except tk.TclError:
                        pass
            
        # 更新Data Processing菜单
        if hasattr(self, 'process_menu'):
            if has_files:
                self.data_processing_button.config(state=tk.NORMAL)
                # 启用/禁用菜单项
                for i in range(self.process_menu.index('end') + 1):
                    try:
                        label = self.process_menu.entrycget(i, 'label')
                        if label == 'Data Processing':
                            self.process_menu.entryconfig(i, state=tk.NORMAL)
                        elif label == 'Data Re-Processing':
                            # 只有在有已处理数据的情况下才启用Data Re-Processing
                            if has_processed_data:
                                self.process_menu.entryconfig(i, state=tk.NORMAL)
                            else:
                                self.process_menu.entryconfig(i, state=tk.DISABLED)
                    except tk.TclError:
                        pass
            else:
                self.data_processing_button.config(state=tk.DISABLED)
                # 禁用菜单项
                for i in range(self.process_menu.index('end') + 1):
                    try:
                        label = self.process_menu.entrycget(i, 'label')
                        if label == 'Data Processing' or label == 'Data Re-Processing':
                            self.process_menu.entryconfig(i, state=tk.DISABLED)
                    except:
                        pass
        
        # 更新Save_as_Excel按钮状态
        # 只有在有已处理数据的情况下才启用Save_as_Excel按钮
        if hasattr(self, 'temp_excel_button'):
            if has_processed_data:
                self.temp_excel_button.config(state=tk.NORMAL)
            else:
                self.temp_excel_button.config(state=tk.DISABLED)
        
        # 更新Data Analysis菜单
        if hasattr(self, 'analysis_menu'):
            # 设置所有分析菜单项和按钮的状态
            if has_processed_data:
                # 启用分析按钮
                if hasattr(self, 'yield_analysis_button'):
                    self.yield_analysis_button.config(state=tk.NORMAL)
                if hasattr(self, 'top10_button'):
                    self.top10_button.config(state=tk.NORMAL)
                if hasattr(self, 'cpk_button'):
                    self.cpk_button.config(state=tk.NORMAL)

                if hasattr(self, 'color_point_chart_button'):
                    self.color_point_chart_button.config(state=tk.NORMAL)
                
                # 启用分析菜单项
                for i in range(self.analysis_menu.index('end') + 1):
                    try:
                        self.analysis_menu.entryconfig(i, state=tk.NORMAL)
                    except tk.TclError:
                        pass
            else:
                # 禁用分析按钮
                if hasattr(self, 'yield_analysis_button'):
                    self.yield_analysis_button.config(state=tk.DISABLED)
                if hasattr(self, 'top10_button'):
                    self.top10_button.config(state=tk.DISABLED)
                if hasattr(self, 'cpk_button'):
                    self.cpk_button.config(state=tk.DISABLED)

                if hasattr(self, 'color_point_chart_button'):
                    self.color_point_chart_button.config(state=tk.DISABLED)
                
                # 禁用分析菜单项
                for i in range(self.analysis_menu.index('end') + 1):
                    try:
                        self.analysis_menu.entryconfig(i, state=tk.DISABLED)
                    except tk.TclError:
                        pass
        
        # 更新Save菜单
        if hasattr(self, 'save_menu'):
            # CSV保存功能可以在有选中文件时启用
            if has_selected_files:
                if hasattr(self, 'data_save_csv_button'):
                    self.data_save_csv_button.config(state=tk.NORMAL)
                if hasattr(self, 'temp_excel_button'):
                    self.temp_excel_button.config(state=tk.NORMAL)
                # 查找并启用Save as CSV菜单项
                for i in range(self.save_menu.index('end') + 1):
                    try:
                        self.save_menu.entryconfig(i, state=tk.NORMAL)
                    except:
                        pass
            else:
                if hasattr(self, 'data_save_csv_button'):
                    self.data_save_csv_button.config(state=tk.DISABLED)
                if hasattr(self, 'temp_excel_button'):
                    self.temp_excel_button.config(state=tk.DISABLED)
                # 查找并禁用Save as CSV菜单项
                for i in range(self.save_menu.index('end') + 1):
                    try:
                        self.save_menu.entryconfig(i, state=tk.DISABLED)
                    except:
                        pass
            
            # Excel保存功能相关代码已移除

    
    def on_checkbox_change(self, file_path):
        """复选框状态改变时的处理函数"""
        # 不再自动预览文件，等待用户点击刷新按钮
        # 检查是否还有选中的文件
        selected_files = [path for path, var in self.file_vars.items() if var.get()]
        # 无论是否有选中文件，都更新self.selected_file
        self.selected_files = selected_files
        
        # 更新Review Criteria按钮状态
        if hasattr(self, 'review_criteria_button'):
            if selected_files:
                self.review_criteria_button.config(state=tk.NORMAL)
            else:
                self.review_criteria_button.config(state=tk.DISABLED)
        
        # 更新Read ColorPoint Spec按钮状态
        if hasattr(self, 'read_colorpoint_spec_button'):
            if selected_files:
                self.read_colorpoint_spec_button.config(state=tk.NORMAL)
            else:
                self.read_colorpoint_spec_button.config(state=tk.DISABLED)
        
        # 更新菜单和按钮状态
        self.update_menu_status(has_files=bool(self.file_vars), has_selected_files=bool(selected_files))
        
    def refresh_data(self):
        """加载数据预览"""
        # 切换到数据预览选项卡
        if hasattr(self, 'tab_control') and hasattr(self, 'data_preview_tab'):
            self.tab_control.select(self.data_preview_tab)
        # 预览所有选中的文件
        self.preview_all_selected_files()
    
    def preview_all_selected_files(self):
        """预览所有选中的文件"""
        # 获取所有选中的文件
        selected_files = [path for path, var in self.file_vars.items() if var.get()]
        
        # 保存最新的选中文件列表，供Save as CSV功能使用
        self.selected_files = selected_files
        
        if not selected_files:
            self.update_status("No files selected")
            self.reset_data_preview_table()
            self.empty_table_label.config(text="No files selected")
            # 禁用Save as CSV按钮
            if hasattr(self, 'data_save_csv_button'):
                self.data_save_csv_button.config(state=tk.DISABLED)
            if hasattr(self, 'temp_excel_button'):
                self.temp_excel_button.config(state=tk.DISABLED)
            return
        
        # 收集所有文件的数据
        all_files_data = []
        for file_path in selected_files:
            file_name = os.path.basename(file_path)
            try:
                # 更新状态栏信息
                self.update_status(f"Processing file: {file_name}")
                
                # 智能检测CSV文件中的数据标题行
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = []
                    for i in range(20):
                        line = f.readline()
                        if not line:
                            break
                        lines.append(line.strip())
                    
                    # 识别元数据行（通常是键值对格式，如"键:,值"的形式）
                    metadata_lines = []
                    non_metadata_lines = []
                    
                    for i, line in enumerate(lines):
                        # 元数据行的特征: 通常包含键值对格式
                        # 1. 检查是否包含冒号加逗号的模式（如"键:,值"）
                        # 2. 检查行的开头是否包含典型的元数据键名
                        is_metadata = False
                        
                        # 检查标点符号模式
                        if ',:' in line or (':' in line and ',' in line):
                            is_metadata = True
                        
                        # 检查是否为典型的元数据行（有内容但不是标题行）
                        elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                            # 检查行是否有有意义的内容且不是明显的标题行
                            content_parts = [p.strip() for p in line.split(',') if p.strip()]
                            if content_parts and ':' in content_parts[0]:
                                is_metadata = True
                        
                        if is_metadata:
                            metadata_lines.append(i)
                        else:
                            non_metadata_lines.append((i, line))
                    
                    # 如果有非元数据行，在这些行中寻找包含最多逗号的非空行作为标题行
                    max_commas = -1
                    header_line_candidates = []
                    header_line_index = 0
                    
                    # 首先筛选出非空的非元数据行
                    non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                    
                    if non_empty_non_metadata:
                        # 找到所有非空非元数据行中的最大逗号数
                        max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                        # 获取所有具有最大逗号数的候选行
                        header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                        
                        if header_line_candidates:
                            # 增强选择逻辑: 优先选择包含字母字符、内容长度更长的行作为标题行
                            # 计算每个候选行的分数
                            best_score = -1
                            for i, line in header_line_candidates:
                                # 基础分数（保证至少有一个分数）
                                score = 1
                                # 包含字母字符加分（标题行通常包含字段名）
                                if any(c.isalpha() for c in line):
                                    score += 10
                                # 内容长度加分（不含逗号）
                                content_length = len(line.replace(',', ''))
                                score += min(content_length, 50)  # 内容长度最多加50分
                                
                                if score > best_score:
                                    best_score = score
                                    header_line_index = i
                    elif non_metadata_lines:
                        # 如果所有非元数据行都是空的，使用原始方法
                        for i, line in non_metadata_lines:
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    else:
                        # 如果没有非元数据行，使用原始方法寻找最多逗号的行
                        for i, line in enumerate(lines):
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    
                    # 再次检查: 确保找到的标题行不是全空的，且有合理数量的非空列
                    if header_line_index < len(lines) and lines[header_line_index].strip() == '':
                        self.update_status(f"Warning: Empty header line detected (Line {header_line_index+1}). Trying to find the next suitable line.")
                        # 寻找下一个非空行
                        for i in range(header_line_index + 1, min(len(lines), header_line_index + 10)):
                            if lines[i].strip() != '':
                                header_line_index = i
                                max_commas = lines[i].count(',')
                                break
                    
                    # 确定是否使用检测到的标题行
                    if header_line_index < len(lines) and max_commas >= 9:
                        # 如果找到明显的标题行（逗号数足够多），使用它
                        self.update_status(f"Data header line detected (Line {header_line_index+1}) with {max_commas+1} columns.")
                        df = pd.read_csv(file_path, 
                                        skiprows=header_line_index,  # 跳过说明性行
                                        on_bad_lines='skip',    # 跳过有问题的行
                                        engine='python')        # 使用Python解析器
                    else:
                        # 如果没有找到明显的标题行，尝试使用默认方式读取
                        self.update_status("No obvious data header line detected. Trying to read with default settings.")
                        df = pd.read_csv(file_path, 
                                        on_bad_lines='skip',    # 跳过有问题的行
                                        engine='python')        # 使用Python解析器
                
                # 记录文件名和数据
                all_files_data.append((file_path, file_name, df))
                
            except Exception as e:
                # 显示错误信息
                messagebox.showerror("Error", f"Failed to process file {file_name}: {str(e)}")
                self.update_status(f"Failed to process file: {file_name}")
        
        if all_files_data:
            # 更新状态栏信息
            self.update_status(f"Successfully processed {len(all_files_data)} files.")
            # 更新数据预览表格
            self.update_data_preview_table_for_multiple_files(all_files_data)
        else:
            self.update_status("No files were successfully processed.")
            self.reset_data_preview_table()
            self.empty_table_label.config(text="No files were successfully processed.")
    
    def preview_first_selected_file(self):
        """预览第一个选中的文件（保持向后兼容）"""
        # 获取所有选中的文件
        selected_files = [path for path, var in self.file_vars.items() if var.get()]
        
        if not selected_files:
            self.update_status("No files selected.")
            return
        
        # 预览第一个选中的文件
        first_file_path = selected_files[0]
        file_name = os.path.basename(first_file_path)
        
        try:
            # 更新状态栏信息
            self.update_status(f"Previewing file: {file_name}")
            
            # 智能检测CSV文件中的数据标题行
            with open(first_file_path, 'r', encoding='utf-8') as f:
                    lines = []
                    for i in range(20):
                        line = f.readline()
                        if not line:
                            break
                        lines.append(line.strip())
                    
                    # 识别元数据行（通常是键值对格式，如"键:,值"的形式）
                    metadata_lines = []
                    non_metadata_lines = []
                    
                    for i, line in enumerate(lines):
                        # 元数据行的特征: 通常包含键值对格式
                        # 1. 检查是否包含冒号加逗号的模式（如"键:,值"）
                        # 2. 检查行的开头是否包含典型的元数据键名
                        is_metadata = False
                        
                        # 检查标点符号模式
                        if ',:' in line or (':' in line and ',' in line):
                            is_metadata = True
                        
                        # 检查是否为典型的元数据行（有内容但不是标题行）
                        elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                            # 检查行是否有有意义的内容且不是明显的标题行
                            content_parts = [p.strip() for p in line.split(',') if p.strip()]
                            if content_parts and ':' in content_parts[0]:
                                is_metadata = True
                        
                        if is_metadata:
                            metadata_lines.append(i)
                        else:
                            non_metadata_lines.append((i, line))
                    
                    # 如果有非元数据行，在这些行中寻找包含最多逗号的非空行作为标题行
                    max_commas = -1
                    header_line_candidates = []
                    header_line_index = 0
                    
                    # 首先筛选出非空的非元数据行
                    non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                    
                    if non_empty_non_metadata:
                        # 找到所有非空非元数据行中的最大逗号数
                        max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                        # 获取所有具有最大逗号数的候选行
                        header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                        
                        if header_line_candidates:
                            # 增强选择逻辑: 优先选择包含字母字符、内容长度更长的行作为标题行
                            # 计算每个候选行的分数
                            best_score = -1
                            for i, line in header_line_candidates:
                                # 基础分数（保证至少有一个分数）
                                score = 1
                                # 包含字母字符加分（标题行通常包含字段名）
                                if any(c.isalpha() for c in line):
                                    score += 10
                                # 内容长度加分（不含逗号）
                                content_length = len(line.replace(',', ''))
                                score += min(content_length, 50)  # 内容长度最多加50分
                                
                                if score > best_score:
                                    best_score = score
                                    header_line_index = i
                    elif non_metadata_lines:
                        # 如果所有非元数据行都是空的，使用原始方法
                        for i, line in non_metadata_lines:
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    else:
                        # 如果没有非元数据行，使用原始方法寻找最多逗号的行
                        for i, line in enumerate(lines):
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    
                    # 再次检查: 确保找到的标题行不是全空的，且有合理数量的非空列
                    if header_line_index < len(lines) and lines[header_line_index].strip() == '':
                        self.update_status(f"Warning: Empty header line detected (Line {header_line_index+1}). Trying to find the next suitable line.")
                        # 寻找下一个非空行
                        for i in range(header_line_index + 1, min(len(lines), header_line_index + 10)):
                            if lines[i].strip() != '':
                                header_line_index = i
                                max_commas = lines[i].count(',')
                                break
                    
                    # 额外检查: 如果找到的标题行内容太短，考虑是否为真正的标题行
                    if header_line_index < len(lines) and len(lines[header_line_index].strip()) < 20 and max_commas >= 5:
                        self.update_status(f"Warning: Header line content is too short (Line {header_line_index+1}). Please confirm manually.")
                    
                    # 确定是否使用检测到的标题行
                    if max_commas >= 9:
                        # 如果找到明显的标题行（逗号数足够多），使用它
                        self.update_status(f"Data header line detected (Line {header_line_index+1}). Columns: {max_commas+1}")
                        self.update_status(f"Excluded metadata lines: {len(metadata_lines)}")
                        df = pd.read_csv(first_file_path,
                                        skiprows=header_line_index,  # 跳过说明性行
                                        on_bad_lines='skip',    # 跳过有问题的行
                                        engine='python')        # 使用Python解析器
                    else:
                        # 如果没有找到明显的标题行，尝试使用默认方式读取
                        self.update_status("No obvious data header line detected. Trying default read.")
                        df = pd.read_csv(first_file_path,
                                        on_bad_lines='skip',    # 跳过有问题的行
                                        engine='python')        # 使用Python解析器
            
            # 更新数据预览表格，传递当前文件路径和文件名
            self.update_data_preview_table(df, first_file_path, file_name)
            
            # 更新状态栏信息
            self.update_status(f"Previewed file: {file_name}")
        except Exception as e:
            # 显示错误信息
            messagebox.showerror("Error", f"Failed to preview file {file_name}: {str(e)}")
            self.update_status(f"Preview file failed: {file_name}")
        
    def find_file(self):
        """查找文件"""
        pass
        
    def remove_selected(self):
        """移除选中的文件"""
        pass
        
    def unload_selected_files(self):
        """卸载选中的文件"""
        if not hasattr(self, 'file_vars') or not self.file_vars:
            tk.messagebox.showinfo("Warning", "No files to unload!")
            return
        
        # 获取选中的文件路径
        selected_files = [path for path, var in self.file_vars.items() if var.get()]
        
        if not selected_files:
            tk.messagebox.showinfo("Warning", "Please select files to unload!")
            return
        
        # 删除选中的文件
        for file_path in selected_files:
            if file_path in self.file_vars:
                del self.file_vars[file_path]
            if file_path in self.file_checks:
                del self.file_checks[file_path]
            if file_path in self.file_labels:
                del self.file_labels[file_path]
            
        # 重新显示剩余文件
        for widget in self.file_list_container.winfo_children():
            widget.destroy()
        
        # 重新创建文件列表
        self.file_count = 0
        remaining_files = list(self.file_vars.keys())
        
        for i, file_path in enumerate(remaining_files):
            # 只获取文件名（不含路径）
            file_name = os.path.basename(file_path)
            
            # 增加文件计数
            self.file_count += 1
            
            # 计算网格位置（每行4个文件）
            row = (i) // 4
            col = (i) % 4
            
            # 创建文件框架
            file_frame = tk.Frame(self.file_list_container)
            file_frame.grid(row=row, column=col, padx=5, pady=5, sticky="w")
            
            # 创建复选框
            check = tk.Checkbutton(file_frame, variable=self.file_vars[file_path], command=lambda path=file_path: self.on_checkbox_change(path))
            check.pack(side="left", padx=2)
            self.file_checks[file_path] = check
            
            # 创建编号标签
            number_label = tk.Label(file_frame, text=f"{self.file_count}.", width=4, anchor="w")
            number_label.pack(side="left")
            
            # 创建文件名标签
            label = tk.Label(file_frame, text=file_name, anchor="w")
            label.pack(side="left")
            self.file_labels[file_path] = label
        
        # 重新配置滚动区域
        self.file_list_container.update_idletasks()
        self.file_list_canvas.configure(scrollregion=self.file_list_canvas.bbox("all"))
        
        # 清空选中文件列表
        if hasattr(self, 'selected_files'):
            self.selected_files = []
        
        # 检查是否还有剩余文件
        has_remaining_files = bool(self.file_vars)
        
        # 更新菜单状态
        self.update_menu_status(has_files=has_remaining_files, has_selected_files=False)
        
        # 更新状态栏信息
        self.status_var.set(f"Unloaded {len(selected_files)} files")
        
        # 记录日志
        logging.info(f"Unloaded {len(selected_files)} files")
        
        # 如果没有文件了，禁用相关按钮
        if not has_remaining_files:
            self.unload_file_button.config(state=tk.DISABLED)
            self.select_all_button.config(state=tk.DISABLED)
            self.refresh_button.config(state=tk.DISABLED)
    
    def find_file(self):
        """在指定目录中使用通配符查找CSV文件"""
        # 创建查找文件对话框
        find_window = tk.Toplevel(self.root)
        find_window.title("Find CSV Files")
        find_window.geometry("700x500")  # 增大窗口大小
        find_window.resizable(True, True)
        
        # 居中显示
        find_window.transient(self.root)
        find_window.grab_set()
        
        # 目录选择区域
        dir_frame = tk.Frame(find_window)
        dir_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(dir_frame, text="Directory:", width=10).pack(side="left")
        
        dir_var = tk.StringVar(value=os.getcwd())
        dir_entry = tk.Entry(dir_frame, textvariable=dir_var, width=50)
        dir_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        def browse_directory():
            selected_dir = tk.filedialog.askdirectory(initialdir=dir_var.get())
            if selected_dir:
                dir_var.set(selected_dir)
        
        browse_button = tk.Button(dir_frame, text="Browse...", command=browse_directory)
        browse_button.pack(side="left", padx=5)
        self.add_hover_effect(browse_button)
        
        # 通配符输入区域
        pattern_frame = tk.Frame(find_window)
        pattern_frame.pack(fill="x", padx=10, pady=(5, 5))
        
        tk.Label(pattern_frame, text="Pattern:", width=10).pack(side="left")
        pattern_var = tk.StringVar(value="*.csv")
        pattern_entry = tk.Entry(pattern_frame, textvariable=pattern_var, width=50)
        pattern_entry.pack(side="left", fill="x", expand=True, padx=5)
        
        # 搜索结果显示区域
        result_frame = tk.Frame(find_window)
        result_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(result_frame, text="Search Results:", anchor="w").pack(fill="x")
        
        result_canvas = tk.Canvas(result_frame)
        result_canvas.pack(side="left", fill="both", expand=True)
        
        result_scrollbar = tk.Scrollbar(result_frame, orient="vertical", command=result_canvas.yview)
        result_scrollbar.pack(side="right", fill="y")
        
        result_canvas.configure(yscrollcommand=result_scrollbar.set)
        
        result_container = tk.Frame(result_canvas)
        result_window = result_canvas.create_window((0, 0), window=result_container, anchor="nw", width=result_canvas.winfo_width())
        
        # 绑定Canvas配置事件
        def on_result_configure(event):
            result_canvas.itemconfig(result_window, width=event.width)
        result_canvas.bind("<Configure>", on_result_configure)
        
        def on_result_container_configure(event):
            result_canvas.configure(scrollregion=result_canvas.bbox("all"))
        result_container.bind("<Configure>", on_result_container_configure)
        
        # 状态标签
        status_label = tk.Label(find_window, text="Ready", anchor="w")
        status_label.pack(fill="x", padx=10, pady=5)
        
        # 文件变量存储
        found_files = []
        file_vars = []
        
        # 定义内部函数
        def search_files():
            # 清空之前的结果
            for widget in result_container.winfo_children():
                widget.destroy()
            found_files.clear()
            file_vars.clear()
            
            directory = dir_var.get()
            pattern = pattern_var.get()
            
            if not os.path.isdir(directory):
                tk.messagebox.showerror("Error", "Please select a valid directory!")
                return
            
            try:
                # 转换通配符为正则表达式
                regex_pattern = fnmatch.translate(pattern)
                regex = re.compile(regex_pattern)
                
                # 搜索文件
                for root, dirs, files in os.walk(directory):
                    for file in files:
                        if regex.match(file) and file.lower().endswith('.csv'):
                            file_path = os.path.join(root, file)
                            found_files.append(file_path)
                
                # 显示搜索结果
                tk.Label(result_container, text=f"Found {len(found_files)} files:", anchor="w").pack(fill="x", pady=(0, 5))
                
                if found_files:
                    for i, file_path in enumerate(found_files):
                        var = tk.BooleanVar(value=False)
                        file_vars.append(var)
                        
                        frame = tk.Frame(result_container)
                        frame.pack(fill="x", pady=2)
                        
                        check = tk.Checkbutton(frame, variable=var)
                        check.pack(side="left", padx=5)
                        
                        # 显示相对路径
                        rel_path = os.path.relpath(file_path, directory)
                        label = tk.Label(frame, text=rel_path, anchor="w")
                        label.pack(side="left", fill="x", expand=True)
                
                # 更新滚动区域
                result_container.update_idletasks()
                result_canvas.configure(scrollregion=result_canvas.bbox("all"))
                
                # 更新状态标签
                status_label.config(text=f"Found {len(found_files)} files")
                
            except Exception as e:
                tk.messagebox.showerror("Searching error ", f"Searching error: {str(e)}")
        
        def select_all_results():
            for var in file_vars:
                var.set(True)
        
        def clear_selection():
            for var in file_vars:
                var.set(False)
        
        def load_selected_files():
            try:
                selected_paths = [found_files[i] for i, var in enumerate(file_vars) if var.get()]
                if not selected_paths:
                    tk.messagebox.showinfo("Warning", "Selece files to load！")
                    return
                # 使用现有的add_files方法逻辑来添加文件
                # 为新添加的文件创建复选框和标签
                first_file_path = None
                # 计算当前已有文件数量
                current_count = len(self.file_vars)
                
                for i, file_path in enumerate(selected_paths):
                    # 只获取文件名（不含路径）
                    file_name = os.path.basename(file_path)
                    
                    # 如果文件已存在，跳过
                    if file_path in self.file_vars:
                        continue
                    
                    # 记录第一个文件路径用于预览和保存目录
                    if first_file_path is None:
                        first_file_path = file_path
                        # 保存为实例变量，供保存功能使用
                        self.first_file_path = file_path
                    
                    # 增加文件计数
                    self.file_count += 1
                    
                    # 创建变量存储复选框状态
                    var = tk.BooleanVar(value=True)
                    self.file_vars[file_path] = var
                    
                    # 计算网格位置（每行4个文件）
                    total_files = current_count + i + 1  # 包含当前正在添加的文件
                    row = (total_files - 1) // 4
                    col = (total_files - 1) % 4
                    
                    # 创建文件框架
                    file_frame = tk.Frame(self.file_list_container)
                    file_frame.grid(row=row, column=col, padx=5, pady=5, sticky="w")
                    
                    # 创建复选框
                    check = tk.Checkbutton(file_frame, variable=var, command=lambda path=file_path: self.on_checkbox_change(path))
                    check.pack(side="left", padx=2)
                    self.file_checks[file_path] = check
                    
                    # 创建编号标签
                    number_label = tk.Label(file_frame, text=f"{self.file_count}.", width=4, anchor="w")
                    number_label.pack(side="left")
                    
                    # 创建文件名标签
                    label = tk.Label(file_frame, text=file_name, anchor="w")
                    label.pack(side="left")
                    self.file_labels[file_path] = label
                    
                    # 更新状态栏信息
                    self.update_status(f"Successfully added file: {file_name}")
                
                # 重新配置滚动区域
                self.file_list_container.update_idletasks()
                self.file_list_canvas.configure(scrollregion=self.file_list_canvas.bbox("all"))
                
                # 保存选中文件列表，供Save as CSV功能使用
                self.selected_files = [path for path, var in self.file_vars.items() if var.get()]
                
                # 更新菜单和按钮状态
                self.update_menu_status(has_files=True, has_selected_files=True)
                
                # 启用File PreView按钮
                self.refresh_button.config(state="normal")
                
                # 关闭查找窗口
                find_window.destroy()
            except (OSError, tk.TclError) as e:
                tk.messagebox.showerror("Error", f"Failed to load files: {str(e)}")
                self.update_status(f"Error loading files: {str(e)}")
        
        # 按钮区域 - 放在Pattern输入框下方
        button_frame = tk.Frame(find_window)
        button_frame.pack(fill="x", padx=10, pady=5)
        
        search_button = tk.Button(button_frame, text="Search", command=search_files)
        search_button.pack(side="left", padx=5)
        self.add_hover_effect(search_button)
        
        select_all_button = tk.Button(button_frame, text="Select All", command=select_all_results)
        select_all_button.pack(side="left", padx=5)
        self.add_hover_effect(select_all_button)
        
        clear_button = tk.Button(button_frame, text="Clear Selection", command=clear_selection)
        clear_button.pack(side="left", padx=5)
        self.add_hover_effect(clear_button)
        
        load_button = tk.Button(button_frame, text="Load Selected", command=load_selected_files)
        load_button.pack(side="right", padx=5)
        self.add_hover_effect(load_button)
        
        cancel_button = tk.Button(button_frame, text="Cancel", command=find_window.destroy)
        cancel_button.pack(side="right", padx=5)
        self.add_hover_effect(cancel_button)

    def select_all_files(self):
        """全选或取消全选所有文件"""
        if not hasattr(self, 'file_vars') or not self.file_vars:
            return
        
        # 检查当前是否全部选中
        all_selected = all(var.get() for var in self.file_vars.values())
        
        # 设置所有复选框状态为相反值
        target_state = 0 if all_selected else 1
        for var in self.file_vars.values():
            var.set(target_state)
        
        # 更新选中文件列表
        if hasattr(self, 'selected_files'):
            if target_state == 1:
                self.selected_files = list(self.file_vars.keys())
            else:
                self.selected_files = []
            
            # 更新菜单和按钮状态
            self.update_menu_status(has_files=True, has_selected_files=(target_state == 1))
        
        # 更新状态栏信息
        action = "Select all" if target_state == 1 else "Unselect all"
        self.status_var.set(f"{action} all files")
        
        # 记录日志
        logging.info(f"{action} all files")
    
    def clear_all(self):
        """清除所有文件和所有选项卡内容，将程序恢复至初始设置状态"""
        # 清除文件列表
        for widget in self.file_list_container.winfo_children():
            widget.destroy()
        
        # 重置文件相关变量
        self.file_vars = {}
        self.file_checks = {}
        self.file_labels = {}
        self.file_count = 0
        
        # 重置数据预览表格
        self.reset_data_preview_table()
        
        # 清除数据处理部分的数据
        # 清除数据处理选项卡中的表格内容
        if hasattr(self, 'processing_table_frame'):
            for widget in self.processing_table_frame.winfo_children():
                widget.destroy()
        
        # 重置数据处理选项卡的状态栏信息
        if hasattr(self, 'processing_status_var'):
            self.processing_status_var.set("Ready")
        
        # 清除所有选项卡的内容
        
        # 清除数据处理选项卡
        if hasattr(self, 'data_processing_tab'):
            for widget in self.data_processing_tab.winfo_children():
                widget.destroy()
            # 重新设置数据处理选项卡的基本结构
            self._setup_data_processing_tab()
        
        # 清除数据重新处理选项卡
        if hasattr(self, 'data_reprocessing_tab'):
            for widget in self.data_reprocessing_tab.winfo_children():
                widget.destroy()
            # 重新设置数据重新处理选项卡的基本结构
            self._setup_data_reprocessing_tab()
        
        # 清除不良率分析选项卡
        if hasattr(self, 'yield_analysis_tab'):
            for widget in self.yield_analysis_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.yield_analysis_tab, text="Click 'Yield Analysis' button to generate yield analysis", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清除Top 10选项卡
        if hasattr(self, 'top10_tab'):
            for widget in self.top10_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.top10_tab, text="Click 'Top Defects' button to view Top Defect Items Analysis", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清除Read Criteria选项卡
        if hasattr(self, 'review_criteria_tab'):
            for widget in self.review_criteria_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.review_criteria_tab, text="Click 'Read Criteria' button to read criteria data", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清除ColorPointSpec选项卡
        if hasattr(self, 'colorpoint_spec_tab'):
            for widget in self.colorpoint_spec_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.colorpoint_spec_tab, text="Click 'ColorPointSpec' button to view color point specifications", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清除Cpk选项卡
        if hasattr(self, 'cpk_tab'):
            for widget in self.cpk_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.cpk_tab, text="Click 'Cpk' button to view Cpk Analysis", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清除Color Point Chart选项卡
        if hasattr(self, 'color_point_chart_tab'):
            for widget in self.color_point_chart_tab.winfo_children():
                widget.destroy()
            # 添加占位标签
            placeholder_label = tk.Label(self.color_point_chart_tab, text="Click 'Color Point Chart' button to view color point chart", font=('SimHei', 10))
            placeholder_label.pack(pady=20)
        
        # 清空处理后的数据引用和选中文件列表
        self.processed_data = None
        if hasattr(self, 'selected_files'):
            self.selected_files = []
        
        # 更新主状态栏信息
        self.update_status("All files and all tabs content have been cleared. The program is now ready.")
        
        # 确保滚动区域正确更新
        self.file_list_container.update_idletasks()
        self.file_list_canvas.configure(scrollregion=self.file_list_canvas.bbox("all"))
        
        # 强制刷新整个窗口布局
        self.root.update_idletasks()
        
        # 切换到数据预览选项卡
        if hasattr(self, 'tab_control') and hasattr(self, 'data_preview_tab'):
            self.tab_control.select(self.data_preview_tab)
            
        # 使用统一的方法更新所有菜单和按钮的状态
        self.update_menu_status(has_files=False, has_selected_files=False, has_processed_data=False)
    
    def create_progress_window(self, title="Data Processing Progress", max_value=100):
        """创建一个弹出式进度条窗口"""
        # 创建进度条窗口
        progress_window = tk.Toplevel(self.root)
        progress_window.title(title)
        progress_window.geometry("400x100")
        progress_window.resizable(False, False)
        progress_window.transient(self.root)  # 设置为主窗口的子窗口
        progress_window.grab_set()  # 模态窗口，阻止操作主窗口
        
        # 计算位置，使其在主窗口中央
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 200
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
        progress_window.geometry(f"400x100+{x}+{y}")
        
        # 创建进度条
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, length=380, mode='determinate', maximum=max_value)
        progress_bar.pack(pady=(20, 10), padx=10)
        
        # 创建状态标签
        status_var = tk.StringVar(value="Data processing is ready to begin.")
        status_label = tk.Label(progress_window, textvariable=status_var)
        status_label.pack(pady=(0, 10))
        
        # 返回窗口和控制变量
        return progress_window, progress_var, status_var
    
    def data_processing_function(self):
        """数据处理功能"""
        # 切换到数据处理选项卡
        if hasattr(self, 'tab_control') and hasattr(self, 'data_processing_tab'):
            self.tab_control.select(self.data_processing_tab)
            
        # 获取所有选中的文件
        selected_files = [path for path, var in self.file_vars.items() if var.get()]
        
        if not selected_files:
            self.update_status("No files selected for processing.")
            self.show_processing_result(None, "No files selected for processing.", is_reprocessing=False)
            return
        
        # 创建进度条窗口，进度最大值设为选中文件数+5（额外步骤）
        progress_window, progress_var, status_var = self.create_progress_window("Data Processing Progress", len(selected_files) + 5)
        
        try:
            # 创建临时文件夹
            import tempfile
            import shutil
            import os
            
            # 更新进度条
            progress_var.set(1)
            status_var.set("Setting up data processing environment...")
            progress_window.update_idletasks()
            
            # 获取或创建数据处理选项卡中的表格容器
            self._setup_data_processing_tab()
            
            # 设置数据重新处理选项卡
            self._setup_data_reprocessing_tab()
            
            # 创建临时目录
            temp_dir = tempfile.mkdtemp()
            processed_files = []
            
            for i, file_path in enumerate(selected_files):
                file_name = os.path.basename(file_path)
                self.update_status(f"Processing file: {file_name}")
                
                # 更新进度条
                progress_var.set(i + 2)  # +2 因为前面已经有一步了
                status_var.set(f"Processing file {i+1}/{len(selected_files)}: {file_name}")
                progress_window.update_idletasks()
                
                # 解析文件名，优先检查MP/PVT标识
                if "MP" in file_name:
                    extracted_text = "MP"
                elif "PVT" in file_name:
                    extracted_text = "PVT"
                else:
                    # 提取第二和第三空格之间的文本
                    parts = file_name.split(' ')
                    # 确保有足够的空格来提取文本
                    if len(parts) >= 4:
                        # 第二和第三空格之间的文本是parts[2]
                        extracted_text = parts[2]
                    else:
                        extracted_text = "Unknown"
                    
                # 读取文件并处理
                try:
                    # 智能检测CSV文件中的数据标题行
                        with open(file_path, 'r', encoding='utf-8') as f:
                            lines = []
                            for i in range(20):
                                line = f.readline()
                                if not line:
                                    break
                                lines.append(line.strip())
                            
                            # 识别元数据行（通常是键值对格式，如"键:,值"的形式）
                            metadata_lines = []
                            non_metadata_lines = []
                            
                            for i, line in enumerate(lines):
                                # 元数据行的特征: 通常包含键值对格式
                                # 1. 检查是否包含冒号加逗号的模式（如"键:,值"）
                                # 2. 检查行的开头是否包含典型的元数据键名
                                is_metadata = False
                                
                                # 检查标点符号模式
                                if ',:' in line or (':' in line and ',' in line):
                                    is_metadata = True
                                
                                # 检查是否为典型的元数据行（有内容但不是标题行）
                                elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                                    # 检查行是否有有意义的内容且不是明显的标题行
                                    content_parts = [p.strip() for p in line.split(',') if p.strip()]
                                    if content_parts and ':' in content_parts[0]:
                                        is_metadata = True
                                
                                if is_metadata:
                                    metadata_lines.append(i)
                                else:
                                    non_metadata_lines.append((i, line))
                            
                            # 如果有非元数据行，在这些行中寻找包含最多逗号的非空行作为标题行
                            max_commas = -1
                            header_line_candidates = []
                            header_line_index = 0
                            
                            # 首先筛选出非空的非元数据行
                            non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                            
                            if non_empty_non_metadata:
                                # 找到所有非空非元数据行中的最大逗号数
                                max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                                # 获取所有具有最大逗号数的候选行
                                header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                                
                                if header_line_candidates:
                                    # 增强选择逻辑: 优先选择包含字母字符、内容长度更长的行作为标题行
                                    # 计算每个候选行的分数
                                    best_score = -1
                                    for i, line in header_line_candidates:
                                        # 基础分数（保证至少有一个分数）
                                        score = 1
                                        # 包含字母字符加分（标题行通常包含字段名）
                                        if any(c.isalpha() for c in line):
                                            score += 10
                                        # 内容长度加分（不含逗号）
                                        content_length = len(line.replace(',', ''))
                                        score += min(content_length, 50)  # 内容长度最多加50分
                                        
                                        if score > best_score:
                                            best_score = score
                                            header_line_index = i
                            elif non_metadata_lines:
                                # 如果所有非元数据行都是空的，使用原始方法
                                for i, line in non_metadata_lines:
                                    comma_count = line.count(',')
                                    if comma_count > max_commas:
                                        max_commas = comma_count
                                        header_line_index = i
                            else:
                                # 如果没有非元数据行，使用原始方法寻找最多逗号的行
                                for i, line in enumerate(lines):
                                    comma_count = line.count(',')
                                    if comma_count > max_commas:
                                        max_commas = comma_count
                                        header_line_index = i
                            
                            # 再次检查: 确保找到的标题行不是全空的，且有合理数量的非空列
                            if header_line_index < len(lines) and lines[header_line_index].strip() == '':
                                self.update_status(f"Warning: Empty header line detected (Line {header_line_index+1}). Trying to find the next suitable line.")
                                # 寻找下一个非空行
                                for i in range(header_line_index + 1, min(len(lines), header_line_index + 10)):
                                    if lines[i].strip() != '':
                                        header_line_index = i
                                        max_commas = lines[i].count(',')
                                        break
                            
                            # 确定是否使用检测到的标题行
                            if max_commas >= 9:
                                # 如果找到明显的标题行（逗号数足够多），使用它
                                self.update_status(f"Data header line detected (Line {header_line_index+1}) with {max_commas+1} columns of data.")
                                self.update_status(f"Excluded metadata lines: {len(metadata_lines)}")
                                df = pd.read_csv(file_path,
                                                skiprows=header_line_index,  # 跳过说明性行
                                                on_bad_lines='skip',    # 跳过有问题的行
                                                engine='python')        # 使用Python解析器
                            else:
                                # 如果没有找到明显的标题行，尝试使用默认方式读取
                                self.update_status("No obvious data header line detected. Trying to read with default settings.")
                                df = pd.read_csv(file_path,
                                                on_bad_lines='skip',    # 跳过有问题的行
                                                engine='python')        # 使用Python解析器
                        
                        # 直接填入第二列，并将标题改为"Config"
                        if not df.empty:
                            # 确保DataFrame至少有一列
                            if len(df.columns) >= 1:
                                # 将第二列（索引为1）的数据替换为提取的文本
                                if len(df.columns) >= 2:
                                    # 检查第二列标题是否为"Config"
                                    if df.columns[1] != 'Config':
                                        # 标题不是Config，才替换数据
                                        df.iloc[:, 1] = extracted_text
                                    # 将第二列的标题改为"Config"
                                    df.columns.values[1] = 'Config'
                                else:
                                    # 没有第二列，添加一列
                                    df['Config'] = extracted_text
                        
                        # 过滤坏行: 移除Serial Number列文本长度异常的行
                        original_rows = len(df)
                        if 'Serial Number' in df.columns:
                            # 计算Serial Number列文本长度的中位数
                            serial_lengths = df['Serial Number'].astype(str).str.len()
                            median_length = serial_lengths.median()
                            
                            # 过滤掉Serial Number长度与中位数相差5以上的行
                            df = df[abs(serial_lengths - median_length) < 5]
                            filtered_rows = original_rows - len(df)
                            if filtered_rows > 0:
                                self.update_status(f"Filtered out {filtered_rows} rows with Serial Number length difference > 5 from median ({median_length}).")
                        

                        
                        # 保存到临时文件夹
                        temp_file_path = os.path.join(temp_dir, file_name)
                        df.to_csv(temp_file_path, index=False, encoding='utf-8')
                        
                        # 记录处理后的文件
                        processed_files.append((temp_file_path, file_name, df))
                        
                except Exception as e:
                    self.update_status(f"Error processing file {file_name}: {str(e)}")
                    continue
            
            # 更新进度条: 文件处理完成
            progress_var.set(len(selected_files) + 2)
            status_var.set("Files processed. Starting post-processing...")
            progress_window.update_idletasks()
            
            # 如果有两个以上文件被选中，合并它们
            if len(processed_files) >= 2:
                self.update_status(f"Combining {len(processed_files)} files...")
                
                # 更新进度条
                progress_var.set(len(selected_files) + 3)
                status_var.set(f"Combining {len(processed_files)} files...")
                progress_window.update_idletasks()
                
                # 合并所有文件
                combined_df = pd.concat([df for _, _, df in processed_files], ignore_index=True)
                
                # 对合并后的数据进行去重: 以Serial Number列为索引，优先保留Pass/Fail列值为PASS的行
                if 'Serial Number' in combined_df.columns:
                    original_dup_rows = len(combined_df)
                    # 创建一个排序键，Pass/Fail列为PASS的排在前面
                    if 'Pass/Fail' in combined_df.columns:
                        # 给PASS值赋较低的排序值，使其在去重时保留
                        combined_df['_sort_key'] = combined_df['Pass/Fail'].apply(lambda x: 0 if str(x).strip().upper() == 'PASS' else 1)
                        # 按照Serial Number和排序键排序
                        combined_df = combined_df.sort_values(by=['Serial Number', '_sort_key'])
                        # 去重，保留每个Serial Number的第一行（即PASS的行优先）
                        combined_df = combined_df.drop_duplicates(subset=['Serial Number'], keep='first')
                        # 删除临时排序键列
                        combined_df = combined_df.drop(columns=['_sort_key'])
                    else:
                        # 如果没有Pass/Fail列，直接去重
                        combined_df = combined_df.drop_duplicates(subset=['Serial Number'], keep='first')
                    
                    deduped_rows = original_dup_rows - len(combined_df)
                    if deduped_rows > 0:
                        self.update_status(f"Deduplicated {deduped_rows} rows with duplicate Serial Numbers.")
                
                # 修改列标题格式
                combined_df = self.rename_columns(combined_df)
                
                # 删除不需要的列
                combined_df = self.remove_unwanted_columns(combined_df)
                
                # 对W_M和M_M列组内的列进行排序
                combined_df = self.sort_wm_and_mm_columns(combined_df)
                
                # 处理特殊单元格（FAIL值和含#字符的单元格）
                combined_df = self.process_special_cells(combined_df)
                
                # 更新进度条
                progress_var.set(len(selected_files) + 4)
                status_var.set("Post-processing completed. Ready to display results...")
                progress_window.update_idletasks()
                
                # 保存处理后的数据引用
                self.processed_data = combined_df
                # 使用统一的方法启用所有相关菜单和按钮
                # 处理完成后，即使没有原始文件，也应该认为has_files为True
                has_files = (len(self.loaded_files) > 0 if hasattr(self, 'loaded_files') else False) or True
                has_selected_files = len([path for path, var in self.file_vars.items() if var.get()]) > 0
                self.update_menu_status(has_files=has_files, has_selected_files=has_selected_files, has_processed_data=True)
                # 在数据处理选项卡中预览合并后的数据
                self.show_processing_result(combined_df, f"Combined data from {len(processed_files)} files.", is_reprocessing=False)
                
                # 保存合并后的文件到临时文件夹
                combined_file_path = os.path.join(temp_dir, f"Combined_data_{len(processed_files)}_files.csv")
                combined_df.to_csv(combined_file_path, index=False, encoding='utf-8')
                
                self.update_status(f"Successfully combined {len(processed_files)} files and previewed.")
                
                # 进度完成
                progress_var.set(len(selected_files) + 5)
                status_var.set("Processing completed!")
                progress_window.update_idletasks()
                # 延迟关闭进度窗口
                self.root.after(500, progress_window.destroy)
            elif len(processed_files) == 1:
                # 只有一个文件，获取数据
                _, file_name, df = processed_files[0]
                
                # 对单个文件数据进行去重: 以Serial Number列为索引，优先保留Pass/Fail列值为PASS的行
                if 'Serial Number' in df.columns:
                    original_dup_rows = len(df)
                    # 创建一个排序键，Pass/Fail列为PASS的排在前面
                    if 'Pass/Fail' in df.columns:
                        # 给PASS值赋较低的排序值，使其在去重时保留
                        df['_sort_key'] = df['Pass/Fail'].apply(lambda x: 0 if str(x).strip().upper() == 'PASS' else 1)
                        # 按照Serial Number和排序键排序
                        df = df.sort_values(by=['Serial Number', '_sort_key'])
                        # 去重，保留每个Serial Number的第一行（即PASS的行优先）
                        df = df.drop_duplicates(subset=['Serial Number'], keep='first')
                        # 删除临时排序键列
                        df = df.drop(columns=['_sort_key'])
                    else:
                        # 如果没有Pass/Fail列，直接去重
                        df = df.drop_duplicates(subset=['Serial Number'], keep='first')
                    
                    deduped_rows = original_dup_rows - len(df)
                    if deduped_rows > 0:
                        self.update_status(f"Deduplicated {deduped_rows} rows with duplicate Serial Numbers in file {file_name}.")
                
                # 修改列标题格式
                df = self.rename_columns(df)
                
                # 删除不需要的列
                df = self.remove_unwanted_columns(df)
                
                # 对W_M和M_M列组内的列进行排序
                df = self.sort_wm_and_mm_columns(df)
                
                # 处理特殊单元格（FAIL值和含#字符的单元格）
                df = self.process_special_cells(df)
                
                # 更新进度条
                progress_var.set(len(selected_files) + 4)
                status_var.set("Post-processing completed. Ready to display results...")
                progress_window.update_idletasks()
                
                # 保存处理后的数据引用
                self.processed_data = df
                # 使用统一的方法启用所有相关菜单和按钮
                # 处理完成后，即使没有原始文件，也应该认为has_files为True
                has_files = (len(self.loaded_files) > 0 if hasattr(self, 'loaded_files') else False) or True
                has_selected_files = len([path for path, var in self.file_vars.items() if var.get()]) > 0
                self.update_menu_status(has_files=has_files, has_selected_files=has_selected_files, has_processed_data=True)
                # 预览数据
                self.show_processing_result(df, f"Post-processed data from file: {file_name}", is_reprocessing=False)
                self.update_status(f"Successfully processed and previewed file: {file_name}")
                
                # 进度完成
                progress_var.set(len(selected_files) + 5)
                status_var.set("Processing completed!")
                progress_window.update_idletasks()
                # 延迟关闭进度窗口
                self.root.after(500, progress_window.destroy)
            else:
                # 没有成功处理数据，清空数据引用并更新状态
                self.processed_data = None
                has_files = len(self.loaded_files) > 0 if hasattr(self, 'loaded_files') else False
                has_selected_files = len([path for path, var in self.file_vars.items() if var.get()]) > 0
                self.update_menu_status(has_files=has_files, has_selected_files=has_selected_files, has_processed_data=False)
                self.show_processing_result(None, "No files were successfully processed.", is_reprocessing=False)
                self.update_status("No files were successfully processed.")
                
                # 关闭进度窗口
                progress_window.destroy()
            
        except Exception as e:
            self.update_status(f"Processing failed: {str(e)}")
            self.show_processing_result(None, f"Processing failed: {str(e)}", is_reprocessing=False)
            
            # 确保关闭进度窗口
            try:
                progress_window.destroy()
            except:
                pass
            
            # 处理失败时更新菜单和按钮状态
            self.processed_data = None
            has_files = len(self.loaded_files) > 0 if hasattr(self, 'loaded_files') else False
            has_selected_files = len(self.selected_files) > 0 if hasattr(self, 'selected_files') else False
            self.update_menu_status(has_files=has_files, has_selected_files=has_selected_files, has_processed_data=False)
        
    def copy_file_path(self, file_path):
        """复制文件路径到剪贴板"""
        pass
    
    def rename_columns(self, df):
        """修改列标题格式:

           - 将 "White Metric2 Value" 改为 "W_M2_前一列第一行文本"
           - 将 "White Metric15 Value" 改为 "W_M15_前一列第一行文本"
           - 将 "Mixed Metric3 Value" 改为 "M_M3_前一列第一行文本"
        """
        import re
        
        # 创建新的列名映射
        new_columns = {}
        renamed_count = 0
        
        # 首先找出所有需要重命名的列及其位置
        columns_to_rename = []
        for i, col in enumerate(df.columns):
            white_match = re.match(r'White\s+Metric(\d+)\s+Value', col)
            mixed_match = re.match(r'Mixed\s+Metric(\d+)\s+Value', col)
            if white_match or mixed_match:
                columns_to_rename.append((i, col, white_match, mixed_match))
        
        # 先创建一个临时映射，只做基本重命名
        temp_columns = {col: col for col in df.columns}
        
        # 进行基本重命名
        for _, col, white_match, mixed_match in columns_to_rename:
            if white_match:
                metric_num = white_match.group(1)
                temp_columns[col] = f'W_M{metric_num}_'
            elif mixed_match:
                metric_num = mixed_match.group(1)
                temp_columns[col] = f'M_M{metric_num}_'
            renamed_count += 1
        
        # 应用临时重命名
        temp_df = df.rename(columns=temp_columns)
        
        # 重新构建列名映射，添加前一列第一行文本
        final_columns = {}
        for original_col, temp_col in temp_columns.items():
            # 如果是需要重命名的列
            if temp_col.startswith('W_M') or temp_col.startswith('M_M'):
                # 找到该列在原始DataFrame中的位置
                original_index = list(df.columns).index(original_col)
                
                # 确保该列不是第一列
                if original_index > 0:
                    # 获取前一列的名称
                    prev_col = df.columns[original_index - 1]
                    # 获取前一列的第一行文本
                    if not df.empty and prev_col in df.columns:
                        first_row_text = str(df.iloc[0][prev_col]).strip()
                        # 移除任何非字母数字字符，只保留有效文本
                        first_row_text = re.sub(r'[^a-zA-Z0-9]', '', first_row_text)
                        if first_row_text:
                            # 将前一列的第一行文本添加到新标题中
                            final_col = f'{temp_col}{first_row_text}'
                            final_columns[original_col] = final_col
                            continue
                # 如果没有前一列或前一列没有有效文本，使用基本重命名
                final_columns[original_col] = temp_col
            else:
                # 其他列名保持不变
                final_columns[original_col] = original_col
        
        # 应用最终重命名
        if renamed_count > 0:
            df = df.rename(columns=final_columns)
            self.update_status(f"Successfully renamed {renamed_count} column headers, added prefix from previous column's first row text.")
        
        return df
        
    def remove_unwanted_columns(self, df):
        """保留指定的列、W_M/M_M开头的列以及包含点号的列，删除其他所有列，并保持原始列顺序不变"""
        # 定义需要保留的特定列
        columns_to_keep = [
            "Model",
            "Config",
            "Serial Number",
            "Test Station",
            "Position ID",
            "Date/Time",
            "Pass/Fail",
            "White L (cd/m^2)",
            "White U (%)",
            "White dY (%/cm)",
            "White u Avg",
            "White v Avg",
            "White Ru",
            "White Rv",
            "White Du",
            "White Dv",
            "White dL*Min (%/cm)",
            "White dL*Max (%/cm)",
            "White dEMax (%/cm)",
            "White Pass/Fail Criteria",
            "Mixed L (cd/m^2)",
            "Mixed U (%)",
            "Mixed dY (%/cm)",
            "Mixed u Avg",
            "Mixed v Avg",
            "Mixed Ru",
            "Mixed Rv",
            "Mixed Du",
            "Mixed Dv",
            "Mixed dL*Min (%/cm)",
            "Mixed dL*Max (%/cm)",
            "Mixed dEMax (%/cm)",
            "Mixed Pass/Fail Criteria"
        ]
        
        # 创建一个集合用于快速查找
        columns_to_keep_set = set(columns_to_keep)
        
        # 按原始DataFrame中列的顺序，筛选出需要保留的列
        all_columns_to_keep = []
        existing_specific_columns_count = 0
        pattern_columns_count = 0
        
        for col in df.columns:
            if col in columns_to_keep_set:
                all_columns_to_keep.append(col)
                existing_specific_columns_count += 1
            elif col.startswith("W_M") or col.startswith("M_M"):
                all_columns_to_keep.append(col)
                pattern_columns_count += 1
            elif "." in col:
                all_columns_to_keep.append(col)
                pattern_columns_count += 1
        
        # 删除其他所有列
        if all_columns_to_keep:
            df = df[all_columns_to_keep]
            self.update_status(f"Successfully kept {len(all_columns_to_keep)} columns (including {existing_specific_columns_count} specific columns and {pattern_columns_count} pattern-matched columns, including those with dots) in original order.")
        else:
            self.update_status("Warning: No columns to keep were found.")
        
        return df
        
    def sort_wm_and_mm_columns(self, df):
        """对标题格式为"W_M"和"M_M"的列组，按列标题中包含的数字从小到大对组内列进行排序，同时保持列组在表格中的整体位置不变"""
        import re
        
        # 创建一个副本，避免直接修改原始数据
        sorted_df = df.copy()
        
        # 找出所有列，并分别处理不同的列组
        all_columns = list(sorted_df.columns)
        new_column_order = []
        current_group = None
        group_columns = []
        
        # 辅助函数: 从列名中提取数字
        def extract_number(col_name):
            match = re.search(r'(W_M|M_M)(\d+)', col_name)
            if match:
                return int(match.group(2))
            return float('inf')  # 非W_M/M_M列返回无穷大，确保它们排在最后
        
        # 遍历所有列，按顺序分组并排序
        for col in all_columns:
            # 确定列所属的组
            if col.startswith('W_M'):
                col_group = 'W_M'
            elif col.startswith('M_M'):
                col_group = 'M_M'
            else:
                col_group = 'OTHER'
            
            # 如果当前列与前一列属于不同的组，先处理前一组
            if col_group != current_group and current_group is not None:
                if current_group in ['W_M', 'M_M']:
                    # 对W_M和M_M组内的列按数字从小到大排序
                    group_columns.sort(key=extract_number)
                # 将排序后的组添加到新的列顺序中
                new_column_order.extend(group_columns)
                # 重置当前组
                group_columns = []
            
            # 更新当前组并添加列
            current_group = col_group
            group_columns.append(col)
        
        # 处理最后一组
        if group_columns:
            if current_group in ['W_M', 'M_M']:
                group_columns.sort(key=extract_number)
            new_column_order.extend(group_columns)
        
        # 重新排列DataFrame的列
        sorted_df = sorted_df[new_column_order]
        
        # 计算排序后的结果
        wm_count = sum(1 for col in sorted_df.columns if col.startswith('W_M'))
        mm_count = sum(1 for col in sorted_df.columns if col.startswith('M_M'))
        
        self.update_status(f"Successfully sorted W_M group ({wm_count} columns) and M_M group ({mm_count} columns) columns by number in ascending order, while keeping group positions unchanged.")
        
        return sorted_df
        
    def process_special_cells(self, df):
        """处理特殊单元格: 将Pass/Fail列中的Fail单元格和包含#字符的单元格（不良）标记为需要格式化，

        去除单元格中含有的#字符，并将含#字符的列尝试转换为数字数据"""
        # 创建一个字典来存储需要格式化的单元格位置
        self.format_cells = {}
        
        # 创建一个集合来存储包含#字符的列索引
        columns_with_hash = set()
        
        # 创建一个字典来存储包含#字符的单元格位置，用于后续添加淡红色背景
        self.hash_cells = {}
        
        # 创建一个副本，避免直接修改原始数据
        processed_df = df.copy()
        
        # 遍历所有单元格
        for col_idx, column in enumerate(processed_df.columns):
            has_hash_char = False
            # 检查当前列是否为Pass/Fail列（不区分大小写，要求列名完全匹配'pass/fail'）
            col_name_lower = str(column).lower().strip()
            is_pass_fail_column = col_name_lower == 'pass/fail'
            
            for row_idx, value in enumerate(processed_df[column]):
                cell_value = str(value)
                
                # 检查是否需要格式化
                need_format = False
                
                # 只在Pass/Fail列中检查FAIL值（不区分大小写）
                if is_pass_fail_column and cell_value.strip().upper() == 'FAIL':
                    need_format = True
                
                # 检查单元格是否包含#字符（标记为不良的单元格）
                elif '#' in cell_value:
                    has_hash_char = True
                    # 去除#字符
                    clean_value = cell_value.replace('#', '')
                    processed_df.iloc[row_idx, col_idx] = clean_value
                    need_format = True
                    
                    # 记录包含#字符的单元格位置
                    if row_idx not in self.hash_cells:
                        self.hash_cells[row_idx] = []
                    self.hash_cells[row_idx].append(col_idx)
                
                # 如果需要格式化，记录单元格位置
                if need_format:
                    if row_idx not in self.format_cells:
                        self.format_cells[row_idx] = []
                    self.format_cells[row_idx].append(col_idx)
            
            # 如果该列包含#字符，尝试将整列转换为数字
            if has_hash_char:
                columns_with_hash.add(col_idx)
                try:
                    # 尝试将列转换为数字
                    processed_df[column] = pd.to_numeric(processed_df[column], errors='ignore')
                except (ValueError, TypeError):
                    # 如果转换失败，保持原样
                    pass
        
        # 保存包含#字符的列信息，用于预览时的背景色设置
        self.columns_with_hash = columns_with_hash
        
        # 更新状态栏
        total_format_cells = sum(len(cols) for cols in self.format_cells.values())
        total_hash_cells = sum(len(cols) for cols in self.hash_cells.values())
        if total_format_cells > 0:
            self.update_status(f"Successfully processed {total_format_cells} special cells, including {total_hash_cells} cells with # characters, converted {len(columns_with_hash)} columns to numeric data.")
        
        return processed_df
        
    def load_spec_file(self):
        """加载规格文件"""
        
    def create_data_preview_table(self):
        """创建数据预览表格"""
        # 创建Canvas和Scrollbar
        # 先创建一个容器来放置Canvas和滚动条，确保布局正确
        self.preview_container = tk.Frame(self.data_preview_tab)
        self.preview_container.pack(fill="both", expand=True)
        
        # 创建Canvas
        self.preview_canvas = tk.Canvas(self.preview_container)
        
        # 创建垂直滚动条
        self.preview_vscroll = tk.Scrollbar(self.preview_container, orient="vertical", command=self.preview_canvas.yview)
        self.preview_canvas.configure(yscrollcommand=self.preview_vscroll.set)
        
        # 创建水平滚动条
        self.preview_hscroll = tk.Scrollbar(self.preview_container, orient="horizontal", command=self.preview_canvas.xview)
        self.preview_canvas.configure(xscrollcommand=self.preview_hscroll.set)
        
        # 使用grid布局来确保滚动条正确放置
        self.preview_vscroll.grid(row=0, column=1, sticky="ns")
        self.preview_hscroll.grid(row=1, column=0, sticky="ew")
        self.preview_canvas.grid(row=0, column=0, sticky="nsew")
        
        # 设置容器的列和行权重，使Canvas能够正确扩展
        self.preview_container.grid_columnconfigure(0, weight=1)
        self.preview_container.grid_rowconfigure(0, weight=1)
        
        # 创建表格容器
        self.table_frame = tk.Frame(self.preview_canvas)
        self.preview_canvas_window = self.preview_canvas.create_window((0, 0), window=self.table_frame, anchor="nw")
        
        # 绑定事件，确保Canvas窗口正确调整宽度和高度
        def on_configure(event):
            # 对于大量列的文件，确保滚动区域包含所有列
            try:
                # 获取所有内容的边界框
                bbox = self.preview_canvas.bbox("all")
                if bbox:
                    self.preview_canvas.configure(scrollregion=bbox)
                else:
                    # 如果边界框获取失败，使用替代方法
                    table_width = self.table_frame.winfo_width()
                    table_height = self.table_frame.winfo_height()
                    self.preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
            except Exception as e:
                # 发生异常时使用安全的默认值
                self.preview_canvas.configure(scrollregion=(0, 0, 1000, 500))
        self.table_frame.bind("<Configure>", on_configure)
        
        # 添加鼠标滚轮支持
        def on_mousewheel(event):
            if event.state & 0x0008:  # 检查是否按下了Shift键
                self.preview_canvas.xview_scroll(-1 * (event.delta // 120), "units")  # 横向滚动
            else:
                self.preview_canvas.yview_scroll(-1 * (event.delta // 120), "units")  # 纵向滚动
        
        # 为Windows系统绑定鼠标滚轮事件
        self.preview_canvas.bind_all("<MouseWheel>", on_mousewheel)
        # 为Linux系统绑定鼠标滚轮事件
        self.preview_canvas.bind_all("<Button-4>", lambda e: self.preview_canvas.yview_scroll(-1, "units"))
        self.preview_canvas.bind_all("<Button-5>", lambda e: self.preview_canvas.yview_scroll(1, "units"))
        

        
    def reset_data_preview_table(self):
        """重置数据预览表格"""
        # 移除所有现有的表格内容
        for widget in self.table_frame.winfo_children():
            widget.destroy()
            
    def save_data_to_excel(self):
        """将处理后的数据保存为Excel文件，包含3个工作表"""
        # 初始化format_cells字典（如果不存在），避免后续操作出错
        if not hasattr(self, 'format_cells'):
            self.format_cells = {}
            
        # 记录开始保存操作
        try:
            if logger:
                logger.info("开始保存数据到Excel文件")
        except Exception:
            pass
            
        # 切换到数据处理选项卡
        if hasattr(self, 'tab_control') and hasattr(self, 'data_processing_tab'):
            self.tab_control.select(self.data_processing_tab)
        
        try:
            # 检查是否有处理后的数据
            if not hasattr(self, 'processed_data') or self.processed_data is None or self.processed_data.empty:
                self.update_status("No data available to save.")
                return
            
            # 弹出文件保存对话框
            import tkinter.filedialog as filedialog
            import os
            import pandas as pd
            
            # 获取当前日期时间作为默认文件名的一部分
            import datetime
            current_datetime = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
            default_filename = f"ProcessedData_{current_datetime}.xlsx"
            
            # 获取默认保存目录 - 第一个选中的原csv文件的目录
            default_dir = ""
            # 获取选中的文件列表
            selected_files = [path for path, var in self.file_vars.items() if var.get()]
            if selected_files:
                # 获取第一个选中文件的目录
                first_file_path = selected_files[0]
                default_dir = os.path.dirname(first_file_path)
                self.update_status(f"Default save directory set to the directory of the first selected file: {default_dir}")
            else:
                self.update_status("No files selected. Using current directory as default save location.")
            
            # 弹出保存对话框，设置默认目录
            # 添加异常处理来捕获可能的TclError
            try:
                # 验证default_dir是否存在，不存在则使用当前目录
                if default_dir and not os.path.isdir(default_dir):
                    self.update_status(f"Warning: Default directory does not exist: {default_dir}. Using current directory instead.")
                    default_dir = ""
                
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=default_filename,
                    initialdir=default_dir
                )
            except Exception as e:
                self.update_status(f"Error opening save dialog: {str(e)}. Trying without initial directory...")
                # 尝试不带initialdir参数打开对话框
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                        initialfile=default_filename
                    )
                except Exception as e2:
                    self.update_status(f"Failed to open save dialog: {str(e2)}")
                    return
            
            # 如果用户取消保存，返回
            if not file_path:
                return
            
            # 检查是否安装了openpyxl库，如果没有则尝试安装
            try:
                import openpyxl
                from openpyxl.styles import PatternFill
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                self.update_status("Installing required libraries, please wait...")
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import PatternFill
                from openpyxl.utils.dataframe import dataframe_to_rows
                
            # 保存数据到Excel文件
            self.update_status(f"Saving processed data to {os.path.basename(file_path)}...")
            
            # 记录数据基本信息用于调试
            data_shape = self.processed_data.shape
            self.update_status(f"Data shape: {data_shape[0]} rows x {data_shape[1]} columns")
            
            # 检查format_cells字典（如果存在）
            if hasattr(self, 'format_cells'):
                # 添加详细日志
                try:
                    import logging
                    logging.info(f"format_cells exists: {bool(self.format_cells)}")
                except:
                    pass
                    
                if self.format_cells:
                    total_format_cells = sum(len(cols) for cols in self.format_cells.values())
                    self.update_status(f"format_cells包含 {total_format_cells} 个单元格")
                    # 记录所有需要格式化的单元格索引
                    try:
                        import logging
                        logging.info(f"format_cells content: {dict(self.format_cells)}")
                        logging.info(f"Total format_cells: {total_format_cells}")
                    except:
                        pass
                        
                    # 验证format_cells的索引范围
                    max_row_idx = max(self.format_cells.keys()) if self.format_cells else -1
                    max_col_idx = max(max(cols) for cols in self.format_cells.values()) if self.format_cells else -1
                    self.update_status(f"format_cells maximum indices: row={max_row_idx}, column={max_col_idx}")
                else:
                    self.update_status("format_cells dictionary is empty, no cells to format.")
                    # 添加日志
                    try:
                        import logging
                        logging.info("format_cells dictionary is empty")
                    except:
                        pass
            else:
                self.update_status("format_cells attribute does not exist, no format will be applied.")
                # 添加日志
                try:
                    import logging
                    logging.info("format_cells attribute does not exist")
                except:
                    pass
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 创建填充样式
            yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 淡黄色 - 用于FAIL或含#的单元格
            light_red_fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")  # 淡红色 - 备用
            light_blue_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")  # 淡蓝色 - 用于其他需要特殊标记的单元格
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色 - 用于PASS单元格
            
            # 定义列筛选函数
            def create_filtered_df(df, include_func):
                # 保持原始列顺序，只筛选符合条件的列
                filtered_columns = [col for col in df.columns if include_func(col)]
                return df[filtered_columns]
            
            # 处理工作表，取消metric工作表，从color point工作表开始
            sheets_config = [
                {
                    "name": "color point",
                    "filter_func": lambda col, idx=enumerate(self.processed_data.columns): idx[0] < 7 or "Avg" in col,
                    "desc": "保留前七列及标题中包含'Avg'的列"
                },
                {
                    "name": "criteria",
                    "is_criteria": True,
                    "desc": "存储Criteria标签页中的White和Mixed标准数据"
                },
                {
                    "name": "Cpk",
                    "is_cpk": True,
                    "desc": "存储Cpk标签页中的White和Mixed统计数据"
                },
                {
                    "name": "Top Defects",
                    "is_top_defects": True,
                    "desc": "存储Top Defects标签页中的不良项目统计数据"
                },
                {
                    "name": "Yield Analysis",
                    "is_yield_analysis": True,
                    "desc": "存储Yield Analysis标签页中的良率分析数据"
                },
                {
                    "name": "ColorPointSpec",
                    "is_colorpoint_spec": True,
                    "desc": "存储ColorPointSpec选项卡中的u'v'坐标点数据"
                }
            ]
            
            # 为每个工作表创建数据和应用格式
            for config in sheets_config:
                sheet_name = config["name"]
                sheet_desc = config["desc"]
                
                # 检查是否是criteria特殊工作表
                if config.get("is_criteria", False):
                    # 处理criteria特殊工作表
                    self.update_status(f"Creating worksheet '{sheet_name}' ({sheet_desc})...")
                    # 创建新工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 尝试获取criteria数据
                    criteria_data = None
                    try:
                        # 使用已有的_save_criteria_to_temp_file方法获取criteria数据
                        temp_file_path = self._save_criteria_to_temp_file()
                        if temp_file_path:
                            import csv
                            criteria_data = {}
                            with open(temp_file_path, 'r', encoding='utf-8') as f:
                                reader = csv.DictReader(f)
                                for row in reader:
                                    data_type = row['DataType']
                                    if data_type not in criteria_data:
                                        criteria_data[data_type] = []
                                    # 将数据添加到对应的数据类型列表中
                                    criteria_data[data_type].append((
                                        row['StandardType'],
                                        row['LowerLimit'],
                                        row['UpperLimit'],
                                        row['Description']
                                    ))
                            # 添加日志记录
                            try:
                                import logging
                                logging.info(f"Successfully loaded criteria data from CSV file. Data types: {list(criteria_data.keys())}")
                            except:
                                pass
                    except Exception as e:
                        # 添加日志记录
                        try:
                            import logging
                            logging.error(f"Error loading criteria data: {str(e)}")
                            logging.debug(f"Error details: {repr(e)}")
                        except:
                            pass
                        self.update_status(f"Error loading criteria data: {str(e)}")
                    
                    # 如果获取到criteria数据，写入工作表
                    if criteria_data:
                        # 设置表头，增加上下限列
                        headers = ["DataType", "Standard Type", "Lower Limit", "Upper Limit", "Description"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                        
                        # 定义标准类型的描述映射
                        standard_descriptions = {
                            "L": "Luminance (cd/m^2)",
                            "U": "Uniformity (%)",
                            "dY": "dY (%/cm)",
                            "u": "u' Coordinate",
                            "v": "v' Coordinate",
                            "Ru": "Ru",
                            "Rv": "Rv",
                            "Du": "Du",
                            "Dv": "Dv",
                            "dL*Min": "dL*Min (%/cm)",
                            "dL*Max": "dL*Max (%/cm)",
                            "dEMax": "dEMax (%/cm)"
                        }
                        
                        row_idx = 2
                        # 写入White和Mixed数据，按照_display_criteria_data方法的逻辑处理规格值
                        for data_type in ["White", "Mixed"]:
                            if data_type in criteria_data:
                                for std_type, lower_limit, upper_limit, description in criteria_data[data_type]:
                                    # 已经从CSV中获取了解析好的上下限和描述，无需再次解析
                                    # 只需进行最后的检查和修正
                                    
                                    # 修正规格值: uniformity规格上限修正为100，规格下限为-1的修正为0
                                    if std_type == "U" and upper_limit == "0":
                                        upper_limit = "100"
                                    if lower_limit == "-1":
                                        lower_limit = "0"
                                    
                                    # 写入数据，尝试将上下限转换为数字格式
                                    ws.cell(row=row_idx, column=1, value=data_type)
                                    ws.cell(row=row_idx, column=2, value=std_type)
                                    
                                    # 尝试将下限转换为数字
                                    if lower_limit:
                                        try:
                                            ws.cell(row=row_idx, column=3, value=float(lower_limit))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=3, value=lower_limit)
                                    else:
                                        ws.cell(row=row_idx, column=3, value="")
                                    
                                    # 尝试将上限转换为数字
                                    if upper_limit:
                                        try:
                                            ws.cell(row=row_idx, column=4, value=float(upper_limit))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=4, value=upper_limit)
                                    else:
                                        ws.cell(row=row_idx, column=4, value="")
                                    
                                    ws.cell(row=row_idx, column=5, value=description)
                                    row_idx += 1
                        
                        self.update_status(f"Successfully wrote criteria data to worksheet '{sheet_name}'")
                    else:
                        self.update_status(f"No criteria data available for worksheet '{sheet_name}'")
                    
                    continue  # 跳过常规处理逻辑
                
                # 检查是否是Cpk特殊工作表
                if config.get("is_cpk", False):
                    # 处理Cpk特殊工作表
                    self.update_status(f"Creating worksheet '{sheet_name}' ({sheet_desc})...")
                    # 创建新工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 尝试获取Cpk数据
                    cpk_data = {}
                    try:
                        # 分别获取White和Mixed的Cpk数据
                        for data_type in ["White", "Mixed"]:
                            self.update_status(f"Calculating Cpk data for {data_type}...")
                            cpk_data[data_type] = self._calculate_cpk_data(data_type)
                        # 添加日志记录
                        try:
                            import logging
                            logging.info(f"Successfully loaded Cpk data. Data types: {list(cpk_data.keys())}")
                        except:
                            pass
                    except Exception as e:
                        # 添加日志记录
                        try:
                            import logging
                            logging.error(f"Error loading Cpk data: {str(e)}")
                        except:
                            pass
                        self.update_status(f"Error loading Cpk data: {str(e)}")
                    
                    # 如果获取到Cpk数据，写入工作表
                    if cpk_data:
                        # 设置表头
                        headers = ["DataType", "Test Item", "Spec", "Mean", "StdDev", "Cpk", "Total", "Fail", "Fail Rate"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                        
                        row_idx = 2
                        # 写入White和Mixed数据
                        for data_type in ["White", "Mixed"]:
                            if data_type in cpk_data and cpk_data[data_type]:
                                for row in cpk_data[data_type]:
                                    # 获取列名作为测试项
                                    column_name = row["column_name"]
                                    
                                    # 确定要显示的规格值（与show_cpk_tab中保持一致的逻辑）
                                    if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                                        spec_value = row['upper_limit']
                                    elif ('White' in column_name or 'Mixed' in column_name) and ('L' in column_name or 'U' in column_name):
                                        spec_value = row['lower_limit']
                                    else:
                                        if row['lower_limit'] == 0 and row['upper_limit'] != 0:
                                            spec_value = row['upper_limit']
                                        elif row['upper_limit'] == 0 and row['lower_limit'] != 0:
                                            spec_value = row['lower_limit']
                                        else:
                                            spec_value = f"{row['lower_limit']}-{row['upper_limit']}"
                                    
                                    # 格式化规格值
                                    if isinstance(spec_value, (int, float)):
                                        # 检查是否包含特殊关键字来决定小数位数
                                        special_keywords = ["Ru", "Rv", "Du", "Dv"]
                                        is_special = any(keyword in column_name for keyword in special_keywords)
                                        decimal_places = 4 if is_special else 2
                                        formatted_spec = f"{spec_value:.{decimal_places}f}"
                                    else:
                                        formatted_spec = str(spec_value)
                                    
                                    # 写入数据
                                    ws.cell(row=row_idx, column=1, value=data_type)
                                    ws.cell(row=row_idx, column=2, value=column_name)
                                    
                                    # 确保Spec列为数字格式: 如果不是范围格式则尝试保存为数字
                                    if isinstance(spec_value, (int, float)) or (isinstance(formatted_spec, str) and '-' not in formatted_spec):
                                        try:
                                            ws.cell(row=row_idx, column=3, value=float(formatted_spec))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=3, value=formatted_spec)
                                    else:
                                        ws.cell(row=row_idx, column=3, value=formatted_spec)
                                    
                                    # 写入数值数据，尝试转换为数字
                                    try:
                                        ws.cell(row=row_idx, column=4, value=float(row['mean']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=4, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=5, value=float(row['std_dev']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=5, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=6, value=float(row['cpk']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=6, value="")
                                    
                                    ws.cell(row=row_idx, column=7, value=row['total_count'])
                                    ws.cell(row=row_idx, column=8, value=row['fail_count'])
                                    # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                                    cell = ws.cell(row=row_idx, column=9, value=row['fail_rate']/100)  # 存储原始数值
                                    cell.number_format = '0.00%'  # 应用Excel百分比格式
                                    
                                    row_idx += 1
                        
                        self.update_status(f"Successfully wrote Cpk data to worksheet '{sheet_name}'")
                    else:
                        self.update_status(f"No Cpk data available for worksheet '{sheet_name}'")
                    
                    continue  # 跳过常规处理逻辑
                
                # 检查是否是Top Defects特殊工作表
                if config.get("is_top_defects", False):
                    # 处理Top Defects特殊工作表
                    self.update_status(f"Creating worksheet '{sheet_name}' ({sheet_desc})...")
                    # 创建新工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    try:
                        # 生成Top Defects数据
                        # 1. Top Defect Items数据
                        top10_data = []
                        # 计算总数量（总行数）
                        total_count = len(self.processed_data)
                        
                        # 统计从第8列开始的所有列的不良情况
                        column_fail_counts = {}
                        format_cells_exists = hasattr(self, 'format_cells') and self.format_cells
                        
                        for col_idx in range(7, len(self.processed_data.columns)):
                            column_name = self.processed_data.columns[col_idx]
                            
                            # 统计该列中被标记为需要格式化的单元格数量
                            fail_count = 0
                            
                            if format_cells_exists:
                                for row_idx, col_indices in self.format_cells.items():
                                    if col_idx in col_indices:
                                        fail_count += 1
                            else:
                                # 额外的统计逻辑: 直接检查数据中的FAIL值或包含#的单元格
                                for row_idx in range(total_count):
                                    try:
                                        cell_value = str(self.processed_data.iloc[row_idx, col_idx])
                                        if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                            fail_count += 1
                                    except:
                                        continue
                            
                            # 忽略特定的v Avg项目
                            if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                continue
                            
                            # 重命名特定的u Avg项目
                            display_name = column_name
                            if column_name == "White u Avg":
                                display_name = "CAFL0 Color Point"
                            elif column_name == "Mixed u Avg":
                                display_name = "CAFL24 Color Point"
                            
                            if fail_count > 0:
                                fail_rate = (fail_count / total_count) * 100 if total_count > 0 else 0
                                top10_data.append((display_name, fail_count, total_count, fail_rate))
                        
                        # 按不良数量降序排序
                        top10_data.sort(key=lambda x: x[1], reverse=True)
                        
                        # 2. Config Group数据
                        config_data = []
                        if 'Config' in self.processed_data.columns:
                            # 获取所有不同的Config值
                            configs = self.processed_data['Config'].unique()
                            
                            # 遍历每个Config
                            for config in configs:
                                # 获取当前Config的数据
                                config_data_df = self.processed_data[self.processed_data['Config'] == config]
                                config_total = len(config_data_df)
                                
                                # 统计每个不良项目在该Config下的不良数量
                                for col_idx in range(7, len(self.processed_data.columns)):
                                    column_name = self.processed_data.columns[col_idx]
                                    
                                    # 忽略特定的v Avg项目
                                    if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                        continue
                                        
                                    fail_count = 0
                                    
                                    if format_cells_exists:
                                        # 只统计当前Config的数据行
                                        for row_idx, col_indices in self.format_cells.items():
                                            if row_idx < len(self.processed_data) and self.processed_data.iloc[row_idx]['Config'] == config:
                                                if col_idx in col_indices:
                                                    fail_count += 1
                                    else:
                                        # 直接检查当前Config数据中的FAIL值或包含#的单元格
                                        for row_idx in range(len(config_data_df)):
                                            try:
                                                original_row_idx = config_data_df.index[row_idx]
                                                cell_value = str(self.processed_data.iloc[original_row_idx, col_idx])
                                                if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                                    fail_count += 1
                                            except:
                                                continue
                                    
                                    # 处理不良项目名称
                                    display_name = column_name
                                    if column_name == "White u Avg":
                                        display_name = "CAFL0 Color Point"
                                    elif column_name == "Mixed u Avg":
                                        display_name = "CAFL24 Color Point"
                                    
                                    if fail_count > 0:
                                        fail_rate = (fail_count / config_total) * 100 if config_total > 0 else 0
                                        config_data.append((config, display_name, fail_count, config_total, fail_rate))
                        
                        # 写入Top Defect Items数据
                        ws.cell(row=1, column=1, value="Top Defect Items")
                        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                        
                        # 设置表头
                        top10_headers = ["No", "Fail Item", "Fail Count", "Total Count", "Fail Rate"]
                        for c_idx, header in enumerate(top10_headers, 1):
                            ws.cell(row=2, column=c_idx, value=header)
                            ws.cell(row=2, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        
                        # 写入数据
                        row_idx = 3
                        for idx, (item_name, fail_count, total, fail_rate) in enumerate(top10_data, start=1):
                            ws.cell(row=row_idx, column=1, value=str(idx))
                            ws.cell(row=row_idx, column=2, value=item_name)
                            ws.cell(row=row_idx, column=3, value=fail_count)
                            ws.cell(row=row_idx, column=4, value=total)
                            # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                            cell = ws.cell(row=row_idx, column=5, value=fail_rate/100)  # 存储原始数值
                            cell.number_format = '0.00%'  # 应用Excel百分比格式
                            row_idx += 1
                        
                        # 添加空行分隔
                        row_idx += 2
                        
                        # 写入Config Group数据
                        ws.cell(row=row_idx, column=1, value="Config Group Fail Items")
                        ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                        row_idx += 1
                        
                        # 设置表头
                        config_headers = ["No", "Config", "Fail Item", "Fail Count", "Total Count", "Fail Rate"]
                        for c_idx, header in enumerate(config_headers, 1):
                            ws.cell(row=row_idx, column=c_idx, value=header)
                            ws.cell(row=row_idx, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        row_idx += 1
                        
                        # 按Config分组并排序
                        if config_data:
                            # 先按Config分组，再在组内按不良率降序排序
                            config_groups = {}
                            for config, item_name, fail_count, total, fail_rate in config_data:
                                if config not in config_groups:
                                    config_groups[config] = []
                                config_groups[config].append((item_name, fail_count, total, fail_rate))
                            
                            # 为每个Config组内的项目按不良率降序排序
                            for config in config_groups:
                                config_groups[config].sort(key=lambda x: x[3], reverse=True)
                            
                            # 写入数据
                            for config in sorted(config_groups.keys()):
                                config_total = len(self.processed_data[self.processed_data['Config'] == config])
                                # 添加Config分组标题行
                                ws.cell(row=row_idx, column=2, value=f"{config} (Total: {config_total})")
                                ws.cell(row=row_idx, column=2).font = openpyxl.styles.Font(bold=True)
                                row_idx += 1
                                
                                # 添加该Config下的不良项目
                                for idx, (item_name, fail_count, total, fail_rate) in enumerate(config_groups[config], start=1):
                                    ws.cell(row=row_idx, column=1, value=str(idx))
                                    ws.cell(row=row_idx, column=2, value=config)
                                    ws.cell(row=row_idx, column=3, value=item_name)
                                    ws.cell(row=row_idx, column=4, value=fail_count)
                                    ws.cell(row=row_idx, column=5, value=total)
                                    # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                                    cell = ws.cell(row=row_idx, column=6, value=fail_rate/100)  # 存储原始数值
                                    cell.number_format = '0.00%'  # 应用Excel百分比格式
                                    row_idx += 1
                                
                                # 添加一个空行分隔不同Config
                                row_idx += 1
                        else:
                            ws.cell(row=row_idx, column=1, value="No Config data found in the file, cannot perform Config grouping statistics")
                        
                        self.update_status(f"Successfully wrote Top Defects data to worksheet '{sheet_name}'")
                    except Exception as e:
                        self.update_status(f"Error writing Top Defects data: {str(e)}")
                    
                    continue  # 跳过常规处理逻辑
                
                # 检查是否是Yield Analysis特殊工作表
                if config.get("is_yield_analysis", False):
                    # 处理Yield Analysis特殊工作表
                    self.update_status(f"Creating worksheet '{sheet_name}' ({sheet_desc})...")
                    # 创建新工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    try:
                        # 生成Yield Analysis数据
                        # 计算总数量和不良数量
                        total_count = len(self.processed_data)
                        fail_count = 0
                        
                        if 'Pass/Fail' in self.processed_data.columns:
                            # 统计"Pass/Fail"列中"FAIL"的值
                            fail_count = (self.processed_data['Pass/Fail'] == 'FAIL').sum()
                        
                        # 计算总体不良率
                        total_fail_rate = (fail_count / total_count * 100) if total_count > 0 else 0
                        
                        # 按Config分组统计不良率
                        config_stats = []
                        if 'Config' in self.processed_data.columns and 'Pass/Fail' in self.processed_data.columns:
                            # 按Config分组
                            for config, group in self.processed_data.groupby('Config'):
                                config_total = len(group)
                                config_fail = (group['Pass/Fail'] == 'FAIL').sum()
                                config_fail_rate = (config_fail / config_total * 100) if config_total > 0 else 0
                                config_stats.append((config, config_total, config_fail, config_fail_rate))
                            
                            # 按不良率降序排序
                            config_stats.sort(key=lambda x: x[3], reverse=True)
                        
                        # 设置表头
                        headers = ["Config", "Total", "Fail Count", "Fail Rate"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                            ws.cell(row=1, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        
                        # 写入Config分组数据
                        row_idx = 2
                        if config_stats:
                            for config, total, fail, rate in config_stats:
                                ws.cell(row=row_idx, column=1, value=f"Config: {config}")
                                ws.cell(row=row_idx, column=2, value=total)
                                ws.cell(row=row_idx, column=3, value=fail)
                                # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                                cell = ws.cell(row=row_idx, column=4, value=rate/100)  # 存储原始数值
                                cell.number_format = '0.00%'  # 应用Excel百分比格式
                                row_idx += 1
                            
                            # 添加分隔行
                            ws.cell(row=row_idx, column=1, value="-" * 50)
                            row_idx += 1
                        
                        # 写入总体统计数据
                        ws.cell(row=row_idx, column=1, value="Total")
                        ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
                        ws.cell(row=row_idx, column=2, value=total_count)
                        ws.cell(row=row_idx, column=3, value=fail_count)
                        # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                        cell = ws.cell(row=row_idx, column=4, value=total_fail_rate/100)  # 存储原始数值
                        cell.number_format = '0.00%'  # 应用Excel百分比格式
                        
                        self.update_status(f"Successfully wrote Yield Analysis data to worksheet '{sheet_name}'")
                    except Exception as e:
                        self.update_status(f"Error writing Yield Analysis data: {str(e)}")
                    
                    continue  # 跳过常规处理逻辑
                
                # 常规工作表处理
                filter_func = config["filter_func"]
                
                self.update_status(f"Creating worksheet '{sheet_name}' ({sheet_desc})...")
                
                # 如果是第一个工作表，使用默认的活动工作表，否则创建新工作表
                if sheet_name == sheets_config[0]["name"]:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                # 为第一个工作表(metric)使用reprocessed_data，如果存在的话
                data_source = self.processed_data
                if sheet_name == "metric" and hasattr(self, 'reprocessed_data') and self.reprocessed_data is not None and not self.reprocessed_data.empty:
                    data_source = self.reprocessed_data
                    try:
                        import logging
                        logging.info("第一个工作表使用data reprocessing后的数据")
                    except:
                        pass
                
                # 创建索引到列名的映射，以正确处理color point的前七列逻辑
                col_index_map = {i: col for i, col in enumerate(data_source.columns)}
                
                # 应用筛选条件，保持原始列顺序
                if sheet_name == "color point":
                    # 特殊处理color point工作表，保留前七列和包含Avg的列
                    filtered_columns = []
                    # 首先添加前七列（如果有的话）
                    filtered_columns.extend([col for i, col in col_index_map.items() if i < 7])
                    # 然后添加标题中包含Avg的列（但不重复添加）
                    avg_columns = [col for col in data_source.columns if "Avg" in col and col not in filtered_columns]
                    filtered_columns.extend(avg_columns)
                else:
                    # 其他工作表使用普通筛选
                    filtered_columns = [col for col in data_source.columns if filter_func(col)]
                
                # 如果没有符合条件的列，跳过该工作表
                if not filtered_columns:
                    self.update_status(f"Warning: Worksheet '{sheet_name}' has no columns meeting the criteria.")
                    continue
                
                # 创建筛选后的数据框
                filtered_df = data_source[filtered_columns]
                filtered_shape = filtered_df.shape
                self.update_status(f"Filtered data shape: {filtered_shape[0]} rows x {filtered_shape[1]} columns")
                
                # 将DataFrame数据写入工作表
                self.update_status(f"Writing data to worksheet '{sheet_name}'...")
                try:
                    row_count = 0
                    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
                        row_count += 1
                        for c_idx, value in enumerate(row, 1):
                            try:
                                # 避免None值导致的问题
                                if value is None:
                                    value = ""
                                
                                # 对于第一个和第二个工作表，从数据行开始，将第8列及之后的数据列强制转换为数值
                                if (sheet_name == "metric" or sheet_name == "color point") and r_idx > 1 and c_idx >= 8:
                                    try:
                                        # 尝试将值转换为数值
                                        num_value = float(value)
                                        ws.cell(row=r_idx, column=c_idx, value=num_value)
                                    except (ValueError, TypeError):
                                        # 如果转换失败，保持原始值
                                        ws.cell(row=r_idx, column=c_idx, value=value)
                                else:
                                    # 其他情况保持原始值
                                    ws.cell(row=r_idx, column=c_idx, value=value)
                            except Exception as cell_error:
                                self.update_status(f"Error writing cell ({r_idx},{c_idx}): {str(cell_error)}")
                    self.update_status(f"Successfully wrote {row_count} rows of data to worksheet '{sheet_name}', and converted data in columns 8 and onwards to numeric format for 'metric' and 'color point' worksheets.")
                except Exception as write_error:
                    self.update_status(f"Error writing data to worksheet '{sheet_name}': {str(write_error)}")
                    # 尝试只写入数据，不应用格式
                    self.update_status(f"Attempting to write data to worksheet '{sheet_name}' without applying formats...")
                    # 清除当前工作表内容
                    ws.delete_rows(1, ws.max_row)
                    ws.delete_cols(1, ws.max_column)
                    # 重新写入数据
                    for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            if value is None:
                                value = ""
                            ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 如果有需要格式化的单元格，应用颜色填充
                formatted_count = 0
                error_count = 0
                skip_count = 0
                
                # 首先检查是否是第一个工作表(metric)并且有reprocessed_data
                # 注意：由于metric工作表已被取消，此逻辑不再执行
                if False:  # 原逻辑：sheet_name == "metric" and hasattr(self, 'reprocessed_data') and self.reprocessed_data is not None and not self.reprocessed_data.empty:
                    self.update_status(f"Applying cell formats to worksheet '{sheet_name}' using reprocessed data...")
                    
                    # 为第一个工作表完全重写颜色填充逻辑
                    # 遍历所有数据行（跳过表头）
                    for r_idx in range(2, row_count + 1):
                        # 转换为reprocessed_data的索引（Excel行-2）
                        data_row_idx = r_idx - 2
                        
                        # 检查数据行索引是否有效
                        if data_row_idx < len(self.reprocessed_data):
                            # 先检查整行是否包含fail（不区分大小写）
                            row_has_fail = False
                            for col_name in filtered_columns:
                                if col_name in self.reprocessed_data.columns:
                                    cell_value = self.reprocessed_data.iloc[data_row_idx][col_name]
                                    cell_value_str = str(cell_value).strip().lower() if cell_value is not None else ''
                                    if 'fail' in cell_value_str:
                                        row_has_fail = True
                                        break
                            
                            # 遍历所有数据列
                            for c_idx in range(1, filtered_shape[1] + 1):
                                # 获取对应的列名
                                if c_idx - 1 < len(filtered_columns):
                                    col_name = filtered_columns[c_idx - 1]
                                    
                                    try:
                                        # 获取单元格值
                                        cell_value = self.reprocessed_data.iloc[data_row_idx][col_name]
                                        cell_value_str_upper = str(cell_value).strip().upper() if cell_value is not None else ''
                                        cell_value_str_lower = str(cell_value).strip().lower() if cell_value is not None else ''
                                        
                                        # 确定单元格填充颜色
                                        if 'PASS' in cell_value_str_upper:
                                            # PASS结果使用白色填充
                                            ws.cell(row=r_idx, column=c_idx).fill = white_fill
                                            formatted_count += 1
                                        elif 'FAIL' in cell_value_str_upper or '#' in cell_value_str_lower:
                                            # FAIL结果或包含#的单元格使用淡黄色填充
                                            ws.cell(row=r_idx, column=c_idx).fill = yellow_fill
                                            formatted_count += 1
                                        elif row_has_fail:
                                            # 如果整行包含fail且当前单元格不是FAIL或PASS，则应用淡蓝色填充
                                            ws.cell(row=r_idx, column=c_idx).fill = light_blue_fill
                                            formatted_count += 1
                                    except Exception as cell_error:
                                        error_count += 1
                                        try:
                                            import logging
                                            logging.error(f"Error formatting cell ({r_idx},{c_idx}): {str(cell_error)}")
                                        except:
                                            pass
                # 对于其他工作表或没有reprocessed_data的情况，使用原有的format_cells逻辑
                elif hasattr(self, 'format_cells') and self.format_cells:
                    self.update_status(f"Applying cell formats to worksheet '{sheet_name}'...")
                    # 添加详细日志
                    try:
                        import logging
                        logging.info(f"Applying formats to worksheet '{sheet_name}', total cells to format: {sum(len(cols) for cols in self.format_cells.values())}")
                    except:
                        pass
                        
                    # 创建原始列名到筛选后列索引的映射
                    original_to_filtered_col = {col: i for i, col in enumerate(filtered_columns)}
                    # 记录列映射信息
                    try:
                        import logging
                        logging.info(f"Column mapping for worksheet '{sheet_name}': {original_to_filtered_col}")
                    except:
                        pass
                        
                    # 注意: DataFrame的索引从0开始，而Excel行索引从1开始（并且第一行是表头）
                    for row_idx, original_col_indices in self.format_cells.items():
                        # 加2是因为Excel第一行是表头，DataFrame行索引从0开始
                        excel_row = row_idx + 2
                        # 检查行索引是否超出范围
                        if excel_row > row_count:
                            self.update_status(f"Warning: Row index {excel_row} exceeds data range (max {row_count})")
                            try:
                                import logging
                                logging.warning(f"Skipping format for row {excel_row} (exceeds data range)")
                            except:
                                pass
                            skip_count += len(original_col_indices)
                            continue
                        
                        # 遍历原始列索引，检查是否在当前工作表中
                        for original_col_idx in original_col_indices:
                            # 获取原始列名，使用当前数据源检查索引范围，确保一致性
                            if original_col_idx < len(data_source.columns):
                                # 使用正确的数据源获取原始列名，确保索引一致性
                                original_col_name = data_source.columns[original_col_idx]
                                # 检查该列是否在当前工作表中
                                if original_col_name in original_to_filtered_col:
                                    # 转换为当前工作表的列索引
                                    filtered_col_idx = original_to_filtered_col[original_col_name]
                                    # Excel列索引从1开始
                                    excel_col = filtered_col_idx + 1
                                    # 检查列索引是否超出范围
                                    if excel_col > filtered_shape[1]:
                                        self.update_status(f"Warning: Column index {excel_col} exceeds data range (max {filtered_shape[1]})")
                                        try:
                                            import logging
                                            logging.warning(f"Skipping format for column {excel_col} (exceeds data range)")
                                        except:
                                            pass
                                        skip_count += 1
                                        continue
                                    try:
                                        # 获取单元格值以决定填充颜色
                                        data_row_idx = row_idx
                                        if data_row_idx < len(data_source) and original_col_name in data_source.columns:
                                            cell_value = data_source.iloc[data_row_idx][original_col_name]
                                            cell_value_str_upper = str(cell_value).strip().upper() if cell_value is not None else ''
                                            cell_value_str_lower = str(cell_value).strip().lower() if cell_value is not None else ''
                                            
                                            # 首先检查是否是异常单元格（超出规格上下限或不在多边形内部）
                                            is_exception_cell = 'FAIL' in cell_value_str_upper or '#' in cell_value_str_lower
                                            
                                            if 'PASS' in cell_value_str_upper:
                                                ws.cell(row=excel_row, column=excel_col).fill = white_fill
                                            elif is_exception_cell:
                                                # 异常单元格（超出规格上下限或不在多边形内部），始终填充淡黄色
                                                ws.cell(row=excel_row, column=excel_col).fill = yellow_fill
                                            else:
                                                # 检查整行是否包含fail
                                                row_has_fail = False
                                                for col in data_source.columns:
                                                    row_value = data_source.iloc[data_row_idx][col]
                                                    row_value_str_lower = str(row_value).strip().lower() if row_value is not None else ''
                                                    if 'fail' in row_value_str_lower:
                                                        row_has_fail = True
                                                        break
                                                
                                                if row_has_fail:
                                                    # 整行包含fail，但当前单元格不是异常情况，应用淡蓝色填充
                                                    ws.cell(row=excel_row, column=excel_col).fill = light_blue_fill
                                                else:
                                                    # 默认保持原始颜色
                                                    pass
                                        else:
                                            # 默认使用淡黄色填充
                                            ws.cell(row=excel_row, column=excel_col).fill = yellow_fill
                                            
                                        formatted_count += 1
                                        # 添加详细日志
                                        try:
                                            import logging
                                            logging.info(f"Successfully applied color to cell ({excel_row},{excel_col}) in worksheet '{sheet_name}'")
                                        except:
                                            pass
                                    except Exception as format_error:
                                        self.update_status(f"Error applying format to cell ({excel_row},{excel_col}): {str(format_error)}")
                                        error_count += 1
                                        # 添加日志
                                        try:
                                            import logging
                                            logging.error(f"Error applying format: {str(format_error)}")
                                        except:
                                            pass
                                else:
                                    # 列不在当前工作表中
                                    skip_count += 1
                                    try:
                                        import logging
                                        logging.info(f"Column '{original_col_name}' not in current worksheet '{sheet_name}', skipping format")
                                    except:
                                        pass
                            else:
                                # 列索引超出范围
                                skip_count += 1
                                try:
                                    import logging
                                    logging.warning(f"Original column index {original_col_idx} out of range, skipping format")
                                except:
                                    pass
                    
                    self.update_status(f"Cell format application completed: {formatted_count} successful, {error_count} errors, {skip_count} skipped")
                    # 添加详细日志
                    try:
                        import logging
                        logging.info(f"Format application statistics: {formatted_count} successful, {error_count} errors, {skip_count} skipped")
                    except:
                        pass
            
            # 在保存工作簿之前，统一对所有工作表应用颜色填充逻辑
            self.update_status("Applying uniform color formatting to all worksheets...")
            # 获取工作表列表
            sheet_names = list(wb.sheetnames)
            
            for idx, sheet in enumerate(sheet_names):
                ws = wb[sheet]
                # 查找Pass/Fail列的索引（如果存在）
                pass_fail_col_idx = None
                # 假设第一行是表头
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Pass/Fail':
                        pass_fail_col_idx = col_idx
                        break
                
                # 从第二行开始处理数据行
                for row_idx in range(2, ws.max_row + 1):
                    # 首先检查整行是否包含fail（不区分大小写）
                    row_has_fail = False
                    for col_idx in range(1, ws.max_column + 1):
                        row_cell = ws.cell(row=row_idx, column=col_idx)
                        row_cell_value = str(row_cell.value).strip().lower() if row_cell.value is not None else ''
                        if 'fail' in row_cell_value:
                            row_has_fail = True
                            break
                    
                    # 遍历整行的所有单元格应用颜色填充
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell_value = str(cell.value).strip().upper() if cell.value is not None else ''
                        cell_value_lower = str(cell.value).strip().lower() if cell.value is not None else ''
                        
                        # 确定单元格填充颜色
                        # 检查是否是异常单元格（超出规格上下限或不在多边形内部）
                        is_exception_cell = 'FAIL' in cell_value or '#' in cell_value_lower
                        
                        # 对第一个工作表（现在为color point）应用原有逻辑
                        if idx == 0:  # 第一个工作表（color point）
                            # 使用原有逻辑处理第一个工作表
                            if 'PASS' in cell_value:
                                # PASS结果使用白色填充
                                cell.fill = white_fill
                            elif is_exception_cell:
                                # 异常单元格（超出规格上下限或不在多边形内部），始终填充淡黄色
                                cell.fill = yellow_fill
                            elif row_has_fail:
                                # 如果整行包含fail且当前单元格不是FAIL或PASS，则应用淡蓝色填充
                                cell.fill = light_blue_fill
                            # 其他情况保持默认颜色
                        else:  # 其他工作表取消颜色填充逻辑，全部无填充色
                            # 取消颜色填充，保持默认无填充色
                            pass
            
            # 保存工作簿
            self.update_status("Saving workbook...")
            try:
                wb.save(file_path)
            except Exception as save_error:
                self.update_status(f"Error saving workbook: {str(save_error)}")
                # 尝试使用pandas的to_excel作为备选方案
                self.update_status("Attempting to save data using pandas...")
                try:
                    # 使用pandas创建一个包含工作表的Excel文件（取消metric工作表）
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        # 取消创建工作表1: metric
                        # metric_df = self.processed_data[[col for col in self.processed_data.columns if "Criteria" not in col or col == "Fail_Reason"]]
                        # metric_df.to_excel(writer, sheet_name='metric', index=False)
                        
                        # 创建工作表1: color point（原工作表2，现在变为第一个工作表）
                        # 保留前七列
                        color_point_cols = [col for i, col in enumerate(self.processed_data.columns) if i < 7]
                        # 添加包含Avg的列（但不重复添加）
                        avg_cols = [col for col in self.processed_data.columns if "Avg" in col and col not in color_point_cols]
                        color_point_cols.extend(avg_cols)
                        color_point_df = self.processed_data[color_point_cols]
                        color_point_df.to_excel(writer, sheet_name='color point', index=False)
                        
                        # 不再创建spec工作表
                        
                        # 创建工作表2: criteria（原工作表3，现在变为第二个工作表）
                        # 尝试获取criteria数据
                        criteria_data = None
                        try:
                            # 使用已有的_save_criteria_to_temp_file方法获取criteria数据
                            temp_file_path = self._save_criteria_to_temp_file()
                            if temp_file_path:
                                import csv
                                criteria_data = []
                                with open(temp_file_path, 'r', encoding='utf-8') as f:
                                    reader = csv.DictReader(f)
                                    for row in reader:
                                        criteria_data.append(row)
                        except Exception as e:
                            self.update_status(f"Error loading criteria data for pandas backup: {str(e)}")
                        
                        # 如果获取到criteria数据，写入工作表
                        if criteria_data:
                            import pandas as pd
                            # 直接使用从CSV获取的数据，确保包含Criteria选项卡中White和Mixed子窗口的所有数据
                            criteria_rows = []
                            
                            for row in criteria_data:
                                # 尝试将上下限转换为数字
                                lower_limit_num = None
                                if row['LowerLimit']:
                                    try:
                                        lower_limit_num = float(row['LowerLimit'])
                                    except (ValueError, TypeError):
                                        lower_limit_num = row['LowerLimit']
                                
                                upper_limit_num = None
                                if row['UpperLimit']:
                                    try:
                                        upper_limit_num = float(row['UpperLimit'])
                                    except (ValueError, TypeError):
                                        upper_limit_num = row['UpperLimit']
                                
                                criteria_rows.append({
                                   "DataType": row['DataType'],
                                   "Standard Type": row['StandardType'],
                                   "Lower Limit": lower_limit_num,
                                   "Upper Limit": upper_limit_num,
                                   "Description": row['Description']
                               })
                            
                            # 创建DataFrame并保存
                            if criteria_rows:
                                criteria_df = pd.DataFrame(criteria_rows)
                                criteria_df.to_excel(writer, sheet_name='criteria', index=False)
                                self.update_status("Successfully wrote criteria data to worksheet 'criteria' (pandas backup)")
                        
                        # 添加Cpk工作表（pandas备选方案）
                        try:
                            cpk_rows = []
                            # 分别获取White和Mixed的Cpk数据
                            for data_type in ["White", "Mixed"]:
                                cpk_data = self._calculate_cpk_data(data_type)
                                for row in cpk_data:
                                    # 获取列名作为测试项
                                    column_name = row["column_name"]
                                    
                                    # 确定要显示的规格值（与show_cpk_tab中保持一致的逻辑）
                                    if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                                        spec_value = row['upper_limit']
                                    elif ('White' in column_name or 'Mixed' in column_name) and ('L' in column_name or 'U' in column_name):
                                        spec_value = row['lower_limit']
                                    else:
                                        if row['lower_limit'] == 0 and row['upper_limit'] != 0:
                                            spec_value = row['upper_limit']
                                        elif row['upper_limit'] == 0 and row['lower_limit'] != 0:
                                            spec_value = row['lower_limit']
                                        else:
                                            spec_value = f"{row['lower_limit']}-{row['upper_limit']}"
                                    
                                    # 格式化规格值
                                    if isinstance(spec_value, (int, float)):
                                        # 检查是否包含特殊关键字来决定小数位数
                                        special_keywords = ["Ru", "Rv", "Du", "Dv"]
                                        is_special = any(keyword in column_name for keyword in special_keywords)
                                        decimal_places = 4 if is_special else 2
                                        formatted_spec = f"{spec_value:.{decimal_places}f}"
                                    else:
                                        formatted_spec = str(spec_value)
                                    
                                    cpk_rows.append({
                                        "DataType": data_type,
                                        "Test Item": column_name,
                                        "Spec": formatted_spec,
                                        "Mean": row['mean'],
                                        "StdDev": row['std_dev'],
                                        "Cpk": row['cpk'],
                                        "Total": row['total_count'],
                                        "Fail": row['fail_count'],
                                        "Fail Rate": f"{row['fail_rate']:.2f}%"
                                    })
                            
                            # 创建DataFrame并保存
                            if cpk_rows:
                                cpk_df = pd.DataFrame(cpk_rows)
                                cpk_df.to_excel(writer, sheet_name='Cpk', index=False)
                                self.update_status("Successfully wrote Cpk data to worksheet 'Cpk' (pandas backup)")
                        except Exception as cpk_error:
                            self.update_status(f"Error writing Cpk data (pandas backup): {str(cpk_error)}")
                        
                        # 添加Top Defects工作表（pandas备选方案）
                        try:
                            # 生成Top Defect Items数据
                            top10_rows = []
                            total_count = len(self.processed_data)
                            column_fail_counts = {}
                            format_cells_exists = hasattr(self, 'format_cells') and self.format_cells
                            
                            for col_idx in range(7, len(self.processed_data.columns)):
                                column_name = self.processed_data.columns[col_idx]
                                
                                # 统计不良数量
                                fail_count = 0
                                if format_cells_exists:
                                    for row_idx, col_indices in self.format_cells.items():
                                        if col_idx in col_indices:
                                            fail_count += 1
                                else:
                                    for row_idx in range(total_count):
                                        try:
                                            cell_value = str(self.processed_data.iloc[row_idx, col_idx])
                                            if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                                fail_count += 1
                                        except:
                                            continue
                                
                                # 忽略特定项目并处理重命名
                                if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                    continue
                                
                                display_name = column_name
                                if column_name == "White u Avg":
                                    display_name = "CAFL0 Color Point"
                                elif column_name == "Mixed u Avg":
                                    display_name = "CAFL24 Color Point"
                                
                                if fail_count > 0:
                                    fail_rate = (fail_count / total_count) * 100 if total_count > 0 else 0
                                    top10_rows.append({
                                        "Section": "Top Defect Items",
                                        "No": len(top10_rows) + 1,
                                        "Fail Item": display_name,
                                        "Fail Count": fail_count,
                                        "Total Count": total_count,
                                        "Fail Rate": f"{fail_rate:.2f}%"
                                    })
                            
                            # 按不良数量降序排序
                            top10_rows.sort(key=lambda x: x["Fail Count"], reverse=True)
                            # 更新序号
                            for i, row in enumerate(top10_rows):
                                row["No"] = i + 1
                            
                            # 生成Config Group数据
                            config_rows = []
                            if 'Config' in self.processed_data.columns:
                                configs = self.processed_data['Config'].unique()
                                
                                for config in configs:
                                    config_data_df = self.processed_data[self.processed_data['Config'] == config]
                                    config_total = len(config_data_df)
                                    
                                    for col_idx in range(7, len(self.processed_data.columns)):
                                        column_name = self.processed_data.columns[col_idx]
                                        
                                        if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                            continue
                                            
                                        fail_count = 0
                                        if format_cells_exists:
                                            for row_idx, col_indices in self.format_cells.items():
                                                if row_idx < len(self.processed_data) and self.processed_data.iloc[row_idx]['Config'] == config:
                                                    if col_idx in col_indices:
                                                        fail_count += 1
                                        else:
                                            for row_idx in range(len(config_data_df)):
                                                try:
                                                    original_row_idx = config_data_df.index[row_idx]
                                                    cell_value = str(self.processed_data.iloc[original_row_idx, col_idx])
                                                    if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                                        fail_count += 1
                                                except:
                                                    continue
                                        
                                        display_name = column_name
                                        if column_name == "White u Avg":
                                            display_name = "CAFL0 Color Point"
                                        elif column_name == "Mixed u Avg":
                                            display_name = "CAFL24 Color Point"
                                        
                                        if fail_count > 0:
                                            fail_rate = (fail_count / config_total) * 100 if config_total > 0 else 0
                                            config_rows.append({
                                                "Section": "Config Group",
                                                "Config": config,
                                                "Fail Item": display_name,
                                                "Fail Count": fail_count,
                                                "Total Count": config_total,
                                                "Fail Rate": f"{fail_rate:.2f}%"
                                            })
                            
                            # 合并并保存Top Defects数据
                            top_defects_rows = []
                            if top10_rows:
                                top_defects_rows.append({"Section": "Top Defect Items"})  # 添加标题行
                                top_defects_rows.extend(top10_rows)
                                top_defects_rows.append({})  # 添加空行分隔
                            
                            if config_rows:
                                top_defects_rows.append({"Section": "Config Group Fail Items"})  # 添加标题行
                                # 按Config分组并排序
                                config_groups = {}
                                for row in config_rows:
                                    config = row["Config"]
                                    if config not in config_groups:
                                        config_groups[config] = []
                                    config_groups[config].append(row)
                                
                                # 为每个Config组内的项目按不良率降序排序
                                for config in config_groups:
                                    config_groups[config].sort(key=lambda x: float(x["Fail Rate"].replace('%', '')), reverse=True)
                                    # 添加Config分组标题
                                    top_defects_rows.append({"Section": f"Config: {config}"})
                                    # 添加该Config下的不良项目
                                    for i, row in enumerate(config_groups[config], start=1):
                                        row["No"] = i
                                        top_defects_rows.append(row)
                            
                            # 创建DataFrame并保存
                            if top_defects_rows:
                                top_defects_df = pd.DataFrame(top_defects_rows)
                                top_defects_df.to_excel(writer, sheet_name='Top Defects', index=False)
                                self.update_status("Successfully wrote Top Defects data to worksheet 'Top Defects' (pandas backup)")
                        except Exception as top_defects_error:
                            self.update_status(f"Error writing Top Defects data (pandas backup): {str(top_defects_error)}")
                        
                        # 添加Yield Analysis工作表（pandas备选方案）
                        try:
                            # 生成Yield Analysis数据
                            yield_rows = []
                            total_count = len(self.processed_data)
                            fail_count = 0
                            
                            if 'Pass/Fail' in self.processed_data.columns:
                                fail_count = (self.processed_data['Pass/Fail'] == 'FAIL').sum()
                            
                            total_fail_rate = (fail_count / total_count * 100) if total_count > 0 else 0
                            
                            # 按Config分组统计
                            if 'Config' in self.processed_data.columns and 'Pass/Fail' in self.processed_data.columns:
                                config_stats = []
                                for config, group in self.processed_data.groupby('Config'):
                                    config_total = len(group)
                                    config_fail = (group['Pass/Fail'] == 'FAIL').sum()
                                    config_fail_rate = (config_fail / config_total * 100) if config_total > 0 else 0
                                    config_stats.append({
                                        "Config": f"Config: {config}",
                                        "Total": config_total,
                                        "Fail Count": config_fail,
                                        "Fail Rate": f"{config_fail_rate:.2f}%"
                                    })
                                
                                # 按不良率降序排序
                                config_stats.sort(key=lambda x: float(x["Fail Rate"].replace('%', '')), reverse=True)
                                yield_rows.extend(config_stats)
                                
                                if config_stats:  # 如果有Config数据，添加分隔行
                                    yield_rows.append({"Config": "-" * 50})
                            
                            # 添加总体统计
                            yield_rows.append({
                                "Config": "Total",
                                "Total": total_count,
                                "Fail Count": fail_count,
                                "Fail Rate": f"{total_fail_rate:.2f}%"
                            })
                            
                            # 创建DataFrame并保存
                            if yield_rows:
                                yield_df = pd.DataFrame(yield_rows)
                                yield_df.to_excel(writer, sheet_name='Yield Analysis', index=False)
                                self.update_status("Successfully wrote Yield Analysis data to worksheet 'Yield Analysis' (pandas backup)")
                        except Exception as yield_error:
                            self.update_status(f"Error writing Yield Analysis data (pandas backup): {str(yield_error)}")
                        
                    self.update_status(f"Data successfully saved to {file_path} (using pandas)")
                    return
                except Exception as pandas_error:
                    self.update_status(f"Error saving data using pandas: {str(pandas_error)}")
                    raise
            
            # 更新状态栏信息
            self.update_status(f"Data successfully saved to {file_path}")
            
            # 自动打开保存目录
            try:
                import subprocess
                import sys
                # 确保正确获取文件的完整目录路径
                directory = os.path.abspath(os.path.dirname(file_path))
                
                # 记录完整路径信息用于调试
                self.update_status(f"Full path of saved file: {os.path.abspath(file_path)}")
                self.update_status(f"Full path of save directory: {directory}")
                
                if directory:
                    # 根据操作系统类型选择打开命令
                    if os.name == 'nt':  # Windows
                        # 使用完整的explorer命令格式，确保正确打开指定目录
                        subprocess.Popen(f'explorer "{directory}"', shell=True)
                    else:  # macOS或Linux
                        open_cmd = 'open' if sys.platform == 'darwin' else 'xdg-open'
                        subprocess.Popen([open_cmd, directory])
                    self.update_status(f"Successfully opened save directory: {directory}")
                else:
                    self.update_status("Error: Could not determine a valid save directory path")
            except Exception as open_dir_error:
                self.update_status(f"Error: Failed to automatically open save directory: {str(open_dir_error)}")
                self.update_status(f"Full error details: {traceback.format_exc()}")
            
        except Exception as e:
            # 处理保存过程中可能出现的错误
            import traceback
            error_details = traceback.format_exc()
            
            # 记录详细错误到日志文件
            try:
                if logger:
                    logger.error(f"Failed saving data to {file_path}: {str(e)}")
                    logger.debug(f"Detailed error information:\n{error_details}")
            except Exception:
                pass
                
            self.update_status(f"Error: Failed to save data: {str(e)}")
            # 尝试创建一个简单的错误日志文件
            try:
                with open("save_error_log.txt", "w", encoding="utf-8") as log_file:
                    log_file.write(f"Error: Failed to save data: {str(e)}\n\nDetailed error information:\n{error_details}")
                self.update_status(f"Error: Failed to save data: {str(e)}，Error log saved to save_error_log.txt")
            except:
                pass
        
    def update_data_preview_table_for_multiple_files(self, all_files_data):
        """更新数据预览表格以显示多个文件的数据"""
        # 重置表格
        self.reset_data_preview_table()
        
        if not all_files_data:
            return
        
        # 添加一个变量来跟踪当前是否显示所有行
        if not hasattr(self, 'show_all_rows'):
            self.show_all_rows = False
        
        # 限制显示行数，显示所有列
        max_rows_per_file = 3  # 默认每个文件只显示前3行
        
        # 批量创建所有控件但先不进行布局计算
        all_headers = []
        all_cells = []
        current_row = 1  # 从1开始，因为0行将用于显示文件信息
        
        # 处理第一个文件，用于获取表头信息
        first_file_path, first_file_name, first_df = all_files_data[0]
        
        # 添加行号列的表头
        row_num_header = tk.Label(self.table_frame, text="Row#", font=('Courier New', 10, 'bold'), relief=tk.RAISED, padx=5, pady=2, anchor="center", wraplength=0)
        all_headers.append(row_num_header)
        # 为行号列设置宽度
        self.table_frame.columnconfigure(0, minsize=60, weight=0)
        
        # 统一所有列的表头设置，所有列使用相同的字体和样式
        for j, column in enumerate(first_df.columns, 1):  # 从1开始编号
            # 所有列使用完全相同的字体设置
            header = tk.Label(self.table_frame, text=f"{j}. {column}", 
                              font=('Courier New', 10, 'bold'), 
                              relief=tk.RAISED, 
                              padx=5, 
                              pady=2, 
                              anchor="w", 
                              wraplength=0, 
                              justify='left')
            # 所有列使用相同的宽度设置，移除特殊列差异化处理
            self.table_frame.columnconfigure(j, minsize=140, weight=1)
            all_headers.append(header)
        
        # 一次性应用表头布局
        for j, header in enumerate(all_headers):
            header.grid(row=current_row, column=j, sticky="nsew", ipady=0)
        
        current_row += 1  # 表头占了一行，内容从下一行开始
        
        # 处理每个文件的数据
        for file_idx, (file_path, file_name, df) in enumerate(all_files_data):
            # 创建文件信息框架
            file_info_frame = tk.Frame(self.table_frame, bg="lightgray")
            file_info_frame.grid(row=current_row, column=0, columnspan=len(all_headers), sticky="we", pady=3)
            
            # 获取当前文件的复选框变量和编号
            file_var = self.file_vars.get(file_path)
            file_number = ""
            for i, (path, var) in enumerate(self.file_vars.items(), 1):
                if path == file_path:
                    file_number = str(i)
                    break
            
            # 添加复选框（如果有文件被选中）
            if file_var:
                # 创建复选框，与文件列表中的复选框联动
                # 移除command参数，使数据预览只在点击File PreView按钮时刷新
                preview_check = tk.Checkbutton(file_info_frame, variable=file_var)
                preview_check.pack(side="left", padx=(5, 5))
            
            # 添加文件编号
            number_label = tk.Label(file_info_frame, text=f"{file_number}." if file_number else "", 
                                   font=('Courier New', 10), width=4, anchor="w", bg="lightgray")
            number_label.pack(side="left", padx=(0, 5))
            
            # 添加文件名
            name_label = tk.Label(file_info_frame, text=f"File: {file_name}", 
                                 font=('Courier New', 10, 'bold'), anchor="w", bg="lightgray")
            name_label.pack(side="left", padx=(0, 10))
            
            # 添加文件的总行数和总列数
            data_info_label = tk.Label(file_info_frame, text=f"({len(df)} rows × {len(df.columns)} columns)", 
                                      font=('Courier New', 10), fg="blue", anchor="w", bg="lightgray")
            data_info_label.pack(side="left")
            
            # 添加提示文本，显示是否仅显示前几行
            提示文本 = []
            if not self.show_all_rows and len(df) > max_rows_per_file:
                提示文本.append(f"Only showing first {max_rows_per_file} rows")
            elif self.show_all_rows:
                提示文本.append(f"Showing all {len(df)} rows")
                
            if 提示文本:
                # 创建普通提示文本（非点击）
                more_label_text = "，".join(提示文本)
                more_label = tk.Label(file_info_frame, text=more_label_text, font=('Courier New', 10, 'italic'), fg="gray", anchor="w")
                more_label.pack(side="left", padx=(10, 0))
            
            current_row += 1
            
            # 限制行数
            if self.show_all_rows:
                df_preview = df  # 显示所有行
            else:
                if len(df) > max_rows_per_file:
                    df_preview = df.head(max_rows_per_file)
                else:
                    df_preview = df
            
            # 创建表格内容，所有列使用统一的设置
            for i, row in df_preview.iterrows():
                row_cells = []
                
                # 添加行号单元格（从1开始）- 统一使用Courier New字体
                row_num = i + 1  # 从1开始计数
                row_num_cell = tk.Label(self.table_frame, text=str(row_num), font=('Courier New', 10), relief=tk.RAISED, padx=5, pady=1, anchor="center")
                row_cells.append(row_num_cell)
                
                for j, value in enumerate(row):
                    # 检查值是否为空且不是第二列（索引为1，因为索引从0开始）
                    display_value = str(value)
                    if j != 1 and (value is None or value == '' or pd.isna(value)):
                        display_value = "NA"
                    
                    # 对所有列的文本进行截断处理，确保内容不会导致行高增加
                    max_display_length = 30  # 设置最大显示长度
                    if len(display_value) > max_display_length:
                        display_value = display_value[:max_display_length] + "..."
                    
                    # 设置单元格背景色，如果包含"#"字符则使用淡黄色
                    bg_color = "lightyellow" if "#" in display_value else None
                    
                    # 所有列使用完全相同的字体设置和样式
                    cell = tk.Label(self.table_frame, text=display_value, 
                                    font=('Courier New', 10), 
                                    relief=tk.RAISED, 
                                    padx=5, 
                                    pady=1, 
                                    anchor="w", 
                                    height=1, 
                                    wraplength=0, 
                                    justify='left', 
                                    takefocus=0,
                                    bg=bg_color if bg_color else None)
                    
                    row_cells.append(cell)
                
                # 应用当前行的布局
                for j, cell in enumerate(row_cells):
                    cell.grid(row=current_row, column=j, sticky="nsew", ipady=0)
                    # 为所有行设置固定高度
                    self.table_frame.rowconfigure(current_row, minsize=20, weight=0)
                
                current_row += 1
        
        # 强制刷新整个表格框架
        self.table_frame.update_idletasks()
    
    def update_data_preview_table(self, df, current_file_path=None, current_file_name=None):
        """更新数据预览表格"""
        # 重置表格
        self.reset_data_preview_table()
        
        if df is None or df.empty:
            return
        
        # 更新选项卡状态
        self.update_tab_status("Data Preview")
        
        # 添加一个变量来跟踪当前是否显示所有行
        if not hasattr(self, 'show_all_rows'):
            self.show_all_rows = False
        
        # 限制显示行数，显示所有列
        max_rows = 5  # 默认只显示前5行
        
        # 限制行数
        if self.show_all_rows:
            df_preview = df  # 显示所有行
            show_more_rows = False
        else:
            if len(df) > max_rows:
                df_preview = df.head(max_rows)
                show_more_rows = True
            else:
                df_preview = df
                show_more_rows = False
        
        # 显示所有列，不再限制列数
        show_more_cols = False
        
        # 保存原始数据的总行数和总列数
        total_rows = len(df)
        total_cols = len(df.columns)
        
        # 批量创建所有控件但先不进行布局计算
        headers = []
        
        # 添加行号列的表头
        row_num_header = tk.Label(self.table_frame, text="行号", font=('Courier New', 10, 'bold'), relief=tk.RAISED, padx=5, pady=2, anchor="center", wraplength=0)
        headers.append(row_num_header)
        # 为行号列设置宽度
        self.table_frame.columnconfigure(0, minsize=60, weight=0)
        
        # 统一所有列的表头设置，所有列使用相同的字体和样式
        for j, column in enumerate(df_preview.columns, 1):  # 从1开始编号
            # 所有列使用完全相同的字体设置
            header = tk.Label(self.table_frame, text=f"{j}. {column}", 
                              font=('Courier New', 10, 'bold'), 
                              relief=tk.RAISED, 
                              padx=5, 
                              pady=2, 
                              anchor="w", 
                              wraplength=0, 
                              justify='left')
            # 所有列使用相同的宽度设置，移除特殊列差异化处理
            self.table_frame.columnconfigure(j, minsize=140, weight=1)
            headers.append(header)
        
        # 创建表格内容，所有列使用统一的设置
        cells = []
        for i, row in df_preview.iterrows():
            row_cells = []
            
            # 添加行号单元格（从1开始）- 统一使用Courier New字体
            row_num = i + 1  # 从1开始计数
            row_num_cell = tk.Label(self.table_frame, text=str(row_num), font=('Courier New', 10), relief=tk.RAISED, padx=5, pady=1, anchor="center")
            row_cells.append(row_num_cell)
            
            for j, value in enumerate(row):
                # 检查值是否为空且不是第二列（索引为1，因为索引从0开始）
                display_value = str(value)
                if j != 1 and (value is None or value == '' or pd.isna(value)):
                    display_value = "NA"
                
                # 对所有列的文本进行截断处理，确保内容不会导致行高增加
                max_display_length = 30  # 设置最大显示长度
                if len(display_value) > max_display_length:
                    display_value = display_value[:max_display_length] + "..."
                
                # 设置单元格背景色，如果包含"#"字符则使用淡黄色
                bg_color = "lightyellow" if "#" in display_value else None
                
                # 所有列使用完全相同的字体设置和样式
                cell = tk.Label(self.table_frame, text=display_value, 
                                font=('Courier New', 10), 
                                relief=tk.RAISED, 
                                padx=5, 
                                pady=1, 
                                anchor="w", 
                                height=1, 
                                wraplength=0, 
                                justify='left', 
                                takefocus=0,
                                bg=bg_color if bg_color else None)
                
                row_cells.append(cell)
            cells.append(row_cells)
        
        # 一次性进行所有控件的布局，减少重绘次数
        # 应用表头布局（行号从1开始，因为文件信息框架在0行）
        for j, header in enumerate(headers):
            # 统一所有表头的布局
            header.grid(row=1, column=j, sticky="nsew", ipady=0)
        
        # 应用表格内容布局（行号从2开始，因为文件信息框架在0行，表头在1行）
        for i, row_cells in enumerate(cells):
            for j, cell in enumerate(row_cells):
                # 统一所有单元格的布局
                cell.grid(row=i+2, column=j, sticky="nsew", ipady=0)
                # 为所有行设置固定高度
                self.table_frame.rowconfigure(i+2, minsize=20, weight=0)
        
        # 强制刷新整个表格框架
        self.table_frame.update_idletasks()
        
        # 显示数据量提示信息
        提示文本 = []
        if show_more_rows:
            提示文本.append(f"Only showing first {max_rows} rows")
        elif self.show_all_rows:
            提示文本.append(f"Showing all {total_rows} rows")
        if show_more_cols:
            提示文本.append(f"Only showing first {max_cols} columns")
        
        # 创建一个框架用于显示文件信息（复选框、编号、文件名和行列数）
        # 将文件信息框架移到表格上方
        file_info_frame = tk.Frame(self.table_frame)
        file_info_frame.grid(row=0, column=0, columnspan=len(df_preview.columns)+1, pady=5, sticky="w")
        
        # 获取当前文件的复选框变量
        file_var = None
        file_number = ""
        if current_file_path:
            file_var = self.file_vars.get(current_file_path)
            # 查找当前文件的编号
            for i, (path, var) in enumerate(self.file_vars.items(), 1):
                if path == current_file_path:
                    file_number = str(i)
                    break
        
        # 添加复选框（如果有文件被选中）
        if file_var:
            # 创建复选框，与文件列表中的复选框联动
            # 移除command参数，使数据预览只在点击File PreView按钮时刷新
            preview_check = tk.Checkbutton(file_info_frame, variable=file_var)
            preview_check.pack(side="left", padx=(0, 5))
        
        # 添加文件编号
        number_label = tk.Label(file_info_frame, text=f"{file_number}." if file_number else "", 
                               font=('Courier New', 10), width=4, anchor="w")
        number_label.pack(side="left", padx=(0, 5))
        
        # 添加文件名
        name_label = tk.Label(file_info_frame, text=current_file_name, 
                             font=('Courier New', 10), anchor="w")
        name_label.pack(side="left", padx=(0, 10))
        
        # 添加文件的总行数和总列数
        data_info_label = tk.Label(file_info_frame, text=f"{total_rows} rows × {total_cols} columns", 
                                  font=('Courier New', 10), fg="blue", anchor="w")
        data_info_label.pack(side="left")
        
        # 将提示文本放在同一行的末尾
        if 提示文本:
            # 创建普通提示文本（非点击）
            more_label_text = "，".join(提示文本)
            more_label = tk.Label(file_info_frame, text=more_label_text, font=('Courier New', 10, 'italic'), fg="gray", anchor="w")
            more_label.pack(side="left", padx=(10, 0))
        
        # 优化滚动区域更新和布局调整逻辑
        def final_layout_refresh():
            """最终布局刷新，确保所有列正确显示"""
            # 所有列已经使用统一的样式设置，无需额外调整
            # 强制刷新整个表格框架和Canvas
            self.table_frame.update_idletasks()
            table_width = self.table_frame.winfo_width()
            table_height = self.table_frame.winfo_height()
            self.preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
        
        def secondary_refresh():
            """二次刷新，确保布局稳定"""
            self.table_frame.update_idletasks()
            # 再次更新滚动区域，确保包含所有内容
            table_width = self.table_frame.winfo_width()
            table_height = self.table_frame.winfo_height()
            self.preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
            
            # 最后进行一次最终布局刷新
            self.table_frame.after(150, final_layout_refresh)
        
        def enhanced_refresh():
            """增强的刷新逻辑，专门针对大量列的情况"""
            self.table_frame.update_idletasks()
            # 重新计算表格宽度，确保包含所有列
            # 强制计算所有列的总宽度
            total_width = 0
            # 至少检查前200列，确保处理大量列的情况
            for j in range(min(200, total_cols + 1)):  # +1 因为有行号列
                try:
                    # 获取每列的实际宽度
                    col_width = self.table_frame.grid_bbox(j, 0, j, 10)[2]  # 检查前10行
                    if col_width > 0:
                        total_width = max(total_width, j * 140)  # 保守估计每列宽度
                except:
                    pass
            
            # 强制更新滚动区域，确保包含所有列
            table_height = self.table_frame.winfo_height()
            self.preview_canvas.configure(scrollregion=(0, 0, max(total_width, self.table_frame.winfo_width()), table_height))
            
            # 对于大量列文件，增加额外的刷新保障
            if total_cols > 100:  # 如果列数超过100，增加额外的刷新
                self.table_frame.after(200, final_layout_refresh)
        
        # 确保表格框架完全渲染后再更新滚动区域
        def update_scroll_region():
            # 强制更新表格框架的尺寸信息
            self.table_frame.update_idletasks()
            # 获取表格框架的实际宽度和高度
            table_width = self.table_frame.winfo_width()
            table_height = self.table_frame.winfo_height()
            # 更新Canvas的滚动区域
            self.preview_canvas.configure(scrollregion=(0, 0, table_width, table_height))
            
            # 第一次延迟刷新
            self.table_frame.after(100, secondary_refresh)
            # 对于大量列文件，增加专门的增强刷新
            if total_cols > 100:
                self.table_frame.after(300, enhanced_refresh)
        
        # 使用after_idle延迟更新，让所有控件先绘制完成
        self.table_frame.after_idle(update_scroll_region)
        
        
    def _standardize_criteria_type(self, std_type):
        """标准化标准类型名称，保留窗口类型信息
        
        Args:
            std_type: 原始标准类型名称
            
        Returns:
            tuple: (window_type, metric_type) - 窗口类型和度量类型
        """
        std_type_lower = std_type.strip().lower()
        window_type = None
        metric_type = std_type
        
        # 检查是否包含窗口类型前缀
        if 'white' in std_type_lower:
            window_type = 'white'
            # 移除窗口类型前缀，提取度量类型
            metric_type = std_type_lower.replace('white', '').strip()
        elif 'mixed' in std_type_lower:
            window_type = 'mixed'
            # 移除窗口类型前缀，提取度量类型
            metric_type = std_type_lower.replace('mixed', '').strip()
        
        # 度量类型标准化
        metric_mapping = {
            'l': 'L', 'l value': 'L', 'lightness': 'L',
            'u': 'U', 'uniformity': 'U',
            'dy': 'dY', 'delta y': 'dY', 'ydelta': 'dY',
            'ru': 'Ru', 'r uniformity': 'Ru',
            'rv': 'Rv', 'v uniformity': 'Rv',
            'du': 'Du', 'd uniformity': 'Du',
            'dv': 'Dv', 'v delta uniformity': 'Dv',
            'dl*min': 'dL*Min', 'dl min': 'dL*Min', 'delta l min': 'dL*Min',
            'dl*max': 'dL*Max', 'dl max': 'dL*Max', 'delta l max': 'dL*Max',
            'demax': 'dEMax', 'de max': 'dEMax', 'delta e max': 'dEMax'
        }
        
        # 标准化度量类型
        metric_type = metric_mapping.get(metric_type, metric_type)
        
        return window_type, metric_type
    
    def _get_best_matching_column(self, std_type_info, available_columns):
        """智能匹配标准类型到最合适的列名，考虑窗口类型和带单位的名称
        
        Args:
            std_type_info: 标准类型信息，格式为(window_type, metric_type)
            available_columns: 可用的列名列表
            
        Returns:
            str or None: 最佳匹配的列名，如果没有匹配则返回None
        """
        import re
        
        # 如果传入的是字符串而不是元组（兼容旧代码）
        if isinstance(std_type_info, str):
            window_type = None
            metric_type = std_type_info
        else:
            window_type, metric_type = std_type_info
        
        # 预处理文本，移除特殊字符、单位并转换为小写
        def preprocess_text(text):
            # 移除括号中的单位信息，如 (cd/m^2), (%), (%/cm) 等
            text = re.sub(r'\([^)]*\)', '', text)
            # 移除特殊字符，只保留字母、数字和空格
            text = re.sub(r'[^a-zA-Z0-9\s]', '', text)
            # 标准化空格
            text = re.sub(r'\s+', ' ', text).strip().lower()
            return text
        
        # 定义窗口类型到前缀的映射
        window_prefix_mapping = {
            'white': ['white', 'w_'],
            'mixed': ['mixed', 'm_']
        }
        
        # 定义度量类型到关键词的映射
        metric_keywords = {
            'L': ['l', 'lightness'],
            'U': ['u', 'uniformity'],
            'dY': ['dy', 'ydelta', 'ydiff'],
            'Ru': ['ru', 'runiformity'],
            'Rv': ['rv', 'rvuniformity'],
            'Du': ['du', 'duuniformity'],
            'Dv': ['dv', 'dvuniformity'],
            'dL*Min': ['dl*min', 'dlmin', 'dl_min'],
            'dL*Max': ['dl*max', 'dlmax', 'dl_max'],
            'dEMax': ['demax', 'de_max', 'deltemax']
        }
        
        # 特殊处理Metric类型（如Metric1, Metric2等）
        metric_match = re.match(r'metric(\d+)', metric_type.lower())
        if metric_match:
            metric_num = metric_match.group(1)
            if window_type:
                # 根据窗口类型选择对应的前缀
                if window_type == 'white':
                    target_pattern = f'w_m{metric_num}'  # White窗口的Metric对应W_M1, W_M2等
                else:  # mixed
                    target_pattern = f'm_m{metric_num}'  # Mixed窗口的Metric对应M_M1, M_M2等
                
                # 查找匹配的列
                for col in available_columns:
                    col_lower = col.lower()
                    if target_pattern in col_lower:
                        return col
            return None
        
        # 预处理标准类型信息
        processed_metric_type = preprocess_text(metric_type)
        
        # 为每个列计算匹配分数
        best_score = 0
        best_match = None
        
        for col in available_columns:
            # 原始列名和处理后的列名
            original_col = col
            col_lower = preprocess_text(col)
            score = 0
            
            # 1. 检查窗口类型匹配（最高优先级）
            if window_type and window_type in window_prefix_mapping:
                for prefix in window_prefix_mapping[window_type]:
                    if col_lower.startswith(prefix) or f' {prefix}' in col_lower:
                        score += 5  # 窗口类型匹配（权重最高）
                        break
            
            # 2. 检查度量类型匹配
            # 首先尝试精确匹配处理后的完整度量类型
            if processed_metric_type in col_lower:
                score += 10  # 完整度量类型匹配给较高分数
            else:
                # 然后尝试使用关键词列表匹配
                base_metric_type = metric_type.split('(')[0].strip()  # 提取不带单位的基本类型
                metric_keywords_list = metric_keywords.get(base_metric_type, [base_metric_type.lower()])
                for keyword in metric_keywords_list:
                    if keyword in col_lower:
                        score += 5  # 度量类型关键词匹配
                        
                        # 如果关键词是列名的主要部分，额外加分
                        if col_lower.endswith(keyword) or keyword + ' ' in col_lower:
                            score += 3
                        break
            
            # 3. 完全匹配检查（同时考虑原始名称和处理后名称）
            if (metric_type.lower() == original_col.lower() or 
                processed_metric_type == col_lower):
                score += 20  # 完全匹配给最高分
            
            # 更新最佳匹配
            if score > best_score:
                best_score = score
                best_match = col
        
        # 只有当匹配分数足够高时才返回匹配结果
        if best_score >= 9:  # 阈值保持不变
            return best_match
        
        return None
    
    def _evaluate_record_against_criteria(self, record, criteria_dict):
        """根据Review Criteria中的阈值判断记录的各项指标是否合格
        
        Args:
            record: 单条测试记录
            criteria_dict: 标准字典，格式为 {std_type: (lower_limit, upper_limit)}
            
        Returns:
            tuple: (is_pass, failed_criteria, matched_columns)
                is_pass: 布尔值，表示记录是否通过所有标准
                failed_criteria: 列表，包含所有未通过的标准项和详细信息
                matched_columns: 字典，记录每个标准类型匹配的列名
        """
        is_pass = True
        failed_criteria = []
        matched_columns = {}
        
        # 遍历所有标准
        for std_type, limits in criteria_dict.items():
            # 确保limits是包含两个元素的元组或列表
            try:
                # 如果是列表且长度大于2，只取前两个元素
                if isinstance(limits, list) and len(limits) >= 2:
                    lower_str, upper_str = limits[0], limits[1]
                # 如果是元组且长度大于2，只取前两个元素
                elif isinstance(limits, tuple) and len(limits) >= 2:
                    lower_str, upper_str = limits[0], limits[1]
                # 非列表或元组格式，跳过该标准
                else:
                    if logger:
                        logger.warning(f"{std_type}的限制值格式无效: {limits}，跳过该标准")
                    continue
            except (ValueError, TypeError):
                # 捕获可能的解包错误或类型错误
                if logger:
                    logger.warning(f"无法解析{std_type}的限制值: {limits}，跳过该标准")
                continue
            # 跳过空值
            if not lower_str and not upper_str:
                continue
                
            # 获取标准类型信息（窗口类型和度量类型）
            std_info = self._standardize_criteria_type(std_type)
            
            # 使用智能配对算法查找最佳匹配的列
            column_name = self._get_best_matching_column(std_info, record.index)
            
            # 如果找不到匹配列，尝试使用原始标准类型
            if column_name is None:
                column_name = self._get_best_matching_column(std_type, record.index)
            
            # 如果仍然找不到匹配列，跳过此标准
            if column_name is None:
                continue
            
            # 记录匹配关系
            matched_columns[std_type] = column_name
            
            # 获取记录值
            value = record[column_name]
            
            # 跳过空值
            if pd.isna(value):
                continue
            
            # 尝试转换为数值
            try:
                value_float = float(value)
            except (ValueError, TypeError):
                continue
            
            # 检查下限
            if lower_str:
                try:
                    lower_limit = float(lower_str)
                    if value_float < lower_limit:
                        is_pass = False
                        # 添加更详细的失败原因信息
                        failed_criteria.append({
                            "std_type": std_type,
                            "std_info": std_info,
                            "column": column_name,
                            "value": value_float,
                            "condition": "lower",
                            "limit": lower_limit,
                            "description": f"{column_name} ({std_type}) = {value_float} < {lower_limit}"
                        })
                except ValueError:
                    pass
            
            # 检查上限
            if upper_str:
                try:
                    upper_limit = float(upper_str)
                    if value_float > upper_limit:
                        is_pass = False
                        # 添加更详细的失败原因信息
                        failed_criteria.append({
                            "std_type": std_type,
                            "std_info": std_info,
                            "column": column_name,
                            "value": value_float,
                            "condition": "upper",
                            "limit": upper_limit,
                            "description": f"{column_name} ({std_type}) = {value_float} > {upper_limit}"
                        })
                except ValueError:
                    pass
        
        return is_pass, failed_criteria, matched_columns
    
    def _load_criteria_from_temp_files(self):
        """从临时文件读取Review Criteria子窗口数据并解析上下限
        
        Returns:
            dict: 标准字典，格式为 {std_type: (lower_limit, upper_limit)}
        """
        criteria_dict = {}
        
        try:
            import os
            import csv
            global logger
            
            # 检查临时文件是否存在
            white_file_exists = hasattr(self, '_white_criteria_file') and os.path.exists(self._white_criteria_file)
            mixed_file_exists = hasattr(self, '_mixed_criteria_file') and os.path.exists(self._mixed_criteria_file)
            
            # 从临时文件读取数据的函数
            def read_criteria_from_file(file_path, window_type):
                results = {}
                if os.path.exists(file_path):
                    with open(file_path, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            std_name = row.get('Standard', '')
                            lower = self._parse_numeric_value(row.get('Lower_Limit', ''))
                            upper = self._parse_numeric_value(row.get('Upper_Limit', ''))
                            
                            # 只有当标准名称有效时才添加到结果中
                            if std_name:
                                full_name = f"{window_type} {std_name}"
                                results[full_name] = (lower, upper)
                return results
            
            # 读取White窗口数据
            if white_file_exists:
                white_data = read_criteria_from_file(self._white_criteria_file, 'White')
                criteria_dict.update(white_data)
                
                if logger:
                    logger.info(f"从临时文件读取White窗口数据: {len(white_data)} 条规格")
            
            # 读取Mixed窗口数据
            if mixed_file_exists:
                mixed_data = read_criteria_from_file(self._mixed_criteria_file, 'Mixed')
                criteria_dict.update(mixed_data)
                
                if logger:
                    logger.info(f"从临时文件读取Mixed窗口数据: {len(mixed_data)} 条规格")
            
            # 如果没有读取到数据，记录警告
            if not criteria_dict and logger:
                logger.warning("未能从临时文件读取Review Criteria数据")
        
        except Exception as e:
            # 记录错误但不中断流程
            if logger:
                logger.error(f"从临时文件读取Review Criteria数据时出错: {str(e)}")
        
        return criteria_dict
    
    def _parse_numeric_value(self, value):
        """将字符串值解析为数值
        
        Args:
            value: 要解析的值
            
        Returns:
            float或int: 解析后的数值，如果无法解析则返回None
        """
        if isinstance(value, (int, float)):
            return value
            
        if value is None:
            return None
            
        try:
            # 尝试移除可能的特殊字符
            clean_value = str(value).strip().replace('#', '')
            if clean_value == '':
                return None
            # 尝试转换为浮点数
            num_value = float(clean_value)
            # 如果是整数，转换为int类型
            if num_value.is_integer():
                return int(num_value)
            return num_value
        except (ValueError, TypeError):
            return None
    
    def _get_criteria_dict(self):
        """获取Review Criteria中的阈值数据，优先从临时规格文件读取
        
        Returns:
            dict: 标准字典，格式为 {std_type: (lower_limit, upper_limit)}
        """
        # 声明全局logger变量 - 必须放在函数最开始，在任何使用logger之前
        global logger
        
        # 优先尝试从TestLogAnalyzer_Criteria.json临时文件读取数据
        try:
            import os
            import json
            import tempfile
            
            # 获取临时文件路径
            temp_dir = tempfile.gettempdir()
            temp_file_path = os.path.join(temp_dir, 'TestLogAnalyzer_Criteria.json')
            
            # 检查文件是否存在且可读
            if os.path.exists(temp_file_path) and os.access(temp_file_path, os.R_OK):
                if logger:
                    logger.info(f"尝试从临时规格文件读取数据: {temp_file_path}")
                
                # 读取并解析JSON文件
                with open(temp_file_path, 'r', encoding='utf-8') as f:
                    criteria_data = json.load(f)
                
                if logger:
                    logger.info(f"成功从临时规格文件读取到数据，数据类型: {type(criteria_data).__name__}")
                
                # 将数据转换为标准格式 {std_type: (lower_limit, upper_limit)}
                criteria_dict = {}
                
                # 处理list格式的数据
                if isinstance(criteria_data, list):
                    if logger:
                        logger.info(f"检测到list格式数据，包含{len(criteria_data)}个元素")
                    
                    for idx, item in enumerate(criteria_data):
                        try:
                            # 检查每个元素是否为字典
                            if isinstance(item, dict):
                                # 检查必要的字段
                                if 'std_type' in item and 'lower' in item and 'upper' in item:
                                    std_type = item['std_type']
                                    # 获取并转换上下限
                                    lower = float(item['lower']) if item['lower'] is not None else 0
                                    upper = float(item['upper']) if item['upper'] is not None else float('inf')
                                    
                                    # 根据std_type判断窗口类型
                                    if std_type.startswith('White'):
                                        criteria_dict[std_type] = (lower, upper)
                                    elif std_type.startswith('Mixed'):
                                        criteria_dict[std_type] = (lower, upper)
                                    else:
                                        # 如果没有窗口类型前缀，默认为通用类型
                                        criteria_dict[std_type] = (lower, upper)
                                else:
                                    if logger:
                                        logger.warning(f"list元素[{idx}]缺少必要字段: {item}")
                            else:
                                if logger:
                                    logger.warning(f"list元素[{idx}]不是字典类型: {type(item).__name__}")
                        except Exception as e:
                            if logger:
                                logger.warning(f"处理list元素[{idx}]时出错: {str(e)}")
                
                # 处理原有的字典格式数据
                elif isinstance(criteria_data, dict):
                    # 处理White窗口数据
                    if 'White' in criteria_data:
                        for std_type, values in criteria_data['White'].items():
                            try:
                                # 处理字典类型的值
                                if isinstance(values, dict):
                                    # 确保值是数值类型
                                    lower = float(values.get('lower', 0)) if values.get('lower') is not None else 0
                                    upper = float(values.get('upper', float('inf'))) if values.get('upper') is not None else float('inf')
                                    criteria_dict[f"White {std_type}"] = (lower, upper)
                                # 新增: 处理列表类型的值
                                elif isinstance(values, list):
                                    if logger:
                                        logger.info(f"检测到White窗口中{std_type}规格数据为list类型，尝试解析: {values}")
                                    # 尝试从列表中提取下限和上限
                                    if len(values) >= 2:
                                        # 假设列表第一个元素是下限，第二个元素是上限
                                        lower = float(values[0]) if values[0] is not None else 0
                                        upper = float(values[1]) if values[1] is not None else float('inf')
                                        criteria_dict[f"White {std_type}"] = (lower, upper)
                                        if logger:
                                            logger.info(f"成功解析White窗口中{std_type}规格数据，下限={lower}，上限={upper}")
                                    else:
                                        if logger:
                                            logger.warning(f"White窗口中{std_type}规格数据列表长度不足，无法解析: {values}")
                                else:
                                    # 处理其他类型的情况
                                    if logger:
                                        logger.warning(f"White窗口中的{std_type}规格数据类型不支持，实际为: {type(values).__name__}")
                            except (ValueError, TypeError) as e:
                                if logger:
                                    logger.warning(f"无法解析White窗口中的{std_type}规格数据: {values}, 错误: {str(e)}")
                    
                    # 处理Mixed窗口数据
                    if 'Mixed' in criteria_data:
                        for std_type, values in criteria_data['Mixed'].items():
                            try:
                                # 处理字典类型的值
                                if isinstance(values, dict):
                                    # 确保值是数值类型
                                    lower = float(values.get('lower', 0)) if values.get('lower') is not None else 0
                                    upper = float(values.get('upper', float('inf'))) if values.get('upper') is not None else float('inf')
                                    criteria_dict[f"Mixed {std_type}"] = (lower, upper)
                                # 新增：处理列表类型的值
                                elif isinstance(values, list):
                                    if logger:
                                        logger.info(f"检测到Mixed窗口中{std_type}规格数据为list类型，尝试解析: {values}")
                                    # 尝试从列表中提取下限和上限
                                    if len(values) >= 2:
                                        # 假设列表第一个元素是下限，第二个元素是上限
                                        lower = float(values[0]) if values[0] is not None else 0
                                        upper = float(values[1]) if values[1] is not None else float('inf')
                                        criteria_dict[f"Mixed {std_type}"] = (lower, upper)
                                        if logger:
                                            logger.info(f"成功解析Mixed窗口中{std_type}规格数据，下限={lower}，上限={upper}")
                                    else:
                                        if logger:
                                            logger.warning(f"Mixed窗口中{std_type}规格数据列表长度不足，无法解析: {values}")
                                else:
                                    # 处理其他类型的情况
                                    if logger:
                                        logger.warning(f"Mixed窗口中的{std_type}规格数据类型不支持，实际为: {type(values).__name__}")
                            except (ValueError, TypeError) as e:
                                if logger:
                                    logger.warning(f"无法解析Mixed窗口中的{std_type}规格数据: {values}, 错误: {str(e)}")
                else:
                    if logger:
                        logger.warning(f"临时规格文件数据格式不支持，数据类型: {type(criteria_data).__name__}")
                
                # 如果成功解析到数据，则直接返回
                if criteria_dict:
                    if logger:
                        logger.info(f"从临时规格文件解析到 {len(criteria_dict)} 条规格数据")
                    return criteria_dict
                elif logger:
                    logger.warning("临时规格文件存在但未解析到有效数据")
            elif logger:
                logger.info(f"临时规格文件不存在或不可读: {temp_file_path}")
        except Exception as e:
            if logger:
                logger.error(f"读取临时规格文件时出错: {str(e)}")
        
        # 如果从临时文件读取失败，则回退到原来的方式获取数据
        if logger:
            logger.info("回退到从UI组件获取规格数据")
            
        # 同时调用两种方法并合并结果
        improved_criteria = self._get_improved_criteria_dict()
        original_criteria = self._get_original_criteria_dict()
        
        # 合并两个字典，优先使用改进方法的数据
        criteria_dict = {}
        
        # 添加带前缀的改进方法数据
        if improved_criteria:
            criteria_dict.update(improved_criteria)
        
        # 添加原始方法数据（如果没有在改进方法中捕获到）
        for std_type, limits in original_criteria.items():
            # 检查是否已经有对应的带前缀版本
            has_white_version = f"White {std_type}" in criteria_dict
            has_mixed_version = f"Mixed {std_type}" in criteria_dict
            
            # 如果没有对应的带前缀版本，则添加原始版本
            if not has_white_version and not has_mixed_version:
                criteria_dict[std_type] = limits
        
        # 创建临时文件存储Review Criteria子窗口数据
        try:
            import tempfile
            import os
            import csv
            import time
            
            # 数据验证函数
            def validate_criteria_data(criteria_dict, window_type):
                """验证规格数据的有效性"""
                # 内部函数不需要再次声明global logger
                valid_data = {}
                invalid_data = {}
                warnings = []
                errors = []
                
                for std_name, values in criteria_dict.items():
                    try:
                        # 检查数据结构
                        if not isinstance(values, (list, tuple)) or len(values) != 2:
                            errors.append(f"{std_name}: 数据格式错误，应为(lower, upper)元组")
                            invalid_data[std_name] = values
                            continue
                        
                        lower, upper = values
                        
                        # 数据类型检查
                        original_lower, original_upper = lower, upper
                        
                        # 转换为数值类型
                        if isinstance(lower, str):
                            lower = lower.strip()
                            if lower == '':
                                lower = 0.0  # 空值默认为0
                                warnings.append(f"{std_name}: 下限为空，已设置为默认值0")
                            else:
                                try:
                                    lower = float(lower)
                                except ValueError:
                                    errors.append(f"{std_name}: 下限值 '{original_lower}' 不是有效的数值")
                                    invalid_data[std_name] = values
                                    continue
                        elif lower is None:
                            lower = 0.0
                            warnings.append(f"{std_name}: 下限为None，已设置为默认值0")
                        
                        if isinstance(upper, str):
                            upper = upper.strip()
                            if upper == '':
                                upper = float('inf')  # 空上限默认为无穷大
                                warnings.append(f"{std_name}: 上限为空，已设置为默认值无穷大")
                            else:
                                try:
                                    upper = float(upper)
                                except ValueError:
                                    errors.append(f"{std_name}: 上限值 '{original_upper}' 不是有效的数值")
                                    invalid_data[std_name] = values
                                    continue
                        elif upper is None:
                            upper = float('inf')
                            warnings.append(f"{std_name}: 上限为None，已设置为默认值无穷大")
                        
                        # 上下限关系检查
                        if lower > upper and upper != float('inf'):
                            # 自动交换上下限
                            lower, upper = upper, lower
                            warnings.append(f"{std_name}: 下限({original_lower})大于上限({original_upper})，已自动交换")
                        
                        # 数值合理性检查
                        if lower < 0 and 'percentage' in std_name.lower() or 'rate' in std_name.lower():
                            warnings.append(f"{std_name}: 百分比/比率值({lower})为负数，可能不合理")
                        
                        # 特殊边界值检查
                        if abs(lower) > 1e6 or (upper != float('inf') and abs(upper) > 1e6):
                            warnings.append(f"{std_name}: 值范围过大({lower} to {upper})，请确认是否合理")
                        
                        # 检查NaN值
                        import math
                        if math.isnan(lower) or (upper != float('inf') and math.isnan(upper)):
                            errors.append(f"{std_name}: 包含无效的NaN值")
                            invalid_data[std_name] = values
                            continue
                        
                        # 所有检查通过，添加到有效数据
                        valid_data[std_name] = (lower, upper)
                        
                    except Exception as e:
                        errors.append(f"{std_name}: 验证过程发生错误: {str(e)}")
                        invalid_data[std_name] = values
                
                # 返回验证结果
                return {
                    'valid_data': valid_data,
                    'invalid_data': invalid_data,
                    'warnings': warnings,
                    'errors': errors
                }
            
            # 按窗口类型分类数据
            white_criteria = {k.replace('White ', ''): v for k, v in criteria_dict.items() if k.startswith('White')}
            mixed_criteria = {k.replace('Mixed ', ''): v for k, v in criteria_dict.items() if k.startswith('Mixed')}
            
            # 验证收集到的数据并记录验证开始
            if logger:
                logger.info("\n===== 数据验证开始 =====")
            
            white_validation = validate_criteria_data(white_criteria, "White")
            mixed_validation = validate_criteria_data(mixed_criteria, "Mixed")
            
            # 记录验证结果
            if logger:
                # White窗口验证结果
                logger.info(f"\nWhite窗口数据验证结果:")
                logger.info(f"  有效数据: {len(white_validation['valid_data'])}")
                logger.info(f"  无效数据: {len(white_validation['invalid_data'])}")
                logger.info(f"  警告: {len(white_validation['warnings'])}")
                logger.info(f"  错误: {len(white_validation['errors'])}")
                
                if white_validation['warnings']:
                    logger.warning("  警告详情:")
                    for w in white_validation['warnings']:
                        logger.warning(f"    - {w}")
                
                if white_validation['errors']:
                    logger.error("  错误详情:")
                    for e in white_validation['errors']:
                        logger.error(f"    - {e}")
                
                # Mixed窗口验证结果
                logger.info(f"\nMixed窗口数据验证结果:")
                logger.info(f"  有效数据: {len(mixed_validation['valid_data'])}")
                logger.info(f"  无效数据: {len(mixed_validation['invalid_data'])}")
                logger.info(f"  警告: {len(mixed_validation['warnings'])}")
                logger.info(f"  错误: {len(mixed_validation['errors'])}")
                
                if mixed_validation['warnings']:
                    logger.warning("  警告详情:")
                    for w in mixed_validation['warnings']:
                        logger.warning(f"    - {w}")
                
                if mixed_validation['errors']:
                    logger.error("  错误详情:")
                    for e in mixed_validation['errors']:
                        logger.error(f"    - {e}")
                
                logger.info("===== 数据验证完成 =====")
            
            # 使用验证后的数据
            white_criteria = white_validation['valid_data']
            mixed_criteria = mixed_validation['valid_data']
            
            # 如果没有数据，记录日志并返回
            if not white_criteria and not mixed_criteria:
                if logger:
                    logger.warning("没有提取到White或Mixed窗口的规格数据，跳过保存操作")
                return
            
            # 定义多个可能的临时目录，按优先级尝试
            temp_dirs = []
            
            # 首选: 标准临时目录
            try:
                system_temp = tempfile.gettempdir()
                if os.path.exists(system_temp) and os.access(system_temp, os.W_OK):
                    temp_dirs.append(system_temp)
                    if logger:
                        logger.debug(f"添加系统临时目录: {system_temp}")
            except Exception as e:
                if logger:
                    logger.warning(f"获取系统临时目录失败: {str(e)}")
            
            # 备选: 程序当前目录
            try:
                current_dir = os.path.dirname(os.path.abspath(__file__))
                if os.path.exists(current_dir) and os.access(current_dir, os.W_OK):
                    temp_dirs.append(current_dir)
                    if logger:
                        logger.debug(f"添加程序当前目录: {current_dir}")
            except Exception as e:
                if logger:
                    logger.warning(f"获取程序当前目录失败: {str(e)}")
            
            # 备选: 用户桌面目录
            try:
                import getpass
                user = getpass.getuser()
                desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
                if os.path.exists(desktop_dir) and os.access(desktop_dir, os.W_OK):
                    temp_dirs.append(desktop_dir)
                    if logger:
                        logger.debug(f"添加桌面目录: {desktop_dir}")
            except Exception as e:
                if logger:
                    logger.warning(f"获取用户桌面目录失败: {str(e)}")
            
            # 检查是否有可用的目录
            if not temp_dirs:
                raise Exception("未找到可写入的临时目录")
            
            # 使用第一个可用目录
            temp_dir = temp_dirs[0]
            if logger:
                logger.info(f"使用临时目录: {temp_dir}")
            
            # 创建带时间戳的文件名，避免冲突
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            white_temp_file = os.path.join(temp_dir, f"review_criteria_white_{timestamp}.csv")
            mixed_temp_file = os.path.join(temp_dir, f"review_criteria_mixed_{timestamp}.csv")
            
            # 存储数据的辅助函数，包含写入验证
            def save_data_to_csv(file_path, data_dict):
                # 只有当有数据时才创建文件
                if not data_dict:
                    if logger:
                        logger.info(f"没有{os.path.basename(file_path).split('_')[2]}窗口数据，跳过创建文件")
                    return False
                
                # 写入文件
                try:
                    with open(file_path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(['Standard', 'Lower_Limit', 'Upper_Limit'])
                        written_count = 0
                        for std_name, (lower, upper) in data_dict.items():
                            writer.writerow([std_name, lower, upper])
                            written_count += 1
                    
                    # 验证文件是否成功创建
                    if not os.path.exists(file_path):
                        raise Exception(f"文件创建失败，路径不存在: {file_path}")
                    
                    # 验证文件大小是否合理
                    file_size = os.path.getsize(file_path)
                    if file_size == 0:
                        raise Exception(f"文件创建成功但为空: {file_path}")
                    
                    if logger:
                        logger.info(f"成功保存{written_count}条数据到 {file_path} (大小: {file_size} 字节)")
                    return True
                except Exception as e:
                    raise Exception(f"写入文件失败: {str(e)}")
            
            # 尝试保存两个窗口的数据
            white_saved = False
            mixed_saved = False
            
            if white_criteria:
                try:
                    white_saved = save_data_to_csv(white_temp_file, white_criteria)
                except Exception as e:
                    if logger:
                        logger.error(f"保存White窗口数据失败: {str(e)}")
            
            if mixed_criteria:
                try:
                    mixed_saved = save_data_to_csv(mixed_temp_file, mixed_criteria)
                except Exception as e:
                    if logger:
                        logger.error(f"保存Mixed窗口数据失败: {str(e)}")
            
            # 只有成功保存时才记录文件路径
            if white_saved:
                self._white_criteria_file = white_temp_file
            
            if mixed_saved:
                self._mixed_criteria_file = mixed_temp_file
            
            # 用户反馈函数
            def show_save_status_dialog(success, details):
                """显示保存状态对话框给用户"""
                try:
                    import tkinter as tk
                    from tkinter import messagebox, scrolledtext
                    import threading
                    
                    # 创建新线程显示对话框，避免阻塞主界面
                    def create_dialog():
                        # 检查是否已存在Tk实例
                        root = None
                        try:
                            root = tk.Tk()
                        except:
                            # 如果无法创建Tk实例，尝试使用messagebox
                            msg = "数据保存" + ("成功" if success else "完成，但有问题") + "\n\n" + details
                            messagebox.showinfo("保存状态", msg)
                            return
                        
                        # 设置窗口属性
                        root.title("保存状态")
                        root.geometry("500x350")
                        root.resizable(True, True)
                        
                        # 添加图标
                        try:
                            icon_label = tk.Label(root, text="📊" if success else "⚠️", font=("Arial", 48))
                            icon_label.pack(pady=10)
                        except:
                            pass
                        
                        # 添加标题
                        status_text = "✅ 数据保存成功" if success else "⚠️ 数据保存完成，但有警告/错误"
                        status_label = tk.Label(root, text=status_text, font=("Arial", 14, "bold"), 
                                              fg="green" if success else "orange")
                        status_label.pack(pady=5)
                        
                        # 添加详情文本框
                        detail_text = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=50, height=10, font=("Arial", 10))
                        detail_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
                        detail_text.insert(tk.END, details)
                        detail_text.config(state=tk.DISABLED)  # 设置为只读
                        
                        # 添加关闭按钮
                        close_btn = tk.Button(root, text="关闭", command=root.destroy, font=("Arial", 10), 
                                             width=15, bg="#4CAF50" if success else "#FF9800", fg="white")
                        close_btn.pack(pady=10)
                        
                        # 添加悬浮效果
                        original_bg = close_btn.cget('bg')
                        def on_enter(event):
                            # 按钮悬浮时背景色稍微加深
                            if success:
                                close_btn.config(bg="#388E3C")
                            else:
                                close_btn.config(bg="#F57C00")
                        
                        def on_leave(event):
                            # 按钮离开时恢复原始背景色
                            close_btn.config(bg=original_bg)
                        
                        # 绑定事件
                        close_btn.bind("<Enter>", on_enter)
                        close_btn.bind("<Leave>", on_leave)
                        
                        # 使窗口居中
                        root.update_idletasks()
                        width = root.winfo_width()
                        height = root.winfo_height()
                        x = (root.winfo_screenwidth() // 2) - (width // 2)
                        y = (root.winfo_screenheight() // 2) - (height // 2)
                        root.geometry('{}x{}+{}+{}'.format(width, height, x, y))
                        
                        # 显示窗口
                        root.mainloop()
                    
                    # 在新线程中启动对话框
                    dialog_thread = threading.Thread(target=create_dialog, daemon=True)
                    dialog_thread.start()
                    
                except Exception as e:
                    # 如果对话框创建失败，记录到日志
                    if logger:
                        logger.warning(f"无法创建保存状态对话框: {str(e)}")
            
            # 准备用户反馈信息
            feedback_details = []
            feedback_success = (white_saved or mixed_saved) and not (white_validation['errors'] or mixed_validation['errors'])
            
            # 添加验证结果
            feedback_details.append("📋 数据验证结果:")
            feedback_details.append(f"  - White窗口: 有效{len(white_validation['valid_data'])}条, 无效{len(white_validation['invalid_data'])}条")
            feedback_details.append(f"  - Mixed窗口: 有效{len(mixed_validation['valid_data'])}条, 无效{len(mixed_validation['invalid_data'])}条")
            
            # 添加保存结果
            feedback_details.append("\n💾 文件保存结果:")
            if white_saved:
                feedback_details.append(f"  - White窗口数据: ✅ 成功 (保存到: {os.path.basename(white_temp_file)})")
            else:
                feedback_details.append(f"  - White窗口数据: ❌ 失败")
                
            if mixed_saved:
                feedback_details.append(f"  - Mixed窗口数据: ✅ 成功 (保存到: {os.path.basename(mixed_temp_file)})")
            else:
                feedback_details.append(f"  - Mixed窗口数据: ❌ 失败")
            
            # 添加警告信息
            if white_validation['warnings'] or mixed_validation['warnings']:
                feedback_details.append("\n⚠️ 警告信息:")
                if white_validation['warnings']:
                    feedback_details.append(f"  White窗口 ({len(white_validation['warnings'])}个):")
                    for w in white_validation['warnings'][:3]:  # 只显示前3个警告
                        feedback_details.append(f"    - {w}")
                    if len(white_validation['warnings']) > 3:
                        feedback_details.append(f"    ... 还有{len(white_validation['warnings'])-3}个警告")
                
                if mixed_validation['warnings']:
                    feedback_details.append(f"  Mixed窗口 ({len(mixed_validation['warnings'])}个):")
                    for w in mixed_validation['warnings'][:3]:  # 只显示前3个警告
                        feedback_details.append(f"    - {w}")
                    if len(mixed_validation['warnings']) > 3:
                        feedback_details.append(f"    ... 还有{len(mixed_validation['warnings'])-3}个警告")
            
            # 添加错误信息
            if white_validation['errors'] or mixed_validation['errors']:
                feedback_details.append("\n❌ 错误信息:")
                if white_validation['errors']:
                    feedback_details.append(f"  White窗口 ({len(white_validation['errors'])}个):")
                    for e in white_validation['errors'][:3]:  # 只显示前3个错误
                        feedback_details.append(f"    - {e}")
                    if len(white_validation['errors']) > 3:
                        feedback_details.append(f"    ... 还有{len(white_validation['errors'])-3}个错误")
                
                if mixed_validation['errors']:
                    feedback_details.append(f"  Mixed窗口 ({len(mixed_validation['errors'])}个):")
                    for e in mixed_validation['errors'][:3]:  # 只显示前3个错误
                        feedback_details.append(f"    - {e}")
                    if len(mixed_validation['errors']) > 3:
                        feedback_details.append(f"    ... 还有{len(mixed_validation['errors'])-3}个错误")
            
            # 添加保存位置信息
            if white_saved or mixed_saved:
                save_dir = os.path.dirname(white_temp_file if white_saved else mixed_temp_file)
                feedback_details.append(f"\n📂 保存位置: {save_dir}")
            
            # 添加时间戳
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            feedback_details.append(f"\n🕒 操作时间: {timestamp}")
            
            # 显示用户反馈
            feedback_text = "\n".join(feedback_details)
            show_save_status_dialog(feedback_success, feedback_text)
            
            # 同时记录到日志
            if logger:
                logger.info("\n===== 临时文件保存结果 =====")
                logger.info(f"White窗口数据: {'成功' if white_saved else '失败'} (计划保存: {len(white_criteria)}条)")
                if white_saved:
                    logger.info(f"  - 文件路径: {white_temp_file}")
                
                logger.info(f"Mixed窗口数据: {'成功' if mixed_saved else '失败'} (计划保存: {len(mixed_criteria)}条)")
                if mixed_saved:
                    logger.info(f"  - 文件路径: {mixed_temp_file}")
                
                if white_saved or mixed_saved:
                    logger.info("保存操作完成")
                else:
                    logger.warning("警告: 所有保存操作均失败")
                    
        except Exception as e:
            # 记录详细错误但不中断流程
            if logger:
                logger.error(f"保存Review Criteria子窗口数据到临时文件时发生严重错误: {str(e)}")
                import traceback
                logger.debug(f"错误详情: {traceback.format_exc()}")
            # 创建备选简单保存方案
            try:
                simple_temp_dir = tempfile.gettempdir()
                simple_white_file = os.path.join(simple_temp_dir, "review_criteria_white.csv")
                simple_mixed_file = os.path.join(simple_temp_dir, "review_criteria_mixed.csv")
                
                # 简化版保存逻辑，作为最后的备选
                if white_criteria:
                    with open(simple_white_file, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(['Standard', 'Lower_Limit', 'Upper_Limit'])
                        for std_name, (lower, upper) in white_criteria.items():
                            writer.writerow([std_name, lower, upper])
                    self._white_criteria_file = simple_white_file
                
                if mixed_criteria:
                    with open(simple_mixed_file, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(['Standard', 'Lower_Limit', 'Upper_Limit'])
                        for std_name, (lower, upper) in mixed_criteria.items():
                            writer.writerow([std_name, lower, upper])
                    self._mixed_criteria_file = simple_mixed_file
                    
                if logger:
                    logger.info("已使用简化方案保存数据")
            except Exception as backup_e:
                if logger:
                    logger.critical(f"简化保存方案也失败: {str(backup_e)}")
        
        # 尝试从临时文件读取规格数据（新的方式）
        try:
            if logger:
                logger.info("尝试从临时文件读取Review Criteria规格数据...")
                
            temp_criteria = self._load_criteria_from_temp_files()
            
            # 如果从临时文件读取到数据，则优先使用这些数据
            if temp_criteria:
                if logger:
                    logger.info(f"成功从临时文件读取 {len(temp_criteria)} 条规格数据")
                criteria_dict = temp_criteria
            else:
                if logger:
                    logger.info("未能从临时文件读取到数据，使用原始方式获取的数据")
                    
        except Exception as e:
            # 记录错误但使用原始数据继续
            if logger:
                logger.error(f"从临时文件读取Review Criteria数据时出错: {str(e)}")
        
        # 确保日志输出包含所有捕获的数据
        if logger:
            logger.info("\n=== Review Criteria规格数据汇总 ===")
            logger.info(f"总共读取到 {len(criteria_dict)} 条规格标准")
            
            # 按子窗口类型分类输出
            white_criteria = {k: v for k, v in criteria_dict.items() if k.startswith('White')}
            mixed_criteria = {k: v for k, v in criteria_dict.items() if k.startswith('Mixed')}
            original_criteria_only = {k: v for k, v in criteria_dict.items() if not (k.startswith('White') or k.startswith('Mixed'))}
            
            # 详细输出White窗口规格，显示没有前缀的名称
            logger.info(f"\n------ White子窗口规格 ------")
            if white_criteria:
                logger.info(f"从White窗口读取到 {len(white_criteria)} 条规格:")
                for std_type, (lower, upper) in sorted(white_criteria.items()):
                    # 显示没有前缀的规格名称，使输出更清晰
                    display_name = std_type.replace('White ', '')
                    logger.info(f"  {display_name}: 下限={lower}, 上限={upper}")
            else:
                logger.info("  未从White子窗口读取到规格数据")
            
            # 详细输出Mixed窗口规格，显示没有前缀的名称
            logger.info(f"\n------ Mixed子窗口规格 ------")
            if mixed_criteria:
                logger.info(f"从Mixed窗口读取到 {len(mixed_criteria)} 条规格:")
                for std_type, (lower, upper) in sorted(mixed_criteria.items()):
                    # 显示没有前缀的规格名称，使输出更清晰
                    display_name = std_type.replace('Mixed ', '')
                    logger.info(f"  {display_name}: 下限={lower}, 上限={upper}")
            else:
                logger.info("  未从Mixed子窗口读取到规格数据")
            
            # 输出未分类规格
            if original_criteria_only:
                logger.info(f"\n------ 未分类规格 ------")
                logger.info(f"读取到 {len(original_criteria_only)} 条未分类规格:")
                for std_type, (lower, upper) in sorted(original_criteria_only.items()):
                    logger.info(f"  {std_type}: 下限={lower}, 上限={upper}")
            
            logger.info("\n==========================\n")
            
        return criteria_dict
        
    def _get_improved_criteria_dict(self):
        """改进的方法，专门用于读取Review Criteria选项卡中White和Mixed两个子窗口的所有规格数据"""
        criteria_dict = {}
        white_data_found = False
        mixed_data_found = False
        
        try:
            # 使用logger记录信息
            global logger
            if logger:
                logger.info("\n--- 开始从Review Criteria选项卡读取规格数据 ---")
            
            # 尝试直接从UI组件获取数据
            # 方法1: 检查是否有left_frame和right_frame属性
            if hasattr(self, 'left_frame') and hasattr(self, 'right_frame'):
                if logger:
                    logger.info("[方法1] 发现独立的left_frame和right_frame，尝试从这两个框架读取数据...")
                
                # 从左框架(White)读取数据
                white_frame_data = {}
                self._extract_criteria_from_ui_hierarchy(self.left_frame, 'White', white_frame_data)
                if white_frame_data:
                    criteria_dict.update(white_frame_data)
                    white_data_found = True
                    if logger:
                        logger.info(f"[方法1] 从left_frame成功提取到 {len(white_frame_data)} 条White窗口数据")
                else:
                    if logger:
                        logger.info("[方法1] 未能从left_frame提取到White窗口数据")
                        
                # 从右框架(Mixed)读取数据
                mixed_frame_data = {}
                self._extract_criteria_from_ui_hierarchy(self.right_frame, 'Mixed', mixed_frame_data)
                if mixed_frame_data:
                    criteria_dict.update(mixed_frame_data)
                    mixed_data_found = True
                    if logger:
                        logger.info(f"[方法1] 从right_frame成功提取到 {len(mixed_frame_data)} 条Mixed窗口数据")
                else:
                    if logger:
                        logger.info("[方法1] 未能从right_frame提取到Mixed窗口数据")
            
            # 方法2: 检查Review Criteria选项卡
            elif hasattr(self, 'review_criteria_tab'):
                tab_children = self.review_criteria_tab.winfo_children()
                
                if logger:
                    logger.info(f"[方法2] 从review_criteria_tab读取数据，发现{len(tab_children)}个子组件")
                
                # 窗口类型映射（左框架为White，右框架为Mixed）
                window_types = ['White', 'Mixed']
                
                # 遍历Review Criteria选项卡中的所有组件
                for idx, child in enumerate(tab_children):
                    # 根据框架位置确定窗口类型
                    window_type = window_types[idx % len(window_types)] if idx < len(window_types) else f'Unknown_{idx}'
                    
                    if logger:
                        logger.info(f"[方法2] 处理 {window_type} 组件: {str(type(child))}")
                    
                    # 递归搜索所有子框架中的规格数据
                    component_data = {}
                    self._extract_criteria_from_ui_hierarchy(child, window_type, component_data)
                    
                    if component_data:
                        criteria_dict.update(component_data)
                        if window_type == 'White':
                            white_data_found = True
                        elif window_type == 'Mixed':
                            mixed_data_found = True
                        if logger:
                            logger.info(f"[方法2] 从{window_type}组件成功提取到 {len(component_data)} 条数据")
                    else:
                        if logger:
                            logger.info(f"[方法2] 未能从{window_type}组件提取到数据")
                            
                # 特殊处理: 尝试从review_criteria_tab的Canvas组件中查找数据
                canvas_found = False
                for child in tab_children:
                    if isinstance(child, tk.Canvas):
                        canvas_found = True
                        if logger:
                            logger.info("[方法2-扩展] 发现Canvas组件，尝试从中提取数据")
                        
                        # 为Canvas创建单独的数据字典
                        canvas_data = {}
                        # 获取Canvas中的所有项目
                        for item_id in child.find_all():
                            try:
                                window = child.itemcget(item_id, "window")
                                if window and isinstance(window, (tk.Frame, ttk.Frame)):
                                    # 尝试从Canvas中的窗口提取White和Mixed数据
                                    canvas_white_data = {}
                                    canvas_mixed_data = {}
                                    self._extract_criteria_from_ui_hierarchy(window, 'White', canvas_white_data)
                                    self._extract_criteria_from_ui_hierarchy(window, 'Mixed', canvas_mixed_data)
                                    
                                    if canvas_white_data:
                                        criteria_dict.update(canvas_white_data)
                                        white_data_found = True
                                        if logger:
                                            logger.info(f"[方法2-扩展] 从Canvas窗口提取到 {len(canvas_white_data)} 条White数据")
                                    if canvas_mixed_data:
                                        criteria_dict.update(canvas_mixed_data)
                                        mixed_data_found = True
                                        if logger:
                                            logger.info(f"[方法2-扩展] 从Canvas窗口提取到 {len(canvas_mixed_data)} 条Mixed数据")
                            except Exception as e:
                                if logger:
                                    logger.debug(f"[方法2-扩展] 处理Canvas项目时出错: {str(e)}")
                
                if not canvas_found and logger:
                    logger.info("[方法2-扩展] 在review_criteria_tab中未发现Canvas组件")
            
            # 方法3: 搜索所有可能的属性名变体
            else:
                if logger:
                    logger.info("[方法3] 尝试通过搜索可能的属性名查找规格数据组件...")
                
                # 定义可能包含窗口数据的属性名模式
                patterns = [
                    ('_white_', 'White'),
                    ('white_', 'White'),
                    ('_mixed_', 'Mixed'),
                    ('mixed_', 'Mixed'),
                    ('_frame', None),  # 通用框架模式，稍后确定类型
                    ('frame_', None)
                ]
                
                # 尝试查找所有可能包含框架的属性
                for attr_name in dir(self):
                    # 检查是否为私有属性或方法
                    if attr_name.startswith('__') and attr_name.endswith('__'):
                        continue
                    
                    # 检查是否为方法
                    attr_value = getattr(self, attr_name)
                    if callable(attr_value):
                        continue
                    
                    # 尝试匹配预定义的模式
                    matched_pattern = None
                    window_type = None
                    for pattern, wtype in patterns:
                        if pattern in attr_name.lower():
                            matched_pattern = pattern
                            window_type = wtype
                            break
                    
                    # 如果找到匹配的模式且是框架类型
                    if matched_pattern and isinstance(attr_value, (tk.Frame, ttk.Frame)):
                        # 如果没有预定义窗口类型，尝试根据属性名推断
                        if window_type is None:
                            window_type = 'White' if 'white' in attr_name.lower() else 'Mixed' if 'mixed' in attr_name.lower() else 'Unknown'
                        
                        if logger:
                            logger.info(f"[方法3] 从属性 {attr_name} 尝试提取 {window_type} 窗口数据")
                        
                        # 创建临时字典来存储从此组件提取的数据
                        attr_data = {}
                        self._extract_criteria_from_ui_hierarchy(attr_value, window_type, attr_data)
                        
                        if attr_data:
                            criteria_dict.update(attr_data)
                            if window_type == 'White':
                                white_data_found = True
                            elif window_type == 'Mixed':
                                mixed_data_found = True
                            if logger:
                                logger.info(f"[方法3] 从{attr_name}成功提取到 {len(attr_data)} 条{window_type}数据")
                        else:
                            if logger:
                                logger.info(f"[方法3] 未能从{attr_name}提取到数据")
            
            # 方法4: 检查是否有专门的white_frame和mixed_frame属性
            if not (white_data_found and mixed_data_found):
                # 尝试查找white_frame
                if hasattr(self, 'white_frame') and isinstance(self.white_frame, (tk.Frame, ttk.Frame)):
                    if logger:
                        logger.info("[方法4] 发现专门的white_frame属性，尝试从中提取数据")
                    
                    white_data = {}
                    self._extract_criteria_from_ui_hierarchy(self.white_frame, 'White', white_data)
                    if white_data:
                        criteria_dict.update(white_data)
                        white_data_found = True
                        if logger:
                            logger.info(f"[方法4] 从white_frame成功提取到 {len(white_data)} 条数据")
                
                # 尝试查找mixed_frame
                try:
                    if hasattr(self, 'mixed_frame') and isinstance(self.mixed_frame, (tk.Frame, ttk.Frame)):
                        if logger:
                            logger.info("[方法4] 发现专门的mixed_frame属性，尝试从中提取数据")
                        
                        mixed_data = {}
                        self._extract_criteria_from_ui_hierarchy(self.mixed_frame, 'Mixed', mixed_data)
                        if mixed_data:
                            criteria_dict.update(mixed_data)
                            mixed_data_found = True
                            if logger:
                                logger.info(f"[方法4] 从mixed_frame成功提取到 {len(mixed_data)} 条数据")
                except Exception as e:
                    if logger:
                        logger.debug(f"[方法4] 处理mixed_frame时出错: {str(e)}")
            
            # 统计最终结果
            white_count = sum(1 for k in criteria_dict.keys() if k.startswith('White'))
            mixed_count = sum(1 for k in criteria_dict.keys() if k.startswith('Mixed'))
            
            # 如果成功获取了数据，记录详情
            if criteria_dict:
                if logger:
                    logger.info(f"\n[总结] 改进方法成功读取到 {len(criteria_dict)} 条带前缀的规格标准！")
                    logger.info(f"[总结] 其中White窗口数据: {white_count} 条")
                    logger.info(f"[总结] Mixed窗口数据: {mixed_count} 条")
                    
                    # 详细记录部分关键数据以便调试
                    sample_data = dict(list(criteria_dict.items())[:5])  # 只记录前5条作为示例
                    for key, value in sample_data.items():
                        logger.info(f"[总结] 示例数据 - {key}: {value}")
            else:
                if logger:
                    logger.info("[总结] 改进方法未从UI读取到规格数据")
                    logger.info("[总结] 检查点:")
                    logger.info(f"[总结] - 是否找到White窗口数据: {white_data_found}")
                    logger.info(f"[总结] - 是否找到Mixed窗口数据: {mixed_data_found}")
                
        except Exception as e:
            error_msg = f"[错误] 读取规格数据时出错: {str(e)}"
            if logger:
                logger.error(error_msg)
            else:
                print(error_msg)
                
        return criteria_dict
        
    def _extract_criteria_from_ui_hierarchy(self, widget, window_type, criteria_dict):
        """递归从UI层次结构中提取规格数据，增强组件识别能力"""
        global logger
        
        try:
            # 添加详细的组件类型和名称日志
            widget_type = type(widget).__name__
            widget_name = getattr(widget, 'winfo_name', lambda: 'unknown')()
            if logger and hasattr(widget, 'winfo_class'):
                widget_class = widget.winfo_class()
                logger.debug(f"处理组件: {widget_type} (名称: {widget_name}, 类: {widget_class})")
            
            if isinstance(widget, (tk.Frame, ttk.Frame)):
                # 为当前窗口类型创建临时计数器
                window_counter = {}
                extracted_count = 0
                
                # 遍历子组件
                children = widget.winfo_children()
                if logger:
                    logger.debug(f"框架组件包含 {len(children)} 个子组件")
                
                # 方法3: 首先检查当前框架是否包含任何标签和输入框组合
                # 这是一个更宽松的检查，可以捕获更多可能的数据结构
                labels = [c for c in children if isinstance(c, (tk.Label, ttk.Label))]
                entries = [c for c in children if isinstance(c, (ttk.Entry, tk.Entry))]
                
                if labels and len(entries) >= 2:
                    try:
                        # 尝试从同一框架中的标签和输入框提取数据
                        std_type = "".join([l.cget("text").strip() for l in labels if l.cget("text").strip()])
                        if std_type:
                            lower_limit = entries[0].get().strip() if entries else ""
                            upper_limit = entries[1].get().strip() if len(entries) > 1 else ""
                            
                            if lower_limit or upper_limit:
                                full_std_type = f"{window_type} {std_type}"
                                
                                # 确保数值类型正确
                                lower_limit_val = float(lower_limit) if lower_limit and self._is_valid_number(lower_limit) else -1
                                upper_limit_val = float(upper_limit) if upper_limit and self._is_valid_number(upper_limit) else -1
                                
                                criteria_dict[full_std_type] = (lower_limit_val, upper_limit_val)
                                window_counter[full_std_type] = True
                                extracted_count += 1
                                
                                if logger:
                                    logger.info(f"从{window_type}窗口提取到(格式3): {full_std_type} - 下限={lower_limit_val}, 上限={upper_limit_val}")
                    except Exception as e:
                        if logger:
                            logger.debug(f"处理直接标签输入框组合时出错: {str(e)}")
                
                for i, child in enumerate(children):
                    # 特殊处理Canvas组件，增强Canvas数据提取
                    if isinstance(child, tk.Canvas):
                        if logger:
                            logger.debug(f"[组件{i+1}/{len(children)}] 发现Canvas组件，尝试提取内嵌窗口数据...")
                        try:
                            # 改进Canvas处理，确保获取所有窗口项目
                            items = child.find_all()
                            if logger:
                                logger.debug(f"Canvas包含 {len(items)} 个项目")
                            
                            for item_id in items:
                                try:
                                    window = child.itemcget(item_id, "window")
                                    if window:
                                        if isinstance(window, (tk.Frame, ttk.Frame)):
                                            if logger:
                                                logger.debug(f"从Canvas项目 {item_id} 发现窗口组件，递归处理")
                                            # 递归处理Canvas中的窗口
                                            prev_count = len(criteria_dict)
                                            self._extract_criteria_from_ui_hierarchy(window, window_type, criteria_dict)
                                            new_count = len(criteria_dict) - prev_count
                                            if new_count > 0 and logger:
                                                logger.info(f"从Canvas窗口成功提取到 {new_count} 条{window_type}数据")
                                        else:
                                            if logger:
                                                logger.debug(f"Canvas项目 {item_id} 包含非框架窗口: {type(window).__name__}")
                                    else:
                                        if logger:
                                            logger.debug(f"Canvas项目 {item_id} 不包含窗口")
                                except Exception as e:
                                    if logger:
                                        logger.debug(f"处理Canvas项目 {item_id} 时出错: {str(e)}")
                        except Exception as e:
                            if logger:
                                logger.error(f"处理Canvas组件时发生严重错误: {str(e)}")
                    # 特殊处理PanedWindow组件
                    elif isinstance(child, (tk.PanedWindow, ttk.PanedWindow)):
                        if logger:
                            logger.debug(f"[组件{i+1}/{len(children)}] 发现PanedWindow组件，尝试提取其内容...")
                        try:
                            # 处理PanedWindow中的窗格
                            panes = child.panes()
                            for pane in panes:
                                if isinstance(pane, (tk.Frame, ttk.Frame)):
                                    self._extract_criteria_from_ui_hierarchy(pane, window_type, criteria_dict)
                        except Exception as e:
                            if logger:
                                logger.debug(f"处理PanedWindow组件时出错: {str(e)}")
                    else:
                        # 递归处理其他子组件
                        if logger:
                            logger.debug(f"[组件{i+1}/{len(children)}] 递归处理 {type(child).__name__} 组件")
                        self._extract_criteria_from_ui_hierarchy(child, window_type, criteria_dict)
                    
                    # 检查当前框架是否包含规格行数据 - 增强版
                    if isinstance(child, (tk.Frame, ttk.Frame)):
                        child_children = child.winfo_children()
                        
                        # 方法1: 尝试典型的规格行格式 (5+组件, 位置1是标签, 位置2和3是输入框)
                        if len(child_children) >= 5 and isinstance(child_children[1], (tk.Label, ttk.Label)) and \
                           all(isinstance(child_children[i], (ttk.Entry, tk.Entry)) for i in range(2, 4)):
                            try:
                                self._process_criteria_row(child, window_type, criteria_dict, window_counter)
                                extracted_count += 1
                            except Exception as e:
                                if logger:
                                    logger.debug(f"处理典型规格行时出错: {str(e)}")
                        
                        # 方法2: 尝试更灵活的规格行格式 (任何标签+两个输入框的组合)
                        elif len(child_children) >= 3:
                            try:
                                # 扩展标签和输入框类型识别
                                labels = [i for i, c in enumerate(child_children) if isinstance(c, (tk.Label, ttk.Label))]
                                entries = [i for i, c in enumerate(child_children) if isinstance(c, (ttk.Entry, tk.Entry))]
                                
                                # 方法2A: 确保找到标签和至少两个输入框
                                if labels and len(entries) >= 2:
                                    # 找到第一个标签和前两个输入框
                                    std_type = child_children[labels[0]].cget("text").strip()
                                    lower_limit = child_children[entries[0]].get().strip()
                                    upper_limit = child_children[entries[1]].get().strip()
                                    
                                    # 放宽验证条件，提高数据提取成功率
                                    if std_type:  # 只需要标准类型有值，上下限可以为空
                                        # 对所有类型都添加窗口类型前缀
                                        full_std_type = f"{window_type} {std_type}"
                                        
                                        # 确保数值类型正确
                                        lower_limit_val = float(lower_limit) if lower_limit and self._is_valid_number(lower_limit) else -1
                                        upper_limit_val = float(upper_limit) if upper_limit and self._is_valid_number(upper_limit) else -1
                                        
                                        # 记录到结果字典
                                        criteria_dict[full_std_type] = (lower_limit_val, upper_limit_val)
                                        window_counter[full_std_type] = True
                                        extracted_count += 1
                                        
                                        if logger:
                                            logger.info(f"从{window_type}窗口提取到(格式2): {full_std_type} - 下限={lower_limit_val}, 上限={upper_limit_val}")
                            except Exception as e:
                                if logger:
                                    logger.debug(f"处理灵活格式规格行时出错: {str(e)}")
                
                # 记录当前窗口提取结果
                if logger:
                    logger.info(f"{window_type}窗口({widget_name})本次提取到{extracted_count}条规格数据")
            else:
                # 对于非框架组件，仅记录类型（可用于调试）
                if logger:
                    logger.debug(f"跳过非框架组件: {widget_type} (名称: {widget_name})")
        except Exception as e:
            # 捕获所有异常，确保方法不会崩溃
            if logger:
                logger.error(f"_extract_criteria_from_ui_hierarchy方法执行时出错: {str(e)}")
                
    def _process_criteria_row(self, row_frame, window_type, criteria_dict, window_counter):
        """处理单行规格数据，增强数据提取和验证能力"""
        global logger
        
        try:
            # 获取框架的所有子组件
            children = row_frame.winfo_children()
            
            # 扩展逻辑1: 首先尝试查找任何标签和两个输入框的组合，不限于固定位置
            if len(children) >= 3:
                # 查找所有标签和输入框
                labels = [c for c in children if isinstance(c, (tk.Label, ttk.Label))]
                entries = [c for c in children if isinstance(c, (tk.Entry, ttk.Entry))]
                
                # 如果找到至少一个标签和两个输入框，尝试提取数据
                if labels and len(entries) >= 2:
                    try:
                        # 获取第一个标签的文本作为标准类型
                        std_type = "".join([l.cget("text").strip() for l in labels if l.cget("text").strip()])
                        
                        if std_type:
                            # 获取前两个输入框的值作为上下限
                            lower_limit_text = entries[0].get().strip()
                            upper_limit_text = entries[1].get().strip()
                            
                            # 放宽验证条件，只需要标准类型有值
                            full_std_type = f"{window_type} {std_type}"
                            
                            # 确保数值类型正确
                            lower_limit_val = float(lower_limit_text) if lower_limit_text and self._is_valid_number(lower_limit_text) else -1
                            upper_limit_val = float(upper_limit_text) if upper_limit_text and self._is_valid_number(upper_limit_text) else -1
                            
                            # 添加数据合理性检查
                            if lower_limit_val > -1 and upper_limit_val > -1 and lower_limit_val > upper_limit_val:
                                if logger:
                                    logger.warning(f"数据合理性警告: {full_std_type} 下限({lower_limit_val})大于上限({upper_limit_val})")
                                # 交换上下限，确保数据合理性
                                lower_limit_val, upper_limit_val = upper_limit_val, lower_limit_val
                                if logger:
                                    logger.warning(f"已自动交换上下限值")
                            
                            # 记录到结果字典
                            criteria_dict[full_std_type] = (lower_limit_val, upper_limit_val)
                            window_counter[full_std_type] = True
                            
                            # 详细记录数据提取结果
                            if logger:
                                logger.info(f"成功从{window_type}窗口提取到(灵活匹配): {full_std_type} - 下限={lower_limit_val}, 上限={upper_limit_val}")
                            return  # 成功提取后返回，避免重复处理
                    except Exception as flexible_e:
                        if logger:
                            logger.debug(f"尝试灵活匹配数据时出错: {str(flexible_e)}")
            
            # 原始逻辑: 检查特定位置的组件
            # 增强索引验证，确保子组件数量足够
            if len(children) < 4:
                if logger:
                    logger.debug(f"规格行框架子组件不足，仅有 {len(children)} 个组件")
                # 不再直接返回，而是继续尝试其他方法
            
            # 检查标签组件是否有效
            elif not isinstance(children[1], (tk.Label, ttk.Label)):
                if logger:
                    logger.debug(f"位置1的组件不是标签类型: {type(children[1]).__name__}")
            
            # 检查输入框组件是否有效
            elif not (isinstance(children[2], (tk.Entry, ttk.Entry)) and isinstance(children[3], (tk.Entry, ttk.Entry))):
                if logger:
                    logger.debug(f"位置2或3的组件不是输入框类型: {type(children[2]).__name__}/{type(children[3]).__name__}")
            
            else:
                # 提取标准类型和上下限数据
                try:
                    std_type = children[1].cget("text").strip()
                    if not std_type:
                        if logger:
                            logger.debug(f"标准类型名称为空")
                        return
                    
                    lower_limit_text = children[2].get().strip()
                    upper_limit_text = children[3].get().strip()
                    
                    # 放宽数据验证条件，只需要标准类型有值
                    # 对所有类型都添加窗口类型前缀
                    full_std_type = f"{window_type} {std_type}"
                    
                    # 确保数值类型正确
                    lower_limit_val = float(lower_limit_text) if lower_limit_text and self._is_valid_number(lower_limit_text) else -1
                    upper_limit_val = float(upper_limit_text) if upper_limit_text and self._is_valid_number(upper_limit_text) else -1
                    
                    # 添加数据合理性检查
                    if lower_limit_val > -1 and upper_limit_val > -1 and lower_limit_val > upper_limit_val:
                        if logger:
                            logger.warning(f"数据合理性警告: {full_std_type} 下限({lower_limit_val})大于上限({upper_limit_val})")
                        # 交换上下限，确保数据合理性
                        lower_limit_val, upper_limit_val = upper_limit_val, lower_limit_val
                        if logger:
                            logger.warning(f"已自动交换上下限值")
                    
                    # 记录到结果字典
                    criteria_dict[full_std_type] = (lower_limit_val, upper_limit_val)
                    window_counter[full_std_type] = True
                    
                    # 详细记录数据提取结果
                    if logger:
                        logger.info(f"成功从{window_type}窗口提取到: {full_std_type} - 下限={lower_limit_val}, 上限={upper_limit_val}")
                except Exception as inner_e:
                    if logger:
                        logger.error(f"提取数据时出错: {str(inner_e)}")
                        # 添加更详细的组件信息以便调试
                        child_types = [type(c).__name__ for c in children]
                        logger.debug(f"子组件类型列表: {child_types}")
                        logger.debug(f"子组件数量: {len(children)}")
        except Exception as e:
            # 捕获所有异常，确保方法不会崩溃
            if logger:
                logger.error(f"处理规格行框架时发生未预期错误: {str(e)}")
            # 不向上层抛出异常，确保其他数据行仍能处理
                
    def _is_valid_number(self, text):
        """检查文本是否是有效的数字格式"""
        try:
            float(text)
            return True
        except:
            return False
            
    def _get_original_criteria_dict(self):
        """原始的方法，获取Review Criteria中的阈值数据"""
        global logger
        criteria_dict = {}
        
        if logger:
            logger.info("开始尝试从UI提取规格数据...")
        
        # 方法1: 尝试我们的增强版数据提取函数
        enhanced_criteria = self._extract_criteria_enhanced()
        if enhanced_criteria:
            criteria_dict = enhanced_criteria
            if logger:
                logger.info(f"增强版数据提取成功，获取到 {len(criteria_dict)} 条规格数据")
            # 直接返回成功提取的数据
            self._print_criteria_dict(criteria_dict)
            return criteria_dict
        
        # 方法2: 尝试原始的数据提取逻辑
        if hasattr(self, 'review_criteria_tab'):
            if logger:
                logger.info("尝试使用原始方法从Review Criteria选项卡中提取数据...")
            # 遍历Review Criteria选项卡中的所有窗口组件
            for child in self.review_criteria_tab.winfo_children():
                if isinstance(child, ttk.Frame):
                    # 查找包含标准数据的子框架
                    for sub_child in child.winfo_children():
                        if isinstance(sub_child, tk.Frame):
                            # 查找canvas
                            for canvas_child in sub_child.winfo_children():
                                if isinstance(canvas_child, tk.Canvas):
                                    # 获取canvas中的窗口
                                    canvas_windows = canvas_child.find_all()
                                    for window_id in canvas_windows:
                                        window = canvas_child.itemcget(window_id, "window")
                                        if window and isinstance(window, tk.Frame):
                                            # 查找行框架
                                            for row_frame in window.winfo_children():
                                                if isinstance(row_frame, tk.Frame):
                                                    # 查找类型标签和输入框
                                                    widgets = row_frame.winfo_children()
                                                    if len(widgets) >= 5:
                                                        type_label = widgets[1]
                                                        lower_entry = widgets[2]
                                                        upper_entry = widgets[3]
                                                        
                                                        # 获取标准类型和上下限
                                                        std_type = type_label.cget("text").strip()
                                                        lower_limit = lower_entry.get().strip()
                                                        upper_limit = upper_entry.get().strip()
                                                        
                                                        if std_type:
                                                            criteria_dict[std_type] = (lower_limit, upper_limit)
                                                            if logger:
                                                                logger.info(f"原始方法提取成功: {std_type} - 下限={lower_limit}, 上限={upper_limit}")
        
        # 方法3: 尝试从整个UI层次结构中搜索White和Mixed窗口
        if not criteria_dict and hasattr(self, 'root'):
            if logger:
                logger.info("尝试从整个UI层次结构中搜索White和Mixed窗口数据...")
            # 从根窗口开始搜索
            white_dict = {}
            mixed_dict = {}
            
            # 搜索White窗口数据
            self._search_ui_for_window_type(self.root, "White", white_dict)
            # 搜索Mixed窗口数据
            self._search_ui_for_window_type(self.root, "Mixed", mixed_dict)
            
            # 合并结果
            if white_dict:
                criteria_dict.update(white_dict)
                if logger:
                    logger.info(f"从整个UI找到 {len(white_dict)} 条White窗口数据")
            
            if mixed_dict:
                criteria_dict.update(mixed_dict)
                if logger:
                    logger.info(f"从整个UI找到 {len(mixed_dict)} 条Mixed窗口数据")
        
        # 如果没有从UI获取到数据，尝试使用默认阈值
        if not criteria_dict and hasattr(self, 'processed_data') and not self.processed_data.empty:
            # 尝试从数据中提取标准信息
            criteria_columns = ['White Pass/Fail Criteria', 'Mixed Pass/Fail Criteria',
                              'Pass/Fail Criteria', 'Criteria', 'Mixed PassFail Criteria',
                              'White PassFail Criteria']
            
            for col in criteria_columns:
                if col in self.processed_data.columns:
                    # 获取第一个非空的值
                    values = self.processed_data[col].dropna()
                    if not values.empty:
                        criteria_string = values.iloc[0]
                        # 解析标准字符串
                        criteria_items = self._parse_criteria_string(criteria_string)
                        # 转换为字典格式
                        for item in criteria_items:
                            # 处理可能有3个值的元组（std_type, std_value, number）或2个值的元组
                            if len(item) == 3:
                                std_type, std_value, _ = item  # 忽略排序数字
                            else:
                                std_type, std_value = item
                                
                            if '|' in std_value:
                                lower, upper = std_value.split('|', 1)
                            else:
                                lower = std_value
                                upper = ''
                            criteria_dict[std_type] = (lower.strip(), upper.strip())
                        break
        
        # 打印提取结果
        self._print_criteria_dict(criteria_dict)
        
        return criteria_dict
    
    def _print_criteria_dict(self, criteria_dict):
        """打印规格数据字典"""
        global logger
        
        if criteria_dict:
            print("\n=== Review Criteria规格数据 ===")
            print(f"总共读取到 {len(criteria_dict)} 条规格标准:")
            for std_type, (lower_limit, upper_limit) in criteria_dict.items():
                if lower_limit and upper_limit:
                    print(f"  {std_type}: 下限={lower_limit}, 上限={upper_limit}")
                elif lower_limit:
                    print(f"  {std_type}: 下限={lower_limit}, 无上限")
                elif upper_limit:
                    print(f"  {std_type}: 无下限, 上限={upper_limit}")
                else:
                    print(f"  {std_type}: 无上下限")
            print("==========================\n")
            
            if logger:
                logger.info(f"总共成功提取 {len(criteria_dict)} 条规格标准")
        else:
            print("\n=== Review Criteria规格数据 ===")
            print("未读取到任何规格标准")
            print("==========================\n")
            
            if logger:
                logger.warning("未能从UI中提取任何规格标准")
    
    def _extract_criteria_enhanced(self):
        """增强版数据提取方法，使用多种策略尝试从UI中提取规格数据"""
        global logger
        criteria_dict = {}
        
        if logger:
            logger.info("开始使用增强版数据提取方法...")
        
        # 检查是否有Review Criteria选项卡
        if not hasattr(self, 'review_criteria_tab'):
            if logger:
                logger.warning("未找到Review Criteria选项卡")
            return criteria_dict
        
        # 策略1: 直接从Review Criteria选项卡开始递归搜索
        white_dict = {}
        mixed_dict = {}
        
        # 从Review Criteria选项卡开始，针对White窗口搜索
        self._extract_criteria_from_ui_hierarchy(self.review_criteria_tab, "White", white_dict)
        
        # 从Review Criteria选项卡开始，针对Mixed窗口搜索
        self._extract_criteria_from_ui_hierarchy(self.review_criteria_tab, "Mixed", mixed_dict)
        
        # 合并结果
        if white_dict:
            criteria_dict.update(white_dict)
            if logger:
                logger.info(f"增强方法从White组件提取到 {len(white_dict)} 条数据")
        else:
            if logger:
                logger.warning("[方法2] 未能从White组件提取到数据")
        
        if mixed_dict:
            criteria_dict.update(mixed_dict)
            if logger:
                logger.info(f"增强方法从Mixed组件提取到 {len(mixed_dict)} 条数据")
        else:
            if logger:
                logger.warning("[方法2] 未能从Mixed组件提取到数据")
        
        # 如果没有找到数据，尝试策略2: 查找所有Frame组件并检查是否包含White或Mixed文本
        if not criteria_dict:
            if logger:
                logger.info("尝试策略2: 查找所有包含White/Mixed文本的Frame组件...")
            
            # 递归查找包含特定文本的组件
            white_frames = []
            mixed_frames = []
            self._find_frames_by_text(self.review_criteria_tab, "white", white_frames)
            self._find_frames_by_text(self.review_criteria_tab, "mixed", mixed_frames)
            
            # 从找到的Frame中提取数据
            for frame in white_frames:
                temp_dict = {}
                self._extract_criteria_from_ui_hierarchy(frame, "White", temp_dict)
                if temp_dict:
                    criteria_dict.update(temp_dict)
                    if logger:
                        logger.info(f"从包含White文本的Frame中提取到 {len(temp_dict)} 条数据")
            
            for frame in mixed_frames:
                temp_dict = {}
                self._extract_criteria_from_ui_hierarchy(frame, "Mixed", temp_dict)
                if temp_dict:
                    criteria_dict.update(temp_dict)
                    if logger:
                        logger.info(f"从包含Mixed文本的Frame中提取到 {len(temp_dict)} 条数据")
        
        if not criteria_dict:
            if logger:
                logger.warning("[总结] 改进方法未从UI读取到规格数据")
        
        return criteria_dict
    
    def _search_ui_for_window_type(self, widget, window_type, criteria_dict):
        """从整个UI层次结构中搜索特定窗口类型的数据"""
        global logger
        
        try:
            # 检查当前组件是否包含目标窗口类型的文本
            if hasattr(widget, 'cget'):
                try:
                    text = widget.cget("text").lower()
                    if window_type.lower() in text:
                        if logger:
                            logger.info(f"发现包含{window_type}文本的组件: {text}")
                except Exception:
                    pass
            
            # 检查组件名称
            if hasattr(widget, 'winfo_name'):
                name = widget.winfo_name().lower()
                if window_type.lower() in name:
                    if logger:
                        logger.info(f"发现名称包含{window_type}的组件: {name}")
                    # 尝试从这个组件提取数据
                    temp_dict = {}
                    self._extract_criteria_from_ui_hierarchy(widget, window_type, temp_dict)
                    if temp_dict:
                        criteria_dict.update(temp_dict)
            
            # 递归处理所有子组件
            if hasattr(widget, 'winfo_children'):
                children = widget.winfo_children()
                for child in children:
                    self._search_ui_for_window_type(child, window_type, criteria_dict)
        except Exception as e:
            if logger:
                logger.error(f"搜索UI时出错: {str(e)}")
    
    def _find_frames_by_text(self, widget, search_text, frame_list):
        """递归查找包含特定文本的Frame组件"""
        try:
            # 检查当前组件是否是Frame
            if isinstance(widget, (tk.Frame, ttk.Frame)):
                # 检查组件自身的文本
                if hasattr(widget, 'cget'):
                    try:
                        text = widget.cget("text").lower()
                        if search_text in text:
                            frame_list.append(widget)
                    except Exception:
                        pass
                
                # 检查组件名称
                if hasattr(widget, 'winfo_name'):
                    name = widget.winfo_name().lower()
                    if search_text in name:
                        frame_list.append(widget)
            
            # 递归处理所有子组件
            if hasattr(widget, 'winfo_children'):
                children = widget.winfo_children()
                for child in children:
                    self._find_frames_by_text(child, search_text, frame_list)
        except Exception:
            pass  # 静默忽略错误，继续搜索
    
    def yield_analysis(self):
        """执行不良率分析功能"""
        # 切换到不良率分析选项卡
        if hasattr(self, 'tab_control') and hasattr(self, 'yield_analysis_tab'):
            self.tab_control.select(self.yield_analysis_tab)
        
        # 检查是否有处理后的数据
        if not hasattr(self, 'processed_data') or self.processed_data is None or self.processed_data.empty:
            self.update_status("No data available for analysis")
            return
        
        self.update_status("Running yield analysis...")
        
        # 创建进度条窗口
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 200
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Yield Analysis in Progress")
        progress_window.geometry(f"400x100+{x}+{y}")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # 创建进度条
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, length=380, mode='determinate', maximum=100)
        progress_bar.pack(pady=(20, 10), padx=10)
        
        # 创建状态标签
        status_var = tk.StringVar(value="Initializing yield analysis...")
        status_label = tk.Label(progress_window, textvariable=status_var)
        status_label.pack(pady=(0, 10))
        
        progress_window.update_idletasks()
        
        # 更新进度条
        progress_var.set(10)
        status_var.set("Running criteria matching tests...")
        progress_window.update_idletasks()
        
        # 运行测试函数验证优化后的配对逻辑
        try:
            test_results = self._test_criteria_matching()
            # 输出测试结果到控制台，便于调试和验证
            print("\n--- Criteria Matching Test Results ---")
            print("1. Standardization Test:")
            for original, standardized in test_results['standardization'].items():
                print(f"   {original} -> {standardized}")
            
            print("\n2. Column Matching Test:")
            for std_type, matched_col in test_results['column_matching'].items():
                print(f"   {std_type} -> {matched_col}")
            
            print("\n3. Threshold Judgment Test:")
            for result in test_results['threshold_judgment']:
                print(f"   Record {result['record_index']}: Pass={result['is_pass']}, " 
                      f"Failed={result['failed_criteria_count']}, " 
                      f"Matched={result['matched_columns_count']}")
            print("--- End Test Results ---")
        except Exception as e:
            print(f"Test execution error: {str(e)}")
        
        # 更新进度条
        progress_var.set(30)
        status_var.set("Generating yield analysis table...")
        progress_window.update_idletasks()
        
        # 生成不良汇总表格，传入进度条参数
        self.update_yield_analysis_table(progress_var, status_var, progress_window)
        
        # 更新选项卡状态
        self.update_tab_status("Yield Analysis")
        
        # 更新进度条并关闭窗口
        progress_var.set(100)
        status_var.set("Yield analysis completed!")
        progress_window.update_idletasks()
        self.root.after(500, progress_window.destroy)  # 短暂显示完成状态后关闭窗口
        
    def _test_criteria_matching(self):
        """测试优化后的阈值配对和判断逻辑
        
        Returns:
            dict: 测试结果统计
        """
        import pandas as pd
        
        # 创建测试数据
        test_data = {
            'Mixed L': [50.5, 45.2, 60.1],
            'White U': [2.1, 3.5, 1.8],
            'Mixed dY': [0.3, 0.5, 0.2],
            'Luminance': [48.0, 46.0, 52.0],  # 测试语义相似性匹配
            'Upper Limit': [2.2, 3.0, 1.7],   # 测试语义相似性匹配
            'Config': ['A', 'B', 'C']
        }
        
        test_df = pd.DataFrame(test_data)
        record = test_df.iloc[0]
        
        # 创建测试标准字典
        test_criteria = {
            'L': ('45.0', '55.0'),
            'U': ('', '3.0'),
            'dY': ('0.2', '0.4'),
            'luminance': ('40.0', '60.0')  # 测试大小写不敏感
        }
        
        results = {
            'standardization': {},
            'column_matching': {},
            'threshold_judgment': []
        }
        
        # 测试标准类型标准化
        for std_type in test_criteria.keys():
            standardized = self._standardize_criteria_type(std_type)
            results['standardization'][std_type] = standardized
        
        # 测试列标题匹配
        for std_type in test_criteria.keys():
            matched_col = self._get_best_matching_column(std_type, test_df.columns)
            results['column_matching'][std_type] = matched_col
        
        # 测试阈值判断
        for idx, row in test_df.iterrows():
            is_pass, failed_criteria, matched_columns = self._evaluate_record_against_criteria(row, test_criteria)
            results['threshold_judgment'].append({
                'record_index': idx,
                'is_pass': is_pass,
                'failed_criteria_count': len(failed_criteria),
                'matched_columns_count': len(matched_columns)
            })
        
        return results
    
    def update_yield_analysis_table(self, progress_var=None, status_var=None, progress_window=None):
        """更新不良率分析表格，使用Treeview创建类似Excel的可复制表格
        根据Review Criteria中的阈值判断不良品"""
        # 清空良率分析选项卡中的现有内容
        for widget in self.yield_analysis_tab.winfo_children():
            widget.destroy()
        
        # 创建主容器
        container = tk.Frame(self.yield_analysis_tab)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建标题
        title_label = tk.Label(container, text="Yield Summary Table", font=("SimHei", 12, "bold"))
        title_label.pack(pady=10)
        
        # 更新进度条
        if progress_var and status_var and progress_window:
            progress_var.set(40)
            status_var.set("Getting criteria data...")
            progress_window.update_idletasks()
        
        # 获取Review Criteria中的阈值数据
        criteria_dict = self._get_criteria_dict()
        
        # 添加统计摘要面板
        if criteria_dict:
            summary_frame = tk.Frame(container, relief="groove", bd=1)
            summary_frame.pack(fill="x", pady=5, padx=5)
            
            summary_title = tk.Label(summary_frame, text="Criteria Summary", font=("SimHei", 10, "bold"))
            summary_title.pack(anchor="w", padx=5, pady=2)
            
            summary_text = f"Total criteria: {len(criteria_dict)}"
            
            # 计算有上下限的阈值数量
            has_lower = sum(1 for _, (lower, _) in criteria_dict.items() if lower)
            has_upper = sum(1 for _, (_, upper) in criteria_dict.items() if upper)
            has_both = sum(1 for _, (lower, upper) in criteria_dict.items() if lower and upper)
            
            summary_text += f" | With lower limits: {has_lower}"
            summary_text += f" | With upper limits: {has_upper}"
            summary_text += f" | With both limits: {has_both}"
            
            summary_label = tk.Label(summary_frame, text=summary_text, font=("SimHei", 9))
            summary_label.pack(anchor="w", padx=5, pady=2)
        
        # 如果没有获取到阈值数据，提示用户
        if not criteria_dict:
            no_criteria_label = tk.Label(container, text="No criteria data found in Review Criteria. Using Pass/Fail column instead.", 
                                      font=("SimHei", 10, "italic"), fg="red")
            no_criteria_label.pack(pady=10)
        
        # 创建Treeview表格组件（类似Excel）
        # 增加'Fail Rate'列，用于显示按Config分组的不良率
        columns = ("Config", "Total", "Fail Count", "Fail Rate")
        tree = ttk.Treeview(container, columns=columns, show="headings", selectmode="extended")
        
        # 设置列宽和表头
        tree.column("Config", width=250, anchor="w")
        tree.column("Total", width=100, anchor="center")  # Total列居中对齐
        tree.column("Fail Count", width=100, anchor="center")
        tree.column("Fail Rate", width=100, anchor="center")
        
        # 记录原始数据，用于排序
        self.yield_analysis_data = []
        self.yield_sort_column = None
        self.yield_sort_reverse = False
        
        # 创建排序函数
        def sort_by_column(col, reverse):
            """根据点击的列对表格进行排序，确保分隔行和总数行不参与排序"""
            # 定义排序函数
            def sort_key(item):
                if col == "Config":
                    return str(item[0]).lower()  # 按Config名称排序
                elif col == "Total":
                    return int(item[1])      # 按Total数量排序
                elif col == "Fail Count":
                    return int(item[2])      # 按Fail Count数量排序
                elif col == "Fail Rate":
                    # 处理numpy.float64类型
                    if hasattr(item[3], 'strip'):
                        # 如果是字符串类型，去除百分号并转换为浮点数
                        return float(item[3].strip('%'))
                    else:
                        # 如果是numpy.float64类型，直接返回
                        return float(item[3])
                return str(item[0])
            
            # 清空当前表格中的所有行
            for child in tree.get_children():
                tree.delete(child)
            
            # 对Config分组数据进行排序
            if self.yield_analysis_data:
                sorted_data = sorted(self.yield_analysis_data, key=sort_key, reverse=reverse)
                
                # 插入排序后的数据行
                for config, total, fail, rate in sorted_data:
                    tree.insert("", "end", values=(f"                          Config: {config}", str(total), str(fail), f"{rate:.2f}%"))
                
                # 重新插入分隔行（不参与排序）
                tree.insert("", "end", values=("-" * 50, "", "", ""))
            
            # 重新插入总数行（不参与排序，始终在最后）
            tree.insert("", "end", values=("Total", str(total_count), str(fail_count), f"{total_fail_rate:.2f}%"), tags=('total_row'))
        
        # 创建带排序功能的表头点击处理函数
        def on_heading_click(col):
            """表头点击事件处理"""
            # 如果点击的是同一列，则切换排序方向
            if self.yield_sort_column == col:
                self.yield_sort_reverse = not self.yield_sort_reverse
            else:
                self.yield_sort_column = col
                self.yield_sort_reverse = False
            
            # 执行排序
            sort_by_column(col, self.yield_sort_reverse)
        
        # 设置表头样式和点击事件
        tree.heading("Config", text="Config", command=lambda: on_heading_click("Config"))
        tree.heading("Total", text="Total", command=lambda: on_heading_click("Total"))
        tree.heading("Fail Count", text="Fail Count", command=lambda: on_heading_click("Fail Count"))
        tree.heading("Fail Rate", text="Fail Rate", command=lambda: on_heading_click("Fail Rate"))
        
        # 配置Treeview样式使其更像Excel
        style = ttk.Style()
        # 设置表格行高
        style.configure("Treeview", rowheight=25, font=("SimHei", 10))
        # 设置表头样式
        style.configure("Treeview.Heading", font=("SimHei", 10, "bold"), relief="solid", borderwidth=1)
        
        # 更新进度条
        if progress_var and status_var and progress_window:
            progress_var.set(50)
            status_var.set("Calculating failure statistics...")
            progress_window.update_idletasks()
        
        # 计算总数量和不良数量
        total_count = len(self.processed_data)
        fail_count = 0
        
        # 创建一个副本用于标记Pass/Fail状态
        df_copy = self.processed_data.copy()
        
        # 更新进度条
        if progress_var and status_var and progress_window:
            progress_var.set(60)
            status_var.set("Evaluating records against criteria...")
            progress_window.update_idletasks()
        
        # 根据Review Criteria中的阈值判断不良品
        if criteria_dict:
            # 在数据副本中添加基于标准的判断结果列和匹配信息
            df_copy['Criteria_Pass/Fail'] = 'PASS'
            df_copy['Failed_Details'] = ''
            df_copy['Matched_Columns'] = ''
            
            # 遍历每条记录，根据阈值判断是否合格
            for idx, row in df_copy.iterrows():
                is_pass, failed_criteria, matched_columns = self._evaluate_record_against_criteria(row, criteria_dict)
                if not is_pass:
                    fail_count += 1
                    df_copy.at[idx, 'Criteria_Pass/Fail'] = 'FAIL'
                    
                    # 保存详细的失败原因和匹配信息
                    if failed_criteria:
                        # 如果失败原因是字典格式，提取描述信息
                        if isinstance(failed_criteria[0], dict):
                            failed_descriptions = [fc['description'] for fc in failed_criteria]
                            df_copy.at[idx, 'Failed_Details'] = ' | '.join(failed_descriptions)
                        else:
                            df_copy.at[idx, 'Failed_Details'] = ', '.join(failed_criteria)
                    
                    if matched_columns:
                        # 保存匹配信息
                        match_info = [f"{k}→{v}" for k, v in matched_columns.items()]
                        df_copy.at[idx, 'Matched_Columns'] = ' | '.join(match_info)
                
                # 更新进度条（每处理10%的数据更新一次）
                if progress_var and status_var and progress_window and idx % max(1, total_count // 10) == 0:
                    progress = 60 + (idx / total_count) * 20
                    progress_var.set(min(progress, 80))
                    status_var.set(f"Evaluated {idx+1}/{total_count} records...")
                    progress_window.update_idletasks()
        else:
            # 如果没有标准数据，使用原来的Pass/Fail列
            if 'Pass/Fail' in df_copy.columns:
                fail_count = (df_copy['Pass/Fail'] == 'FAIL').sum()
                # 使用原来的Pass/Fail列作为判断依据
                df_copy['Criteria_Pass/Fail'] = df_copy['Pass/Fail']
            else:
                # 如果也没有Pass/Fail列，所有记录都视为通过
                df_copy['Criteria_Pass/Fail'] = 'PASS'
        
        # 更新进度条
        if progress_var and status_var and progress_window:
            progress_var.set(80)
            status_var.set("Calculating failure rates...")
            progress_window.update_idletasks()
        
        # 计算总体不良率
        total_fail_rate = 0
        if total_count > 0:
            total_fail_rate = (fail_count / total_count) * 100
        
        # 按Config分组统计不良率
        if 'Config' in df_copy.columns:
            # 添加按Config分组的统计数据
            config_stats = []
            
            # 按Config分组
            for config, group in df_copy.groupby('Config'):
                config_total = len(group)
                # 使用基于标准的判断结果统计不良数
                config_fail = (group['Criteria_Pass/Fail'] == 'FAIL').sum()
                config_fail_rate = (config_fail / config_total * 100) if config_total > 0 else 0
                config_stats.append((config, config_total, config_fail, config_fail_rate))
            
            # 按ID从小到大排序（Config名称升序）
            config_stats.sort(key=lambda x: str(x[0]))
            
            # 保存数据到实例变量供排序使用
            self.yield_analysis_data = config_stats.copy()
            
            # 更新进度条
            if progress_var and status_var and progress_window:
                progress_var.set(90)
                status_var.set("Populating results table...")
                progress_window.update_idletasks()
            
            # 添加Config分组数据行
            for config, total, fail, rate in config_stats:
                tree.insert("", "end", values=(f"                          Config: {config}", str(total), str(fail), f"{rate:.2f}%"))
            
            # 添加分隔行
            tree.insert("", "end", values=("-" * 50, "", "", ""))
        
        # 更新进度条
        if progress_var and status_var and progress_window:
            progress_var.set(95)
            status_var.set("Finalizing results...")
            progress_window.update_idletasks()
        
        # 添加总体统计数据行并设置所有单元格居中对齐
        tree.tag_configure('total_row', anchor='center', font=('SimHei', 10))
        tree.insert("", "end", values=("Total", str(total_count), str(fail_count), f"{total_fail_rate:.2f}%"), tags=('total_row'))
        
        # 添加垂直滚动条
        yscrollbar = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        tree.configure(yscroll=yscrollbar.set)
        
        # 添加水平滚动条
        xscrollbar = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(xscroll=xscrollbar.set)
        
        # 布局组件
        xscrollbar.pack(side="bottom", fill="x")
        yscrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)
        
        # 添加复制功能支持
        def copy_selection(event):
            """复制选中的单元格内容，同时包含表头信息"""
            selected_items = tree.selection()
            if not selected_items:
                return
            
            # 构建复制文本，先添加表头
            copy_text = []
            # 获取表头文本
            headers = [tree.heading(col, "text") for col in tree['columns']]
            copy_text.append("\t".join(headers))
            
            # 添加选中的数据行
            for item in selected_items:
                values = tree.item(item, "values")
                copy_text.append("\t".join(values))
            
            # 将文本复制到剪贴板
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(copy_text))
            self.update_status("Copied to clipboard")
        
        # 绑定Ctrl+C快捷键复制
        tree.bind("<Control-c>", copy_selection)
        tree.bind("<Control-C>", copy_selection)
        
        # 添加右键菜单支持复制
        def show_context_menu(event):
            """显示右键菜单"""
            # 只在点击单元格或行时显示右键菜单，避免点击表头时也触发
            if tree.identify_region(event.x, event.y) in ("cell", "row"):
                # 选中点击的项目
                item = tree.identify_row(event.y)
                if item:
                    tree.selection_set(item)
                # 显示右键菜单
                menu.post(event.x_root, event.y_root)
        
        # 创建右键菜单
        menu = tk.Menu(container, tearoff=0)
        menu.add_command(label="Copy", command=lambda: copy_selection(None))
        
        # 绑定右键菜单
        tree.bind("<Button-3>", show_context_menu)
        
        # 如果使用Review Criteria中的阈值判断，添加详细分析按钮
        if criteria_dict and 'Criteria_Pass/Fail' in df_copy.columns:
            details_frame = tk.Frame(container)
            details_frame.pack(pady=5, fill="x", anchor="e")
            
            details_btn = ttk.Button(details_frame, text="Show Detailed Failure Analysis", style="HoverButton.TButton", 
                                    command=lambda: self.show_detailed_failure_analysis(df_copy, criteria_dict))
            details_btn.pack(side="right", padx=10)
        
        # 添加提示信息
        hint_label = tk.Label(container, text="Tip: Click on column headers to sort data, use Ctrl+C to copy selected content, or right-click menu to copy", 
                             font=("SimHei", 9), fg="gray")
        hint_label.pack(pady=5)
        
        # 更新状态栏
        if criteria_dict:
            analysis_method = "Review Criteria thresholds"
        else:
            analysis_method = "Pass/Fail column"
        
        self.update_status(f"Fail Rate Analysis Completed ({analysis_method}): Total={total_count}, Fail Count={fail_count}, Fail Rate={total_fail_rate:.2f}%")
    
    def show_detailed_failure_analysis(self, data_df, criteria_dict):
        """显示详细的不良分析，包括每条记录的具体不良项、判断依据和匹配信息
        
        Args:
            data_df: 包含'Criteria_Pass/Fail'、'Failed_Details'和'Matched_Columns'列的数据框
            criteria_dict: 标准字典
        """
        # 创建新窗口
        details_window = tk.Toplevel(self.root)
        details_window.title("Detailed Failure Analysis")
        details_window.geometry("1200x700")
        details_window.minsize(1000, 600)
        
        # 创建主容器
        container = tk.Frame(details_window)
        container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 创建标题
        title_label = tk.Label(container, text="Detailed Failure Analysis", font=("SimHei", 12, "bold"))
        title_label.pack(pady=10)
        
        # 添加筛选选项
        filter_frame = tk.Frame(container)
        filter_frame.pack(fill="x", pady=5)
        
        tk.Label(filter_frame, text="Filter by Status:", font=("SimHei", 10)).pack(side="left", padx=5)
        
        status_var = tk.StringVar(value="All")
        status_filter = ttk.Combobox(filter_frame, textvariable=status_var, values=["All", "PASS", "FAIL"], width=10)
        status_filter.pack(side="left", padx=5)
        
        # 创建Treeview表格组件
        columns = ("ID", "Serial Number", "Config", "Status", "Failed Criteria", "Failed Details", "Matched Columns")
        tree = ttk.Treeview(container, columns=columns, show="headings", selectmode="extended")
        
        # 设置列宽和表头
        tree.column("ID", width=80, anchor="center")
        tree.column("Serial Number", width=120, anchor="center")
        tree.column("Config", width=150, anchor="center")
        tree.column("Status", width=100, anchor="center")
        tree.column("Failed Criteria", width=150, anchor="center")
        tree.column("Failed Details", width=300, anchor="w")
        tree.column("Matched Columns", width=300, anchor="w")
        
        # 设置表头
        tree.heading("ID", text="ID")
        tree.heading("Serial Number", text="Serial Number")
        tree.heading("Config", text="Config")
        tree.heading("Status", text="Status")
        tree.heading("Failed Criteria", text="Failed Criteria Count")
        tree.heading("Failed Details", text="Detailed Failure Reasons")
        tree.heading("Matched Columns", text="Standard to Column Mapping")
        
        # 配置Treeview样式
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("SimHei", 10))
        style.configure("Treeview.Heading", font=("SimHei", 10, "bold"), relief="solid", borderwidth=1)
        
        # 添加数据行并记录筛选数据
        all_items = []
        
        # 按ID（索引）从小到大排序DataFrame
        sorted_data_df = data_df.sort_index()
        
        for idx, row in sorted_data_df.iterrows():
            status = row.get('Criteria_Pass/Fail', 'UNKNOWN')
            config = row.get('Config', '-')
            
            # 获取失败详情
            failed_details = row.get('Failed_Details', '')
            
            # 获取匹配列信息
            matched_columns = row.get('Matched_Columns', '')
            
            # 计算失败标准数量
            if failed_details:
                if ' | ' in failed_details:
                    fail_count = len(failed_details.split(' | '))
                else:
                    fail_count = len(failed_details.split(', '))
            else:
                fail_count = 0
            
            # 获取Serial Number（检查不同可能的列名）
            serial_number = row.get('Serial Number', '-')
            
            # 如果未找到，尝试其他可能的列名
            if serial_number == '-':
                for col in ['Serial_Number', 'SN', 'serial', 'serial_number']:
                    if col in data_df.columns:
                        serial_number = row.get(col, '-')
                        break
            
            # 插入行数据，添加Serial Number列
            item = tree.insert("", "end", values=(idx+1, serial_number, config, status, str(fail_count), failed_details, matched_columns))
            
            # 根据状态设置行颜色
            if status == 'FAIL':
                tree.tag_configure('fail', background='#ffcccc')
                tree.item(item, tags=('fail',))
            elif status == 'PASS':
                tree.tag_configure('pass', background='#ccffcc')
                tree.item(item, tags=('pass',))
            
            all_items.append((item, status))
        
        # 筛选函数
        def filter_tree():
            filter_value = status_var.get()
            
            for item, status in all_items:
                if filter_value == "All" or status == filter_value:
                    tree.item(item, open=True, values=tree.item(item, "values"))
                else:
                    # 临时隐藏不匹配的行
                    tree.detach(item)
            
            # 重新插入可见的行
            if filter_value != "All":
                visible_items = [(item, status) for item, status in all_items if status == filter_value]
                for item, _ in visible_items:
                    try:
                        tree.reattach(item, "", "end")
                    except:
                        pass  # 忽略已经附加的项目
        
        # 绑定筛选事件
        status_filter.bind("<<ComboboxSelected>>", lambda event: filter_tree())
        
        # 添加垂直滚动条
        yscrollbar = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        tree.configure(yscroll=yscrollbar.set)
        
        # 添加水平滚动条
        xscrollbar = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(xscroll=xscrollbar.set)
        
        # 布局组件
        xscrollbar.pack(side="bottom", fill="x")
        yscrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)
        
        # 添加复制功能支持
        def copy_selection(event):
            """复制选中的单元格内容，同时包含表头信息"""
            selected_items = tree.selection()
            if not selected_items:
                return
            
            # 构建复制文本，先添加表头
            copy_text = []
            # 获取表头文本
            headers = [tree.heading(col, "text") for col in tree['columns']]
            copy_text.append("\t".join(headers))
            
            # 添加选中的数据行
            for item in selected_items:
                values = tree.item(item, "values")
                copy_text.append("\t".join(values))
            
            # 将文本复制到剪贴板
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(copy_text))
            self.update_status("Copied to clipboard")
        
        # 绑定Ctrl+C快捷键复制
        tree.bind("<Control-c>", copy_selection)
        tree.bind("<Control-C>", copy_selection)
        
        # 添加关闭按钮
        button_frame = tk.Frame(container)
        button_frame.pack(pady=10, fill="x", anchor="e")
        
        close_btn = ttk.Button(button_frame, text="Close", style="HoverButton.TButton", command=details_window.destroy)
        close_btn.pack(side="right", padx=10)
    
    def create_yield_analysis_chart(self):
        """创建不良率分析图表"""
        # 创建一个占位标签
        placeholder_label = tk.Label(self.yield_analysis_tab, text="Click 'Fail Rate Analysis' button to generate fail rate summary table")
        placeholder_label.pack(pady=20)
    
    def create_top10_content(self):
        """创建Top 10选项卡的内容"""
        # 创建一个占位标签
        placeholder_label = tk.Label(self.top10_tab, text="Click 'Top Defects' button to view Top Defect Items Analysis")
        placeholder_label.pack(pady=20)
    
    def show_top10_tab(self):
        """显示Top Defects选项卡并生成不良项目统计报表，数据来源为Data Re-Processing Tab"""
        try:
            # 切换到Top Defects选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'top10_tab'):
                self.tab_control.select(self.top10_tab)
            else:
                self.update_status("Top Defects tab not found")
                return
            
            # 检查是否有Data Re-Processing Tab的数据（优先使用）
            if hasattr(self, 'reprocessed_data') and self.reprocessed_data is not None and not self.reprocessed_data.empty:
                data_source = self.reprocessed_data
                self.update_status("Using data from Data Re-Processing Tab for analysis")
            # 如果没有reprocessed_data，检查是否有原始处理数据作为备选
            elif hasattr(self, 'processed_data') and self.processed_data is not None and not self.processed_data.empty:
                data_source = self.processed_data
                self.update_status("No Data Re-Processing data found, using original processed data")
            else:
                self.update_status("No data available for analysis")
                # 清空选项卡内容并显示提示
                for widget in self.top10_tab.winfo_children():
                    widget.destroy()
                hint_label = tk.Label(self.top10_tab, text="Please run Data Re-Processing first", font=("SimHei", 12))
                hint_label.grid(pady=20)
                return
            
            # 创建进度条窗口，设置最大值为10（根据处理步骤）
            progress_window, progress_var, status_var = self.create_progress_window("Top Defects Analysis Progress", 10)
            
            # 更新进度条初始状态
            progress_var.set(1)
            status_var.set("Setting up analysis environment...")
            progress_window.update_idletasks()
            
            self.update_status("Generating Top Defect Items Analysis...")
            
            # 更新选项卡状态
            self.update_tab_status("Top Defects")
            
            # 更新进度条
            progress_var.set(2)
            status_var.set("Clearing previous content...")
            progress_window.update_idletasks()
            
            # 清空Top Defects选项卡中的现有内容
            for widget in self.top10_tab.winfo_children():
                widget.destroy()
            
            # 更新进度条
            progress_var.set(3)
            status_var.set("Creating UI components...")
            progress_window.update_idletasks()
            
            # 创建主容器
            main_container = tk.Frame(self.top10_tab)
            main_container.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 创建标题
            title_label = tk.Label(main_container, text="Top Defect Items Analysis", font=("SimHei", 12, "bold"))
            title_label.pack(pady=(0, 10))
            
            # 创建横向排列的容器
            h_container = tk.Frame(main_container)
            h_container.pack(fill="both", expand=True)
            
            # =======================================
            # 1. Top 10不良项目统计部分
            # =======================================
            top10_frame = tk.LabelFrame(h_container, text="Top Defect Items", font=("SimHei", 10))
            top10_frame.pack(side="left", fill="both", expand=True, padx=(0, 5))
            
            # 创建Treeview表格组件 - Top 10统计
            top10_columns = ("No", "Fail Item", "Fail Count", "Total Count", "Fail Rate")
            top10_tree = ttk.Treeview(top10_frame, columns=top10_columns, show="headings", selectmode="extended")
            
            # 设置列宽和表头
            top10_tree.column("No", width=60, anchor="center")
            top10_tree.column("Fail Item", width=300, anchor="w")
            top10_tree.column("Fail Count", width=100, anchor="center")
            top10_tree.column("Total Count", width=100, anchor="center")
            top10_tree.column("Fail Rate", width=100, anchor="center")
            
            # 设置表头
            top10_tree.heading("No", text="No")
            top10_tree.heading("Fail Item", text="Fail Item")
            top10_tree.heading("Fail Count", text="Fail Count")
            top10_tree.heading("Total Count", text="Total Count")
            top10_tree.heading("Fail Rate", text="Fail Rate")
            
            # 配置Treeview样式
            style = ttk.Style()
            style.configure("Treeview", rowheight=25, font=("SimHei", 10))
            style.configure("Treeview.Heading", font=("SimHei", 10, "bold"), relief="solid", borderwidth=1)
            
            # 更新进度条
            progress_var.set(4)
            status_var.set("Calculating total count...")
            progress_window.update_idletasks()
            
            # 计算总数量（总行数）
            total_count = len(data_source)
            
            # 更新进度条
            progress_var.set(5)
            status_var.set("Analyzing defect data...")
            progress_window.update_idletasks()
            
            # 统计从第8列开始的所有列的不良情况（使用format_cells字典）
            # 注意：format_cells的键是行索引，值是该行列索引的列表
            column_fail_counts = {}
            
            # 检查format_cells是否存在
            format_cells_exists = hasattr(self, 'format_cells') and self.format_cells
            
            # 从第8列开始统计（索引从0开始，所以是7）
            valid_columns_found = False
            # 添加日志信息
            debug_info = f"Total Columns: {len(data_source.columns)}, Starting from Column 8"
            
            # 计算总列数，用于进度计算
            total_columns = len(data_source.columns) - 7  # 从第8列开始的列数
            
            # 使用data_source和format_cells字典进行统计
            for i, col_idx in enumerate(range(7, len(data_source.columns))):
                # 更新进度条
                progress = 5 + (i / total_columns) * 2  # 5到7之间的进度
                progress_var.set(progress)
                status_var.set(f"Analyzing column {i+1}/{total_columns}...")
                progress_window.update_idletasks()
                
                column_name = data_source.columns[col_idx]
                valid_columns_found = True
                
                # 统计该列中被标记为需要格式化的单元格数量（只统计淡黄色单元格）
                fail_count = 0
                
                if format_cells_exists:
                    # 使用format_cells统计淡黄色单元格 - 不良判断标准：标记为淡黄色的单元格
                    for row_idx, col_indices in self.format_cells.items():
                        # 检查该列是否在当前行的格式化列表中
                        if col_idx in col_indices:
                            # 直接计入统计，因为format_cells中的单元格就是淡黄色的
                            fail_count += 1
                
                # 无论是否有不良，都将项目添加到统计中
                column_fail_counts[column_name] = fail_count
            
            # 更新进度条
            progress_var.set(7)
            status_var.set("Processing defect items...")
            progress_window.update_idletasks()
            
            # 处理不良项目名称并重命名/过滤特定项目
            processed_items = []
            for item_name, fail_count in column_fail_counts.items():
                # 忽略特定的v Avg项目
                if item_name == "White v Avg" or item_name == "Mixed v Avg":
                    continue
                    
                # 重命名特定的u Avg项目
                if item_name == "White u Avg":
                    item_name = "CAFL0 Color Point"
                elif item_name == "Mixed u Avg":
                    item_name = "CAFL24 Color Point"
                
                processed_items.append((item_name, fail_count))
            
            # 更新进度条
            progress_var.set(8)
            status_var.set("Filtering and sorting results...")
            progress_window.update_idletasks()
            
            # 过滤不良数量为零的条目，并按不良数量降序排序
            filtered_items = [(item_name, fail_count) for item_name, fail_count in processed_items if fail_count > 0]
            sorted_items = sorted(filtered_items, key=lambda x: x[1], reverse=True)
            
            # 更新进度条
            progress_var.set(9)
            status_var.set("Populating table with results...")
            progress_window.update_idletasks()
            
            # 添加数据行到Top 10表格
            for idx, (item_name, fail_count) in enumerate(sorted_items, start=1):
                # 计算不良率
                fail_rate = (fail_count / total_count) * 100 if total_count > 0 else 0
                top10_tree.insert("", "end", values=(str(idx), item_name, str(fail_count), str(total_count), f"{fail_rate:.2f}%"))
            
            # 更新进度条
            progress_var.set(10)
            status_var.set("Analysis completed!")
            progress_window.update_idletasks()
            
            # 延迟关闭进度窗口
            self.root.after(500, progress_window.destroy)
            
            # 添加滚动条和表格 - Top 10统计
            # 添加垂直滚动条
            top10_yscrollbar = ttk.Scrollbar(top10_frame, orient="vertical", command=top10_tree.yview)
            top10_tree.configure(yscroll=top10_yscrollbar.set)
            
            # 添加水平滚动条
            top10_xscrollbar = ttk.Scrollbar(top10_frame, orient="horizontal", command=top10_tree.xview)
            top10_tree.configure(xscroll=top10_xscrollbar.set)
            
            # 使用grid布局来放置表格和滚动条
            top10_tree.grid(row=0, column=0, sticky="nsew")
            top10_yscrollbar.grid(row=0, column=1, sticky="ns")
            top10_xscrollbar.grid(row=1, column=0, sticky="ew")
            
            # 配置容器的权重，使表格可以随窗口大小调整
            top10_frame.grid_rowconfigure(0, weight=1)
            top10_frame.grid_columnconfigure(0, weight=1)
            
            # =======================================
            # 2. Config分组统计部分
            # =======================================
            config_frame = tk.LabelFrame(h_container, text="Config Group Fail Items", font=("SimHei", 10))
            config_frame.pack(side="left", fill="both", expand=True, padx=(5, 0))
            
            # 创建Treeview表格组件 - Config统计（增加序号列）
            config_columns = ("No", "Config", "Fail Item", "Fail Count", "Total Count", "Fail Rate")
            config_tree = ttk.Treeview(config_frame, columns=config_columns, show="headings", selectmode="extended")
            
            # 设置列宽和表头
            config_tree.column("No", width=60, anchor="center")
            config_tree.column("Config", width=120, anchor="center")
            config_tree.column("Fail Item", width=200, anchor="w")
            config_tree.column("Fail Count", width=80, anchor="center")
            config_tree.column("Total Count", width=100, anchor="center")
            config_tree.column("Fail Rate", width=80, anchor="center")
            
            # 设置表头
            config_tree.heading("No", text="No")
            config_tree.heading("Config", text="Config")
            config_tree.heading("Fail Item", text="Fail Item")
            config_tree.heading("Fail Count", text="Fail Count")
            config_tree.heading("Total Count", text="Total Count")
            config_tree.heading("Fail Rate", text="Fail Rate")
            
            # 按Config分组统计不良
            config_stats = []
            
            # 检查是否存在Config列
            has_config_column = 'Config' in data_source.columns
            
            if has_config_column:
                # 获取所有不同的Config值
                configs = data_source['Config'].unique()
                
                # 遍历每个Config
                for config in configs:
                    # 获取当前Config的数据
                    config_data = data_source[data_source['Config'] == config]
                    config_total = len(config_data)
                    
                    # 统计每个不良项目在该Config下的不良数量
                    for col_idx in range(7, len(data_source.columns)):
                        column_name = data_source.columns[col_idx]
                        
                        # 忽略特定的v Avg项目
                        if column_name == "White v Avg" or column_name == "Mixed v Avg":
                            continue
                            
                        fail_count = 0
                        
                        # 使用format_cells统计淡黄色单元格 - 不良判断标准：标记为淡黄色的单元格
                        if format_cells_exists:
                            # 只统计当前Config的数据行中被标记为淡黄色的单元格
                            for row_idx, col_indices in self.format_cells.items():
                                # 检查行索引是否在当前Config的数据范围内
                                if row_idx < len(data_source) and data_source.iloc[row_idx]['Config'] == config:
                                    if col_idx in col_indices:
                                        # 直接计入统计，因为format_cells中的单元格就是淡黄色的
                                        fail_count += 1
                        
                        # 计算不良率
                        fail_rate = (fail_count / config_total) * 100 if config_total > 0 else 0
                        config_stats.append((config, column_name, fail_count, config_total, fail_rate))
                
                # 先按Config分组，再在组内按不良率降序排序
                # 过滤不良数量为零的条目
                filtered_config_stats = [(config, item_name, fail_count, config_total, fail_rate) 
                                        for config, item_name, fail_count, config_total, fail_rate in config_stats 
                                        if fail_count > 0]
                
                # 按Config分组并在组内排序
                config_groups = {}
                for config, item_name, fail_count, config_total, fail_rate in filtered_config_stats:
                    if config not in config_groups:
                        config_groups[config] = []
                    config_groups[config].append((item_name, fail_count, config_total, fail_rate))
                
                # 为每个Config组内的项目按不良率降序排序
                for config in config_groups:
                    config_groups[config].sort(key=lambda x: x[3], reverse=True)
                
                # 添加数据行到Config统计表格，按Config分组显示，每个分组增加序号
                for config in sorted(config_groups.keys()):
                    # 添加Config分组标题行
                    config_total = len(self.processed_data[self.processed_data['Config'] == config])
                    config_tree.insert("", "end", values=("", f"{config} (Total: {config_total})", "", "", "", ""))
                    
                    # 添加该Config下的不良项目，每个分组内的项目重新编号
                    for idx, (item_name, fail_count, _, fail_rate) in enumerate(config_groups[config], start=1):
                        # 处理不良项目名称
                        display_name = item_name
                        if item_name == "White u Avg":
                            display_name = "CAFL0 Color Point"
                        elif item_name == "Mixed u Avg":
                            display_name = "CAFL24 Color Point"
                        
                        config_tree.insert("", "end", values=(str(idx), config, display_name, str(fail_count), str(config_total), f"{fail_rate:.2f}%"))
                    
                    # 添加一个空行分隔不同Config
                    config_tree.insert("", "end", values=("", "", "", "", "", ""))
            
            # 添加滚动条和表格 - Config统计
            # 添加垂直滚动条
            config_yscrollbar = ttk.Scrollbar(config_frame, orient="vertical", command=config_tree.yview)
            config_tree.configure(yscroll=config_yscrollbar.set)
            
            # 添加水平滚动条
            config_xscrollbar = ttk.Scrollbar(config_frame, orient="horizontal", command=config_tree.xview)
            config_tree.configure(xscroll=config_xscrollbar.set)
            
            # 使用grid布局来放置表格和滚动条
            config_tree.grid(row=0, column=0, sticky="nsew")
            config_yscrollbar.grid(row=0, column=1, sticky="ns")
            config_xscrollbar.grid(row=1, column=0, sticky="ew")
            
            # 配置容器的权重，使表格可以随窗口大小调整
            config_frame.grid_rowconfigure(0, weight=1)
            config_frame.grid_columnconfigure(0, weight=1)
            
            # 如果没有Config列，显示提示
            if not has_config_column:
                no_config_label = tk.Label(config_frame, text="No config data found in the file, cannot perform Config grouping statistics", font=("SimHei", 10), fg="red")
                no_config_label.grid(row=0, column=0, columnspan=3, pady=20, sticky="nsew")
            elif not config_stats:
                no_data_label = tk.Label(config_frame, text="Current file has no Config grouping data with failed items", font=("SimHei", 10), fg="orange")
                no_data_label.grid(row=0, column=0, columnspan=3, pady=20, sticky="nsew")
            
            # =======================================
            # 复制功能 - Top 10统计表格
            # =======================================
            def copy_top10_selected(event):
                try:
                    # 获取选中的项
                    selected_items = top10_tree.selection()
                    if not selected_items:
                        return
                    
                    # 构建复制文本（制表符分隔）
                    copy_text = []
                    # 添加表头
                    headers = [top10_tree.heading(col, "text") for col in top10_columns]
                    copy_text.append("\t".join(headers))
                    
                    # 添加选中行的数据
                    for item in selected_items:
                        values = top10_tree.item(item, "values")
                        copy_text.append("\t".join(values))
                    
                    # 将文本复制到剪贴板
                    top10_tree.clipboard_clear()
                    top10_tree.clipboard_append("\n".join(copy_text))
                    self.update_status(f"Copied {len(selected_items)} rows from Top Defect Items Table to Clipboard")
                except Exception as e:
                    self.update_status(f"Copy Failed: {str(e)}")
            
            # 绑定Ctrl+C快捷键
            top10_tree.bind("<Control-c>", copy_top10_selected)
            
            # 添加右键菜单支持
            def show_top10_popup_menu(event):
                popup_menu = tk.Menu(top10_tree, tearoff=0)
                popup_menu.add_command(label="Copy Selected", command=lambda: copy_top10_selected(None))
                popup_menu.post(event.x_root, event.y_root)
            
            top10_tree.bind("<Button-3>", show_top10_popup_menu)
            
            # =======================================
            # 复制功能 - Config统计表格
            # =======================================
            def copy_config_selected(event):
                try:
                    # 获取选中的项
                    selected_items = config_tree.selection()
                    if not selected_items:
                        return
                    
                    # 构建复制文本（制表符分隔）
                    copy_text = []
                    # 添加表头
                    headers = [config_tree.heading(col, "text") for col in config_columns]
                    copy_text.append("\t".join(headers))
                    
                    # 添加选中行的数据
                    for item in selected_items:
                        values = config_tree.item(item, "values")
                        copy_text.append("\t".join(values))
                    
                    # 将文本复制到剪贴板
                    config_tree.clipboard_clear()
                    config_tree.clipboard_append("\n".join(copy_text))
                    self.update_status(f"Copied {len(selected_items)} rows from Config Table to Clipboard")
                except Exception as e:
                    self.update_status(f"Copy Failed: {str(e)}")
            
            # 绑定Ctrl+C快捷键
            config_tree.bind("<Control-c>", copy_config_selected)
            
            # 添加右键菜单支持
            def show_config_popup_menu(event):
                popup_menu = tk.Menu(config_tree, tearoff=0)
                popup_menu.add_command(label="Copy Selected", command=lambda: copy_config_selected(None))
                popup_menu.post(event.x_root, event.y_root)
            
            config_tree.bind("<Button-3>", show_config_popup_menu)
            
            # 添加提示信息
            hint_label = tk.Label(main_container, text="Tip: Use Ctrl+C to copy selected content, or right-click menu to copy", 
                                 font=("SimHei", 9), fg="gray")
            hint_label.pack(pady=(5, 0), anchor="w")
            
            # 如果没有数据，显示无数据提示 - Top 10统计
            if not valid_columns_found:
                no_data_label = tk.Label(top10_frame, text="No valid project columns found (starting from column 8)", font=("SimHei", 10), fg="red")
                no_data_label.grid(row=2, column=0, columnspan=2, pady=10, sticky="nsew")
            elif not sorted_items:
                no_data_label = tk.Label(top10_frame, text="No bad project data found", font=("SimHei", 10), fg="orange")
                no_data_label.grid(row=2, column=0, columnspan=2, pady=10, sticky="nsew")
            
            # 调整提示标签的位置，确保它在横向排列的布局下方
            hint_label.pack(pady=(5, 0), anchor="w")
            
            # 更新状态栏，显示更准确的信息
            if valid_columns_found:
                # 添加调试信息到状态栏
                debug_info += f", format_cells status: {'Exist and Not Empty' if format_cells_exists else 'Not Exist or Empty'}"
                # 只统计有不良的项目数量
                status_message = f"Top Defect Items Statistics Completed, Displaying {len(sorted_items)} Defect Items"  
                if has_config_column:
                    # 统计有不良的Config分组记录数
                    filtered_config_count = len([(config, item_name, fail_count, config_total, fail_rate) 
                                               for config, item_name, fail_count, config_total, fail_rate in config_stats 
                                               if fail_count > 0])
                    status_message += f"；Config Group Statistics Completed, Displaying {filtered_config_count} Bad Records (Grouped)"
                self.update_status(status_message)
            else:
                self.update_status("No valid project columns found (starting from column 8)")
        except Exception as e:
            # 捕获所有异常，确保UI不会崩溃
            self.update_status(f"Top Defect Items Statistics Generation Error: {str(e)}")
            # 清空选项卡内容并显示错误信息
            try:
                if hasattr(self, 'top10_tab'):
                    for widget in self.top10_tab.winfo_children():
                        widget.destroy()
                    error_label = tk.Label(self.top10_tab, text=f"Top Defect Items Statistics Generation Error: {str(e)}", font=('SimHei', 10), fg="red")
                    error_label.grid(pady=20)
            except Exception as inner_error:
                print(f"Error displaying error message: {inner_error}")
        
    def create_cpk_content(self):
        """创建Cpk选项卡的内容"""
        # 创建一个占位标签
        placeholder_label = tk.Label(self.cpk_tab, text="Click 'Cpk' button to view Cpk Analysis", font=('SimHei', 10))
        placeholder_label.pack(pady=20)
        
    def _save_criteria_to_temp_file(self):
        """将Criteria标签页中的数据保存为CSV临时文件（当前目录）
        实现规格值的正确处理：
        - 使用"|"分隔符将规格值拆分为下限和上限
        - 修正uniformity规格笔误（当U类型的上限为"0"时自动修正为"100"）
        - 将规格下限为"-1"的值自动修正为"0"
        - 添加完整的描述信息
        
        Returns:
            str: 临时CSV文件路径，如果保存失败则返回None
        """
        # 声明全局logger变量
        global logger
        
        try:
            import os
            import csv
            import time
            
            # 在当前目录创建临时CSV文件，使用时间戳避免文件名冲突
            current_dir = os.path.dirname(os.path.abspath(__file__))
            timestamp = int(time.time())
            temp_file_path = os.path.join(current_dir, f"criteria_temp_data_{timestamp}.csv")
            
            # 收集并处理Criteria数据
            processed_criteria_data = []
            
            # 定义标准类型的描述映射
            standard_descriptions = {
                "L": "Luminance (cd/m^2)",
                "U": "Uniformity (%)",
                "dY": "dY (%/cm)",
                "u": "u' Coordinate",
                "v": "v' Coordinate",
                "Ru": "Ru",
                "Rv": "Rv",
                "Du": "Du",
                "Dv": "Dv",
                "dL*Min": "dL*Min (%/cm)",
                "dL*Max": "dL*Max (%/cm)",
                "dEMax": "dEMax (%/cm)"
            }
            
            # 检查是否有Criteria标签页中的Treeview组件
            if hasattr(self, 'criteria_tab'):
                # 查找所有可能的Treeview组件（White和Mixed的表格）
                for widget in self.criteria_tab.winfo_children():
                    # 查找包含Treeview的容器
                    if isinstance(widget, tk.Frame) or isinstance(widget, tk.LabelFrame):
                        for child in widget.winfo_children():
                            if isinstance(child, ttk.Treeview):
                                # 获取Treeview数据
                                tree = child
                                items = tree.get_children()
                                
                                # 确定是White还是Mixed数据
                                data_type = "Unknown"
                                parent_text = widget.cget("text") if isinstance(widget, tk.LabelFrame) else ""
                                if "White" in parent_text:
                                    data_type = "White"
                                elif "Mixed" in parent_text:
                                    data_type = "Mixed"
                                
                                # 提取表格数据
                                for item in items:
                                    values = tree.item(item, "values")
                                    if values and data_type != "Unknown":
                                        # 确保有足够的数据列
                                        standard_type = values[0] if len(values) > 0 else ""
                                        specs = values[1] if len(values) > 1 else ""
                                        
                                        # 处理规格值：使用"|"分隔符拆分
                                        lower_limit = ""
                                        upper_limit = ""
                                        
                                        if '|' in specs:
                                            # | 是规格上限和规格下限的分隔符
                                            limits = specs.split('|', 1)
                                            lower_limit = limits[0].strip()
                                            upper_limit = limits[1].strip() if len(limits) > 1 else ""
                                        else:
                                            # 如果没有|分隔符，尝试其他分隔符
                                            if '~' in specs:
                                                limits = specs.split('~')
                                            elif '-' in specs and not specs.startswith('-'):
                                                limits = specs.split('-', 1)
                                            elif ' ' in specs:
                                                limits = specs.split()
                                            elif ',' in specs:
                                                limits = specs.split(',')
                                            else:
                                                limits = [specs, ""]
                                                
                                            lower_limit = limits[0].strip()
                                            upper_limit = limits[1].strip() if len(limits) > 1 else ""
                                        
                                        # 修正规格值
                                        # 1. 修正uniformity规格笔误
                                        if standard_type == "U" and upper_limit == "0":
                                            upper_limit = "100"
                                        # 2. 将规格下限为"-1"的值自动修正为"0"
                                        if lower_limit == "-1":
                                            lower_limit = "0"
                                        # 对于Uniformity类型，确保上限不超过100
                                        if standard_type == "U" or "_U" in standard_type or "Uniformity" in standard_type:
                                            try:
                                                if upper_limit:
                                                    upper_num = float(upper_limit)
                                                    if upper_num > 100.0:
                                                        upper_limit = "100"
                                            except ValueError:
                                                pass
                                        
                                        # 获取描述信息
                                        description = standard_descriptions.get(standard_type, f"{data_type} {standard_type}")
                                        
                                        # 添加处理后的数据到列表
                                        processed_criteria_data.append([
                                            data_type, 
                                            standard_type, 
                                            lower_limit, 
                                            upper_limit,
                                            description
                                        ])
            
            # 如果没有从Treeview获取到数据，尝试从reprocessed_data中获取
            if not processed_criteria_data and hasattr(self, 'reprocessed_data') and self.reprocessed_data is not None and not self.reprocessed_data.empty:
                for data_type in ["White", "Mixed"]:
                    criteria_column = f'{data_type} Pass/Fail Criteria'
                    if criteria_column in self.reprocessed_data.columns:
                        criteria_values = self.reprocessed_data[criteria_column].dropna()
                        if not criteria_values.empty:
                            first_criteria = criteria_values.iloc[0]
                            parsed_data = self._parse_criteria_string(first_criteria)
                            if parsed_data:
                                for item in parsed_data:
                                    if isinstance(item, (list, tuple)) and len(item) >= 2:
                                        standard_type = item[0]
                                        specs = item[1] if len(item) > 1 else ""
                                        
                                        # 处理规格值：使用"|"分隔符拆分
                                        lower_limit = ""
                                        upper_limit = ""
                                        
                                        if '|' in specs:
                                            # | 是规格上限和规格下限的分隔符
                                            limits = specs.split('|', 1)
                                            lower_limit = limits[0].strip()
                                            upper_limit = limits[1].strip() if len(limits) > 1 else ""
                                        else:
                                            # 如果没有|分隔符，尝试其他分隔符
                                            if '~' in specs:
                                                limits = specs.split('~')
                                            elif '-' in specs and not specs.startswith('-'):
                                                limits = specs.split('-', 1)
                                            elif ' ' in specs:
                                                limits = specs.split()
                                            elif ',' in specs:
                                                limits = specs.split(',')
                                            else:
                                                limits = [specs, ""]
                                                
                                            lower_limit = limits[0].strip()
                                            upper_limit = limits[1].strip() if len(limits) > 1 else ""
                                        
                                        # 修正规格值
                                        if standard_type == "U" and upper_limit == "0":
                                            upper_limit = "100"
                                        if lower_limit == "-1":
                                            lower_limit = "0"
                                        # 对于Uniformity类型，确保上限不超过100
                                        if standard_type == "U" or "_U" in standard_type or "Uniformity" in standard_type:
                                            try:
                                                if upper_limit:
                                                    upper_num = float(upper_limit)
                                                    if upper_num > 100.0:
                                                        upper_limit = "100"
                                            except ValueError:
                                                pass
                                        
                                        # 获取描述信息
                                        description = standard_descriptions.get(standard_type, f"{data_type} {standard_type}")
                                        
                                        # 添加处理后的数据到列表
                                        processed_criteria_data.append([
                                            data_type, 
                                            standard_type, 
                                            lower_limit, 
                                            upper_limit,
                                            description
                                        ])
            
            # 如果有数据，尝试保存到CSV文件
            if processed_criteria_data:
                # 尝试保存文件，如果失败则重试几次
                max_retries = 3
                retry_count = 0
                saved = False
                
                while not saved and retry_count < max_retries:
                    try:
                        # 先检查文件是否存在，如果存在则尝试删除
                        if os.path.exists(temp_file_path):
                            try:
                                os.remove(temp_file_path)
                                print(f"Removed existing file: {temp_file_path}")
                            except Exception as remove_error:
                                print(f"Warning: Could not remove existing file: {remove_error}")
                                # 使用不同的文件名重试
                                temp_file_path = os.path.join(current_dir, f"criteria_temp_data_{timestamp}_{retry_count}.csv")
                        
                        # 尝试保存文件
                        with open(temp_file_path, 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            # 写入表头（包含处理后的上下限和描述）
                            writer.writerow(["DataType", "StandardType", "LowerLimit", "UpperLimit", "Description"])
                            # 写入处理后的数据
                            writer.writerows(processed_criteria_data)
                        
                        saved = True
                        print(f"Successfully saved criteria data to: {temp_file_path}")
                        
                    except PermissionError as e:
                        retry_count += 1
                        print(f"Permission error saving file (attempt {retry_count}/{max_retries}): {e}")
                        # 等待一小段时间后重试
                        if retry_count < max_retries:
                            time.sleep(1)
                        # 使用不同的文件名
                        temp_file_path = os.path.join(current_dir, f"criteria_temp_data_{timestamp}_{retry_count}.csv")
                    except Exception as e:
                        print(f"Error saving file: {e}")
                        break
                
                if saved:
                    self.update_status(f"Processed criteria data saved to CSV file: {temp_file_path}")
                    return temp_file_path
                else:
                    # 如果仍然失败，尝试使用临时目录
                    import tempfile
                    try:
                        with tempfile.NamedTemporaryFile(suffix='.csv', prefix='criteria_temp_', delete=False, mode='w', newline='', encoding='utf-8') as f:
                            temp_file_path = f.name
                            writer = csv.writer(f)
                            writer.writerow(["DataType", "StandardType", "LowerLimit", "UpperLimit", "Description"])
                            writer.writerows(processed_criteria_data)
                        
                        self.update_status(f"Processed criteria data saved to temporary directory: {temp_file_path}")
                        return temp_file_path
                    except Exception as temp_error:
                        print(f"Failed to save to temporary directory: {temp_error}")
            
            # 如果没有数据
            self.update_status("No criteria data to save")
            return None
        except Exception as e:
            self.update_status(f"Error saving criteria data to CSV file: {str(e)}")
            print(f"Error saving criteria data: {e}")
            return None
    
    def show_cpk_tab(self):
        """显示Cpk选项卡并生成Cpk统计表格"""
        try:
            # 切换到Cpk选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'cpk_tab'):
                self.tab_control.select(self.cpk_tab)
            else:
                self.update_status("Cpk tab not found")
                return
            
            # 清空选项卡内容
            for widget in self.cpk_tab.winfo_children():
                widget.destroy()
            
            # 检查是否有Data Re-Processing数据可供分析
            if not hasattr(self, 'reprocessed_data') or self.reprocessed_data is None or self.reprocessed_data.empty:
                self.update_status("No Data Re-Processing data available for Cpk analysis")
                hint_label = tk.Label(self.cpk_tab, text="Please perform Data Re-Processing first", font=("SimHei", 12))
                hint_label.pack(pady=20)
                return
            
            # 优化：先将Criteria标签页的数据保存为临时文件
            self.update_status("Saving Criteria data to temporary file...")
            temp_file_path = self._save_criteria_to_temp_file()
            
            # 存储临时文件路径供_calculate_cpk_data使用
            self._criteria_temp_file = temp_file_path
            
            self.update_status("Generating Cpk Analysis...")
            
            # 更新选项卡状态
            self.update_tab_status("Cpk")
            
            # 创建标题
            title_label = tk.Label(self.cpk_tab, text="Cpk Statistical Analysis", font=("SimHei", 12, "bold"))
            title_label.pack(pady=10)
            
            # 创建横向排列的容器
            h_container = tk.Frame(self.cpk_tab)
            h_container.pack(fill="both", expand=True, padx=10, pady=5)
            
            # 设置grid布局使两列平均分配空间
            h_container.grid_columnconfigure(0, weight=1)
            h_container.grid_columnconfigure(1, weight=1)
            h_container.grid_rowconfigure(0, weight=1)
            
            # 分别处理White和Mixed数据
            self._create_cpk_table(h_container, "White", grid_row=0, grid_col=0)
            self._create_cpk_table(h_container, "Mixed", grid_row=0, grid_col=1)
            
            # 添加提示信息
            hint_label = tk.Label(self.cpk_tab, text="Tip: Click on column headers to sort data, use Ctrl+C to copy selected content, or right-click menu to copy",
                                font=("SimHei", 10), fg="gray")
            hint_label.pack(pady=5)
            
        except Exception as e:
            self.update_status(f"Error displaying Cpk tab: {str(e)}")
            error_label = tk.Label(self.cpk_tab, text=f"Error: {str(e)}", font=("SimHei", 10), fg="red")
            error_label.pack(pady=20)
    
    def _create_cpk_table(self, parent, data_type, grid_row=0, grid_col=0):
        """创建Cpk统计表格
        
        Args:
            parent: 父容器
            data_type: 数据类型 ("White" 或 "Mixed")
            grid_row: grid布局的行号
            grid_col: grid布局的列号
        """
        try:
            # 创建选项卡容器
            frame = tk.LabelFrame(parent, text=f"{data_type} Cpk Statistics", font=("SimHei", 10))
            frame.grid(row=grid_row, column=grid_col, sticky="nsew", padx=5, pady=5)
            
            # 定义表格列 - 移除LSL和USL列，增加Spec列
            columns = ("No", "Test Item", "Spec", 
                      "Mean", "StdDev", "Cpk", "Total", "Fail", "Fail Rate")
            
            # 创建Treeview表格组件
            tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode="extended")
            
            # 设置列宽和表头 - 统一的列宽设置，确保两个子窗口的列宽度一致
            tree.column("No", width=60, anchor="center")
            tree.column("Test Item", width=250, anchor="w")  # 标题列靠左对齐，宽度250像素
            tree.column("Spec", width=100, anchor="center")  # 新增规格列
            tree.column("Mean", width=100, anchor="center")
            tree.column("StdDev", width=100, anchor="center")
            tree.column("Cpk", width=80, anchor="center")
            tree.column("Total", width=100, anchor="center")
            tree.column("Fail", width=100, anchor="center")
            tree.column("Fail Rate", width=100, anchor="center")
            
            # 添加排序相关的实例变量
            if not hasattr(self, 'cpk_sort_column'):
                self.cpk_sort_column = ""
                self.cpk_sort_order = ""
                self.cpk_data = {}
            
            # 定义排序函数
            def sort_by_column(tree, col, reverse):
                """根据列对表格数据进行排序，确保Cpk等数值列正确排序"""
                # 获取所有数据行
                items = []
                for k in tree.get_children(''):
                    value = tree.set(k, col)
                    # 保存原始值用于后续处理
                    items.append((value, k))
                
                # 准备用于排序的键值对列表
                sorted_items = []
                for value, item_id in items:
                    try:
                        # 特殊处理Fail Rate列，去除百分号并转换为浮点数
                        if col == "Fail Rate":
                            if isinstance(value, str) and '%' in value:
                                sort_key = float(value.strip('%'))
                            else:
                                sort_key = float(value) if value and value.replace('.', '', 1).isdigit() else 0
                        # 特殊处理数值列，确保正确转换为浮点数
                        elif col in ["No", "Mean", "StdDev", "Cpk", "Total", "Fail"]:
                            # 检查是否是数字字符串并转换
                            if isinstance(value, str):
                                if value.replace('.', '', 1).isdigit():
                                    sort_key = float(value)
                                else:
                                    # 非数字值使用一个极小值或极大值
                                    sort_key = float('-inf') if reverse else float('inf')
                            else:
                                # 已经是数值类型
                                sort_key = float(value)
                        else:
                            # 非数值列按字符串排序
                            sort_key = str(value).lower()
                    except (ValueError, AttributeError):
                        # 转换失败时使用字符串排序
                        sort_key = str(value)
                    
                    sorted_items.append((sort_key, value, item_id))
                
                # 使用排序键进行排序
                sorted_items.sort(key=lambda x: x[0], reverse=reverse)
                
                # 重新排列行
                for index, (_, _, item_id) in enumerate(sorted_items):
                    tree.move(item_id, '', index)
                
                # 更新排序状态
                tree.heading(col, command=lambda _col=col: sort_by_column(tree, _col, not reverse))
                
                # 更新表头显示，添加排序指示符
                for c in tree['columns']:
                    if c == col:
                        sort_indicator = "▼" if reverse else "▲"
                        tree.heading(c, text=f"{c} {sort_indicator}")
                    else:
                        # 移除其他列的排序指示符
                        original_text = c
                        tree.heading(c, text=original_text)
            
            # 表头点击处理函数
            def on_heading_click(col):
                """处理表头点击事件"""
                # 检查当前排序状态
                if self.cpk_sort_column == col:
                    # 如果点击的是当前排序列，则切换排序方向
                    new_reverse = self.cpk_sort_order != "desc"
                else:
                    # 如果点击的是新列，则重置排序状态，默认升序
                    new_reverse = False
                
                # 更新排序状态
                self.cpk_sort_column = col
                self.cpk_sort_order = "desc" if new_reverse else "asc"
                
                # 执行排序
                sort_by_column(tree, col, new_reverse)
            
            # 设置表头并绑定点击事件
            for col in columns:
                tree.heading(col, text=col, command=lambda _col=col: on_heading_click(_col))
            
            # 配置Treeview样式
            style = ttk.Style()
            style.configure("Treeview", rowheight=25, font=("SimHei", 10))
            style.configure("Treeview.Heading", font=("SimHei", 10, "bold"), relief="solid", borderwidth=1)
            
            # 获取Cpk数据
            cpk_data = self._calculate_cpk_data(data_type)
            
            # 如果没有数据，只显示提示标签
            if not cpk_data:
                no_data_label = tk.Label(frame, text=f"No {data_type} data available", font=("SimHei", 10), fg="orange")
                no_data_label.pack(pady=20)
                return
            
            # 保存数据到实例变量供排序使用
            self.cpk_data[data_type] = cpk_data
            
            # 添加数据行
            for idx, row in enumerate(cpk_data, 1):
                # 格式化数据，保留适当的小数位数
                # 检查标题列是否包含特殊关键字
                column_name = row["column_name"]
                special_keywords = ["Ru", "Rv", "Du", "Dv"]
                is_special_row = any(keyword in column_name for keyword in special_keywords)
                
                # 根据是否包含特殊关键字设置小数位数
                decimal_places = 4 if is_special_row else 2
                
                # 确定要显示的规格值
                # 对于White dL Max和Mixed dL Max类型，强制使用上限作为规格值
                if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                    spec_value = row['upper_limit']
                # 对于White L, Mixed L, White U, Mixed U类型，使用下限作为规格值
                elif ('White' in column_name or 'Mixed' in column_name) and ('L' in column_name or 'U' in column_name):
                    spec_value = row['lower_limit']
                else:
                    # 对于其他类型的数据，根据Cpk计算逻辑显示保留的规格值
                    # 检查是否是单边规格（根据之前的Cpk计算逻辑）
                    if row['lower_limit'] == 0 and row['upper_limit'] != 0:
                        # 下限为0，显示上限作为规格值
                        spec_value = row['upper_limit']
                    elif row['upper_limit'] == 0 and row['lower_limit'] != 0:
                        # 上限为0，显示下限作为规格值
                        spec_value = row['lower_limit']
                    else:
                        # 对于双边规格，显示适当的规格值
                        # 这里可以根据具体需求决定显示哪个值或两个值
                        # 为了简单起见，我们可以显示一个组合值
                        spec_value = f"{row['lower_limit']}-{row['upper_limit']}"
                
                # 格式化规格值
                if isinstance(spec_value, (int, float)):
                    # 如果是数字类型，进行数值格式化
                    formatted_spec = f"{spec_value:.{decimal_places}f}"
                else:
                    # 如果是字符串类型（如双边规格的组合字符串），直接使用
                    formatted_spec = str(spec_value)
                
                formatted_row = [
                    str(idx),
                    column_name,  # 使用原始列标题作为Standard Type
                    formatted_spec,
                    f"{row['mean']:.{decimal_places}f}" if isinstance(row['mean'], (int, float)) else "NA",
                    f"{row['std_dev']:.{decimal_places}f}" if isinstance(row['std_dev'], (int, float)) else "NA",
                    f"{row['cpk']:.2f}" if isinstance(row['cpk'], (int, float)) else "NA",
                    str(row['total_count']),
                    str(row['fail_count']),
                    f"{row['fail_rate']:.2f}%" if isinstance(row['fail_rate'], (int, float)) else "NA"
                ]
                item_id = tree.insert("", "end", values=formatted_row)
                
                # 根据Cpk值设置行颜色
                if isinstance(row['cpk'], (int, float)):
                    if row['cpk'] >= 1.33:
                        tree.item(item_id, tags=('good_cpk',))
                    elif 1.0 <= row['cpk'] < 1.33:
                        tree.item(item_id, tags=('warning_cpk',))
                    else:  # Cpk < 1.0
                        tree.item(item_id, tags=('low_cpk',))
            
            # 设置不同Cpk值范围的行样式
            tree.tag_configure('good_cpk', background='#CCFFCC')   # 淡绿色：Cpk ≥ 1.33
            tree.tag_configure('warning_cpk', background='#FFFFCC')  # 淡黄色：1.0 ≤ Cpk < 1.33
            tree.tag_configure('low_cpk', background='#FFCCCC')      # 淡红色：Cpk < 1.0
            
            # 添加垂直滚动条
            yscrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscroll=yscrollbar.set)
            
            # 添加水平滚动条
            xscrollbar = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
            tree.configure(xscroll=xscrollbar.set)
            
            # 布局组件
            tree.pack(side="top", fill="both", expand=True)
            yscrollbar.pack(side="right", fill="y")
            xscrollbar.pack(side="bottom", fill="x")
            
            # 添加复制功能支持
            def copy_selection(event):
                """复制选中的单元格内容，同时包含表头信息"""
                selected_items = tree.selection()
                if not selected_items:
                    return
                
                # 构建复制文本，先添加表头
                copy_text = []
                # 获取表头文本
                headers = [tree.heading(col, "text") for col in tree['columns']]
                copy_text.append("\t".join(headers))
                
                # 添加选中的数据行
                for item in selected_items:
                    values = tree.item(item, "values")
                    copy_text.append("\t".join(values))
                
                # 将文本复制到剪贴板
                self.root.clipboard_clear()
                self.root.clipboard_append("\n".join(copy_text))
                self.update_status(f"Copied {len(selected_items)} rows from {data_type} Cpk Table to Clipboard")
            
            # 绑定Ctrl+C快捷键复制
            tree.bind("<Control-c>", copy_selection)
            tree.bind("<Control-C>", copy_selection)
            
            # 添加右键菜单支持复制
            def show_context_menu(event):
                """显示右键菜单"""
                # 只在点击单元格或行时显示右键菜单
                if tree.identify_region(event.x, event.y) in ("cell", "row"):
                    # 选中点击的项目
                    item = tree.identify_row(event.y)
                    if item:
                        tree.selection_set(item)
                    # 创建右键菜单
                    menu = tk.Menu(frame, tearoff=0)
                    menu.add_command(label="Copy", command=lambda: copy_selection(None))
                    # 显示右键菜单
                    menu.post(event.x_root, event.y_root)
            
            # 绑定右键菜单
            tree.bind("<Button-3>", show_context_menu)
                
        except Exception as e:
            error_label = tk.Label(frame, text=f"Error creating {data_type} Cpk table: {str(e)}", font=("SimHei", 10), fg="red")
            error_label.pack(pady=10)
    
    def _calculate_cpk_data(self, data_type):
        """
        计算Cpk数据
        
        Args:
            data_type: 数据类型 ("White" 或 "Mixed")
            
        Returns:
            list: Cpk统计数据列表
        """
        try:
            import os
            import glob
            import csv
            import json
            import pandas as pd
            import tempfile
            import time
            
            # 将reprocessed_data保存到临时文件
            temp_data_file = None
            working_data = None
            temp_file_path = None  # 初始化temp_file_path变量以避免未定义错误
            
            if hasattr(self, 'reprocessed_data') and self.reprocessed_data is not None and not self.reprocessed_data.empty:
                try:
                    # 创建带时间戳的临时文件名
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    temp_data_file = os.path.join(tempfile.gettempdir(), f'TestLogAnalyzer_ReprocessedData_{timestamp}.csv')
                    
                    # 保存reprocessed_data到CSV临时文件
                    self.reprocessed_data.to_csv(temp_data_file, index=False, encoding='utf-8')
                    self.update_status(f"Reprocessed data saved to temporary file: {temp_data_file}")
                    
                    # 从临时文件读取数据，确保数据一致性
                    working_data = pd.read_csv(temp_data_file, encoding='utf-8')
                    self.update_status(f"Reprocessed data loaded from temporary file for {data_type} analysis")
                    
                except Exception as temp_save_error:
                    self.update_status(f"Warning: Failed to save/load reprocessed data to temporary file: {str(temp_save_error)}")
                    # 如果临时文件操作失败，直接使用reprocessed_data
                    working_data = self.reprocessed_data
            
            # 如果没有可用的数据，返回空列表
            if working_data is None or working_data.empty:
                self.update_status(f"No valid data available for {data_type} Cpk calculation")
                return []
            
            # 定义标准类型的描述映射（用于与criteria表中的描述列匹配）
            standard_descriptions = {
                "L": ["Luminance", "亮度", "L ", "L ("],
                "U": ["Uniformity", "均匀性", "U ", "U ("],
                "dY": ["dY", "Delta Y", "dY ("],
                "u": ["u'", "u ", "u坐标"],
                "v": ["v'", "v ", "v坐标"],
                "Ru": ["Ru", "R_u"],
                "Rv": ["Rv", "R_v"],
                "Du": ["Du", "D_u"],
                "Dv": ["Dv", "D_v"],
                "dL*Min": ["dL*Min", "dL*最小值", "dL* Min", "dL*Min ("],
                "dL*Max": ["dL*Max", "dL*最大值", "dL* Max", "dL*Max ("],
                "dEMax": ["dEMax", "dE最大值", "Delta E Max", "dEMax ("],
                "Metric2": ["M2", "Gradient"],
                "Metric3": ["M3", "MediumRangeUniformity"],
                "Metric7": ["M7", "YellowPatches"],
                "Metric13": ["M13", "GlobalDelta"],
                "Metric14": ["M14", "ShortEdgeBrightLine"],
                "Metric15": ["M15", "LongEdgeHotSpot"],
                "Metric17": ["M17", "LongEdgeDarkBand"]
            }
            
            cpk_results = []
            std_type_to_specs = {}
            
            # 从JSON临时文件读取规格数据（优先方式）
            json_file_loaded = False
            try:
                import tempfile
                # 获取临时目录中的JSON文件路径
                temp_dir = tempfile.gettempdir()
                json_file_path = os.path.join(temp_dir, 'TestLogAnalyzer_Criteria.json')
                
                # 检查JSON文件是否存在
                if os.path.exists(json_file_path) and os.path.getsize(json_file_path) > 0:
                    try:
                        with open(json_file_path, 'r', encoding='utf-8') as f:
                            criteria_data = json.load(f)
                            
                        # 检查是否有当前数据类型的数据
                        if isinstance(criteria_data, dict) and data_type in criteria_data:
                            type_criteria = criteria_data[data_type]
                            if isinstance(type_criteria, dict):
                                for std_type, limits in type_criteria.items():
                                    # 确保limits是一个包含下限和上限的元组或列表
                                    if isinstance(limits, (list, tuple)) and len(limits) >= 2:
                                        lower_str = limits[0].strip() if isinstance(limits[0], str) else str(limits[0])
                                        upper_str = limits[1].strip() if isinstance(limits[1], str) else str(limits[1])
                                        
                                        try:
                                            # 尝试转换为浮点数
                                            lower = float(lower_str) if lower_str else None
                                            upper = float(upper_str) if upper_str else None
                                            
                                            # 确保有有效的上下限
                                            if lower is not None and upper is not None:
                                                # 添加数据类型前缀以区分White和Mixed的规格
                                                qualified_std_type = f"{data_type}_{std_type}"
                                                std_type_to_specs[qualified_std_type] = (lower, upper)
                                                std_type_to_specs[std_type] = (lower, upper)
                                                
                                        except (ValueError, TypeError):
                                            # 解析失败时跳过该规格
                                            print(f"Failed to parse limits from JSON for {std_type}: Lower={lower_str}, Upper={upper_str}")
                        
                        json_file_loaded = True
                        self.update_status(f"Successfully loaded criteria from JSON file: {json_file_path}")
                        
                        # 如果成功从JSON加载了规格数据，则直接继续进行数据处理
                        if std_type_to_specs:
                            self.update_status(f"Loaded {len(std_type_to_specs)} criteria items from JSON for {data_type}")
                        else:
                            self.update_status(f"No valid criteria found in JSON file for {data_type}")
                            # 重置json_file_loaded标志，以便后续尝试其他数据源
                            json_file_loaded = False
                    except json.JSONDecodeError as json_decode_error:
                        self.update_status(f"JSON decode error: {str(json_decode_error)}")
                        print(f"JSON文件格式错误: {str(json_decode_error)}")
                        # 删除损坏的JSON文件，以便下次重新创建
                        try:
                            os.remove(json_file_path)
                            self.update_status(f"Removed corrupted JSON file: {json_file_path}")
                        except Exception as remove_error:
                            print(f"无法删除损坏的JSON文件: {str(remove_error)}")
                    except Exception as json_file_error:
                        self.update_status(f"Error reading JSON file: {str(json_file_error)}")
                        print(f"读取JSON文件时出错: {str(json_file_error)}")
                else:
                    self.update_status(f"JSON file not found or empty: {json_file_path}")
            except Exception as json_error:
                self.update_status(f"Error loading criteria from JSON file: {str(json_error)}")
                print(f"JSON加载错误详情: {str(json_error)}")
                import traceback
                traceback.print_exc()
            
            # 如果JSON文件加载失败或没有数据，则尝试从CSV文件读取（备用方式）
            if not json_file_loaded or not std_type_to_specs:
                self.update_status(f"Attempting to load criteria from CSV files for {data_type}")
                # 获取临时文件路径，支持带时间戳的文件名
                temp_file_path = None
                
                # 1. 首先检查self._criteria_temp_file
                if hasattr(self, '_criteria_temp_file') and self._criteria_temp_file and os.path.exists(self._criteria_temp_file):
                    temp_file_path = self._criteria_temp_file
                    self.update_status(f"Found criteria temp file at: {temp_file_path}")
                else:
                    # 2. 搜索当前目录下的所有criteria_temp_data文件（包括带时间戳的）
                    import glob
                    current_dir = os.path.dirname(os.path.abspath(__file__))
                    pattern = os.path.join(current_dir, "criteria_temp_data*.csv")
                    matching_files = glob.glob(pattern)
                    
                    # 按修改时间排序，选择最新的文件
                    if matching_files:
                        matching_files.sort(key=os.path.getmtime, reverse=True)
                        temp_file_path = matching_files[0]
                        self.update_status(f"Found latest criteria temp file: {temp_file_path}")
                        # 更新self._criteria_temp_file以便下次使用
                        if hasattr(self, '_criteria_temp_file'):
                            self._criteria_temp_file = temp_file_path
            
            # 从临时文件读取Criteria数据
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    import os
                    import csv
                    
                    with open(temp_file_path, 'r', encoding='utf-8') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            # 只处理当前数据类型的规格
                            if row['DataType'] == data_type:
                                std_type = row['StandardType'].strip()
                                
                                # 直接从新格式的CSV中读取上下限
                                lower_str = row.get('LowerLimit', '').strip()
                                upper_str = row.get('UpperLimit', '').strip()
                                
                                # 尝试将上下限转换为浮点数
                                try:
                                    lower = float(lower_str) if lower_str else None
                                    upper = float(upper_str) if upper_str else None
                                    
                                    # 确保有有效的上下限
                                    if lower is not None and upper is not None:
                                        # 添加数据类型前缀以区分White和Mixed的规格
                                        qualified_std_type = f"{data_type}_{std_type}"
                                        std_type_to_specs[qualified_std_type] = (lower, upper)
                                        
                                        # 只在第一次遇到某个标准类型时保存原始映射
                                        if std_type not in std_type_to_specs:
                                            std_type_to_specs[std_type] = (lower, upper)
                                except ValueError:
                                    # 解析失败时跳过该规格
                                    print(f"Failed to parse limits for {std_type}: Lower={lower_str}, Upper={upper_str}")
                    
                    self.update_status(f"Loaded processed criteria data from CSV file for {data_type}")
                except Exception as e:
                    self.update_status(f"Error reading criteria data from CSV file: {str(e)}")
                    print(f"Error reading criteria CSV file: {e}")
            
            # 如果没有从CSV文件读取到规格数据，尝试从working_data中获取作为后备
            if not std_type_to_specs:
                criteria_column = f'{data_type} Pass/Fail Criteria'
                if working_data is not None and not working_data.empty and criteria_column in working_data.columns:
                    # 获取第一个非空的标准值
                    criteria_values = working_data[criteria_column].dropna()
                    if not criteria_values.empty:
                        first_criteria = criteria_values.iloc[0]
                        criteria_data = self._parse_criteria_string(first_criteria)
                        
                        # 处理解析出的criteria数据
                        for item in criteria_data:
                            if isinstance(item, (list, tuple)) and len(item) >= 2:
                                std_type = item[0]
                                specs_str = item[1]
                                
                                # 分割规格字符串
                                specs = []
                                if specs_str:
                                    if '~' in specs_str:
                                        specs = specs_str.split('~')
                                    elif '-' in specs_str and not specs_str.startswith('-'):
                                        specs = specs_str.split('-', 1)
                                    elif ' ' in specs_str:
                                        specs = specs_str.split()
                                    elif ',' in specs_str:
                                        specs = specs_str.split(',')
                                    elif '|' in specs_str:
                                        specs = specs_str.split('|')
                                
                                # 解析规格值
                                if len(specs) >= 2:
                                    try:
                                        lower_str = specs[0].strip()
                                        upper_str = specs[1].strip()
                                        
                                        lower = float(''.join(filter(lambda x: x.isdigit() or x == '.' or x == '-', lower_str)))
                                        upper = float(''.join(filter(lambda x: x.isdigit() or x == '.' or x == '-', upper_str)))
                                        
                                        qualified_std_type = f"{data_type}_{std_type}"
                                        std_type_to_specs[qualified_std_type] = (lower, upper)
                                        std_type_to_specs[std_type] = (lower, upper)
                                    except ValueError:
                                        # 解析失败时跳过
                                        pass
            
            # 从working_data中筛选目标数据列
            # 根据data_type筛选对应的列，并严格排除包含"Avg"或"Criteria"的列
            target_columns = []
            if working_data is not None and not working_data.empty:
                for col in working_data.columns:
                    # 排除不需要的列
                    if "Avg" in col or "Criteria" in col or col.strip() == "" or col.startswith("Unnamed"):
                        continue
                    
                    # 根据data_type筛选对应的列
                    if data_type == "White":
                        # White类型：只处理以"White "开头或"W_M"开头的列
                        if col.startswith("White ") or col.startswith("W_M"):
                            target_columns.append(col)
                    elif data_type == "Mixed":
                        # Mixed类型：只处理以"Mixed "开头或"M_M"开头的列
                        if col.startswith("Mixed ") or col.startswith("M_M"):
                            target_columns.append(col)
            
            # 为每个目标列计算Cpk
            for column_name in target_columns:
                # 尝试匹配标准类型
                matched_std_type = None
                
                # 1. 优先处理White/Mixed开头的标准列，如"White L (cd/m^2)" 或 "Mixed L (cd/m^2)"
                if column_name.startswith("White ") or column_name.startswith("Mixed "):
                    # 提取White/Mixed后面的部分
                    suffix = column_name.split(" ", 1)[1]
                    # 处理带括号的情况，如"L (cd/m^2)" -> "L"
                    if "(" in suffix:
                        potential_std_type = suffix.split("(")[0].strip()
                    else:
                        potential_std_type = suffix.split(" ")[0].strip()
                    
                    # 构建带数据类型前缀的标准类型，优先匹配
                    qualified_potential_std_type = f"{data_type}_{potential_std_type}"
                    
                    # 优先查找带数据类型前缀的标准类型
                    if qualified_potential_std_type in std_type_to_specs:
                        matched_std_type = qualified_potential_std_type
                    else:
                        # 然后查找原始标准类型
                        for std_type in std_type_to_specs.keys():
                            if potential_std_type == std_type or potential_std_type == std_type.replace("*", ""):
                                matched_std_type = std_type
                                break
                
                # 2. 处理Metric类型的列名，严格匹配"W_MX_Xxx" 或 "M_MX_Xxx"格式
                if not matched_std_type:
                    import re
                    # 处理White Metric格式: W_M2_Gradient -> White_Metric2
                    if data_type == "White":
                        w_metric_match = re.search(r'^W_M(\d+)_', column_name)
                        if w_metric_match:
                            metric_number = w_metric_match.group(1)
                            potential_metric = f"Metric{metric_number}"
                            qualified_potential_metric = f"White_{potential_metric}"
                            # 优先使用带数据类型前缀的规格
                            if qualified_potential_metric in std_type_to_specs:
                                matched_std_type = qualified_potential_metric
                            elif potential_metric in std_type_to_specs:
                                matched_std_type = potential_metric
                    
                    # 处理Mixed Metric格式: M_M2_Gradient -> Mixed_Metric2
                    elif data_type == "Mixed":
                        m_metric_match = re.search(r'^M_M(\d+)_', column_name)
                        if m_metric_match:
                            metric_number = m_metric_match.group(1)
                            potential_metric = f"Metric{metric_number}"
                            qualified_potential_metric = f"Mixed_{potential_metric}"
                            # 优先使用带数据类型前缀的规格
                            if qualified_potential_metric in std_type_to_specs:
                                matched_std_type = qualified_potential_metric
                            elif potential_metric in std_type_to_specs:
                                matched_std_type = potential_metric
                
                # 方法1: 直接匹配criteria表中的标准类型（精确匹配优先级最高）
                if not matched_std_type:
                    # 优先匹配带数据类型前缀的标准类型
                    for std_type in std_type_to_specs.keys():
                        # 检查是否是当前数据类型的规格
                        if std_type.startswith(f"{data_type}_"):
                            # 提取原始标准类型
                            original_std_type = std_type[len(data_type)+1:]
                            std_type_clean = original_std_type.replace("*", "")
                            # 更精确的匹配：列名包含完整的标准类型，并且前后不是字母或数字
                            if (f" {original_std_type} " in f" {column_name} " or 
                                f" {std_type_clean} " in f" {column_name} " or
                                column_name.endswith(f" {original_std_type}") or
                                column_name.endswith(f" {std_type_clean}") or
                                column_name.startswith(f"{original_std_type} ") or
                                column_name.startswith(f"{std_type_clean} ") or
                                column_name == original_std_type or
                                column_name == std_type_clean or
                                column_name.endswith(f"({original_std_type}") or
                                column_name.endswith(f"({std_type_clean}")):
                                matched_std_type = std_type
                                break
                    
                    # 如果没有匹配到带前缀的，再尝试匹配原始标准类型
                    if not matched_std_type:
                        for std_type in std_type_to_specs.keys():
                            # 跳过已经检查过的带前缀的标准类型
                            if std_type.startswith("White_") or std_type.startswith("Mixed_"):
                                continue
                            # 处理带*的标准类型
                            std_type_clean = std_type.replace("*", "")
                            # 更精确的匹配：列名包含完整的标准类型，并且前后不是字母或数字
                            if (f" {std_type} " in f" {column_name} " or 
                                f" {std_type_clean} " in f" {column_name} " or
                                column_name.endswith(f" {std_type}") or
                                column_name.endswith(f" {std_type_clean}") or
                                column_name.startswith(f"{std_type} ") or
                                column_name.startswith(f"{std_type_clean} ") or
                                column_name == std_type or
                                column_name == std_type_clean or
                                column_name.endswith(f"({std_type}") or
                                column_name.endswith(f"({std_type_clean}")):
                                matched_std_type = std_type
                                break
                
                # 方法2: 如果没有直接匹配，使用描述关键词匹配
                if not matched_std_type:
                    for std_type, keywords in standard_descriptions.items():
                        for keyword in keywords:
                            if keyword in column_name:
                                # 优先检查带数据类型前缀的标准类型
                                qualified_std_type = f"{data_type}_{std_type}"
                                if qualified_std_type in std_type_to_specs:
                                    matched_std_type = qualified_std_type
                                    break
                                # 然后检查原始标准类型
                                elif std_type in std_type_to_specs:
                                    matched_std_type = std_type
                                    break
                        if matched_std_type:
                            break
                
                # 只有找到匹配的标准类型才进行Cpk计算
                if not matched_std_type:
                    continue
                
                # 移除数据类型前缀，用于显示
                display_std_type = matched_std_type
                if matched_std_type.startswith("White_") or matched_std_type.startswith("Mixed_"):
                    display_std_type = matched_std_type.split("_", 1)[1]
                
                # 获取该列的数据
                col_data = working_data[column_name].dropna()
                
                # 过滤掉非数值数据
                try:
                    col_data = pd.to_numeric(col_data, errors='coerce')
                    col_data = col_data.dropna()
                except:
                    continue
                
                if len(col_data) < 3:  # 需要至少3个数据点才有意义
                    continue
                
                # 从规格映射中获取规格上下限
                if matched_std_type in std_type_to_specs:
                    lower_limit, upper_limit = std_type_to_specs[matched_std_type]
                else:
                    # 如果没有找到匹配的规格，跳过该列
                    continue
                
                # 修正规格值
                if lower_limit == -1:
                    lower_limit = 0
                
                # 对于Uniformity类型，确保上限不超过100
                if "U" == matched_std_type or "_U" in matched_std_type or "Uniformity" in matched_std_type:
                    upper_limit = min(upper_limit, 100.0)
                
                # 计算统计量
                mean_value = col_data.mean()
                std_dev_value = col_data.std()
                total_count = len(col_data)
                
                # 计算不良数量
                fail_count = 0
                # 对于White dL Max和Mixed dL Max类型，只检查是否超过上限
                if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                    for value in col_data:
                        if value > upper_limit:  # 只检查上限
                            fail_count += 1
                else:
                    # 其他类型正常检查上下限
                    for value in col_data:
                        if value < lower_limit or value > upper_limit:
                            fail_count += 1
                
                # 计算不良率
                fail_rate = (fail_count / total_count * 100) if total_count > 0 else 0
                
                # 计算Cpk
                cpk = 0
                if std_dev_value > 0:
                    # 特殊处理1：White dL Max和Mixed dL Max类型，强制使用上限，忽略下限
                    if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                        # 强制使用上限计算CPU
                        cpu = (upper_limit - mean_value) / (3 * std_dev_value) if upper_limit != float('inf') else float('inf')
                        cpk = cpu
                    # 特殊处理2：White L, Mixed L, White U, Mixed U 忽略规格上限，保留规格下限
                    elif ('White' in column_name or 'Mixed' in column_name) and ('L' in column_name or 'U' in column_name):
                        # 忽略上限，只使用下限计算Cpk
                        cpl = (mean_value - lower_limit) / (3 * std_dev_value) if lower_limit != -float('inf') else float('inf')
                        cpk = cpl
                    # 根据规格是否为0来决定使用单边规格算法
                    elif lower_limit == 0 and upper_limit != 0:
                        # 下限为0，使用单边规格算法（只考虑上限）
                        cpu = (upper_limit - mean_value) / (3 * std_dev_value) if upper_limit != float('inf') else float('inf')
                        cpk = cpu
                    elif upper_limit == 0 and lower_limit != 0:
                        # 上限为0，使用单边规格算法（只考虑下限）
                        cpl = (mean_value - lower_limit) / (3 * std_dev_value) if lower_limit != -float('inf') else float('inf')
                        cpk = cpl
                    else:
                        # 对于dL*Max类型，确保正确处理规格
                        if 'dL*Max' in matched_std_type or 'dL*Max' in column_name:
                            # 确保只使用非零的规格进行计算
                            if lower_limit != 0 and upper_limit != 0:
                                # 双边规格算法
                                cpl = (mean_value - lower_limit) / (3 * std_dev_value) if lower_limit != -float('inf') else float('inf')
                                cpu = (upper_limit - mean_value) / (3 * std_dev_value) if upper_limit != float('inf') else float('inf')
                                cpk = min(cpl, cpu)
                            elif lower_limit != 0:
                                # 只有下限非零，使用下限计算
                                cpl = (mean_value - lower_limit) / (3 * std_dev_value) if lower_limit != -float('inf') else float('inf')
                                cpk = cpl
                            elif upper_limit != 0:
                                # 只有上限非零，使用上限计算
                                cpu = (upper_limit - mean_value) / (3 * std_dev_value) if upper_limit != float('inf') else float('inf')
                                cpk = cpu
                        else:
                            # 其他类型的双边规格算法
                            cpl = (mean_value - lower_limit) / (3 * std_dev_value) if lower_limit != -float('inf') else float('inf')
                            cpu = (upper_limit - mean_value) / (3 * std_dev_value) if upper_limit != float('inf') else float('inf')
                            cpk = min(cpl, cpu)
                
                # 添加到结果列表
                cpk_results.append({
                    "standard_type": matched_std_type,  # 直接使用匹配的标准类型
                    "lower_limit": lower_limit,
                    "upper_limit": upper_limit,
                    "mean": mean_value,
                    "std_dev": std_dev_value,
                    "cpk": cpk,
                    "total_count": total_count,
                    "fail_count": fail_count,
                    "fail_rate": fail_rate,
                    "column_name": column_name  # 保存原始列名以便参考
                })
            
            # 保持原始列顺序（不按standard_type排序）
            return cpk_results
            
        except Exception as e:
            print(f"Error calculating Cpk data for {data_type}: {str(e)}")
            return []
    
    def _save_criteria_to_temp_file(self):
        """将White和Mixed的标准数据整合并保存为临时规格文件"""
        try:
            import json
            import tempfile
            import os
            import csv
            import time
            import logging
            
            # 获取logger实例
            logger = logging.getLogger("TestLogAnalyzer")
            
            # 创建保存所有标准数据的字典
            all_criteria = {}
            processed_criteria_data = []
            
            # 定义标准类型的描述映射
            standard_descriptions = {
                "L": "Luminance", "U": "Uniformity", "dY": "dY", "u": "u'", "v": "v'",
                "Ru": "Ru", "Rv": "Rv", "Du": "Du", "Dv": "Dv",
                "dL*Min": "dL*Min", "dL*Max": "dL*Max", "dEMax": "dEMax",
                "Metric2": "Metric2", "Metric3": "Metric3", "Metric7": "Metric7",
                "Metric13": "Metric13", "Metric14": "Metric14", "Metric15": "Metric15",
                "Metric17": "Metric17"
            }
            
            # 初始化criteria_data字典如果它不存在
            if not hasattr(self, 'criteria_data'):
                self.criteria_data = {}
                logger.warning("初始化新的criteria_data字典")
            
            # 优先使用已更新的self.criteria_data（从save_changes函数中获取的用户修改）
            logger.info("正在保存用户修改的规格数据")
            
            # 遍历criteria_data字典生成processed_criteria_data
            for data_type, types_dict in self.criteria_data.items():
                for standard_type, limits in types_dict.items():
                    # 确保limits是有效的列表或元组
                    if isinstance(limits, (list, tuple)) and len(limits) >= 2:
                        lower_limit = str(limits[0]) if limits[0] is not None else ""
                        upper_limit = str(limits[1]) if limits[1] is not None else ""
                        
                        # 添加描述
                        description = standard_descriptions.get(standard_type, f"{data_type} {standard_type}")
                        
                        # 添加到processed_criteria_data
                        processed_criteria_data.append([data_type, standard_type, lower_limit, upper_limit, description])
            
            # 如果criteria_data为空，则尝试从Treeview获取数据
            if not processed_criteria_data and hasattr(self, 'tree_criteria'):
                logger.info("criteria_data为空，尝试从Treeview获取数据")
                # 遍历Treeview中的所有项目
                for item in self.tree_criteria.get_children():
                    values = self.tree_criteria.item(item, "values")
                    if len(values) >= 5:
                        data_type = values[0]  # DataType
                        standard_type = values[1]  # StandardType
                        lower_limit = values[2]  # LowerLimit
                        upper_limit = values[3]  # UpperLimit
                        
                        # 添加到processed_criteria_data
                        processed_criteria_data.append([data_type, standard_type, lower_limit, upper_limit, values[4]])
                        
                        # 更新criteria_data字典
                        if data_type not in self.criteria_data:
                            self.criteria_data[data_type] = {}
                        self.criteria_data[data_type][standard_type] = [lower_limit, upper_limit]
            
            # 如果没有从Treeview获取到数据，尝试从processed_data中获取
            if not processed_criteria_data and hasattr(self, 'processed_data') and self.processed_data is not None and not self.processed_data.empty:
                for data_type in ["White", "Mixed"]:
                    criteria_column = f'{data_type} Pass/Fail Criteria'
                    if criteria_column in self.processed_data.columns:
                        criteria_values = self.processed_data[criteria_column].dropna()
                        if not criteria_values.empty:
                            first_criteria = criteria_values.iloc[0]
                            parsed_data = self._parse_criteria_string(first_criteria)
                            if parsed_data:
                                for item in parsed_data:
                                    if isinstance(item, (list, tuple)) and len(item) >= 2:
                                        standard_type = item[0]
                                        specs = item[1] if len(item) > 1 else ""
                                        
                                        # 处理规格值：使用多种分隔符尝试拆分
                                        lower_limit = ""
                                        upper_limit = ""
                                        
                                        for sep in ['|', '~', '-', ' ', ',']:
                                            if sep in specs and (sep != '-' or not specs.startswith('-')):
                                                limits = specs.split(sep, 1)
                                                lower_limit = limits[0].strip()
                                                upper_limit = limits[1].strip() if len(limits) > 1 else ""
                                                break
                                        
                                        # 修正规格值
                                        if standard_type == "U" and upper_limit == "0":
                                            upper_limit = "100"
                                        if lower_limit == "-1":
                                            lower_limit = "0"
                                        
                                        # 添加到processed_criteria_data
                                        description = standard_descriptions.get(standard_type, f"{data_type} {standard_type}")
                                        processed_criteria_data.append([data_type, standard_type, lower_limit, upper_limit, description])
                                        
                                        # 更新criteria_data字典
                                        if data_type not in self.criteria_data:
                                            self.criteria_data[data_type] = {}
                                        self.criteria_data[data_type][standard_type] = [lower_limit, upper_limit]
            
            # 创建临时文件路径
            temp_dir = tempfile.gettempdir()
            temp_file_path = os.path.join(temp_dir, 'TestLogAnalyzer_Criteria.json')
            
            # 确保criteria_data不为空，包含有效数据
            if not self.criteria_data:
                logger.warning("criteria_data为空，无法保存有效规格数据")
                self.update_status("Warning: No criteria data to save")
                return None
            
            # 将数据保存为JSON文件，添加更健壮的错误处理
            try:
                # 先写入临时文件
                temp_file_path_temp = f"{temp_file_path}.tmp"
                with open(temp_file_path_temp, 'w', encoding='utf-8') as f:
                    json.dump(self.criteria_data, f, indent=4, ensure_ascii=False)
                
                # 如果临时文件写入成功，再替换正式文件
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                os.rename(temp_file_path_temp, temp_file_path)
                
                logger.info(f"成功将规格数据保存到临时JSON文件: {temp_file_path}")
                logger.debug(f"保存的JSON数据内容: {self.criteria_data}")
            except Exception as json_error:
                logger.error(f"保存JSON临时文件时出错: {json_error}")
                raise
            
            # 同时创建CSV文件作为备用
            csv_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"criteria_temp_data_{int(time.time())}.csv")
            if processed_criteria_data:
                try:
                    with open(csv_file_path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f)
                        writer.writerow(["DataType", "StandardType", "LowerLimit", "UpperLimit", "Description"])
                        writer.writerows(processed_criteria_data)
                    logger.info(f"标准数据已保存到CSV文件: {csv_file_path}")
                except Exception as csv_error:
                    logger.error(f"保存CSV文件时出错: {csv_error}")
            
            # 使用logger替代print输出
            logger.info(f"标准数据已保存到临时文件: {temp_file_path}")
            logger.debug(f"保存的数据内容: {self.criteria_data}")
            self.update_status(f"成功保存规格数据到临时文件")
            
            # 返回CSV文件路径用于兼容旧逻辑
            return csv_file_path if processed_criteria_data else temp_file_path
            
        except Exception as e:
            self.update_status(f"Error saving criteria to temporary file: {str(e)}")
            print(f"保存标准数据时出错: {e}")
            import traceback
            traceback.print_exc()
            return None
            
    def _read_criteria_from_temp_file(self):
        """从临时文件读取保存的规格数据，如果文件存在且有效则返回数据，否则返回None"""
        try:
            import os
            import json
            import tempfile
            import logging
            logger = logging.getLogger("TestLogAnalyzer")
            
            # 构建临时文件路径
            temp_file_path = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer_Criteria.json")
            
            # 检查文件是否存在
            if not os.path.exists(temp_file_path):
                logger.info(f"临时规格文件不存在: {temp_file_path}")
                return None
            
            # 读取文件内容
            with open(temp_file_path, 'r', encoding='utf-8') as f:
                criteria_data = json.load(f)
            
            # 验证数据格式是否正确
            if isinstance(criteria_data, dict) and ('White' in criteria_data or 'Mixed' in criteria_data):
                logger.info(f"成功从临时文件读取规格数据: {len(criteria_data)}组数据")
                return criteria_data
            else:
                logger.warning(f"临时规格文件格式不正确，跳过读取")
                return None
        except Exception as e:
            logger.error(f"读取临时规格文件时发生错误: {str(e)}")
            return None
    
    def review_criteria(self):
        """从Data PreView中的第一个文件提取标准数据，在Review Criteria选项卡中显示，优先使用已保存的临时规格文件"""
        try:
            import logging
            logger = logging.getLogger("TestLogAnalyzer")
            
            # 尝试从临时文件读取已保存的规格数据
            saved_criteria_data = self._read_criteria_from_temp_file()
            
            # 获取所有选中的文件
            selected_files = [path for path, var in self.file_vars.items() if var.get()]
            
            if not selected_files:
                self.update_status("No files selected for criteria review.")
                return
            
            # 获取第一个选中的文件
            first_file_path = selected_files[0]
            file_name = os.path.basename(first_file_path)
            self.update_status(f"Reviewing criteria from file: {file_name}")
            
            # 清除Review Criteria选项卡的内容
            for widget in self.review_criteria_tab.winfo_children():
                widget.destroy()
            
            # 更新选项卡状态
            self.update_tab_status("Review Criteria")
            
            # 创建两个并排的框架
            left_frame = ttk.Frame(self.review_criteria_tab)
            right_frame = ttk.Frame(self.review_criteria_tab)
            
            # 使用grid布局使两个框架并排
            left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            right_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
            
            # 设置权重，使两个框架平均分配空间
            self.review_criteria_tab.grid_columnconfigure(0, weight=1)
            self.review_criteria_tab.grid_columnconfigure(1, weight=1)
            self.review_criteria_tab.grid_rowconfigure(0, weight=1)
            
            # 初始化存储标准数据的字典
            self.criteria_data = {"White": {}, "Mixed": {}}
            
            # 智能检测CSV文件中的数据标题行并读取文件
            with open(first_file_path, 'r', encoding='utf-8') as f:
                lines = []
                for i in range(20):
                    line = f.readline()
                    if not line:
                        break
                    lines.append(line.strip())
                
                # 识别元数据行
                metadata_lines = []
                non_metadata_lines = []
                
                for i, line in enumerate(lines):
                    is_metadata = False
                    if ',:' in line or (':' in line and ',' in line):
                        is_metadata = True
                    elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                        content_parts = [p.strip() for p in line.split(',') if p.strip()]
                        if content_parts and ':' in content_parts[0]:
                            is_metadata = True
                    
                    if is_metadata:
                        metadata_lines.append(i)
                    else:
                        non_metadata_lines.append((i, line))
                
                # 寻找标题行
                max_commas = -1
                header_line_candidates = []
                header_line_index = 0
                
                non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                
                if non_empty_non_metadata:
                    max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                    header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                    
                    if header_line_candidates:
                        best_score = -1
                        for i, line in header_line_candidates:
                            score = 1
                            if any(c.isalpha() for c in line):
                                score += 10
                            content_length = len(line.replace(',', ''))
                            score += min(content_length, 50)
                            
                            if score > best_score:
                                best_score = score
                                header_line_index = i
                
                # 确定是否使用检测到的标题行
                if max_commas >= 9:
                    df = pd.read_csv(first_file_path,
                                    skiprows=header_line_index,
                                    on_bad_lines='skip',
                                    engine='python')
                else:
                    df = pd.read_csv(first_file_path,
                                    on_bad_lines='skip',
                                    engine='python')
            
            # 尝试从不同的列名中提取标准信息
            white_criteria = None
            mixed_criteria = None
            
            # 检查可能的列名
            criteria_columns = ['White Pass/Fail Criteria', 'Mixed Pass/Fail Criteria',
                              'Pass/Fail Criteria', 'Criteria', 'Mixed PassFail Criteria',
                              'White PassFail Criteria']
            
            for col in criteria_columns:
                if col in df.columns:
                    # 获取第一个非空的值
                    values = df[col].dropna()
                    if not values.empty:
                        if 'White' in col:
                            white_criteria = values.iloc[0]
                        elif 'Mixed' in col:
                            mixed_criteria = values.iloc[0]
                        else:
                            # 如果没有明确的类型标识，尝试根据内容判断
                            if mixed_criteria is None:
                                mixed_criteria = values.iloc[0]
                            elif white_criteria is None:
                                white_criteria = values.iloc[0]
            
            # 显示White标准数据
            if white_criteria:
                self._display_criteria_data(left_frame, white_criteria, "White")
            else:
                no_data_label = tk.Label(left_frame, text="No White criteria data found in the file.", font=('SimHei', 10))
                no_data_label.pack(pady=20)
            
            # 显示Mixed标准数据
            if mixed_criteria:
                self._display_criteria_data(right_frame, mixed_criteria, "Mixed")
            else:
                no_data_label = tk.Label(right_frame, text="No Mixed criteria data found in the file.", font=('SimHei', 10))
                no_data_label.pack(pady=20)
            
            # 切换到Review Criteria选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'review_criteria_tab'):
                self.tab_control.select(self.review_criteria_tab)
                self.root.update_idletasks()
                self.review_criteria_tab.focus_set()
                self.review_criteria_tab.lift()
            
            # 更新状态栏信息
            self.update_status("Criteria review completed.")
            
            # 保存初始标准数据到临时文件
            try:
                self.update_status("开始保存Review Criteria设置...")
                success = self._save_criteria_to_temp_file()
                if success:
                    self.update_status("首次保存Review Criteria设置成功")
                else:
                    self.update_status("首次保存Review Criteria设置失败")
            except Exception as save_error:
                self.update_status(f"首次保存时发生错误: {str(save_error)}")
            
            # 自动执行两个子窗口的"保存修改"按钮功能
            # 添加短暂延迟确保UI元素完全加载，并增加重试机制
            def save_with_retry(attempt=0):
                try:
                    self.update_status(f"尝试自动保存Review Criteria设置 (尝试 {attempt+1}/3)...")
                    success = self._save_criteria_to_temp_file()
                    if success:
                        self.update_status("自动保存Review Criteria设置完成")
                        # 验证文件是否成功创建
                        temp_file_path = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer_Criteria.json")
                        if os.path.exists(temp_file_path):
                            self.update_status(f"确认: 标准数据已成功保存到 {temp_file_path}")
                    else:
                        if attempt < 2:  # 最多重试3次
                            self.update_status("保存失败，稍后重试...")
                            self.root.after(1000, lambda: save_with_retry(attempt + 1))
                        else:
                            self.update_status("多次尝试后保存仍然失败，请手动保存")
                except Exception as save_error:
                    self.update_status(f"自动保存时发生错误: {str(save_error)}")
                    if attempt < 2:
                         self.root.after(1000, lambda: save_with_retry(attempt + 1))
            
            # 增加5秒延迟以确保UI完全加载
            self.root.after(1000, save_with_retry)
        except Exception as e:
            self.update_status(f"Error reviewing criteria: {str(e)}")
            error_label = tk.Label(self.review_criteria_tab, text=f"Error: {str(e)}", font=('SimHei', 10), fg="red")
            error_label.pack(pady=20)
            
    def read_colorpoint_spec(self):
        '''Extract ColorPoint specification data from the first file in Data PreView and display it in the ColorPointSpec tab'''
        try:
            # 立即切换到ColorPointSpec选项卡，确保界面切换流畅无延迟
            if hasattr(self, 'tab_control') and hasattr(self, 'colorpoint_spec_tab'):
                # 先切换选项卡，再处理数据
                self.tab_control.select(self.colorpoint_spec_tab)
                # 强制刷新UI
                self.root.update_idletasks()
                self.colorpoint_spec_tab.focus_set()
                self.colorpoint_spec_tab.lift()
            
            # 获取所有选中的文件
            selected_files = [path for path, var in self.file_vars.items() if var.get()]
            
            if not selected_files:
                self.update_status("No files selected for ColorPointSpec review.")
                # 显示无文件选中的提示信息
                no_file_label = ttk.Label(self.colorpoint_spec_tab, 
                                        text="No files selected for ColorPointSpec review.\nPlease select a file in Data PreView first.", 
                                        foreground="red", font=('SimHei', 10))
                no_file_label.pack(pady=20, padx=10)
                return
            
            # 获取第一个选中的文件
            first_file_path = selected_files[0]
            file_name = os.path.basename(first_file_path)
            self.update_status(f"Reviewing ColorPointSpec from file: {file_name}")
            
            # 清除ColorPointSpec选项卡的内容
            for widget in self.colorpoint_spec_tab.winfo_children():
                widget.destroy()
            
            # 更新选项卡状态
            self.update_tab_status("ColorPointSpec")
            
            # 创建一个框架来容纳内容
            main_frame = ttk.Frame(self.colorpoint_spec_tab)
            main_frame.pack(fill="both", expand=True, padx=5, pady=5)
            
            # 初始化存储ColorPoint规格数据的字典
            self.colorpoint_spec_data = {"White": {}, "Mixed": {}}
            
            # 添加加载中的提示
            loading_label = ttk.Label(main_frame, text="Loading ColorPointSpec data...", font=('SimHei', 10))
            loading_label.pack(pady=20)
            self.root.update_idletasks()
            
            # 智能检测CSV文件中的数据标题行并读取文件
            try:
                with open(first_file_path, 'r', encoding='utf-8') as f:
                    lines = []
                    for i in range(20):
                        line = f.readline()
                        if not line:
                            break
                        lines.append(line.strip())
                    
                    # 识别元数据行
                    metadata_lines = []
                    non_metadata_lines = []
                    
                    for i, line in enumerate(lines):
                        is_metadata = False
                        if ',:' in line or (':' in line and ',' in line):
                            is_metadata = True
                        elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                            content_parts = [p.strip() for p in line.split(',') if p.strip()]
                            if content_parts and ':' in content_parts[0]:
                                is_metadata = True
                        
                        if is_metadata:
                            metadata_lines.append(i)
                        else:
                            non_metadata_lines.append((i, line))
                    
                    # 寻找标题行
                    max_commas = -1
                    header_line_candidates = []
                    header_line_index = 0
                    
                    non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                    
                    if non_empty_non_metadata:
                        max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                        header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                        
                        if header_line_candidates:
                            best_score = -1
                            for i, line in header_line_candidates:
                                score = 1
                                if any(c.isalpha() for c in line):
                                    score += 10
                                content_length = len(line.replace(',', ''))
                                score += min(content_length, 50)
                                
                                if score > best_score:
                                    best_score = score
                                    header_line_index = i
                    
                    # 确定是否使用检测到的标题行
                    if max_commas >= 9:
                        df = pd.read_csv(first_file_path,
                                        skiprows=header_line_index,
                                        on_bad_lines='skip',
                                        engine='python')
                    else:
                        df = pd.read_csv(first_file_path,
                                        on_bad_lines='skip',
                                        engine='python')
            except Exception as e:
                self.update_status(f"Error reading CSV file: {str(e)}")
                loading_label.destroy()
                error_label = ttk.Label(main_frame, text=f"Error reading CSV file: {str(e)}", foreground="red", font=('SimHei', 10))
                error_label.pack(pady=20)
                return
            
            # 移除加载中的提示
            loading_label.destroy()
            
            # 尝试从特定列中提取ColorPoint规格信息
            white_colorpoint = None
            mixed_colorpoint = None
            
            # 从White Pass/Fail Criteria列提取CAFL0标准坐标点
            if "White Pass/Fail Criteria" in df.columns:
                values = df["White Pass/Fail Criteria"].dropna()
                if not values.empty:
                    white_colorpoint = values.iloc[0]
            
            # 从Mixed Pass/Fail Criteria列提取CALF24标准坐标点
            if "Mixed Pass/Fail Criteria" in df.columns:
                values = df["Mixed Pass/Fail Criteria"].dropna()
                if not values.empty:
                    mixed_colorpoint = values.iloc[0]
            
            # 创建两个并排的框架来显示White和Mixed的ColorPoint规格
            white_frame = ttk.LabelFrame(main_frame, text="CAFL0 (White) ColorPointSpec")
            mixed_frame = ttk.LabelFrame(main_frame, text="CALF24 (Mixed) ColorPointSpec")
            
            # 使用grid布局使两个框架并排
            white_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            mixed_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
            
            # 设置权重，使两个框架平均分配空间
            main_frame.grid_columnconfigure(0, weight=1)
            main_frame.grid_columnconfigure(1, weight=1)
            main_frame.grid_rowconfigure(0, weight=1)
            
            # 保存对框架的引用，用于后续刷新多边形
            self.white_frame = white_frame
            self.mixed_frame = mixed_frame
            
            # 显示White ColorPoint规格数据
            if white_colorpoint:
                # 添加CAFL0标准坐标点解析状态提示
                status_label = ttk.Label(white_frame, text="Parsing CAFL0 (White) ColorPointSpec...", font=('SimHei', 10))
                status_label.pack(pady=5, padx=5, anchor='w')
                self.root.update_idletasks()
                
                # 尝试解析为u'v'坐标格式
                try:
                    # 清理输入数据，确保格式正确
                    cleaned_data = str(white_colorpoint).strip()
                    
                    # 使用增强的坐标点解析功能
                    colorpoint_items = self.parse_color_criteria(cleaned_data)
                    
                    # 验证解析结果
                    valid_points = []
                    for point in colorpoint_items:
                        if isinstance(point, tuple) and len(point) == 2:
                            u_val, v_val = point
                            # 过滤有效范围的坐标点 (0.0-1.0)
                            if isinstance(u_val, (int, float)) and isinstance(v_val, (int, float)):
                                if 0.0 <= u_val <= 1.0 and 0.0 <= v_val <= 1.0:
                                    valid_points.append((u_val, v_val))
                    
                    # 限制最多8个坐标点，确保预留8个坐标输入空位
                    if len(valid_points) > 8:
                        valid_points = valid_points[:8]
                    # 如果坐标点少于8个，补充空坐标点（使用None表示未获取的坐标）以确保显示8个输入框
                    while len(valid_points) < 8:
                        valid_points.append((None, None))
                    
                    status_label.config(text=f"Successfully parsed {len(valid_points)} CAFL0 (White) ColorPointSpec values")
                    
                    # 创建一个子框架，用于坐标显示
                    coord_frame = ttk.LabelFrame(white_frame, text="CAFL0 Coordinate Points")
                    
                    # 显示坐标点并允许编辑
                    self._display_editable_colorpoint_data(coord_frame, valid_points, "White", self.white_frame)
                    self.colorpoint_spec_data["White"] = valid_points
                    
                    # 只显示坐标框架，取消多边形绘图区
                    coord_frame.pack(fill="both", expand=True, padx=5, pady=5)
                except Exception as e:
                    # 如果解析失败，显示原始数据并使用默认值
                    status_label.config(text="Failed to parse CAFL0 ColorPointSpec, using default values", foreground="red")
                    self.logger.error(f"Error parsing White ColorPoint: {str(e)}")
                    
                    # 解析失败时创建8个默认坐标点，确保显示完整的8个输入框
                    empty_points = [(None, None) for _ in range(8)]
                    
                    # 创建框架和显示内容
                    coord_frame = ttk.LabelFrame(white_frame, text="CAFL0 Coordinate Points")
                    
                    # 显示空的坐标点列表
                    self._display_editable_colorpoint_data(coord_frame, empty_points, "White", self.white_frame)
                    self.colorpoint_spec_data["White"] = empty_points
                    
                    # 只显示坐标框架，取消多边形绘图区
                    coord_frame.pack(fill="both", expand=True, padx=5, pady=5)
                    
                    # 添加原始数据显示框
                    raw_frame = ttk.LabelFrame(white_frame, text="Original CAFL0 Data (Parsing Failed)")
                    raw_frame.pack(fill="both", expand=True, padx=5, pady=5)
                    raw_data_label = tk.Label(raw_frame, text=f"{white_colorpoint}", font=('SimHei', 9), justify="left", wraplength=300)
                    raw_data_label.pack(padx=5, pady=5, fill="both", expand=True)
            else:
                no_data_label = tk.Label(white_frame, text="No CAFL0 (White) ColorPointSpec found in the file.", font=('SimHei', 10))
                no_data_label.pack(pady=20)
            
            # 显示Mixed ColorPoint规格数据
            if mixed_colorpoint:
                # 添加CALF24标准坐标点解析状态提示
                status_label = ttk.Label(mixed_frame, text="Parsing CALF24 (Mixed) ColorPointSpec...", font=('SimHei', 10))
                status_label.pack(pady=5, padx=5, anchor='w')
                self.root.update_idletasks()
                
                # 尝试解析为u'v'坐标格式
                try:
                    # 清理输入数据，确保格式正确
                    cleaned_data = str(mixed_colorpoint).strip()
                    
                    # 使用增强的坐标点解析功能
                    colorpoint_items = self.parse_color_criteria(cleaned_data)
                    
                    # 验证解析结果
                    valid_points = []
                    for point in colorpoint_items:
                        if isinstance(point, tuple) and len(point) == 2:
                            u_val, v_val = point
                            # 过滤有效范围的坐标点 (0.0-1.0)
                            if isinstance(u_val, (int, float)) and isinstance(v_val, (int, float)):
                                if 0.0 <= u_val <= 1.0 and 0.0 <= v_val <= 1.0:
                                    valid_points.append((u_val, v_val))
                    
                    # 限制最多8个坐标点，确保预留8个坐标输入空位
                    if len(valid_points) > 8:
                        valid_points = valid_points[:8]
                    # 如果坐标点少于8个，补充空坐标点（使用None表示未获取的坐标）以确保显示8个输入框
                    while len(valid_points) < 8:
                        valid_points.append((None, None))
                    
                    status_label.config(text=f"Successfully parsed {len(valid_points)} CALF24 (Mixed) ColorPointSpec values")
                    
                    # 创建一个子框架，用于坐标显示
                    coord_frame = ttk.LabelFrame(mixed_frame, text="CALF24 Coordinate Points")
                    
                    # 显示坐标点并允许编辑
                    self._display_editable_colorpoint_data(coord_frame, valid_points, "Mixed", self.mixed_frame)
                    self.colorpoint_spec_data["Mixed"] = valid_points
                    
                    # 只显示坐标框架，取消多边形绘图区
                    coord_frame.pack(fill="both", expand=True, padx=5, pady=5)
                except Exception as e:
                    # 如果解析失败，显示原始数据并使用默认值
                    status_label.config(text="Failed to parse CALF24 ColorPointSpec, using default values", foreground="red")
                    self.logger.error(f"Error parsing Mixed ColorPoint: {str(e)}")
                    
                    # 解析失败时创建8个默认坐标点，确保显示完整的8个输入框
                    empty_points = [(None, None) for _ in range(8)]
                    
                    # 创建框架和显示内容
                    coord_frame = ttk.LabelFrame(mixed_frame, text="CALF24 Coordinate Points")
                    
                    # 显示空的坐标点列表
                    self._display_editable_colorpoint_data(coord_frame, empty_points, "Mixed", self.mixed_frame)
                    self.colorpoint_spec_data["Mixed"] = empty_points
                    
                    # 只显示坐标框架，取消多边形绘图区
                    coord_frame.pack(fill="both", expand=True, padx=5, pady=5)
                    
                    # 添加原始数据显示框
                    raw_frame = ttk.LabelFrame(mixed_frame, text="Original CALF24 Data (Parsing Failed)")
                    raw_frame.pack(fill="both", expand=True, padx=5, pady=5)
                    raw_data_label = tk.Label(raw_frame, text=f"{mixed_colorpoint}", font=('SimHei', 9), justify="left", wraplength=300)
                    raw_data_label.pack(padx=5, pady=5, fill="both", expand=True)
            else:
                no_data_label = tk.Label(mixed_frame, text="No CALF24 (Mixed) ColorPointSpec found in the file.", font=('SimHei', 10))
                no_data_label.pack(pady=20)
            
            # 确保选项卡已经切换（再次确认）
            if hasattr(self, 'tab_control') and hasattr(self, 'colorpoint_spec_tab'):
                self.tab_control.select(self.colorpoint_spec_tab)
                self.root.update_idletasks()
            
            # 更新状态栏信息
            self.update_status("ColorPoint spec review completed.")
            
            # 自动保存颜色坐标多边形规格到临时文件
            try:
                self.logger.info("正在自动保存用户修改的规格数据")
                saved_file = self._save_colorpoint_spec_to_temp_file()
                if saved_file:
                    self.logger.info(f"成功将规格数据保存到临时JSON文件: {saved_file}")
                    self.update_status(f"颜色坐标规格已自动保存到: {os.path.basename(saved_file)}")
                else:
                    self.logger.error("自动保存规格数据失败")
                    self.update_status("警告: 颜色坐标规格自动保存失败")
            except Exception as save_error:
                self.logger.error(f"自动保存规格数据时发生错误: {str(save_error)}")
                self.update_status(f"警告: 颜色坐标规格自动保存发生错误")
            
        except Exception as e:
            self.update_status(f"Error reading ColorPoint spec: {str(e)}")
            # 清除界面内容
            for widget in self.colorpoint_spec_tab.winfo_children():
                widget.destroy()
            # 显示错误信息
            error_label = tk.Label(self.colorpoint_spec_tab, text=f"Error: {str(e)}", font=('SimHei', 10), fg="red")
            error_label.pack(pady=20)
            self.logger.error(f"ColorPointSpec error: {str(e)}")
            error_label.pack(pady=20)
    
    def _display_editable_colorpoint_data(self, tab, colorpoint_items, type_name, parent_frame):
        """Display editable ColorPoint specifications data in specified tab, using the same style as Read Criteria tab"""
        try:
            # 清空选项卡内容
            for widget in tab.winfo_children():
                widget.destroy()
            
            # 创建容器框架，与Read Criteria选项卡保持一致的内边距
            container = tk.Frame(tab)
            container.pack(fill="both", expand=True, padx=15, pady=15)
            
            # 添加标题和按钮区域 - 同一行布局
            title_button_frame = tk.Frame(container)
            title_button_frame.pack(fill="x", pady=(0, 10))
            
            # 左侧标题框架
            title_frame = tk.Frame(title_button_frame)
            title_frame.pack(side="left", fill="x", expand=True)
            
            # 标题标签 - 左对齐，使用与Read Criteria相同的字体样式
            title_label = tk.Label(title_frame, text=f"{type_name} ColorPoint Specifications", 
                                 font=('SimHei', 12, 'bold'), anchor="w")
            title_label.pack(padx=5, pady=5, fill="x")
            
            # 右侧按钮框架
            button_container = tk.Frame(title_button_frame)
            button_container.pack(side="right", padx=5, pady=3)
            
            # 创建中文按钮标签 - 右对齐，使用HoverButton.TButton样式
            save_button = ttk.Button(button_container, text="Save Changes", style="HoverButton.TButton", command=lambda: save_changes())
            save_button.pack(side="right", padx=(5, 0))
            
            cancel_button = ttk.Button(button_container, text="Cancel Changes", style="HoverButton.TButton", command=lambda: cancel_changes())
            cancel_button.pack(side="right", padx=(0, 5))
            
            # 创建固定表头框架
            header_container = tk.Frame(container)
            header_container.pack(fill="x", pady=(0, 2))
            
            # 表头标签定义
            headers = [("Point", 80), ("u' Value", 120), ("v' Value", 120), ("Description", 300)]
            
            # 创建表头框架 - 固定在顶部
            header_frame = tk.Frame(header_container)
            header_frame.pack(fill="x")
            
            # 表头标签 - 使用与Read Criteria相同的样式和对齐
            for i, (text, width) in enumerate(headers):
                anchor = "center" if i < 3 else "w"
                label = tk.Label(header_frame, text=text, font=("SimHei", 10, "bold"), width=width//6, 
                                anchor=anchor, relief="flat", bd=0, padx=8, pady=4, bg="#f0f0f0")
                label.grid(row=0, column=i, sticky="nsew")
                
            # 配置列权重，使描述列可伸展
            header_frame.grid_columnconfigure(3, weight=1)
            
            # 创建响应式滚动框架
            scroll_container = tk.Frame(container)
            scroll_container.pack(fill="both", expand=True, pady=(0, 10))
            
            # 创建横向和纵向滚动条
            vscrollbar = ttk.Scrollbar(scroll_container, orient="vertical")
            hscrollbar = ttk.Scrollbar(scroll_container, orient="horizontal")
            vscrollbar.pack(side="right", fill="y")
            hscrollbar.pack(side="bottom", fill="x")
            
            # 创建canvas，同时绑定横向和纵向滚动条
            canvas = tk.Canvas(scroll_container, 
                             yscrollcommand=vscrollbar.set,
                             xscrollcommand=hscrollbar.set,
                             highlightthickness=0)
            canvas.pack(side="left", fill="both", expand=True)
            
            # 同步表头和内容的水平滚动
            def sync_horizontal_scroll(*args):
                # 当canvas水平滚动时，表头也同步移动
                x = -float(args[0])
                header_frame.place_configure(x=x)
            
            # 配置滚动条与canvas的联动
            vscrollbar.config(command=canvas.yview)
            hscrollbar.config(command=canvas.xview)
            
            # 绑定水平滚动条事件到表头同步函数
            canvas.xview_scroll = lambda *args: (tk.Canvas.xview_scroll(canvas, *args), sync_horizontal_scroll(canvas.xview()))
            
            # 创建可滚动的内容框架
            scrollable_frame = tk.Frame(canvas)
            
            # 创建窗口对象，将scrollable_frame嵌入canvas
            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            
            # 配置滚动区域更新事件
            def update_scrollregion(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
            
            # 绑定滚动区域更新事件
            scrollable_frame.bind("<Configure>", update_scrollregion)
            
            # 绑定鼠标滚轮事件
            def _on_mousewheel(event):
                # Windows鼠标滚轮事件
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                # 防止事件冒泡到其他窗口
                return "break"
            
            canvas.bind("<MouseWheel>", _on_mousewheel)
            scrollable_frame.bind("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1*(event.delta/120)), "units"))
            
            # 配置Canvas窗口跟随框架大小变化而调整宽度
            def _on_canvas_configure(event):
                canvas.itemconfig(canvas_window, width=event.width)
            
            canvas.bind("<Configure>", _on_canvas_configure)
            
            # 存储原始数据和Entry引用
            original_data = {}
            entry_data = {}
            
            # 移除了验证和格式化功能，现在允许用户输入任何值
            
            # 显示坐标点并创建可编辑的输入框
            if colorpoint_items:
                for i, point in enumerate(colorpoint_items):
                    # 创建行框架
                    row_frame = tk.Frame(scrollable_frame)
                    row_frame.pack(fill="x", padx=0, pady=1)
                    
                    # 配置行框架的列权重
                    row_frame.grid_columnconfigure(3, weight=1)
                    
                    # 点编号
                    tk.Label(row_frame, text=f"Point {i+1}", font=("SimHei", 10), width=headers[0][1]//6, 
                            anchor="center", padx=10, pady=4).grid(row=0, column=0, sticky="nsew")
                    
                    # u'坐标输入框 - 不再格式化显示，直接显示原始值
                    u_value = "" if point[0] is None else str(point[0])
                    u_entry = ttk.Entry(row_frame, font=("SimHei", 10), width=headers[1][1]//6, justify="center")
                    u_entry.insert(0, u_value)
                    u_entry.grid(row=0, column=1, sticky="nsew", padx=5, pady=3)
                    
                    # v'坐标输入框 - 不再格式化显示，直接显示原始值
                    v_value = "" if point[1] is None else str(point[1])
                    v_entry = ttk.Entry(row_frame, font=("SimHei", 10), width=headers[2][1]//6, justify="center")
                    v_entry.insert(0, v_value)
                    v_entry.grid(row=0, column=2, sticky="nsew", padx=5, pady=3)
                    
                    # 描述标签
                    description = f"{type_name} ColorPoint #{i+1} - CIE 1976 u'v' Chromaticity Coordinates"
                    desc_label = tk.Label(row_frame, text=description, font=("SimHei", 10), width=headers[3][1]//6, 
                                        anchor="w", padx=10, pady=4, wraplength=0, justify="left")
                    desc_label.grid(row=0, column=3, sticky="nsew")
                    
                    # 存储原始数据和Entry引用
                    original_data[i] = (point[0], point[1])
                    entry_data[i] = {
                        "u": u_entry,
                        "v": v_entry
                    }
            else:
                # 如果没有坐标点数据，显示提示信息
                no_data_frame = tk.Frame(container, padx=20, pady=20)
                no_data_frame.pack(fill="x", pady=10, padx=0)
                no_data_label = tk.Label(no_data_frame, 
                                     text="未找到坐标点数据或无法解析规格字符串。",
                                     font=("SimHei", 10))
                no_data_label.pack(anchor="center")
            
            # 移除了验证函数，不再对输入进行验证和格式化
            
            # 保存更改的函数
            def save_changes():
                try:
                    import logging
                    import os
                    import tempfile
                    logger = logging.getLogger("TestLogAnalyzer")
                    logger.info(f"开始保存{type_name}类型的ColorPoint规格参数更改")
                    
                    # 获取固定文件路径进行一致性检查
                    temp_dir = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer")
                    fixed_file_path = os.path.join(temp_dir, "ColorPointSpec_Current.json")
                    
                    # 一致性校验：在保存前先检查文件是否存在且与内存数据一致
                    if os.path.exists(fixed_file_path):
                        try:
                            with open(fixed_file_path, 'r', encoding='utf-8') as f:
                                file_data = json.load(f)
                            
                            # 如果文件数据与内存数据不一致，记录警告
                            if "data" in file_data and file_data["data"] != self.colorpoint_spec_data:
                                logger.warning("保存前检测到文件数据与内存数据不一致，可能有并发修改")
                                self.update_status("警告: 检测到潜在的数据不一致，正在保存当前修改")
                        except Exception as e:
                            logger.warning(f"保存前一致性检查失败: {str(e)}")
                    
                    # 处理当前类型的坐标点数据
                    updated_points = []
                    # 使用sorted确保按顺序处理坐标点
                    for i in sorted(entry_data.keys()):
                        u_entry = entry_data[i]["u"]
                        v_entry = entry_data[i]["v"]
                        
                        u_text = u_entry.get().strip()
                        v_text = v_entry.get().strip()
                        
                        # 尝试转换为浮点数，失败则保留原始字符串
                        if u_text and v_text:
                            try:
                                u_value = float(u_text)
                                v_value = float(v_text)
                                updated_points.append((u_value, v_value))
                            except ValueError:
                                # 如果无法转换为浮点数，仍然保存原始值
                                updated_points.append((u_text, v_text))
                        else:
                            # 对于未填写的坐标，添加None值标记
                            updated_points.append((None, None))
                    
                    # 过滤掉None值坐标点，只保留有效的坐标点
                    valid_points = [point for point in updated_points if point[0] is not None and point[1] is not None]
                    
                    # 确保数据格式与系统期望的一致
                    # 创建正确的数据结构，包含coordinates字段
                    formatted_data = {
                        'coordinates': valid_points
                    }
                    
                    # 更新全局数据
                    self.colorpoint_spec_data[type_name] = formatted_data
                    
                    # 确保整个colorpoint_spec_data包含所有必要的数据
                    # 如果另一种类型不存在，确保它有默认结构
                    other_type = "Mixed" if type_name == "White" else "White"
                    if other_type not in self.colorpoint_spec_data:
                        self.colorpoint_spec_data[other_type] = {'coordinates': []}
                    
                    # 保存到临时文件，完整更新所有数据
                    saved_file = self._save_colorpoint_spec_to_temp_file()
                    
                    # 保存后一致性验证
                    if saved_file and os.path.exists(saved_file):
                        logger.info(f"成功保存到固定文件路径: {saved_file}")
                        # 确保所有引用都指向固定文件
                        self.last_saved_colorpoint_file = fixed_file_path
                        if "White" in self.colorpoint_spec_data:
                            self._white_criteria_file = fixed_file_path
                        if "Mixed" in self.colorpoint_spec_data:
                            self._mixed_criteria_file = fixed_file_path
                    
                    # 刷新多边形显示
                    self._refresh_polygon_display(parent_frame, valid_points, type_name)
                    
                    # 提供详细的成功反馈
                    success_message = f"成功保存{type_name}的ColorPoint规格参数更改，共{len(valid_points)}个有效坐标点到标准文件"
                    if saved_file:
                        success_message += f"，已保存到文件: {os.path.basename(saved_file)}"
                    self.update_status(success_message)
                    logger.info(success_message)
                    logger.debug(f"完整的ColorPoint规格数据: {json.dumps(self.colorpoint_spec_data, ensure_ascii=False, default=str)}")
                    
                except Exception as e:
                    import traceback
                    error_msg = f"保存ColorPoint规格参数时发生错误: {str(e)}"
                    self.update_status(error_msg)
                    logger.error(error_msg)
                    logger.debug(traceback.format_exc())
            
            # 取消修改的函数
            def cancel_changes():
                """Cancel all unsaved changes and restore original settings"""
                # 恢复所有输入框的值为原始数据
                for i, (original_u, original_v) in original_data.items():
                    if i in entry_data:
                        u_entry = entry_data[i]["u"]
                        v_entry = entry_data[i]["v"]
                        
                        u_value = "" if original_u is None else f"{original_u:.6f}"
                        v_value = "" if original_v is None else f"{original_v:.6f}"
                        
                        u_entry.delete(0, tk.END)
                        u_entry.insert(0, u_value)
                        v_entry.delete(0, tk.END)
                        v_entry.insert(0, v_value)
                        
                        # 重置验证状态和样式
                        entry_data[i]["valid"] = True
                        u_entry.configure(style="TEntry")
                        v_entry.configure(style="TEntry")
                
                # 更新状态栏
                self.update_status(f"已取消{type_name}的ColorPoint规格参数更改")
            
            # 添加操作按钮区域 - 与Read Criteria选项卡保持一致
            button_frame = tk.Frame(container, bd=1, relief="flat", bg="#f5f5f5", pady=5)
            button_frame.pack(fill="x", pady=10)
            
            # 添加额外说明标签 - 放在底部，样式与Read Criteria一致
            info_label = tk.Label(container, text="说明: u'v'坐标范围为0.0-1.0，输入值将自动验证和格式化。", font=('SimHei', 9), fg="blue")
            info_label.pack(pady=2, padx=5, anchor='w')
            
            # 绑定快捷键
            container.bind_all("<Control-s>", lambda event: save_changes())
            container.bind_all("<Escape>", lambda event: cancel_changes())
            
            # 添加快捷键支持
            container.bind("<Control-s>", lambda event: save_changes())
            container.bind("<Control-S>", lambda event: save_changes())
            container.bind("<Escape>", lambda event: cancel_changes())
            # 设置焦点使快捷键生效
            container.focus_set()
            
        except Exception as e:
            error_label = tk.Label(tab, text=f"创建ColorPoint表格时发生错误: {str(e)}", font=('SimHei', 10), fg="red")
            error_label.pack(pady=10)
    
    def _validate_and_format_entry(self, idx, var, entry_data=None, type_name=None, parent_frame=None):
        # 验证并格式化输入值，确保数值格式正确
        try:
            # 检查输入是否为空
            input_value = var.get().strip()
            if not input_value:
                # 输入为空，不做处理，保持空白
                return
                
            # 输入不为空，进行正常的验证和格式化
            value = float(input_value)
            # 限制范围并格式化为6位小数
            clamped_value = max(0.0, min(1.0, value))
            var.set(f"{clamped_value:.6f}")
            
            # 如果提供了足够的参数，自动刷新多边形
            if entry_data and type_name and parent_frame:
                self._auto_refresh_polygon(entry_data, type_name, parent_frame)
        except ValueError:
            # 如果不是有效数字，不做处理，让用户继续编辑
            pass
            
    def _auto_refresh_polygon(self, entry_data, type_name, parent_frame):
        # 根据当前编辑的坐标值自动刷新多边形显示
        try:
            # 获取当前所有坐标点的值
            updated_points = []
            for i in sorted(entry_data.keys()):
                try:
                    u_var, v_var = entry_data[i]
                    u_value = float(u_var.get())
                    v_value = float(v_var.get())
                    updated_points.append((u_value, v_value))
                except (ValueError, TypeError):
                    # 如果有无效值，跳过本次刷新
                    return
                    
            # 不再调用刷新多边形方法，因为多边形绘图区已取消
            pass
        except Exception as e:
            # 自动刷新错误不要影响用户使用
            self.logger.debug(f"Auto-refresh error: {str(e)}")
            
    def _display_colorpoint_polygon(self, tab, colorpoint_items, type_name):
        # Display ColorPoint polygon in specified frame
        try:
            # 添加说明文本
            desc_text = f"基于CIELUV色彩空间的u'v'坐标系，显示{type_name}标准的通过区域"
            desc_label = tk.Label(tab, text=desc_text, font=('SimHei', 9), fg="gray")
            desc_label.pack(pady=2, padx=5, anchor='w')
            
            # 导入必要的库
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from matplotlib.patches import Polygon
            import numpy as np
            
            # 创建图形
            fig, ax = plt.subplots(figsize=(6, 5), dpi=100)
            
            # 定义函数：将u'v'值转换为RGB颜色
            def uv_to_rgb(u, v, luminance=100):
                # CIELUV到XYZ转换
                u0 = 0.19783000664283681  # D65的u'值
                v0 = 0.46831999493879114  # D65的v'值
                
                # 计算Y（亮度）
                Y = luminance / 100.0  # 转换为0-1范围
                
                # 计算u和v的实际值
                if v != 0:
                    d = Y * (5 * v - 12 * v0 * u) / (v * 3)
                else:
                    d = 0
                
                u_prime = u
                v_prime = v
                
                # 计算XYZ坐标
                if v_prime > 0:
                    Y = Y
                    X = Y * 9 * u_prime / (4 * v_prime)
                    Z = Y * (12 - 3 * u_prime - 20 * v_prime) / (4 * v_prime)
                else:
                    X, Y, Z = 0, Y, 0
                
                # XYZ到RGB转换（使用sRGB矩阵）
                r = X * 3.2406 + Y * -1.5372 + Z * -0.4986
                g = X * -0.9689 + Y * 1.8758 + Z * 0.0415
                b = X * 0.0557 + Y * -0.2040 + Z * 1.0570
                
                # 应用gamma校正
                def gamma_correction(x):
                    if x <= 0.0031308:
                        return 12.92 * x
                    else:
                        return 1.055 * (x ** (1/2.4)) - 0.055
                
                r = gamma_correction(r)
                g = gamma_correction(g)
                b = gamma_correction(b)
                
                # 确保RGB值在0-1范围内
                r = max(0, min(1, r))
                g = max(0, min(1, g))
                b = max(0, min(1, b))
                
                return r, g, b
            
            # 创建颜色背景
            def add_color_background(ax, xlim, ylim, resolution=100):
                # 创建网格
                u = np.linspace(xlim[0], xlim[1], resolution)
                v = np.linspace(ylim[0], ylim[1], resolution)
                u_grid, v_grid = np.meshgrid(u, v)
                
                # 创建RGB颜色数组
                rgb_array = np.zeros((resolution, resolution, 3))
                
                # 为每个点计算RGB颜色
                rgb_flat = np.array([uv_to_rgb(u, v) for u, v in zip(u_grid.flatten(), v_grid.flatten())])
                rgb_array = rgb_flat.reshape(resolution, resolution, 3)
                
                # 显示颜色背景
                ax.imshow(rgb_array, extent=[xlim[0], xlim[1], ylim[0], ylim[1]], 
                         origin='lower', aspect='auto', alpha=0.8)
            
            # 设置坐标轴标签
            ax.set_xlabel("u'")
            ax.set_ylabel("v'")
            
            # 如果有有效的坐标点，绘制多边形
            if len(colorpoint_items) >= 3:
                # 调整坐标轴范围以适应多边形点
                all_u_values = [p[0] for p in colorpoint_items]
                all_v_values = [p[1] for p in colorpoint_items]
                
                u_margin = (max(all_u_values) - min(all_u_values)) * 0.15 if max(all_u_values) != min(all_u_values) else 0.02
                v_margin = (max(all_v_values) - min(all_v_values)) * 0.15 if max(all_v_values) != min(all_v_values) else 0.02
                xlim = [min(all_u_values) - u_margin, max(all_u_values) + u_margin]
                ylim = [min(all_v_values) - v_margin, max(all_v_values) + v_margin]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax, xlim, ylim)
                
                # 设置坐标轴范围
                ax.set_xlim(xlim)
                ax.set_ylim(ylim)
                
                # 绘制多边形
                color = 'blue' if type_name == 'White' else 'red'
                polygon = Polygon(colorpoint_items, closed=True, fill=False, 
                                edgecolor=color, linewidth=2, label=f'{type_name} Region')
                ax.add_patch(polygon)
                
                # 添加顶点标签
                for i, (u, v) in enumerate(colorpoint_items):
                    ax.text(u, v, f'P{i+1}', fontsize=8, ha='right', va='bottom',
                           bbox=dict(boxstyle="round,pad=0.3", fc="white", alpha=0.7))
                
                # 添加图例
                ax.legend(loc='upper right')
            else:
                # 设置默认范围
                xlim = [0.18, 0.22]
                ylim = [0.48, 0.52]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax, xlim, ylim)
                
                # 设置坐标轴范围
                ax.set_xlim(xlim)
                ax.set_ylim(ylim)
                
                # 根据坐标点数量显示相应提示信息
                if len(colorpoint_items) == 0:
                    ax.text(0.5, 0.5, 'No valid polygon data to display', 
                           ha='center', va='center', transform=ax.transAxes,
                           fontsize=12, color='red', fontstyle='italic')
                elif len(colorpoint_items) < 3:
                    ax.text(0.5, 0.5, f'Need at least 3 points to draw polygon (current: {len(colorpoint_items)})', 
                           ha='center', va='center', transform=ax.transAxes,
                           fontsize=10, color='orange', fontstyle='italic')
                
                # 添加说明文本在底部
                if type_name == 'White':
                    criteria_name = 'CAFL0'
                else:
                    criteria_name = 'CALF24'
                
                note_text = f"{type_name} {criteria_name} Standard: "
                note_text += "No valid polygon" if len(colorpoint_items) == 0 else \
                             f"{len(colorpoint_items)} point{'s' if len(colorpoint_items) > 1 else ''} (need at least 3 for polygon)"
                
                ax.text(0.5, 0.01, note_text, 
                       ha='center', va='bottom', transform=ax.transAxes,
                       fontsize=8, color='gray')
            
            # 添加网格线
            ax.grid(True, linestyle='--', alpha=0.7)
            
            # 调整布局
            plt.tight_layout()
            
            # 创建画布并添加到Tkinter窗口
            canvas = FigureCanvasTkAgg(fig, master=tab)
            canvas.draw()
            # 添加边框效果
            canvas_frame = ttk.Frame(tab, relief="sunken")
            canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, in_=canvas_frame)
            
            # 保存画布引用，用于后续刷新
            if type_name == 'White':
                self.white_canvas = canvas
            else:
                self.mixed_canvas = canvas
                
        except Exception as e:
            error_label = tk.Label(tab, text=f"Error displaying ColorPoint polygon: {str(e)}", fg="red")
            error_label.pack(pady=10)
    
    def _refresh_polygon_display(self, parent_frame, updated_points, type_name):
        # 多边形绘图区已取消，此函数不再执行任何操作
        pass
    
    def _save_colorpoint_spec_to_temp_file(self):
        # Save ColorPoint specification data to a consistent file with validation
        try:
            import tempfile
            import os
            import datetime
            
            # 使用固定文件名而不是时间戳，确保所有操作都指向同一文件
            temp_dir = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer")
            
            # 确保目录存在
            if not os.path.exists(temp_dir):
                os.makedirs(temp_dir)
            
            # 创建固定文件路径
            temp_file = os.path.join(temp_dir, "ColorPointSpec_Current.json")
            
            # 一致性检查：如果文件已存在，先读取并验证是否与当前内存中的数据一致
            if hasattr(self, 'last_saved_colorpoint_file') and self.last_saved_colorpoint_file and os.path.exists(self.last_saved_colorpoint_file):
                # 如果之前保存的文件与当前要保存的文件不同，提供警告
                if self.last_saved_colorpoint_file != temp_file:
                    self.logger.warning(f"文件路径不一致: 之前保存的文件 '{self.last_saved_colorpoint_file}', 当前要保存到 '{temp_file}'")
                    self.update_status("警告: 检测到文件路径变更，可能导致数据不一致")
                
                # 读取现有文件进行一致性检查
                try:
                    with open(temp_file, 'r', encoding='utf-8') as f:
                        existing_data = json.load(f)
                    
                    # 比较数据结构是否一致
                    if "data" in existing_data and existing_data["data"] != self.colorpoint_spec_data:
                        # 记录数据不一致警告
                        self.logger.warning("检测到内存数据与文件数据不一致，将覆盖现有文件")
                except Exception as e:
                    self.logger.warning(f"一致性检查失败: {str(e)}")
            
            # 准备要保存的数据（添加元数据）
            save_data = {
                "metadata": {
                    "timestamp": datetime.datetime.now().isoformat(),
                    "version": "1.0",
                    "description": "TestLogAnalyzer ColorPoint Specification Data"
                },
                "data": self.colorpoint_spec_data
            }
            
            # 保存数据到临时文件
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, ensure_ascii=False, indent=2)
            
            # 保存最近的文件路径，便于用户访问
            self.last_saved_colorpoint_file = temp_file
            
            # 更新状态栏
            file_name = os.path.basename(temp_file)
            self.update_status(f"ColorPointSpec data saved: {file_name}")
            
            # 返回保存的文件路径
            return temp_file
            
        except Exception as e:
            error_msg = f"Error saving ColorPointSpec to temporary file: {str(e)}"
            self.update_status(error_msg)
            self.logger.error(error_msg)
            return None
            
    def _load_colorpoint_spec_from_temp_file(self, file_path=None):
        # Load ColorPoint specification data from a consistent file with enhanced validation
        try:
            import os
            
            # 使用固定的文件路径，确保与保存函数使用同一个文件
            import tempfile
            temp_dir = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer")
            fixed_file_path = os.path.join(temp_dir, "ColorPointSpec_Current.json")
            
            # 如果没有提供文件路径，优先使用固定文件名
            if file_path is None:
                # 检查固定文件是否存在
                if os.path.exists(fixed_file_path):
                    file_path = fixed_file_path
                # 其次使用最近保存的文件
                elif hasattr(self, 'last_saved_colorpoint_file') and self.last_saved_colorpoint_file and os.path.exists(self.last_saved_colorpoint_file):
                    file_path = self.last_saved_colorpoint_file
                    # 如果之前保存的文件与固定文件不同，提供警告
                    self.logger.warning(f"使用非标准路径: '{file_path}'，建议使用固定路径 '{fixed_file_path}'")
                    self.update_status("警告: 使用了非标准路径的ColorPoint规格文件")
                else:
                    # 尝试查找最新的ColorPointSpec临时文件（向后兼容）
                    import glob
                    if os.path.exists(temp_dir):
                        colorpoint_files = glob.glob(os.path.join(temp_dir, "ColorPointSpec_*.json"))
                        if colorpoint_files:
                            # 按修改时间排序，取最新的
                            colorpoint_files.sort(key=os.path.getmtime, reverse=True)
                            file_path = colorpoint_files[0]
                            # 如果找到的文件不是固定文件，提供警告
                            if file_path != fixed_file_path:
                                self.logger.warning(f"使用旧格式文件: '{file_path}'，建议迁移到固定路径 '{fixed_file_path}'")
                                self.update_status("警告: 使用了旧格式的ColorPoint规格文件")
                    
            # 验证文件是否存在
            if not file_path or not os.path.exists(file_path):
                error_msg = "错误: ColorPoint规格文件不存在或已被删除"
                self.update_status(error_msg)
                self.logger.error(error_msg)
                return False
            
            # 验证文件格式
            if not file_path.endswith(".json"):
                error_msg = "错误: ColorPoint规格文件格式不正确，需要JSON格式文件"
                self.update_status(error_msg)
                self.logger.error(error_msg)
                return False
            
            # 读取和解析文件
            with open(file_path, 'r', encoding='utf-8') as f:
                loaded_data = json.load(f)
            
            # 验证文件内容
            if "data" not in loaded_data or "metadata" not in loaded_data:
                error_msg = "错误: ColorPoint规格文件格式无效，缺少必要的数据结构"
                self.update_status(error_msg)
                self.logger.error(error_msg)
                return False
            
            # 更新类的colorpoint_spec_data属性
            self.colorpoint_spec_data = loaded_data["data"]
            
            # 更新状态栏
            file_name = os.path.basename(file_path)
            self.update_status(f"成功加载ColorPoint规格数据: {file_name}")
            self.logger.info(f"Loaded ColorPointSpec from {file_path}")
            
            # 一致性处理：如果加载的文件不是固定文件，自动更新引用到固定文件
            import tempfile
            temp_dir = os.path.join(tempfile.gettempdir(), "TestLogAnalyzer")
            fixed_file_path = os.path.join(temp_dir, "ColorPointSpec_Current.json")
            
            # 强制使用固定文件路径更新所有引用
            self.last_saved_colorpoint_file = fixed_file_path
            
            # 更新颜色标准文件引用，确保都指向同一个文件
            if "White" in self.colorpoint_spec_data:
                self._white_criteria_file = fixed_file_path
            if "Mixed" in self.colorpoint_spec_data:
                self._mixed_criteria_file = fixed_file_path
            
            # 如果加载的文件与固定文件不同，执行自动同步
            if file_path != fixed_file_path:
                self.logger.info(f"自动同步文件: 从 '{file_path}' 同步到 '{fixed_file_path}'")
                self.update_status("信息: 正在同步ColorPoint规格数据到标准文件")
                # 重新保存到固定文件以确保同步
                try:
                    self._save_colorpoint_spec_to_temp_file()
                except Exception as e:
                    self.logger.error(f"自动同步失败: {str(e)}")
                    self.update_status(f"警告: 自动同步失败: {str(e)}")
                
            return True
            
        except json.JSONDecodeError as e:
            error_msg = f"错误: ColorPoint规格文件解析失败，格式可能损坏: {str(e)}"
            self.update_status(error_msg)
            self.logger.error(error_msg)
            return False
        except Exception as e:
            error_msg = f"加载ColorPoint规格数据时发生错误: {str(e)}"
            self.update_status(error_msg)
            self.logger.error(error_msg)
            return False
            
    def _open_colorpoint_file_location(self):
        # 打开保存的临时文件所在目录
        try:
            import os
            import subprocess
            
            if hasattr(self, 'last_saved_colorpoint_file') and self.last_saved_colorpoint_file:
                file_dir = os.path.dirname(self.last_saved_colorpoint_file)
                # 根据操作系统打开文件浏览器
                if os.name == 'nt':  # Windows
                    subprocess.Popen(['explorer', file_dir])
                elif os.name == 'posix':  # macOS or Linux
                    if 'darwin' in os.sys.platform:  # macOS
                        subprocess.Popen(['open', file_dir])
                    else:  # Linux
                        subprocess.Popen(['xdg-open', file_dir])
                
                self.update_status(f"Opened directory containing saved ColorPointSpec files")
            else:
                self.update_status("No ColorPointSpec file has been saved yet")
                
        except Exception as e:
            error_msg = f"Error opening file location: {str(e)}"
            self.update_status(error_msg)
            self.logger.error(error_msg)

    def _display_colorpoint_data(self, tab, colorpoint_items, type_name):
        # Display ColorPoint specifications data in specified frame (legacy method)
        try:
            # 创建一个文本框来显示解析后的数据
            text_widget = tk.Text(tab, wrap="word", font=('SimHei', 10))
            text_widget.pack(fill="both", expand=True, padx=10, pady=10)
            
            # 添加滚动条
            scrollbar = ttk.Scrollbar(text_widget, command=text_widget.yview)
            scrollbar.pack(side="right", fill="y")
            text_widget.config(yscrollcommand=scrollbar.set)
            
            # 显示解析后的数据
            text_widget.insert("end", f"{type_name} ColorPoint Specification:\n\n")
            
            if isinstance(colorpoint_items, list):
                for i, point in enumerate(colorpoint_items):
                    text_widget.insert("end", f"Point {i+1}: u' = {point[0]}, v' = {point[1]}\n")
            elif isinstance(colorpoint_items, dict):
                for key, value in colorpoint_items.items():
                    text_widget.insert("end", f"{key}: {value}\n")
            else:
                text_widget.insert("end", str(colorpoint_items))
            
            # 使文本框只读
            text_widget.config(state="disabled")
            
        except Exception as e:
            error_label = tk.Label(tab, text=f"Error displaying data: {str(e)}", font=('SimHei', 10), fg="red")
            error_label.pack(pady=10)

    def _display_criteria_data(self, tab, criteria_string, type_name):
        """Display criteria data in the specified tab, support editing upper/lower parameters, and implement validation, save and cancel functions"""
        import re
        
        try:
            # 清空选项卡内容
            for widget in tab.winfo_children():
                widget.destroy()
                
            # 创建容器框架
            container = tk.Frame(tab)
            container.pack(fill="both", expand=True, padx=15, pady=15)
            
            # 解析标准字符串
            criteria_items = self._parse_criteria_string(criteria_string)
            
            # 实现MetricX按照数字大小升序排序，并保持相对位置区域不变
            sorted_criteria_items = []
            current_group = []
            
            for item in criteria_items:
                if len(item) >= 2:
                    std_type = item[0]
                    # 检查是否为MetricX格式
                    if std_type.startswith('Metric') and len(std_type) > 6:
                        # 尝试提取数字部分
                        try:
                            # 提取'Metric'后面的数字
                            metric_num = int(std_type[6:])
                            # 将项目添加到当前MetricX组
                            current_group.append((metric_num, item))
                        except ValueError:
                            # 如果不是有效的数字后缀，先处理当前组，然后单独添加此项
                            if current_group:
                                # 对当前组按数字升序排序并添加到结果
                                sorted_criteria_items.extend([item_tuple[1] for item_tuple in sorted(current_group)])
                                current_group = []
                            sorted_criteria_items.append(item)
                    else:
                        # 非MetricX项，先处理当前组，然后添加此项
                        if current_group:
                            # 对当前组按数字升序排序并添加到结果
                            sorted_criteria_items.extend([item_tuple[1] for item_tuple in sorted(current_group)])
                            current_group = []
                        sorted_criteria_items.append(item)
                else:
                    # 格式不正确的项，先处理当前组，然后添加此项
                    if current_group:
                        sorted_criteria_items.extend([item_tuple[1] for item_tuple in sorted(current_group)])
                        current_group = []
                    sorted_criteria_items.append(item)
            
            # 处理最后一组MetricX项
            if current_group:
                sorted_criteria_items.extend([item_tuple[1] for item_tuple in sorted(current_group)])
            
            # 定义标准类型的描述映射（更完整的描述信息）
            standard_descriptions = {
                "L": f"{type_name} 亮度 (Luminance) - cd/m²",
                "U": f"{type_name} 均匀度 (Uniformity) - %",
                "dY": f"{type_name} 亮度梯度 (dY) - %/cm",
                "u": f"{type_name} u' 色度坐标",
                "v": f"{type_name} v' 色度坐标",
                "Ru": f"{type_name} Ru - u坐标一致性",
                "Rv": f"{type_name} Rv - v坐标一致性",
                "Du": f"{type_name} Du - u坐标偏差",
                "Dv": f"{type_name} Dv - v坐标偏差",
                "dL*Min": f"{type_name} dL*最小值 - %/cm",
                "dL*Max": f"{type_name} dL*最大值 - %/cm",
                "dEMax": f"{type_name} dE最大值 - %/cm"
            }
            
            # 移除分类说明映射
            
            # 添加标题和按钮区域 - 同一行布局
            title_button_frame = tk.Frame(container)
            title_button_frame.pack(fill="x", pady=(0, 10))
            
            # 左侧标题框架
            title_frame = tk.Frame(title_button_frame)
            title_frame.pack(side="left", fill="x", expand=True)
            
            # 标题标签 - 左对齐
            title_label = tk.Label(title_frame, text=f"{type_name} Pass/Fail Criteria", 
                                 font=('SimHei', 12, 'bold'), anchor="w")
            title_label.pack(padx=5, pady=5, fill="x")
            
            # 右侧按钮框架
            button_container = tk.Frame(title_button_frame)
            button_container.pack(side="right", padx=5, pady=3)
            
            # 创建中文按钮标签 - 右对齐
            save_button = ttk.Button(button_container, text="Save Changes", style="HoverButton.TButton", command=lambda: save_changes())
            save_button.pack(side="right", padx=(5, 0))
            
            cancel_button = ttk.Button(button_container, text="Cancel Changes", style="HoverButton.TButton", command=lambda: cancel_changes())
            cancel_button.pack(side="right", padx=(0, 5))
            
            # 创建固定表头框架
            header_container = tk.Frame(container)
            header_container.pack(fill="x", pady=(0, 2))
            
            # 表头标签定义
            headers = [("No", 50), ("Test Item", 120), ("Lower Limit", 100), ("Upper Limit", 100), ("Description", 300)]
            
            # 创建表头框架 - 固定在顶部
            header_frame = tk.Frame(header_container)
            header_frame.pack(fill="x")
            
            # 表头标签 - 改进样式和对齐
            for i, (text, width) in enumerate(headers):
                anchor = "center" if i < 4 else "w"
                label = tk.Label(header_frame, text=text, font=("SimHei", 10, "bold"), width=width//6, 
                                anchor=anchor, relief="flat", bd=0, padx=8, pady=4, bg="#f0f0f0")
                label.grid(row=0, column=i, sticky="nsew")
                
            # 配置列权重，使描述列可伸展
            header_frame.grid_columnconfigure(4, weight=1)
            
            # 创建响应式滚动框架
            # 使用框架来容纳canvas和scrollbar，实现更好的布局控制
            scroll_container = tk.Frame(container)
            scroll_container.pack(fill="both", expand=True, pady=(0, 10))
            
            # 创建横向和纵向滚动条
            vscrollbar = ttk.Scrollbar(scroll_container, orient="vertical")
            hscrollbar = ttk.Scrollbar(scroll_container, orient="horizontal")
            vscrollbar.pack(side="right", fill="y")
            hscrollbar.pack(side="bottom", fill="x")
            
            # 创建canvas，同时绑定横向和纵向滚动条
            canvas = tk.Canvas(scroll_container, 
                             yscrollcommand=vscrollbar.set,
                             xscrollcommand=hscrollbar.set,
                             highlightthickness=0)
            canvas.pack(side="left", fill="both", expand=True)
            
            # 同步表头和内容的水平滚动
            def sync_horizontal_scroll(*args):
                # 当canvas水平滚动时，表头也同步移动
                x = -float(args[0])
                header_frame.place_configure(x=x)
            
            # 配置滚动条与canvas的联动
            vscrollbar.config(command=canvas.yview)
            hscrollbar.config(command=canvas.xview)
            
            # 绑定水平滚动条事件到表头同步函数
            canvas.xview_scroll = lambda *args: (tk.Canvas.xview_scroll(canvas, *args), sync_horizontal_scroll(canvas.xview()))
            
            # 创建可滚动的内容框架
            scrollable_frame = tk.Frame(canvas)
            
            # 创建窗口对象，将scrollable_frame嵌入canvas
            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            
            # 配置滚动区域更新事件
            def update_scrollregion(event):
                canvas.configure(scrollregion=canvas.bbox("all"))
            
            # 绑定滚动区域更新事件
            scrollable_frame.bind("<Configure>", update_scrollregion)
            
            # 绑定鼠标滚轮事件，支持直接在canvas上滚动
            def _on_mousewheel(event):
                # Windows鼠标滚轮事件
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                # 防止事件冒泡到其他窗口
                return "break"
            
            # 使用bind而不是bind_all，确保在当前canvas上响应
            canvas.bind("<MouseWheel>", _on_mousewheel)
            # 为scrollable_frame也添加鼠标滚轮绑定，确保在任何子元素上都能滚动
            scrollable_frame.bind("<MouseWheel>", lambda event: canvas.yview_scroll(int(-1*(event.delta/120)), "units"))
            
            # 配置Canvas窗口跟随框架大小变化而调整宽度
            def _on_canvas_configure(event):
                # 当canvas大小改变时，调整内部窗口宽度
                canvas.itemconfig(canvas_window, width=event.width)
            
            canvas.bind("<Configure>", _on_canvas_configure)
            
            # 布局已经在响应式滚动框架部分完成
            
            # 创建数据存储结构
            # 保存原始数据
            original_data = {}
            # 保存Entry引用和验证状态
            entry_data = {}
            
            # 表头已移至滚动区域外部，保持固定显示
            
            # 直接按原始顺序显示所有标准项，不分组
            row_idx = 1  # 全局行号计数器
            
            # 遍历所有标准项
            for item in sorted_criteria_items:
                # 处理可能有3个值的元组（std_type, std_value, number）或2个值的元组
                if len(item) == 3:
                    std_type, std_value, _ = item  # 忽略排序数字
                else:
                    std_type, std_value = item
                    # 解析标准值中的|分隔符，将其分为下限和上限
                    lower_limit = ""
                    upper_limit = ""
                    
                    if '|' in std_value:
                        # | 是规格上限和规格下限的分隔符
                        limits = std_value.split('|', 1)
                        lower_limit = limits[0].strip()
                        upper_limit = limits[1].strip() if len(limits) > 1 else ""
                    else:
                        # 如果没有|分隔符，将整个值作为单个值显示在下限列
                        lower_limit = std_value
                    
                    # 修正规格值：uniformity规格上限修正为100，规格下限为-1的修正为0
                    if std_type == "U" and upper_limit == "0":
                        upper_limit = "100"
                    if lower_limit == "-1":
                        lower_limit = "0"
                    
                    # 描述 - 增强描述信息
                    description = standard_descriptions.get(std_type, f"{type_name} {std_type}")
                    
                    # 移除类别额外说明
                    
                    # 保存原始数据
                    original_data[std_type] = (lower_limit, upper_limit)
                    
                    # 自动保存原始数据到self.criteria_data字典，确保自动保存时有数据
                    if not hasattr(self, 'criteria_data'):
                        self.criteria_data = {}
                    self.criteria_data[type_name] = original_data.copy()
                    print(f"已自动保存{type_name}标准数据到criteria_data字典")
                    
                    # 创建行框架
                    row_frame = tk.Frame(scrollable_frame)
                    row_frame.pack(fill="x", padx=0, pady=1)
                    
                    # 配置行框架的列权重
                    row_frame.grid_columnconfigure(4, weight=1)
                    
                    # 添加行数据 - 优化内容显示和添加工具提示
                    # 序号标签
                    tk.Label(row_frame, text=str(row_idx), font=("SimHei", 10), width=headers[0][1]//6, 
                            anchor="center", padx=10, pady=4).grid(row=0, column=0, sticky="nsew")
                    
                    # 测试类型标签 - 添加工具提示
                    type_label = tk.Label(row_frame, text=std_type, font=("SimHei", 10), width=headers[1][1]//6, 
                                        anchor="center", padx=10, pady=4)
                    type_label.grid(row=0, column=1, sticky="nsew")
                    
                    # 创建下限输入框 - 优化配置和显示
                    lower_entry = ttk.Entry(row_frame, font=("SimHei", 10), width=headers[2][1]//6, justify="center")
                    lower_entry.insert(0, lower_limit)
                    lower_entry.grid(row=0, column=2, sticky="nsew", padx=5, pady=3)
                    
                    # 添加输入验证绑定
                    lower_entry.bind("<FocusOut>", lambda event, std_type=std_type, which="lower": validate_input(event, std_type, which))
                    lower_entry.bind("<KeyRelease>", lambda event, std_type=std_type, which="lower": reset_status_colors(std_type, which))
                    
                    # 创建上限输入框 - 优化配置和显示
                    upper_entry = ttk.Entry(row_frame, font=("SimHei", 10), width=headers[3][1]//6, justify="center")
                    upper_entry.insert(0, upper_limit)
                    upper_entry.grid(row=0, column=3, sticky="nsew", padx=5, pady=3)
                    
                    # 添加输入验证绑定
                    upper_entry.bind("<FocusOut>", lambda event, std_type=std_type, which="upper": validate_input(event, std_type, which))
                    upper_entry.bind("<KeyRelease>", lambda event, std_type=std_type, which="upper": reset_status_colors(std_type, which))
                    
                    # 描述标签 - 优化显示确保内容完整
                    desc_label = tk.Label(row_frame, text=description, font=("SimHei", 10), width=headers[4][1]//6, 
                                        anchor="w", padx=10, pady=4, wraplength=0, justify="left")
                    desc_label.grid(row=0, column=4, sticky="nsew")
                    
                    # 移除所有提示说明
                    
                    # 保存Entry引用
                    entry_data[std_type] = {
                        "lower": lower_entry,
                        "upper": upper_entry,
                        "valid": True
                    }
                    
                    row_idx += 1
            
            # 初始化空的grouped_items变量以避免未定义错误
            grouped_items = []
            
            # 如果没有标准项，显示提示信息 - 简化样式
            if not grouped_items and not sorted_criteria_items:
                no_data_frame = tk.Frame(container, padx=20, pady=20)
                no_data_frame.pack(fill="x", pady=10, padx=0)
                no_data_label = tk.Label(no_data_frame, 
                                     text="未找到标准项或无法解析标准字符串。",
                                     font=("SimHei", 10))
                no_data_label.pack(anchor="center")
            
            # 定义验证函数
            def validate_input(event, std_type, which):
                pass
            
            def reset_status_colors(std_type, which):
                pass
            
            def save_changes():
                pass
            
            def cancel_changes():
                # 简化函数实现，移除对status_var的引用
                pass
            def is_valid_number(value):
                """验证输入值是否为有效的数字格式"""
                if not value.strip():
                    return True  # 允许空值
                try:
                    float(value)
                    return True
                except ValueError:
                    return False
            
            def reset_status_colors():
                """移除状态标签后的空函数"""
                pass
            
            def validate_input(entry, std_type, is_lower=True):
                """验证单个输入框的值"""
                # math模块已在文件顶部导入
                value = entry.get().strip()
                
                # 重置样式
                entry.configure(style="TEntry")
                
                # 验证数字格式
                if not is_valid_number(value):
                    entry.configure(style="Invalid.TEntry")
                    entry_data[std_type]["valid"] = False
                    return False
                
                # 如果两个输入框都有值，则验证上下限关系
                lower_val = entry_data[std_type]["lower"].get().strip()
                upper_val = entry_data[std_type]["upper"].get().strip()
                
                if lower_val and upper_val:
                    try:
                        lower_float = float(lower_val)
                        upper_float = float(upper_val)
                        
                        # 使用math.isclose处理浮点数精度问题，只有当下限值明显大于上限值时才判定为无效
                        if not math.isclose(lower_float, upper_float) and lower_float >= upper_float:
                            entry_data[std_type]["valid"] = False
                            
                            # 高亮两个输入框
                            entry_data[std_type]["lower"].configure(style="Invalid.TEntry")
                            entry_data[std_type]["upper"].configure(style="Invalid.TEntry")
                            return False
                    except ValueError:
                        # 如果转换失败，已经被前面的验证捕获
                        pass
                
                # 如果通过验证，检查是否所有项都有效
                all_valid = True
                for item_data in entry_data.values():
                    if not item_data["valid"]:
                        all_valid = False
                        break
                
                if all_valid:
                    entry_data[std_type]["valid"] = True
                
                return True
            
            # 为输入框添加验证
            for std_type, data in entry_data.items():
                lower_entry = data["lower"]
                upper_entry = data["upper"]
                
                # 绑定输入变化事件
                lower_entry.bind("<FocusOut>", lambda event, st=std_type, low=True: validate_input(event.widget, st, low))
                upper_entry.bind("<FocusOut>", lambda event, st=std_type, low=False: validate_input(event.widget, st, low))
                
                # 绑定按键释放事件进行实时验证
                lower_entry.bind("<KeyRelease>", lambda event, st=std_type, low=True: validate_input(event.widget, st, low))
                upper_entry.bind("<KeyRelease>", lambda event, st=std_type, low=False: validate_input(event.widget, st, low))
            
            # 配置无效输入的样式
            style = ttk.Style()
            style.configure("Invalid.TEntry", foreground="red", fieldbackground="#ffe6e6")
            
            # 添加操作按钮区域 - 增强视觉设计和交互体验
            button_frame = tk.Frame(container, bd=1, relief="flat", bg="#f5f5f5", pady=5)
            button_frame.pack(fill="x", pady=10)
            
            # 移除操作提示标签
            
            # 保存修改的函数
            def save_changes():
                """保存所有修改后的上下限设置"""
                try:
                    import logging
                    logger = logging.getLogger("TestLogAnalyzer")
                    logger.info(f"开始保存{type_name}类型的规格参数更改")
                    
                    # 检查是否有任何无效输入
                    for std_type, data in entry_data.items():
                        if not data["valid"]:
                            error_msg = f"错误: 请在保存前修复所有无效输入"
                            self.update_status(error_msg)
                            logger.error(error_msg)
                            return
                    
                    # 检查是否有任何输入框的值为无效数字
                    for std_type, data in entry_data.items():
                        lower_val = data["lower"].get().strip()
                        upper_val = data["upper"].get().strip()
                        
                        if lower_val and not is_valid_number(lower_val):
                            error_msg = f"错误: {std_type}下限值的数字格式无效"
                            self.update_status(error_msg)
                            data["lower"].configure(style="Invalid.TEntry")
                            logger.error(error_msg)
                            return
                        
                        if upper_val and not is_valid_number(upper_val):
                            error_msg = f"错误: {std_type}上限值的数字格式无效"
                            self.update_status(error_msg)
                            data["upper"].configure(style="Invalid.TEntry")
                            logger.error(error_msg)
                            return
                        
                        # 再次验证上下限关系
                        if lower_val and upper_val:
                            try:
                                # math模块已在文件顶部导入
                                lower_float = float(lower_val)
                                upper_float = float(upper_val)
                                
                                # 使用math.isclose处理浮点数精度问题，只有当下限值明显大于上限值时才判定为无效
                                if not math.isclose(lower_float, upper_float) and lower_float >= upper_float:
                                    error_msg = f"错误: {std_type}的下限值必须小于上限值"
                                    self.update_status(error_msg)
                                    data["lower"].configure(style="Invalid.TEntry")
                                    data["upper"].configure(style="Invalid.TEntry")
                                    logger.error(error_msg)
                                    return
                            except ValueError:
                                pass
                    
                    # 更新原始数据
                    for std_type, data in entry_data.items():
                        lower_val = data["lower"].get().strip()
                        upper_val = data["upper"].get().strip()
                        original_data[std_type] = (lower_val, upper_val)
                    
                    # 确保criteria_data字典存在并更新
                    if not hasattr(self, 'criteria_data'):
                        self.criteria_data = {}
                        logger.info("创建新的criteria_data字典")
                    
                    # 更新全局标准数据字典
                    self.criteria_data[type_name] = original_data.copy()  # 使用copy确保数据完整性
                    logger.info(f"已更新criteria_data[{type_name}]，包含{len(original_data)}项规格参数")
                    
                    # 刷新并保存临时规格文件
                    logger.info("开始调用_save_criteria_to_temp_file保存规格数据")
                    result = self._save_criteria_to_temp_file()
                    
                    if result is not None:
                        # 重置输入框样式
                        for data in entry_data.values():
                            data["lower"].configure(style="TEntry")
                            data["upper"].configure(style="TEntry")
                        
                        # 更新状态栏
                        success_message = f"成功保存{type_name}的规格参数更改"
                        self.update_status(success_message)
                        logger.info(success_message)
                    else:
                        error_msg = f"保存{type_name}规格参数到临时文件失败"
                        self.update_status(error_msg)
                        logger.error(error_msg)
                        
                except Exception as e:
                    import traceback
                    error_msg = f"保存规格参数时发生错误: {str(e)}"
                    self.update_status(error_msg)
                    logger.error(error_msg)
                    logger.debug(traceback.format_exc())
                
                # 移除对status_var的引用
            
            # 取消修改的函数
            def cancel_changes():
                """Cancel all unsaved changes and restore original settings"""
                # 恢复所有输入框的值为原始数据
                for std_type, (lower, upper) in original_data.items():
                    if std_type in entry_data:
                        entry_data[std_type]["lower"].delete(0, tk.END)
                        entry_data[std_type]["lower"].insert(0, lower)
                        entry_data[std_type]["upper"].delete(0, tk.END)
                        entry_data[std_type]["upper"].insert(0, upper)
                        
                        # 重置验证状态和样式
                        entry_data[std_type]["valid"] = True
                        entry_data[std_type]["lower"].configure(style="TEntry")
                        entry_data[std_type]["upper"].configure(style="TEntry")
                
                # 更新状态栏
                self.update_status(f"Cancelled criteria changes for {type_name}")
                
                # 移除对status_var的引用
            
            # 绑定快捷键
            container.bind_all("<Control-s>", lambda event: save_changes())
            container.bind_all("<Escape>", lambda event: cancel_changes())
            
            # 添加快捷键支持
            # 绑定Ctrl+S保存
            container.bind("<Control-s>", lambda event: save_changes())
            container.bind("<Control-S>", lambda event: save_changes())
            # 绑定Esc取消
            container.bind("<Escape>", lambda event: cancel_changes())
            # 设置焦点使快捷键生效
            container.focus_set()
            
        except Exception as e:
            error_label = tk.Label(tab, text=f"Error creating criteria table: {str(e)}", font=('SimHei', 10), fg="red")
            error_label.pack(pady=10)
    
    def _parse_criteria_string(self, criteria_string):
        """Parse criteria string, extract test standards, ignore ColorGamut standards, and sort Metric+number format items"""
        import re
        
        criteria_items = []
        
        # 查找所有<...>格式的标准项
        criteria_matches = re.findall(r'<([^<>]+)>', criteria_string)
        
        for match in criteria_matches:
            # 检查是否使用/分隔不同标准类型
            if '/' in match:
                # 使用/分割多个标准项
                sub_items = match.split('/')
                for sub_item in sub_items:
                    # 检查是否使用|分隔标准类型和值
                    if '|' in sub_item:
                        parts = sub_item.split('|', 1)
                        if len(parts) == 2:
                            std_type = parts[0].strip()
                            # 忽略ColorGamut标准
                            if std_type.lower() != 'colorgamut':
                                std_value = parts[1].strip()
                                criteria_items.append((std_type, std_value))
                    # 也支持传统的冒号分隔格式
                    elif ':' in sub_item:
                        parts = sub_item.split(':', 1)
                        if len(parts) == 2:
                            std_type = parts[0].strip()
                            # 忽略ColorGamut标准
                            if std_type.lower() != 'colorgamut':
                                std_value = parts[1].strip()
                                criteria_items.append((std_type, std_value))
                    else:
                        # 如果无法解析，添加原始子项内容
                        sub_item_stripped = sub_item.strip()
                        if sub_item_stripped.lower() != 'colorgamut':
                            criteria_items.append((sub_item_stripped, ""))
            # 如果没有/分隔符，尝试直接解析整个匹配内容
            else:
                # 优先检查|分隔格式
                if '|' in match:
                    parts = match.split('|', 1)
                    if len(parts) == 2:
                        std_type = parts[0].strip()
                        # 忽略ColorGamut标准
                        if std_type.lower() != 'colorgamut':
                            std_value = parts[1].strip()
                            criteria_items.append((std_type, std_value))
                # 也支持传统的冒号分隔格式
                elif ':' in match:
                    parts = match.split(':', 1)
                    if len(parts) == 2:
                        std_type = parts[0].strip()
                        # 忽略ColorGamut标准
                        if std_type.lower() != 'colorgamut':
                            std_value = parts[1].strip()
                            criteria_items.append((std_type, std_value))
                else:
                    # 如果无法解析，添加原始匹配内容
                    match_stripped = match.strip()
                    if match_stripped.lower() != 'colorgamut':
                        criteria_items.append((match_stripped, ""))
        
        # 如果没有找到<...>格式的标准项，尝试其他格式
        if not criteria_items:
            # 检查是否包含多个标准项，可能用分号分隔
            if ';' in criteria_string:
                items = criteria_string.split(';')
                for item in items:
                    # 同样支持|和:分隔格式
                    if '|' in item:
                        parts = item.split('|', 1)
                        if len(parts) == 2:
                            std_type = parts[0].strip()
                            # 忽略ColorGamut标准
                            if std_type.lower() != 'colorgamut':
                                std_value = parts[1].strip()
                                criteria_items.append((std_type, std_value))
                    elif ':' in item:
                        parts = item.split(':', 1)
                        if len(parts) == 2:
                            std_type = parts[0].strip()
                            # 忽略ColorGamut标准
                            if std_type.lower() != 'colorgamut':
                                std_value = parts[1].strip()
                                criteria_items.append((std_type, std_value))
        
        # 对Metric+数字格式的项目进行排序
        # 1. 将项目分为Metric+数字和其他两类
        metric_items = []  # 存储Metric+数字格式的项目
        non_metric_items = []  # 存储其他项目
        metric_pattern = re.compile(r'^Metric(\d+)$', re.IGNORECASE)
        
        for item in criteria_items:
            # 处理可能有3个值的元组（std_type, std_value, number）或2个值的元组
            if len(item) == 3:
                std_type, std_value, _ = item  # 忽略可能已有的排序数字
            else:
                std_type, std_value = item
                
            match = metric_pattern.match(std_type)
            if match:
                # 提取数字部分作为排序键
                number = int(match.group(1))
                metric_items.append((std_type, std_value, number))
            else:
                non_metric_items.append((std_type, std_value))
        
        # 2. 对Metric+数字格式的项目按数字从小到大排序
        metric_items.sort(key=lambda x: x[2])
        
        # 3. 合并结果，保持非Metric项的相对顺序，Metric项按排序后顺序插入
        # 这里我们不能简单地合并，需要保持原始的整体位置关系
        # 更好的方法是在_display_criteria_data方法中处理排序
        
        # 为了简单起见，我们先返回原始顺序，排序将在_display_criteria_data中进行
        
        return criteria_items
            
    # create_distribution_chart方法已移除，因为Data Distribution功能已被移除
        
    def merge_data(self):
        """合并数据"""
        pass
        
    def remove_duplicates(self):
        """去重数据"""
        pass
        
    def remove_bad_rows(self):
        """去除坏行"""
        pass
        
    def process_all(self):
        """执行所有处理"""
        pass
        
    def calculate_yield(self):
        """计算不良率"""
        pass
        
    # data_distribution_analysis方法已移除，因为Data Distribution功能已被移除
        
    def generate_report(self):
        """生成统计报告"""
        pass
        
    def _process_data_for_csv(self, selected_files):
        '''Process data for CSV export'''
        import tempfile
        import os
        import pandas as pd
        
        processed_files = []
        
        for i, file_path in enumerate(selected_files):
            file_name = os.path.basename(file_path)
            self.update_status(f"处理文件: {file_name}")
            
            # 1. 解析文件名提取配置信息
            if "MP" in file_name:
                extracted_text = "MP"
            elif "PVT" in file_name:
                extracted_text = "PVT"
            else:
                # 提取第二和第三空格之间的文本
                parts = file_name.split(' ')
                if len(parts) >= 4:
                    extracted_text = parts[2]
                else:
                    extracted_text = "Unknown"
                
            # 2. 智能检测CSV文件中的数据标题行
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = []
                    for i in range(20):
                        line = f.readline()
                        if not line:
                            break
                        lines.append(line.strip())
                    
                    # 识别元数据行
                    metadata_lines = []
                    non_metadata_lines = []
                    
                    for i, line in enumerate(lines):
                        is_metadata = False
                        if ',:' in line or (':' in line and ',' in line):
                            is_metadata = True
                        elif line.strip() and i < 10 and not (line.startswith('Model') or line.startswith('Serial Number')):
                            content_parts = [p.strip() for p in line.split(',') if p.strip()]
                            if content_parts and ':' in content_parts[0]:
                                is_metadata = True
                        
                        if is_metadata:
                            metadata_lines.append(i)
                        else:
                            non_metadata_lines.append((i, line))
                    
                    # 寻找标题行
                    max_commas = -1
                    header_line_candidates = []
                    header_line_index = 0
                    
                    non_empty_non_metadata = [(i, line) for i, line in non_metadata_lines if line.strip()]
                    
                    if non_empty_non_metadata:
                        max_commas = max(line.count(',') for i, line in non_empty_non_metadata)
                        header_line_candidates = [(i, line) for i, line in non_empty_non_metadata if line.count(',') == max_commas]
                        
                        if header_line_candidates:
                            best_score = -1
                            for i, line in header_line_candidates:
                                score = 1
                                if any(c.isalpha() for c in line):
                                    score += 10
                                content_length = len(line.replace(',', ''))
                                score += min(content_length, 50)
                                
                                if score > best_score:
                                    best_score = score
                                    header_line_index = i
                    elif non_metadata_lines:
                        for i, line in non_metadata_lines:
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    else:
                        for i, line in enumerate(lines):
                            comma_count = line.count(',')
                            if comma_count > max_commas:
                                max_commas = comma_count
                                header_line_index = i
                    
                    # 确保标题行不为空
                    if header_line_index < len(lines) and lines[header_line_index].strip() == '':
                        for i in range(header_line_index + 1, min(len(lines), header_line_index + 10)):
                            if lines[i].strip() != '':
                                header_line_index = i
                                max_commas = lines[i].count(',')
                                break
                    
                    # 读取数据
                    if max_commas >= 9:
                        df = pd.read_csv(file_path,
                                        skiprows=header_line_index,
                                        on_bad_lines='skip',
                                        engine='python')
                    else:
                        df = pd.read_csv(file_path,
                                        on_bad_lines='skip',
                                        engine='python')
                    
                    # 3. 添加Config列（放入第二列）
                    if not df.empty:
                        if len(df.columns) >= 1:
                            if len(df.columns) >= 2:
                                df.iloc[:, 1] = extracted_text
                            else:
                                df['Config'] = extracted_text
                            if len(df.columns) >= 2:
                                df.columns.values[1] = 'Config'
                    
                    # 4. 过滤异常行（通过Serial Number列文本长度过滤）
                    original_rows = len(df)
                    if 'Serial Number' in df.columns:
                        serial_lengths = df['Serial Number'].astype(str).str.len()
                        median_length = serial_lengths.median()
                        
                        df = df[abs(serial_lengths - median_length) < 5]
                        filtered_rows = original_rows - len(df)
                        if filtered_rows > 0:
                            self.update_status(f"Remove {filtered_rows} rows with Serial Number length difference > 5 from median ({median_length})。")
                    
                    processed_files.append((file_path, file_name, df))
                    
            except Exception as e:
                self.update_status(f"File process {file_name} wrong: {str(e)}")
                continue
        
        # 5. 数据合并后处理
        if len(processed_files) >= 2:
            self.update_status(f"Combined {len(processed_files)} files...")
            
            # 合并所有文件
            combined_df = pd.concat([df for _, _, df in processed_files], ignore_index=True)
            
            # 6. 智能去重
            if 'Serial Number' in combined_df.columns:
                original_dup_rows = len(combined_df)
                if 'Pass/Fail' in combined_df.columns:
                    combined_df['_sort_key'] = combined_df['Pass/Fail'].apply(lambda x: 0 if str(x).strip().upper() == 'PASS' else 1)
                    combined_df = combined_df.sort_values(by=['Serial Number', '_sort_key'])
                    combined_df = combined_df.drop_duplicates(subset=['Serial Number'], keep='first')
                    combined_df = combined_df.drop(columns=['_sort_key'])
                else:
                    combined_df = combined_df.drop_duplicates(subset=['Serial Number'], keep='first')
                
                deduped_rows = original_dup_rows - len(combined_df)
                if deduped_rows > 0:
                    self.update_status(f"Remove {deduped_rows} rows with duplicate Serial Numbers。")
            
            return combined_df
        elif len(processed_files) == 1:
            _, file_name, df = processed_files[0]
            
            # 对单个文件也进行去重
            if 'Serial Number' in df.columns:
                original_dup_rows = len(df)
                if 'Pass/Fail' in df.columns:
                    df['_sort_key'] = df['Pass/Fail'].apply(lambda x: 0 if str(x).strip().upper() == 'PASS' else 1)
                    df = df.sort_values(by=['Serial Number', '_sort_key'])
                    df = df.drop_duplicates(subset=['Serial Number'], keep='first')
                    df = df.drop(columns=['_sort_key'])
                else:
                    df = df.drop_duplicates(subset=['Serial Number'], keep='first')
                
                deduped_rows = original_dup_rows - len(df)
                if deduped_rows > 0:
                    self.update_status(f"Remove {deduped_rows} rows with duplicate Serial Numbers from {file_name}。")
            
            return df
        else:
            return None
    
    def _open_directory(self, file_path):
        """Open the directory containing the file"""
        import subprocess
        import sys
        import traceback
        
        try:
            directory = os.path.abspath(os.path.dirname(file_path))
            self.update_status(f"File path saved successfully: {os.path.abspath(file_path)}")
            self.update_status(f"Save directory path: {directory}")
            
            if directory:
                if os.name == 'nt':
                    subprocess.Popen(f'explorer "{directory}"', shell=True)
                else:
                    open_cmd = 'open' if sys.platform == 'darwin' else 'xdg-open'
                    subprocess.Popen([open_cmd, directory])
                self.update_status(f"Open save directory successfully: {directory}")
            else:
                self.update_status("Error: Could not determine a valid save directory path")
        except Exception as open_dir_error:
            self.update_status(f"Error: Failed to automatically open save directory: {str(open_dir_error)}")
            self.update_status(f"Full error details: {traceback.format_exc()}")
    
    def save_processed_data(self):
        """保存处理后数据为CSV格式 - 使用用户要求的数据处理逻辑"""
        import os
        import datetime
        import traceback
        import pandas as pd
        from tkinter import filedialog
        from tkinter import messagebox
        
        try:
            # 记录日志
            self.update_status("Start save processed data to CSV format...")
            
            # 检查是否有选中的文件
            if not hasattr(self, 'selected_files') or not self.selected_files:
                messagebox.showerror("Error", "Select file to process!")
                self.update_status("Error: No files selected for processing!")
                return
            
            # 使用自定义的数据处理逻辑处理文件
            self.update_status("Start process data...")
            processed_data = self._process_data_for_csv(self.selected_files)
            
            if processed_data is None or processed_data.empty:
                messagebox.showerror("Error", "No data was successfully processed!")
                self.update_status("Error: No data was successfully processed!")
                return
            
            # 设置默认文件名（包含日期时间）
            now = datetime.datetime.now()
            default_filename = f"ProcessedData_{now.strftime('%Y%m%d-%H%M%S')}.csv"
            
            # 优先使用第一个被选中文件所在的目录
            default_dir = ""
            try:
                if hasattr(self, 'selected_files') and self.selected_files:
                    default_dir = os.path.dirname(os.path.abspath(self.selected_files[0]))
                    self.update_status(f"Use the folder of the first selected file as default save location: {default_dir}")
                elif hasattr(self, 'last_save_dir') and self.last_save_dir:
                    default_dir = self.last_save_dir
            except Exception as e:
                self.update_status(f"Error determining default directory: {str(e)}")
                default_dir = ""
            
            try:
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                    initialfile=default_filename,
                    initialdir=default_dir
                )
            except Exception as e:
                self.update_status(f"Error opening save dialog: {str(e)}. Trying without initial directory...")
                # 尝试不带initialdir参数打开对话框
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".csv",
                        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                        initialfile=default_filename
                    )
                except tk.TclError as e2:
                    self.update_status(f"Failed to open save dialog: {str(e2)}")
                    return
            
            # 如果用户取消保存，返回
            if not file_path:
                return
            
            # 保存目录以便下次使用
            self.last_save_dir = os.path.dirname(file_path)
            
            # 保存数据为CSV
            try:
                # 使用utf-8编码并包含BOM，以确保Excel正确识别
                processed_data.to_csv(file_path, index=False, encoding='utf-8-sig')
                
                # 更新状态栏信息
                self.update_status(f"Data saved successfully to {file_path}")
                
                # 自动打开保存目录
                self._open_directory(file_path)
                    
            except Exception as save_error:
                # 处理保存过程中可能出现的错误
                error_details = traceback.format_exc()
                self.update_status(f"Error: Failed to save data: {str(save_error)}")
                # 尝试创建一个简单的错误日志文件
                try:
                    with open("save_error_log.txt", "w", encoding="utf-8") as log_file:
                        log_file.write(f"Error: Failed to save data: {str(save_error)}\n\nDetailed error information:\n{error_details}")
                    self.update_status(f"Error: Failed to save data: {str(save_error)}，Error log saved to save_error_log.txt")
                except (OSError, IOError):
                    pass
                    
        except Exception as e:
            # 捕获其他所有可能的异常
            self.update_status(f"Error in save_processed_data: {str(e)}")
            # 尝试创建错误日志
            try:
                with open("save_processed_data_error.txt", "w", encoding="utf-8") as log_file:
                    log_file.write(f"Error: {str(e)}\n\nTraceback:\n{traceback.format_exc()}")
            except:
                pass
        
    def save_processed_data_to_excel(self, file_path=None, config=None):
        """将Data Re-Processing选项卡中的数据保存为Excel文件，包含color point、criteria、Cpk、Top Defects、Yield Analysis工作表"""
        import os
        import datetime
        import traceback
        import pandas as pd
        from tkinter import filedialog
        from tkinter import messagebox
        
        try:
            # 记录日志
            self.update_status("开始将Data Re-Processing数据保存为Excel格式...")
            
            # 检查是否有重新处理的数据
            if not hasattr(self, 'reprocessed_data') or self.reprocessed_data is None or self.reprocessed_data.empty:
                messagebox.showerror("错误", "没有可保存的重新处理数据！")
                self.update_status("错误：没有重新处理的数据可供保存！")
                return
            
            # 切换到Data Re-Processing选项卡
            if hasattr(self, 'tab_control') and hasattr(self, 'data_reprocessing_tab'):
                self.tab_control.select(self.data_reprocessing_tab)
            
            # 如果未提供文件路径，显示保存对话框
            if file_path is None:
                # 设置默认文件名（包含日期时间）
                now = datetime.datetime.now()
                default_filename = f"ReprocessedData_{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
                
                # 文件存储目录默认为第一个CSV文件所在的目录
                default_dir = ""
                try:
                    # 获取选中的文件列表
                    selected_files = [path for path, var in self.file_vars.items() if var.get()]
                    if selected_files:
                        # 获取第一个选中文件的目录
                        first_file_path = selected_files[0]
                        default_dir = os.path.dirname(first_file_path)
                        self.update_status(f"使用第一个CSV文件的目录作为默认保存位置：{default_dir}")
                    elif hasattr(self, 'last_save_dir') and self.last_save_dir:
                        default_dir = self.last_save_dir
                except Exception as e:
                    self.update_status(f"确定默认目录时出错：{str(e)}")
                    default_dir = ""
                
                # 弹出保存对话框
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                        initialfile=default_filename,
                        initialdir=default_dir
                    )
                except Exception as e:
                    self.update_status(f"打开保存对话框时出错：{str(e)}。尝试不使用初始目录...")
                    # 尝试不带initialdir参数打开对话框
                    try:
                        file_path = filedialog.asksaveasfilename(
                            defaultextension=".xlsx",
                            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                            initialfile=default_filename
                        )
                    except tk.TclError as e2:
                        self.update_status(f"无法打开保存对话框：{str(e2)}")
                        return
                
                # 如果用户取消保存，返回
                if not file_path:
                    return
                
                # 保存目录以便下次使用
                self.last_save_dir = os.path.dirname(file_path)
            
            # 检查是否安装了openpyxl库，如果没有则尝试安装
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                self.update_status("正在安装所需库，请稍候...")
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 保存数据到Excel文件
            self.update_status(f"正在将重新处理的数据保存到 {os.path.basename(file_path)}...")
            
            # 记录数据基本信息用于调试
            data_shape = self.reprocessed_data.shape
            self.update_status(f"数据形状：{data_shape[0]} 行 x {data_shape[1]} 列")
            
            # 检查format_cells字典（如果存在）
            format_cells_info = ""
            if hasattr(self, 'format_cells') and self.format_cells:
                total_format_cells = sum(len(cols) for cols in self.format_cells.values())
                format_cells_info = f"，包含 {total_format_cells} 个需要格式化的单元格"
                self.update_status(f"format_cells字典{format_cells_info}")
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            # 删除默认工作表，稍后会创建新的工作表
            wb.remove(wb.active)
            
            # 创建填充样式（用于所有工作表）
            yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 淡黄色 - 用于FAIL或超限单元格
            light_blue_fill = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")  # 淡蓝色 - 用于Fail行
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色 - 用于PASS单元格
            
            # 首先创建Reprocessed Data工作表，保存重新处理数据（过滤掉包含Criteria的列）
            self.update_status("正在创建'Reprocessed Data'工作表，过滤掉包含Criteria的列...")
            ws_reprocessed = wb.create_sheet(title="Reprocessed Data")
            
            # 过滤掉列标题包含"Criteria"的列
            filtered_columns = [col for col in self.reprocessed_data.columns if "Criteria" not in col]
            
            # 写入Reprocessed Data工作表的表头
            for col_idx, column_name in enumerate(filtered_columns, 1):
                ws_reprocessed.cell(row=1, column=col_idx, value=column_name)
            
            # 写入Reprocessed Data工作表的数据，并应用与Data Reprocessing选项卡完全一致的颜色填充
            for row_idx, (data_index, row_data) in enumerate(self.reprocessed_data.iterrows(), 2):
                # 首先检查该行是否包含Fail值
                is_fail_row = False
                for col_name, value in row_data.items():
                    if 'Pass/Fail' in str(col_name) and str(value).upper() == 'FAIL':
                        is_fail_row = True
                        break
                
                # 写入行数据并应用颜色填充（只处理过滤后的列）
                for col_idx, column_name in enumerate(filtered_columns, 1):
                    value = row_data[column_name]
                    cell = ws_reprocessed.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 确定单元格背景色 - 与Data Reprocessing选项卡完全一致
                    bg_fill = white_fill  # 默认白色
                    
                    # 检查单元格是否需要特殊背景色
                    is_pass_fail_cell = 'Pass/Fail' in str(column_name)
                    is_fail_cell = is_pass_fail_cell and str(value).upper() == 'FAIL'
                    
                    # 使用format_cells字典判断单元格是否需要淡黄色填充
                    # 首先获取原始列索引
                    original_col_index = list(self.reprocessed_data.columns).index(column_name)
                    if hasattr(self, 'format_cells') and data_index in self.format_cells and original_col_index in self.format_cells[data_index]:
                        bg_fill = yellow_fill  # 设置淡黄色背景
                    # 对于Pass/Fail列中的FAIL单元格，也设置淡黄色背景
                    elif is_fail_cell:
                        bg_fill = yellow_fill  # 设置淡黄色背景
                    # Fail行设置淡蓝色背景
                    elif is_fail_row:
                        bg_fill = light_blue_fill
                    
                    # 应用背景色
                    cell.fill = bg_fill
            
            # 自动调整Reprocessed Data工作表的列宽
            for column in ws_reprocessed.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
                ws_reprocessed.column_dimensions[column_letter].width = adjusted_width
            
            self.update_status("'Reprocessed Data'工作表创建完成")
            
            # 定义工作表配置
            sheets_config = [
                {
                    "name": "color point",
                    "desc": "保留前七列及标题中包含'Avg'的列",
                    "apply_color": True  # 应用颜色填充
                },
                {
                    "name": "criteria",
                    "is_criteria": True,
                    "desc": "存储Criteria标签页中的White和Mixed标准数据"
                },
                {
                    "name": "Cpk",
                    "is_cpk": True,
                    "desc": "存储Cpk标签页中的White和Mixed统计数据"
                },
                {
                    "name": "Top Defects",
                    "is_top_defects": True,
                    "desc": "存储Top Defects标签页中的不良项目统计数据"
                },
                {
                    "name": "Yield Analysis",
                    "is_yield_analysis": True,
                    "desc": "存储Yield Analysis标签页中的良率分析数据"
                },
                {
                    "name": "ColorPointSpec",
                    "is_colorpoint_spec": True,
                    "desc": "存储ColorPointSpec标签页中的色点规格数据"
                }
            ]
            
            # 为每个工作表创建数据
            for sheet_config in sheets_config:
                sheet_name = sheet_config["name"]
                sheet_desc = sheet_config["desc"]
                
                # 如果提供了config参数，检查是否应该创建此工作表
                if config and isinstance(config, dict):
                    # 检查config中是否指定了特定的工作表
                    if (sheet_config.get("is_criteria") and not config.get("is_criteria", False)) or \
                       (sheet_config.get("is_cpk") and not config.get("is_cpk", False)) or \
                       (sheet_config.get("is_top_defects") and not config.get("is_top_defects", False)) or \
                       (sheet_config.get("is_yield_analysis") and not config.get("is_yield_analysis", False)) or \
                       (sheet_config.get("is_colorpoint_spec") and not config.get("is_colorpoint_spec", False)) or \
                       (not any(sheet_config.get(key) for key in ["is_criteria", "is_cpk", "is_top_defects", "is_yield_analysis", "is_colorpoint_spec"]) and not config.get("is_color_point", False)):
                        continue
                
                self.update_status(f"正在创建工作表 '{sheet_name}' ({sheet_desc})...")
                
                # 检查是否是color point工作表
                if not sheet_config.get("is_criteria", False) and not sheet_config.get("is_cpk", False) and not sheet_config.get("is_top_defects", False) and not sheet_config.get("is_yield_analysis", False) and not sheet_config.get("is_colorpoint_spec", False):
                    # 处理color point工作表
                    # 筛选符合条件的列
                    color_point_columns = []
                    for col_idx, column_name in enumerate(self.reprocessed_data.columns):
                        # 直接使用列索引和列名进行筛选，不传入lambda函数
                        if col_idx < 7 or "Avg" in column_name:
                            color_point_columns.append(column_name)
                    
                    if color_point_columns:
                        # 创建新工作表
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # 写入数据到工作表
                        color_point_df = self.reprocessed_data[color_point_columns]
                        
                        # 写入表头
                        for col_idx, column_name in enumerate(color_point_columns, 1):
                            cell = ws.cell(row=1, column=col_idx, value=column_name)
                        
                        # 写入数据行
                        for row_idx, (data_index, row_data) in enumerate(color_point_df.iterrows(), 2):
                            # 首先检查该行是否包含Fail值
                            is_fail_row = False
                            for col_name, value in row_data.items():
                                if 'Pass/Fail' in str(col_name) and str(value).upper() == 'FAIL':
                                    is_fail_row = True
                                    break
                            
                            # 写入行数据
                            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                                
                                # 应用颜色填充（仅对color point工作表）
                                if sheet_config.get("apply_color", False):
                                    # 确定单元格背景色
                                    bg_fill = white_fill  # 默认白色
                                    
                                    # 检查单元格是否需要特殊背景色
                                    is_pass_fail_cell = 'Pass/Fail' in str(col_name)
                                    is_fail_cell = is_pass_fail_cell and str(value).upper() == 'FAIL'
                                    
                                    # 使用format_cells字典判断单元格是否需要淡黄色填充
                                    if hasattr(self, 'format_cells') and data_index in self.format_cells and (list(self.reprocessed_data.columns).index(col_name)) in self.format_cells[data_index]:
                                        bg_fill = yellow_fill  # 设置淡黄色背景
                                    # 对于Pass/Fail列中的FAIL单元格，也设置淡黄色背景
                                    elif is_fail_cell:
                                        bg_fill = yellow_fill  # 设置淡黄色背景
                                    # Fail行设置淡蓝色背景
                                    elif is_fail_row:
                                        bg_fill = light_blue_fill
                                    
                                    # 应用背景色
                                    cell.fill = bg_fill
                        
                        # 自动调整列宽
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        self.update_status(f"工作表 '{sheet_name}' 创建完成，包含 {len(color_point_columns)} 列")
                    else:
                        self.update_status(f"没有符合条件的列用于工作表 '{sheet_name}'")
                
                # 检查是否是criteria特殊工作表
                elif sheet_config.get("is_criteria", False):
                    # 处理criteria特殊工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 尝试获取criteria数据
                    criteria_data = None
                    try:
                        # 使用已有的_save_criteria_to_temp_file方法获取criteria数据
                        temp_file_path = self._save_criteria_to_temp_file()
                        if temp_file_path:
                            import csv
                            criteria_data = {}
                            with open(temp_file_path, 'r', encoding='utf-8') as f:
                                reader = csv.DictReader(f)
                                for row in reader:
                                    data_type = row['DataType']
                                    if data_type not in criteria_data:
                                        criteria_data[data_type] = []
                                    # 将数据添加到对应的数据类型列表中
                                    criteria_data[data_type].append((
                                        row['StandardType'],
                                        row['LowerLimit'],
                                        row['UpperLimit'],
                                        row['Description']
                                    ))
                    except Exception as e:
                        self.update_status(f"加载criteria数据时出错：{str(e)}")
                    
                    # 如果获取到criteria数据，写入工作表
                    if criteria_data:
                        # 设置表头，增加上下限列
                        headers = ["DataType", "Standard Type", "Lower Limit", "Upper Limit", "Description"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                        
                        row_idx = 2
                        # 写入White和Mixed数据，按照_display_criteria_data方法的逻辑处理规格值
                        for data_type in ["White", "Mixed"]:
                            if data_type in criteria_data:
                                for std_type, lower_limit, upper_limit, description in criteria_data[data_type]:
                                    # 修正规格值: uniformity规格上限修正为100，规格下限为-1的修正为0
                                    if std_type == "U" and upper_limit == "0":
                                        upper_limit = "100"
                                    if lower_limit == "-1":
                                        lower_limit = "0"
                                    
                                    # 写入数据，尝试将上下限转换为数字格式
                                    ws.cell(row=row_idx, column=1, value=data_type)
                                    ws.cell(row=row_idx, column=2, value=std_type)
                                    
                                    # 尝试将下限转换为数字
                                    if lower_limit:
                                        try:
                                            ws.cell(row=row_idx, column=3, value=float(lower_limit))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=3, value=lower_limit)
                                    else:
                                        ws.cell(row=row_idx, column=3, value="")
                                    
                                    # 尝试将上限转换为数字
                                    if upper_limit:
                                        try:
                                            ws.cell(row=row_idx, column=4, value=float(upper_limit))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=4, value=upper_limit)
                                    else:
                                        ws.cell(row=row_idx, column=4, value="")
                                    
                                    ws.cell(row=row_idx, column=5, value=description)
                                    row_idx += 1
                        
                        self.update_status(f"成功将criteria数据写入工作表 '{sheet_name}'")
                    else:
                        self.update_status(f"没有可用的criteria数据用于工作表 '{sheet_name}'")
                
                # 检查是否是Cpk特殊工作表
                elif sheet_config.get("is_cpk", False):
                    # 处理Cpk特殊工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # 尝试获取Cpk数据
                    cpk_data = {}
                    try:
                        # 分别获取White和Mixed的Cpk数据
                        for data_type in ["White", "Mixed"]:
                            self.update_status(f"正在计算 {data_type} 的Cpk数据...")
                            cpk_data[data_type] = self._calculate_cpk_data(data_type)
                    except Exception as e:
                        self.update_status(f"加载Cpk数据时出错：{str(e)}")
                    
                    # 如果获取到Cpk数据，写入工作表
                    if cpk_data:
                        # 设置表头
                        headers = ["DataType", "Test Item", "Spec", "Mean", "StdDev", "Cpk", "Total", "Fail", "Fail Rate"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                        
                        row_idx = 2
                        # 写入White和Mixed数据
                        for data_type in ["White", "Mixed"]:
                            if data_type in cpk_data and cpk_data[data_type]:
                                for row in cpk_data[data_type]:
                                    # 获取列名作为测试项
                                    column_name = row["column_name"]
                                    
                                    # 确定要显示的规格值
                                    if ('White' in column_name or 'Mixed' in column_name) and ('dL Max' in column_name or 'dL*Max' in column_name):
                                        spec_value = row['upper_limit']
                                    elif ('White' in column_name or 'Mixed' in column_name) and ('L' in column_name or 'U' in column_name):
                                        spec_value = row['lower_limit']
                                    else:
                                        if row['lower_limit'] == 0 and row['upper_limit'] != 0:
                                            spec_value = row['upper_limit']
                                        elif row['upper_limit'] == 0 and row['lower_limit'] != 0:
                                            spec_value = row['lower_limit']
                                        else:
                                            spec_value = f"{row['lower_limit']}-{row['upper_limit']}"
                                    
                                    # 格式化规格值
                                    if isinstance(spec_value, (int, float)):
                                        # 检查是否包含特殊关键字来决定小数位数
                                        special_keywords = ["Ru", "Rv", "Du", "Dv"]
                                        is_special = any(keyword in column_name for keyword in special_keywords)
                                        decimal_places = 4 if is_special else 2
                                        formatted_spec = f"{spec_value:.{decimal_places}f}"
                                    else:
                                        formatted_spec = str(spec_value)
                                    
                                    # 写入数据
                                    ws.cell(row=row_idx, column=1, value=data_type)
                                    ws.cell(row=row_idx, column=2, value=column_name)
                                    
                                    # 确保Spec列为数字格式
                                    if isinstance(spec_value, (int, float)) or (isinstance(formatted_spec, str) and '-' not in formatted_spec):
                                        try:
                                            ws.cell(row=row_idx, column=3, value=float(formatted_spec))
                                        except (ValueError, TypeError):
                                            ws.cell(row=row_idx, column=3, value=formatted_spec)
                                    else:
                                        ws.cell(row=row_idx, column=3, value=formatted_spec)
                                    
                                    # 写入数值数据，尝试转换为数字
                                    try:
                                        ws.cell(row=row_idx, column=4, value=float(row['mean']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=4, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=5, value=float(row['std_dev']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=5, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=6, value=float(row['cpk']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=6, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=7, value=int(row['total_count']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=7, value="")

                                    try:
                                        ws.cell(row=row_idx, column=8, value=int(row['fail_count']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=8, value="")
                                    
                                    try:
                                        ws.cell(row=row_idx, column=9, value=float(row['fail_rate']))
                                    except (ValueError, TypeError):
                                        ws.cell(row=row_idx, column=9, value="")
                                    
                                    row_idx += 1
                        
                        self.update_status(f"成功将Cpk数据写入工作表 '{sheet_name}'")
                    else:
                        self.update_status(f"没有可用的Cpk数据用于工作表 '{sheet_name}'")
                
                # 检查是否是Top Defects特殊工作表
                elif sheet_config.get("is_top_defects", False):
                    # 处理Top Defects特殊工作表
                    ws = wb.create_sheet(title=sheet_name)
                    
                    try:
                        # 生成Top Defects数据
                        # 1. Top Defect Items数据
                        top10_data = []
                        # 计算总数量（总行数）
                        total_count = len(self.reprocessed_data)
                        
                        # 统计从第8列开始的所有列的不良情况
                        column_fail_counts = {}
                        format_cells_exists = hasattr(self, 'format_cells') and self.format_cells
                        
                        for col_idx in range(7, len(self.reprocessed_data.columns)):
                            column_name = self.reprocessed_data.columns[col_idx]
                            
                            # 统计该列中被标记为需要格式化的单元格数量
                            fail_count = 0
                            
                            if format_cells_exists:
                                for row_idx, col_indices in self.format_cells.items():
                                    if col_idx in col_indices:
                                        fail_count += 1
                            else:
                                # 额外的统计逻辑: 直接检查数据中的FAIL值或包含#的单元格
                                for row_idx in range(total_count):
                                    try:
                                        cell_value = str(self.reprocessed_data.iloc[row_idx, col_idx])
                                        if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                            fail_count += 1
                                    except:
                                        continue
                            
                            # 忽略特定的v Avg项目
                            if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                continue
                            
                            # 重命名特定的u Avg项目
                            display_name = column_name
                            if column_name == "White u Avg":
                                display_name = "CAFL0 Color Point"
                            elif column_name == "Mixed u Avg":
                                display_name = "CAFL24 Color Point"
                            
                            if fail_count > 0:
                                fail_rate = (fail_count / total_count) * 100 if total_count > 0 else 0
                                top10_data.append((display_name, fail_count, total_count, fail_rate))
                        
                        # 按不良数量降序排序
                        top10_data.sort(key=lambda x: x[1], reverse=True)
                        
                        # 2. Config Group数据
                        config_data = []
                        if 'Config' in self.reprocessed_data.columns:
                            # 获取所有不同的Config值
                            configs = self.reprocessed_data['Config'].unique()
                            
                            # 遍历每个Config
                            for config in configs:
                                # 获取当前Config的数据
                                config_data_df = self.reprocessed_data[self.reprocessed_data['Config'] == config]
                                config_total = len(config_data_df)
                                
                                # 统计每个不良项目在该Config下的不良数量
                                for col_idx in range(7, len(self.reprocessed_data.columns)):
                                    column_name = self.reprocessed_data.columns[col_idx]
                                    
                                    # 忽略特定的v Avg项目
                                    if column_name == "White v Avg" or column_name == "Mixed v Avg":
                                        continue
                                        
                                    fail_count = 0
                                    
                                    if format_cells_exists:
                                        # 只统计当前Config的数据行
                                        for row_idx, col_indices in self.format_cells.items():
                                            if row_idx < len(self.reprocessed_data) and self.reprocessed_data.iloc[row_idx]['Config'] == config:
                                                if col_idx in col_indices:
                                                    fail_count += 1
                                    else:
                                        # 直接检查当前Config数据中的FAIL值或包含#的单元格
                                        for row_idx in range(len(config_data_df)):
                                            try:
                                                original_row_idx = config_data_df.index[row_idx]
                                                cell_value = str(self.reprocessed_data.iloc[original_row_idx, col_idx])
                                                if cell_value.strip().upper() == 'FAIL' or '#' in cell_value:
                                                    fail_count += 1
                                            except:
                                                continue
                                    
                                    # 处理不良项目名称
                                    display_name = column_name
                                    if column_name == "White u Avg":
                                        display_name = "CAFL0 Color Point"
                                    elif column_name == "Mixed u Avg":
                                        display_name = "CAFL24 Color Point"
                                    
                                    if fail_count > 0:
                                        fail_rate = (fail_count / config_total) * 100 if config_total > 0 else 0
                                        config_data.append((config, display_name, fail_count, config_total, fail_rate))
                        
                        # 写入Top Defect Items数据
                        ws.cell(row=1, column=1, value="Top Defect Items")
                        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                        
                        # 设置表头
                        top10_headers = ["No", "Fail Item", "Fail Count", "Total Count", "Fail Rate"]
                        for c_idx, header in enumerate(top10_headers, 1):
                            ws.cell(row=2, column=c_idx, value=header)
                            ws.cell(row=2, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        
                        # 写入数据
                        row_idx = 3
                        for idx, (item_name, fail_count, total, fail_rate) in enumerate(top10_data, start=1):
                            ws.cell(row=row_idx, column=1, value=str(idx))
                            ws.cell(row=row_idx, column=2, value=item_name)
                            ws.cell(row=row_idx, column=3, value=fail_count)
                            ws.cell(row=row_idx, column=4, value=total)
                            # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                            cell = ws.cell(row=row_idx, column=5, value=fail_rate/100)  # 存储原始数值
                            cell.number_format = '0.00%'  # 应用Excel百分比格式
                            row_idx += 1
                        
                        # 添加空行分隔
                        row_idx += 2
                        
                        # 写入Config Group数据
                        ws.cell(row=row_idx, column=1, value="Config Group Fail Items")
                        ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                        row_idx += 1
                        
                        # 设置表头
                        config_headers = ["No", "Config", "Fail Item", "Fail Count", "Total Count", "Fail Rate"]
                        for c_idx, header in enumerate(config_headers, 1):
                            ws.cell(row=row_idx, column=c_idx, value=header)
                            ws.cell(row=row_idx, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        row_idx += 1
                        
                        # 按Config分组并排序
                        if config_data:
                            # 先按Config分组，再在组内按不良率降序排序
                            config_groups = {}
                            for config, item_name, fail_count, total, fail_rate in config_data:
                                if config not in config_groups:
                                    config_groups[config] = []
                                config_groups[config].append((item_name, fail_count, total, fail_rate))
                            
                            # 为每个Config组内的项目按不良率降序排序
                            for config in config_groups:
                                config_groups[config].sort(key=lambda x: x[3], reverse=True)
                            
                            # 写入数据
                            for config in sorted(config_groups.keys()):
                                config_total = len(self.reprocessed_data[self.reprocessed_data['Config'] == config])
                                # 添加Config分组标题行
                                ws.cell(row=row_idx, column=2, value=f"{config} (Total: {config_total})")
                                ws.cell(row=row_idx, column=2).font = openpyxl.styles.Font(bold=True)
                                row_idx += 1
                                
                                # 添加该Config下的不良项目
                                for idx, (item_name, fail_count, total, fail_rate) in enumerate(config_groups[config], start=1):
                                    ws.cell(row=row_idx, column=1, value=str(idx))
                                    ws.cell(row=row_idx, column=2, value=config)
                                    ws.cell(row=row_idx, column=3, value=item_name)
                                    ws.cell(row=row_idx, column=4, value=fail_count)
                                    ws.cell(row=row_idx, column=5, value=total)
                                    # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                                    cell = ws.cell(row=row_idx, column=6, value=fail_rate/100)  # 存储原始数值
                                    cell.number_format = '0.00%'  # 应用Excel百分比格式
                                    row_idx += 1
                                
                                # 添加一个空行分隔不同Config
                                row_idx += 1
                        else:
                            ws.cell(row=row_idx, column=1, value="No Config data found in the file, cannot perform Config grouping statistics")
                        
                        self.update_status(f"成功将Top Defects数据写入工作表 '{sheet_name}'")
                    except Exception as e:
                        self.update_status(f"写入Top Defects数据时出错：{str(e)}")
                
                # 检查是否是Yield Analysis特殊工作表
                elif sheet_config.get("is_yield_analysis", False):
                    # 处理Yield Analysis特殊工作表 - 使用与选项卡相同的逻辑
                    ws = wb.create_sheet(title=sheet_name)
                    
                    try:
                        # 使用与update_yield_analysis_table方法相同的逻辑
                        # 获取Review Criteria中的阈值数据
                        criteria_dict = self._get_criteria_dict()
                        
                        # 计算总数量和不良数量
                        total_count = len(self.processed_data)
                        fail_count = 0
                        
                        # 创建一个副本用于标记Pass/Fail状态
                        df_copy = self.processed_data.copy()
                        
                        # 根据Review Criteria中的阈值判断不良品
                        if criteria_dict:
                            # 在数据副本中添加基于标准的判断结果列
                            df_copy['Criteria_Pass/Fail'] = 'PASS'
                            
                            # 遍历每条记录，根据阈值判断是否合格
                            for idx, row in df_copy.iterrows():
                                is_pass, failed_criteria, matched_columns = self._evaluate_record_against_criteria(row, criteria_dict)
                                if not is_pass:
                                    fail_count += 1
                                    df_copy.at[idx, 'Criteria_Pass/Fail'] = 'FAIL'
                        else:
                            # 如果没有标准数据，使用原来的Pass/Fail列
                            if 'Pass/Fail' in df_copy.columns:
                                fail_count = (df_copy['Pass/Fail'] == 'FAIL').sum()
                                # 使用原来的Pass/Fail列作为判断依据
                                df_copy['Criteria_Pass/Fail'] = df_copy['Pass/Fail']
                            else:
                                # 如果也没有Pass/Fail列，所有记录都视为通过
                                df_copy['Criteria_Pass/Fail'] = 'PASS'
                        
                        # 计算总体不良率
                        total_fail_rate = 0
                        if total_count > 0:
                            total_fail_rate = (fail_count / total_count) * 100
                        
                        # 按Config分组统计不良率
                        config_yield_data = []
                        if 'Config' in df_copy.columns:
                            # 按Config分组 - 使用不同的变量名避免冲突
                            for config_name, group in df_copy.groupby('Config'):
                                config_total = len(group)
                                # 使用基于标准的判断结果统计不良数
                                config_fail = (group['Criteria_Pass/Fail'] == 'FAIL').sum()
                                config_fail_rate = (config_fail / config_total * 100) if config_total > 0 else 0
                                config_yield_data.append((config_name, config_total, config_fail, config_fail_rate))
                            
                            # 按Config名称升序排序（与选项卡一致）
                            config_yield_data.sort(key=lambda x: str(x[0]))
                        
                        # 设置表头（与选项卡一致，但不包含No列）
                        headers = ["Config", "Total", "Fail Count", "Fail Rate"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                            ws.cell(row=1, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        
                        # 写入Config分组数据
                        row_idx = 2
                        for config_name, total, fail, fail_rate in config_yield_data:
                            ws.cell(row=row_idx, column=1, value=config_name)
                            ws.cell(row=row_idx, column=2, value=total)
                            ws.cell(row=row_idx, column=3, value=fail)
                            # 设置Fail Rate列为百分比格式（数值存储，应用Excel百分比格式）
                            cell = ws.cell(row=row_idx, column=4, value=fail_rate/100)  # 存储原始数值
                            cell.number_format = '0.00%'  # 应用Excel百分比格式
                            row_idx += 1
                        
                        # 添加分隔行
                        row_idx += 1
                        
                        # 写入总体统计
                        ws.cell(row=row_idx, column=1, value="Total")
                        ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
                        ws.cell(row=row_idx, column=2, value=total_count)
                        ws.cell(row=row_idx, column=3, value=fail_count)
                        if total_count > 0:
                            cell = ws.cell(row=row_idx, column=4, value=total_fail_rate/100)  # 存储原始数值
                            cell.number_format = '0.00%'  # 应用Excel百分比格式
                        
                        self.update_status(f"成功将Yield Analysis数据写入工作表 '{sheet_name}'")
                    except Exception as e:
                        self.update_status(f"写入Yield Analysis数据时出错：{str(e)}")
                
                # 检查是否是ColorPointSpec特殊工作表
                elif sheet_config.get("is_colorpoint_spec", False):
                    # 处理ColorPointSpec特殊工作表 - 存储ColorPointSpec选项卡的内容
                    ws = wb.create_sheet(title=sheet_name)
                    
                    try:
                        # 获取ColorPointSpec数据
                        colorpoint_data = getattr(self, 'colorpoint_spec_data', {})
                        
                        # 设置表头
                        headers = ["Type", "Point", "u' Value", "v' Value", "Description"]
                        for c_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=c_idx, value=header)
                            ws.cell(row=1, column=c_idx).font = openpyxl.styles.Font(bold=True)
                        
                        row_idx = 2
                        
                        # 处理White (CAFL0) 数据
                        if "White" in colorpoint_data:
                            white_data = colorpoint_data["White"]
                            # 处理不同数据结构
                            if isinstance(white_data, dict) and "coordinates" in white_data:
                                coordinates = white_data["coordinates"]
                            elif isinstance(white_data, list):
                                coordinates = white_data
                            else:
                                coordinates = []
                            
                            for i, point in enumerate(coordinates):
                                if isinstance(point, (list, tuple)) and len(point) >= 2:
                                    u_val, v_val = point[0], point[1]
                                    if u_val is not None and v_val is not None:
                                        ws.cell(row=row_idx, column=1, value="CAFL0 (White)")
                                        ws.cell(row=row_idx, column=2, value=f"Point {i+1}")
                                        ws.cell(row=row_idx, column=3, value=u_val)
                                        ws.cell(row=row_idx, column=4, value=v_val)
                                        ws.cell(row=row_idx, column=5, value=f"CAFL0 ColorPoint #{i+1} - CIE 1976 u'v' Chromaticity Coordinates")
                                        row_idx += 1
                        
                        # 处理Mixed (CALF24) 数据
                        if "Mixed" in colorpoint_data:
                            mixed_data = colorpoint_data["Mixed"]
                            # 处理不同数据结构
                            if isinstance(mixed_data, dict) and "coordinates" in mixed_data:
                                coordinates = mixed_data["coordinates"]
                            elif isinstance(mixed_data, list):
                                coordinates = mixed_data
                            else:
                                coordinates = []
                            
                            for i, point in enumerate(coordinates):
                                if isinstance(point, (list, tuple)) and len(point) >= 2:
                                    u_val, v_val = point[0], point[1]
                                    if u_val is not None and v_val is not None:
                                        ws.cell(row=row_idx, column=1, value="CALF24 (Mixed)")
                                        ws.cell(row=row_idx, column=2, value=f"Point {i+1}")
                                        ws.cell(row=row_idx, column=3, value=u_val)
                                        ws.cell(row=row_idx, column=4, value=v_val)
                                        ws.cell(row=row_idx, column=5, value=f"CALF24 ColorPoint #{i+1} - CIE 1976 u'v' Chromaticity Coordinates")
                                        row_idx += 1
                        
                        # 如果没有数据，显示提示信息
                        if row_idx == 2:
                            ws.cell(row=2, column=1, value="No ColorPointSpec data available")
                            for c_idx in range(2, 6):
                                ws.cell(row=2, column=c_idx, value="-")
                        
                        self.update_status(f"成功将ColorPointSpec数据写入工作表 '{sheet_name}'")
                    except Exception as e:
                        self.update_status(f"写入ColorPointSpec数据时出错：{str(e)}")
            
            # 保存工作簿
            wb.save(file_path)
            
            # 更新状态栏信息
            self.update_status(f"数据成功保存到 {file_path}，包含 {len(wb.sheetnames)} 个工作表")
            
            # 自动打开保存目录
            self._open_directory(file_path)
                
        except Exception as e:
            # 捕获其他所有可能的异常
            error_msg = f"保存重新处理数据时出错：{str(e)}"
            self.update_status(error_msg)
            messagebox.showerror("保存错误", error_msg)
            # 尝试创建错误日志
            try:
                with open("save_reprocessed_data_error.txt", "w", encoding="utf-8") as log_file:
                    log_file.write(f"错误：{str(e)}\n\n跟踪信息：\n{traceback.format_exc()}")
            except:
                pass
    
    def save_charts(self):
        """保存图表"""
        pass
        
    def export_report(self):
        """导出报告"""
        pass
        
    def show_about(self):
        """显示关于信息"""
        about_window = tk.Toplevel(self.root)
        about_window.title("About TestLogAnalyzer")
        about_window.geometry("400x300")
        about_window.resizable(False, False)
        
        # 居中显示
        about_window.geometry(f"+{int(self.root.winfo_x() + self.root.winfo_width()/2 - 200)}+{int(self.root.winfo_y() + self.root.winfo_height()/2 - 150)}")
        
        # 创建内容框架
        content_frame = tk.Frame(about_window, padx=20, pady=20)
        content_frame.pack(fill="both", expand=True)
        
        # 添加关于信息
        title_label = tk.Label(content_frame, text="TestLogAnalyzer", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        version_label = tk.Label(content_frame, text="V 1.50")
        version_label.pack(pady=(0, 15))
        
        desc_label = tk.Label(content_frame, text="Test Log Analyzer ", wraplength=350)
        desc_label.pack(pady=(0, 15))
        
        features_label = tk.Label(content_frame, text="Function：\n- Merage Multi File - Save as CSV  \n- Data Analysis ", justify="left")
        features_label.pack(pady=(0, 15))
        
        # 添加版权信息
        copyright_label = tk.Label(content_frame, text="© 2025 TestLogAnalyzer Team", font=("Arial", 8))
        copyright_label.pack(side="bottom", pady=10)
        
    def show_help(self):
        """显示使用说明"""
        help_window = tk.Toplevel(self.root)
        help_window.title("Helps")
        help_window.geometry("600x500")
        
        # 居中显示
        help_window.geometry(f"+{int(self.root.winfo_x() + self.root.winfo_width()/2 - 300)}+{int(self.root.winfo_y() + self.root.winfo_height()/2 - 250)}")
        
        # 创建滚动文本框
        text_widget = tk.Text(help_window, wrap="word", padx=10, pady=10, font=("Arial", 10))
        scrollbar = tk.Scrollbar(help_window, command=text_widget.yview)
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # 放置文本框和滚动条
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 添加帮助内容
        help_text = """
TestLogAnalyzer Instructions

1. Load Files
   - Load: Click "Load files" button to import CSV files
   - Clear All: Click "Clear All" button to clear all data
   - Exit: Click "Exit" button or use File menu to exit the program

2. Data Processing 
   - Yield Analysis: Calculate and display yield statistics
   - Top Defects: View top defect types statistics
   - Cpk: Calculate process capability index
   - Criteria: Set analysis criteria
   - Color Point Chart: Generate color point chart

3. Data Analysis
   - Yield Analysis: compute the yield 
   - Top Defects: compute the top yield hitters 
   - Cpk: Cpk for all test items
   - Criteria: list all specfications 
   - Color Point Chart: plot color point chart for CAFL0 and CAFL24

4. Data saving
   - Save as CSV: Save data to CSV format, for offline analysis or reload back with oters CSV
   - Data Analysis and Reporting 
"""
        text_widget.insert("1.0", help_text)
        text_widget.config(state="disabled")  # 设置为只读
        
    def on_closing(self):
        # Close window handler, clean up temp files before exit
        try:
            import os, glob
            # 获取当前目录
            current_dir = os.path.dirname(os.path.abspath(__file__))
            # 查找所有以"criteria_temp_data"开头的CSV文件
            temp_files = glob.glob(os.path.join(current_dir, "criteria_temp_data*.csv"))
            # 删除找到的临时文件
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                    print(f"Temp file deleted: {temp_file}")
                except Exception as e:
                    print(f"Failed to delete temp file: {temp_file}: {e}")
            
            # 查找可能在临时目录中创建的文件
            if hasattr(self, '_criteria_temp_file') and self._criteria_temp_file:
                try:
                    if os.path.exists(self._criteria_temp_file):
                        os.remove(self._criteria_temp_file)
                        print(f"Temp file deleted: {self._criteria_temp_file}")
                        # 记录到日志
                        try:
                            if logger:
                                logger.info(f"Temp file deleted: {self._criteria_temp_file}")
                        except (OSError, AttributeError):
                            pass
                except Exception as e:
                    print(f"Failed to delete temp file: {self._criteria_temp_file}: {e}")
                    # 记录到日志
                    try:
                        if logger:
                            logger.warning(f"Failed to delete temp file: {self._criteria_temp_file}: {e}")
                    except Exception:
                        pass
        except Exception as e:
            print(f"Failed to clean temp files: {e}")
            # 记录到日志
            try:
                if logger:
                    logger.warning(f"Failed to clean temp files: {e}")
            except Exception:
                pass
        finally:
            # 记录程序退出
            try:
                if logger:
                    logger.info("Program exited normally")
            except Exception:
                pass
                
            # 无论清理是否成功，都销毁窗口并退出程序
            self.root.destroy()
            
    def parse_color_criteria(self, criteria_text):
        # Parse color standard text and extract coordinate points
        # 添加日志记录功能，用于追踪多边形各顶点坐标的读取状态
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        logging.info(f"[{timestamp}] [Color Point Chart] 开始读取多边形顶点坐标")
        
        try:
            # 尝试多种可能的数据格式提取u'v'坐标
            import re
            
            # 格式1: "[(u1,v1), (u2,v2), ...]"
            points = re.findall(r'\(([^)]+)\)', criteria_text)
            coordinates = []
            
            for point in points:
                try:
                    # 尝试逗号分隔的格式
                    u, v = point.split(',')
                    coordinates.append((float(u.strip()), float(v.strip())))
                except:
                    # 尝试其他可能的格式，提取所有数字
                    values = re.findall(r'[-+]?\d*\.?\d+', point)
                    if len(values) >= 2:
                        coordinates.append((float(values[0]), float(values[1])))
            
            # 如果没有找到有效点，尝试格式2: "u1,v1,u2,v2,..."
            if not coordinates:
                all_values = re.findall(r'[-+]?\d*\.?\d+', criteria_text)
                if len(all_values) >= 4:  # 至少需要两个点
                    for i in range(0, len(all_values) - 1, 2):
                        coordinates.append((float(all_values[i]), float(all_values[i+1])))
            
            # 过滤掉无效的坐标点（例如极大值、极小值或任意坐标为零的点）
            valid_coordinates = []
            for u, v in coordinates:
                # 排除任意坐标为零的点，并设置合理的u'v'坐标范围
                if u != 0 and v != 0 and 0.0 <= u <= 1.0 and 0.0 <= v <= 1.0:
                    valid_coordinates.append((u, v))
            
            # 记录成功状态和数据总量
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            logging.info(f"[{timestamp}] [Color Point Chart] 成功读取多边形顶点坐标，数据总量: {len(valid_coordinates)} 个顶点")
            return valid_coordinates
        except ValueError as e:
            # 处理坐标数据格式错误
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            logging.error(f"[{timestamp}] [Color Point Chart] 失败读取多边形顶点坐标: 坐标数据格式错误 - {str(e)}")
            return []
        except IndexError as e:
            # 处理数据缺失
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            logging.error(f"[{timestamp}] [Color Point Chart] 失败读取多边形顶点坐标: 数据缺失 - {str(e)}")
            return []
        except Exception as e:
            # 处理其他异常
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            logging.error(f"[{timestamp}] [Color Point Chart] 失败读取多边形顶点坐标: 其他异常 - {str(e)}")
            return []
    
    def show_color_point_chart(self):
        # Display color point chart analysis results
        # 切换到Color Point Chart选项卡
        self.tab_control.select(self.color_point_chart_tab)
        
        # 清空选项卡内容
        for widget in self.color_point_chart_tab.winfo_children():
            widget.destroy()
        
        # 创建图表
        self.create_color_point_chart_content()
        
        # 更新选项卡状态
        self.update_tab_status("Color Point Chart")
        
    def create_color_point_chart_content(self):
        # Create color point chart content
        # 创建标题标签
        label = tk.Label(self.color_point_chart_tab, text="Color Point Chart", font=('Arial', 12, 'bold'))
        label.pack(pady=10)
        
        # 检查是否有处理后的数据
        if self.processed_data is None or self.processed_data.empty:
            no_data_label = tk.Label(self.color_point_chart_tab, text="Click \"Color Point Chart button\"")
            no_data_label.pack(pady=20)
            return
        
        try:
            # 优先从ColorPointSpec选项卡中保存的规格文件获取多边形顶点坐标
            # 首先尝试加载最新的规格文件数据
            self.logger.info("尝试加载最新的ColorPointSpec规格文件以获取多边形顶点坐标")
            self._load_colorpoint_spec_from_temp_file()
            
            white_points = []
            mixed_points = []
            
            # 优先从colorpoint_spec_data获取数据
            if hasattr(self, 'colorpoint_spec_data') and self.colorpoint_spec_data:
                # 获取White多边形顶点
                if 'White' in self.colorpoint_spec_data:
                    white_data = self.colorpoint_spec_data['White']
                    # 验证数据结构有效性
                    if isinstance(white_data, dict) and 'coordinates' in white_data:
                        coordinates = white_data['coordinates']
                        if isinstance(coordinates, list) and len(coordinates) >= 3:
                            # 过滤有效的坐标点
                            white_points = []
                            for point in coordinates:
                                if isinstance(point, (list, tuple)) and len(point) == 2:
                                    try:
                                        u, v = float(point[0]), float(point[1])
                                        # 排除无效坐标
                                        if u != 0 and v != 0 and 0.0 <= u <= 1.0 and 0.0 <= v <= 1.0:
                                            white_points.append((u, v))
                                    except (ValueError, TypeError):
                                        continue
                            self.logger.info(f"从ColorPointSpec获取了{len(white_points)}个White多边形顶点")
                
                # 获取Mixed多边形顶点
                if 'Mixed' in self.colorpoint_spec_data:
                    mixed_data = self.colorpoint_spec_data['Mixed']
                    # 验证数据结构有效性
                    if isinstance(mixed_data, dict) and 'coordinates' in mixed_data:
                        coordinates = mixed_data['coordinates']
                        if isinstance(coordinates, list) and len(coordinates) >= 3:
                            # 过滤有效的坐标点
                            mixed_points = []
                            for point in coordinates:
                                if isinstance(point, (list, tuple)) and len(point) == 2:
                                    try:
                                        u, v = float(point[0]), float(point[1])
                                        # 排除无效坐标
                                        if u != 0 and v != 0 and 0.0 <= u <= 1.0 and 0.0 <= v <= 1.0:
                                            mixed_points.append((u, v))
                                    except (ValueError, TypeError):
                                        continue
                            self.logger.info(f"从ColorPointSpec获取了{len(mixed_points)}个Mixed多边形顶点")
            
            # 如果从ColorPointSpec没有获取到有效数据，回退到从表格列解析
            if not white_points and "White Pass/Fail Criteria" in self.processed_data.columns:
                # 获取所有非空的White标准值并尝试解析
                for idx, white_criteria in self.processed_data["White Pass/Fail Criteria"].items():
                    if pd.notna(white_criteria):
                        points = self.parse_color_criteria(str(white_criteria))
                        if points:  # 如果解析到有效点，则使用这个标准
                            white_points = points
                            break
            
            # 如果从ColorPointSpec没有获取到有效数据，回退到从表格列解析
            if not mixed_points and "Mixed Pass/Fail Criteria" in self.processed_data.columns:
                # 获取所有非空的Mixed标准值并尝试解析
                for idx, mixed_criteria in self.processed_data["Mixed Pass/Fail Criteria"].items():
                    if pd.notna(mixed_criteria):
                        points = self.parse_color_criteria(str(mixed_criteria))
                        if points:  # 如果解析到有效点，则使用这个标准
                            mixed_points = points
                            break
                            
            # 提取White u Avg 和 White v Avg 数据点，并保存原始行数据用于分组
            white_avg_points = []
            white_row_data = []  # 保存原始行数据用于分组
            if "White u Avg" in self.processed_data.columns and "White v Avg" in self.processed_data.columns:
                mask = (self.processed_data['White u Avg'].notna() & 
                        self.processed_data['White v Avg'].notna() & 
                        (self.processed_data['White u Avg'] != 0) & 
                        (self.processed_data['White v Avg'] != 0))
                filtered_data = self.processed_data[mask]
                white_avg_points = list(zip(filtered_data['White u Avg'].astype(float), 
                                           filtered_data['White v Avg'].astype(float)))
                white_row_data = [row for _, row in filtered_data.iterrows()]
            
            # 提取Mixed u Avg 和 Mixed v Avg 数据点，并保存原始行数据用于分组
            mixed_avg_points = []
            mixed_row_data = []  # 保存原始行数据用于分组
            if "Mixed u Avg" in self.processed_data.columns and "Mixed v Avg" in self.processed_data.columns:
                mask = (self.processed_data['Mixed u Avg'].notna() & 
                        self.processed_data['Mixed v Avg'].notna() & 
                        (self.processed_data['Mixed u Avg'] != 0) & 
                        (self.processed_data['Mixed v Avg'] != 0))
                filtered_data = self.processed_data[mask]
                mixed_avg_points = list(zip(filtered_data['Mixed u Avg'].astype(float), 
                                           filtered_data['Mixed v Avg'].astype(float)))
                mixed_row_data = [row for _, row in filtered_data.iterrows()]
            
            # 创建包含两个子图的图形布局（水平排列）
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
            
            # 定义函数：将u'v'值转换为RGB颜色（使用100nit亮度）
            def uv_to_rgb(u, v, luminance=100):
                # CIELUV到XYZ转换
                # 使用D65标准光源
                u0 = 0.19783000664283681  # D65的u'值
                v0 = 0.46831999493879114  # D65的v'值
                
                # 计算Y（亮度）
                Y = luminance / 100.0  # 转换为0-1范围
                
                # 计算u和v的实际值
                if v != 0:
                    d = Y * (5 * v - 12 * v0 * u) / (v * 3)
                else:
                    d = 0
                
                u_prime = u
                v_prime = v
                
                # 计算XYZ坐标
                if v_prime > 0:
                    Y = Y
                    X = Y * 9 * u_prime / (4 * v_prime)
                    Z = Y * (12 - 3 * u_prime - 20 * v_prime) / (4 * v_prime)
                else:
                    X, Y, Z = 0, Y, 0
                
                # XYZ到RGB转换（使用sRGB矩阵）
                r = X * 3.2406 + Y * -1.5372 + Z * -0.4986
                g = X * -0.9689 + Y * 1.8758 + Z * 0.0415
                b = X * 0.0557 + Y * -0.2040 + Z * 1.0570
                
                # 应用gamma校正
                def gamma_correction(x):
                    if x <= 0.0031308:
                        return 12.92 * x
                    else:
                        return 1.055 * (x ** (1/2.4)) - 0.055
                
                r = gamma_correction(r)
                g = gamma_correction(g)
                b = gamma_correction(b)
                
                # 确保RGB值在0-1范围内
                r = max(0, min(1, r))
                g = max(0, min(1, g))
                b = max(0, min(1, b))
                
                return r, g, b
            
            # 创建颜色背景的通用函数
            def add_color_background(ax, xlim, ylim, resolution=100):
                import numpy as np
                
                # 创建网格
                u = np.linspace(xlim[0], xlim[1], resolution)
                v = np.linspace(ylim[0], ylim[1], resolution)
                u_grid, v_grid = np.meshgrid(u, v)
                
                # 创建RGB颜色数组
                rgb_array = np.zeros((resolution, resolution, 3))
                
                # 为每个点计算RGB颜色
                rgb_flat = np.array([uv_to_rgb(u, v) for u, v in zip(u_grid.flatten(), v_grid.flatten())])
                rgb_array = rgb_flat.reshape(resolution, resolution, 3)
                
                # 显示颜色背景
                ax.imshow(rgb_array, extent=[xlim[0], xlim[1], ylim[0], ylim[1]], 
                         origin='lower', aspect='auto', alpha=0.8)
            
            # ------------------- 第一个子图：White Range -------------------
            # 设置坐标轴标签和标题
            ax1.set_xlabel("u'")
            ax1.set_ylabel("v'")
            ax1.set_title("CAFL0 (u'v' Coordinates)")
            
            # 如果有有效的White坐标点，绘制多边形
            if len(white_points) >= 3:
                # 调整坐标轴范围以适应多边形点
                all_u_values = [p[0] for p in white_points]
                all_v_values = [p[1] for p in white_points]
                
                u_margin = (max(all_u_values) - min(all_u_values)) * 0.15 if max(all_u_values) != min(all_u_values) else 0.02
                v_margin = (max(all_v_values) - min(all_v_values)) * 0.15 if max(all_v_values) != min(all_v_values) else 0.02
                xlim = [min(all_u_values) - u_margin, max(all_u_values) + u_margin]
                ylim = [min(all_v_values) - v_margin, max(all_v_values) + v_margin]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax1, xlim, ylim)
                
                # 设置坐标轴范围
                ax1.set_xlim(xlim)
                ax1.set_ylim(ylim)
                
                from matplotlib.patches import Polygon
                # 取消多边形填充，只保留边框
                white_poly = Polygon(white_points, closed=True, fill=False, 
                                  edgecolor='blue', label='CAFL0')
                ax1.add_patch(white_poly)
                
                # 添加图例
                ax1.legend(loc='upper right')
                
                # CAFL0 Avg Points数据点将通过plot_data_points函数绘制
            else:
                # 设置默认范围并显示无数据提示
                xlim = [0.18, 0.22]
                ylim = [0.48, 0.52]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax1, xlim, ylim)
                
                # 设置坐标轴范围
                ax1.set_xlim(xlim)
                ax1.set_ylim(ylim)
                
                ax1.text(0.5, 0.5, 'Did not find valid CAFL0 range data', 
                        horizontalalignment='center', verticalalignment='center', 
                        transform=ax1.transAxes, color='gray', fontsize=12)
            
            # 添加网格线
            ax1.grid(True, linestyle='--', alpha=0.7)
            
            # ------------------- 第二个子图：Mixed Range -------------------
            # 设置坐标轴标签和标题
            ax2.set_xlabel("u'")
            ax2.set_ylabel("v'")
            ax2.set_title("CALF24 (u'v' Coordinates)")
            
            # 如果有有效的Mixed坐标点，绘制多边形
            if len(mixed_points) >= 3:
                # 调整坐标轴范围以适应多边形点
                all_u_values = [p[0] for p in mixed_points]
                all_v_values = [p[1] for p in mixed_points]
                
                u_margin = (max(all_u_values) - min(all_u_values)) * 0.15 if max(all_u_values) != min(all_u_values) else 0.02
                v_margin = (max(all_v_values) - min(all_v_values)) * 0.15 if max(all_v_values) != min(all_v_values) else 0.02
                xlim = [min(all_u_values) - u_margin, max(all_u_values) + u_margin]
                ylim = [min(all_v_values) - v_margin, max(all_v_values) + v_margin]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax2, xlim, ylim)
                
                # 设置坐标轴范围
                ax2.set_xlim(xlim)
                ax2.set_ylim(ylim)
                
                from matplotlib.patches import Polygon
                # 取消多边形填充，只保留边框
                mixed_poly = Polygon(mixed_points, closed=True, fill=False, 
                                  edgecolor='red', label='CALF24')
                ax2.add_patch(mixed_poly)
                
                # 添加图例
                ax2.legend(loc='upper right')
                
                # CALF24 Avg Points数据点将通过plot_data_points函数绘制
            else:
                # 设置默认范围并显示无数据提示
                xlim = [0.18, 0.22]
                ylim = [0.48, 0.52]
                
                # 添加基于u'v'的颜色背景
                add_color_background(ax2, xlim, ylim)
                
                # 设置坐标轴范围
                ax2.set_xlim(xlim)
                ax2.set_ylim(ylim)
                
                ax2.text(0.5, 0.5, 'Did not find valid CALF24 range data', 
                        horizontalalignment='center', verticalalignment='center', 
                        transform=ax2.transAxes, color='gray', fontsize=12)
            
            # 添加网格线
            ax2.grid(True, linestyle='--', alpha=0.7)
            
            # 调整子图布局
            plt.tight_layout(pad=3.0)
            
            # 创建画布并添加到Tkinter窗口
            canvas = FigureCanvasTkAgg(fig, master=self.color_point_chart_tab)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # 用于存储分组Z轴顺序的全局字典
            group_z_order_dict = {}
            
            # 定义绘制数据点的函数，支持分组和zorder控制
            def plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, group_column=None):
                from matplotlib.colors import LinearSegmentedColormap
                
                # 保存当前使用的分组列到轴对象，用于右键菜单
                ax.current_group_column = group_column
                
                # 清除现有的散点图
                for artist in ax.collections[:]:
                    # 检查是否为散点图类型
                    if hasattr(artist, '_sizes'):  # 散点图特有的属性
                        artist.remove()
                
                # 更新图例
                ax.legend().set_visible(False)
                
                if not group_column or group_column not in row_data[0].index:
                    # 不分组，使用单一颜色绘制所有点
                    u_values = [p[0] for p in points]
                    v_values = [p[1] for p in points]
                    scatter = ax.scatter(u_values, v_values, color=base_color, 
                                        s=4, alpha=0.40, label=f'{title_prefix} Avg Points', zorder=10)
                    ax.legend(loc='upper right')
                else:
                    # 按指定列分组绘制
                    # 获取唯一分组值
                    groups = sorted(list(set(row[group_column] for row in row_data)))
                    # 生成不同颜色
                    if len(groups) <= 10:
                        # 使用matplotlib默认的分类颜色循环
                        colors = plt.cm.tab10.colors[:len(groups)]
                    else:
                        # 使用连续颜色映射
                        cmap = LinearSegmentedColormap.from_list('group_colors', [base_color, 'dark' + base_color], N=len(groups))
                        colors = [cmap(i/len(groups)) for i in range(len(groups))]
                    
                    # 确保当前分组的zorder字典存在
                    if group_column not in group_z_order_dict:
                        # 默认按照排序顺序设置zorder
                        group_z_order_dict[group_column] = {}
                        for i, group_val in enumerate(groups):
                            group_z_order_dict[group_column][group_val] = i + 10
                    
                    # 按组绘制数据点
                    for i, group_val in enumerate(groups):
                        group_points = [(p[0], p[1]) for j, p in enumerate(points) 
                                      if row_data[j][group_column] == group_val]
                        if group_points:
                            u_values = [p[0] for p in group_points]
                            v_values = [p[1] for p in group_points]
                            # 使用存储的zorder值
                            zorder_val = group_z_order_dict[group_column].get(group_val, i + 10)
                            scatter = ax.scatter(u_values, v_values, color=colors[i], 
                                                s=4, alpha=0.40, label=f'{str(group_val)}', zorder=zorder_val)
                    
                    # 添加分组图例
                    handles, labels = ax.get_legend_handles_labels()
                    by_label = dict(zip(labels, handles))
                    ax.legend(by_label.values(), by_label.keys(), loc='upper right', 
                             title=f'Group by {group_column}')
                
                canvas.draw()
            
            # 初始绘制数据点
            if white_avg_points:
                plot_data_points(ax1, white_avg_points, white_row_data, 'green', 'CAFL0', canvas)
            if mixed_avg_points:
                plot_data_points(ax2, mixed_avg_points, mixed_row_data, 'purple', 'CALF24', canvas)
            
            # 创建右键菜单
            # 为CAFL0和CALF24子图创建单独的右键菜单处理
            # 用于跟踪顶点标签是否显示的字典
            vertex_labels_visible = {ax1: False, ax2: False}  # 初始状态为不显示
            vertex_text_objects = {ax1: [], ax2: []}  # 存储顶点文本对象
            
            # 显示或隐藏顶点坐标的函数
            def toggle_vertex_labels(ax, poly_points):
                # 清除现有标签
                for text_obj in vertex_text_objects[ax]:
                    if text_obj in ax.texts:
                        text_obj.remove()
                vertex_text_objects[ax] = []
                
                # 切换状态
                if vertex_labels_visible[ax]:
                    vertex_labels_visible[ax] = False  # 隐藏标签
                else:
                    # 显示标签
                    vertex_labels_visible[ax] = True
                    for i, (u, v) in enumerate(poly_points):
                        # 添加顶点坐标标签，使用格式化显示保留6位小数
                        text = ax.text(u, v, f'P{i}: ({u:.6f}, {v:.6f})', 
                                     fontsize=8, ha='right', va='bottom',
                                     bbox=dict(boxstyle="round,pad=0.3", fc="white", alpha=0.7))
                        vertex_text_objects[ax].append(text)
                
                # 更新画布
                canvas.draw()
                
            def handle_right_click(ax, points, row_data, base_color, title_prefix, event):
                # 创建主菜单
                main_menu = tk.Menu(self.color_point_chart_tab, tearoff=0)
                
                # 添加取消分组选项
                def reset_action():
                    plot_data_points(ax, points, row_data, base_color, title_prefix, canvas)
                main_menu.add_command(label="Ungroup", command=reset_action)
                
                # 添加显示/隐藏顶点坐标选项
                def toggle_vertices_action():
                    # 获取对应的多边形点
                    if ax == ax1 and len(white_points) >= 3:
                        toggle_vertex_labels(ax, white_points)
                    elif ax == ax2 and len(mixed_points) >= 3:
                        toggle_vertex_labels(ax, mixed_points)
                
                # 根据当前状态设置菜单项文本
                label_text = "Hide Vertex Coordinates" if vertex_labels_visible.get(ax, False) else "Show Vertex Coordinates"
                main_menu.add_command(label=label_text, command=toggle_vertices_action)
                
                # 添加分隔线
                main_menu.add_separator()
                
                # 添加分组子菜单
                if row_data:
                    group_menu = tk.Menu(main_menu, tearoff=0)
                    # 获取所有文本类型的数据列
                    columns = []
                    if len(row_data) > 0 and not row_data[0].empty:
                        for col in row_data[0].index:
                            # 判断列是否包含文本数据（非数值型）或列名为"Position ID"
                            try:
                                # 尝试检查第一行的数据类型
                                sample_value = row_data[0][col]
                                # 条件：1. 字符串类型或NA值 或 2. 列名为"Position ID"
                                if isinstance(sample_value, str) or pd.isna(sample_value) or col == "Position ID":
                                    columns.append(col)
                            except:
                                # 如果出现异常，跳过该列
                                continue
                    
                    # 为每个列创建分组命令
                    for col in columns:
                        # 使用闭包和默认参数正确捕获列名
                        def create_group_command(column_name):
                            def execute_group():
                                plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, column_name)
                            return execute_group
                        
                        group_menu.add_command(label=col, command=create_group_command(col))
                    
                    main_menu.add_cascade(label="Group by", menu=group_menu)
                    
                    # 检查是否已分组，如果是，添加调整层级的子菜单
                    if hasattr(ax, 'current_group_column') and ax.current_group_column:
                        current_group = ax.current_group_column
                        # 获取当前分组的所有值
                        current_groups = sorted(list(set(row[current_group] for row in row_data)))
                        
                        if len(current_groups) > 1:  # 只有多于1个分组时才添加层级调整
                            level_menu = tk.Menu(main_menu, tearoff=0)
                            main_menu.add_separator()
                            main_menu.add_cascade(label="调整分组显示层级", menu=level_menu)
                            
                            # 为每个分组创建提升和降低层级的命令
                            for group_val in current_groups:
                                group_submenu = tk.Menu(level_menu, tearoff=0)
                                level_menu.add_cascade(label=str(group_val), menu=group_submenu)
                                
                                # 提升层级
                                def bring_forward(group):
                                    def action():
                                        if current_group in group_z_order_dict and group in group_z_order_dict[current_group]:
                                            current_z = group_z_order_dict[current_group][group]
                                            # 找到下一个更高的zorder值
                                            all_z = group_z_order_dict[current_group].values()
                                            next_z = max(all_z) + 1 if all_z else 11
                                            group_z_order_dict[current_group][group] = next_z
                                            # 重新绘制图表
                                            plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, current_group)
                                    return action
                                
                                # 降低层级
                                def send_backward(group):
                                    def action():
                                        if current_group in group_z_order_dict and group in group_z_order_dict[current_group]:
                                            current_z = group_z_order_dict[current_group][group]
                                            # 找到下一个更低的zorder值
                                            all_z = group_z_order_dict[current_group].values()
                                            next_z = min(all_z) - 1 if all_z else 9
                                            group_z_order_dict[current_group][group] = next_z
                                            # 重新绘制图表
                                            plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, current_group)
                                    return action
                                
                                # 置于顶层
                                def bring_to_top(group):
                                    def action():
                                        if current_group in group_z_order_dict and group in group_z_order_dict[current_group]:
                                            # 找到当前最大zorder值并+1
                                            all_z = group_z_order_dict[current_group].values()
                                            top_z = max(all_z) + 1 if all_z else 100
                                            group_z_order_dict[current_group][group] = top_z
                                            # 重新绘制图表
                                            plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, current_group)
                                    return action
                                
                                # 置于底层
                                def send_to_bottom(group):
                                    def action():
                                        if current_group in group_z_order_dict and group in group_z_order_dict[current_group]:
                                            # 找到当前最小zorder值并-1
                                            all_z = group_z_order_dict[current_group].values()
                                            bottom_z = min(all_z) - 1 if all_z else 1
                                            group_z_order_dict[current_group][group] = bottom_z
                                            # 重新绘制图表
                                            plot_data_points(ax, points, row_data, base_color, title_prefix, canvas, current_group)
                                    return action
                                
                                group_submenu.add_command(label="提升一层", command=bring_forward(group_val))
                                group_submenu.add_command(label="降低一层", command=send_backward(group_val))
                                group_submenu.add_separator()
                                group_submenu.add_command(label="置于顶层", command=bring_to_top(group_val))
                                group_submenu.add_command(label="置于底层", command=send_to_bottom(group_val))
                
                # 显示菜单
                main_menu.post(event.x_root, event.y_root)
                
                # 设置点击其他地方关闭菜单
                def close_menu(event):
                    main_menu.unpost()
                    # 解绑临时事件
                    for widget in [self.color_point_chart_tab, canvas_widget]:
                        try:
                            widget.unbind('<Button-1>', close_id)
                            widget.unbind('<Button-3>', close_id)
                        except tk.TclError:
                            pass
                
                close_id = self.color_point_chart_tab.bind('<Button-1>', close_menu)
                
            # 创建Tkinter级别的右键事件处理
            def on_canvas_right_click(event):
                # 判断点击位置对应的子图
                x, y = canvas_widget.winfo_width(), canvas_widget.winfo_height()
                
                # 简单的位置判断
                if event.x < x // 2:  # 左侧是CAFL0
                    if white_avg_points and white_row_data:
                        handle_right_click(ax1, white_avg_points, white_row_data, 'green', 'CAFL0', event)
                else:  # 右侧是CALF24
                    if mixed_avg_points and mixed_row_data:
                        handle_right_click(ax2, mixed_avg_points, mixed_row_data, 'purple', 'CALF24', event)
            
            # 获取画布的Tkinter组件并绑定右键事件
            canvas_widget = canvas.get_tk_widget()
            canvas_widget.bind('<Button-3>', on_canvas_right_click)
            
            # 为了确保兼容性，同时绑定<ButtonRelease-3>
            canvas_widget.bind('<ButtonRelease-3>', lambda e: None)  # 只是为了确保右键事件被正确捕获
            
            # 添加信息标签
            info_text = "Color Point Chart - based on Data Processing tab data\n\n"
            
            # 更新的信息标签，包含颜色背景说明和数据点显示
            info_text += "Chart Description: Display color coordinate regions for CAFL0 and CALF24\n"
            info_text += "- Background color is based on u'v' values, default brightness is 100nit\n"
            info_text += "- CAFL0 region is displayed with a blue border polygon\n"
            info_text += "- CALF24 region is displayed with a red border polygon\n"
            info_text += "- Green Points display CAFL0 Avg data points\n"
            info_text += "- Purple Points display CALF24 Avg data points\n"
            info_text += "- Based on data from the Data Processing tab\n\n"
            info_text += "Right-click Menu Functions:\n"
            info_text += "- Ungroup: Restore default display mode\n"
            info_text += "- Show/Hide Vertex Coordinates: Display or hide exact coordinates of polygon vertices\n"
            info_text += "- Group by: Group data points by all text columns in the Data Processing tab\n"
            info_text += "- 调整分组显示层级: 当数据点分组显示时，可调整不同分组之间的重叠显示顺序\n"
            
            
            info_label = tk.Label(self.color_point_chart_tab, text=info_text, justify=tk.LEFT, font=('Arial', 10))
            info_label.pack(padx=10, pady=5, anchor='w')
            
        except Exception as e:
            error_label = tk.Label(self.color_point_chart_tab, text=f"Failed to create Color Point Chart: {str(e)}", fg="red")
            error_label.pack(pady=20)
            logging.error(f"Failed to create Color Point Chart: {str(e)}")
        finally:
            # 确保释放matplotlib资源
            plt.close('all')
        
if __name__ == "__main__":
    root = tk.Tk()
    app = TestLogAnalyzer(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()